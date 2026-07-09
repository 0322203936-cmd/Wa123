"""
Walmex Dashboard — CFBC
Reporte ejecutivo estilo WalmartMX

"""
import hashlib
import json
import base64
import openpyxl
import pickle
import tempfile
import time
import os
import requests
import msal
from collections import defaultdict
from pathlib import Path
import streamlit as st
import streamlit.components.v1 as components
from supabase import create_client, Client

try:
    if "supabase" in st.secrets:
        SUPABASE_URL = st.secrets["supabase"]["url"]
        SUPABASE_KEY = st.secrets["supabase"]["service_role_key"]
    else:
        SUPABASE_URL = st.secrets["SUPABASE_URL"]
        SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
except KeyError:
    import os
    SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
    SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")

@st.cache_resource
def init_supabase() -> Client:
    try:
        return create_client(SUPABASE_URL, SUPABASE_KEY)
    except Exception as e:
        print("Error initializing Supabase:", e)
        return None

def get_supabase_data():
    client = init_supabase()
    if not client: return []
    try:
        resp = client.table('facturas_folios').select('*').execute()
        return resp.data
    except Exception as e:
        print("Error fetching Supabase data:", e)
        return []

def get_devoluciones_data():
    client = init_supabase()
    if not client: return []
    try:
        resp = client.table('devoluciones').select('*').order('created_at', desc=True).execute()
        return resp.data
    except Exception as e:
        print("Error fetching devoluciones data:", e)
        return []


def _get_sharepoint_cfg() -> dict:
    """Lee credenciales de st.secrets (local) o variables de entorno (Render)."""
    try:
        return dict(st.secrets["sharepoint"])
    except Exception:
        return {
            "tenant_id":     os.environ["TENANT_ID"],
            "client_id":     os.environ["CLIENT_ID"],
            "client_secret": os.environ["CLIENT_SECRET"],
            "site_url":      os.environ["SITE_URL"],
            "file_path":     os.environ["FILE_PATH"],
        }


# Caché en disco
_CACHE_DIR = Path(__file__).resolve().parent / ".wa123_cache"


@st.cache_data(ttl=300, show_spinner=False)
def _sharepoint_access_token(tenant_id: str, client_id: str, client_secret: str) -> str:
    app_msal = msal.ConfidentialClientApplication(
        client_id,
        authority=f"https://login.microsoftonline.com/{tenant_id}",
        client_credential=client_secret,
    )
    token = app_msal.acquire_token_for_client(
        scopes=["https://graph.microsoft.com/.default"]
    )
    if "access_token" not in token:
        raise RuntimeError(
            f"Error de autenticación SharePoint: {token.get('error_description', token)}"
        )
    return token["access_token"]


@st.cache_data(ttl=3600, show_spinner=False)
def _sharepoint_site_id(site_url: str, tenant_id: str, client_id: str, client_secret: str) -> str:
    parts     = site_url.rstrip("/").split("/")
    hostname  = parts[2]
    site_path = "/".join(parts[3:])
    headers = {
        "Authorization": f"Bearer {_sharepoint_access_token(tenant_id, client_id, client_secret)}"
    }
    site_resp = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{hostname}:/{site_path}",
        headers=headers,
        timeout=30,
    )
    site_resp.raise_for_status()
    return site_resp.json()["id"]


@st.cache_data(ttl=5, show_spinner=False)
def _sharepoint_file_meta(site_url: str, file_path: str, tenant_id: str, client_id: str, client_secret: str) -> dict:
    token = _sharepoint_access_token(tenant_id, client_id, client_secret)
    site_id = _sharepoint_site_id(site_url, tenant_id, client_id, client_secret)
    headers = {"Authorization": f"Bearer {token}"}
    file_path_encoded = file_path.lstrip("/")
    item_resp = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{file_path_encoded}",
        headers=headers,
        timeout=30,
    )
    item_resp.raise_for_status()
    item = item_resp.json()
    version_seed = "|".join([
        str(item.get("eTag") or item.get("cTag") or ""),
        str(item.get("lastModifiedDateTime") or ""),
        str(item.get("size") or 0),
    ])
    cache_key = hashlib.md5(version_seed.encode("utf-8")).hexdigest()[:16]
    return {
        "cache_key": cache_key,
        "etag": item.get("eTag") or item.get("cTag") or "",
        "last_modified": item.get("lastModifiedDateTime") or "",
        "size": int(item.get("size") or 0),
        "name": item.get("name") or "",
    }


def _obtener_meta_excel_actual() -> dict:
    cfg = _get_sharepoint_cfg()
    return _sharepoint_file_meta(
        cfg["site_url"],
        cfg["file_path"],
        cfg["tenant_id"],
        cfg["client_id"],
        cfg["client_secret"],
    )

def _descargar_excel_sharepoint() -> tuple:
    """
    Descarga el Excel desde SharePoint usando Azure App (client credentials).
    Credenciales en st.secrets["sharepoint"]:
        tenant_id, client_id, client_secret, site_url, file_path
    Retorna (path_local: Path, cache_key: str)
    """
    cfg = _get_sharepoint_cfg()
    token = _sharepoint_access_token(cfg["tenant_id"], cfg["client_id"], cfg["client_secret"])
    site_id = _sharepoint_site_id(cfg["site_url"], cfg["tenant_id"], cfg["client_id"], cfg["client_secret"])
    headers = {"Authorization": f"Bearer {token}"}

    # ── Descargar archivo ──
    file_url = (
        f"https://graph.microsoft.com/v1.0/sites/{site_id}"
        f"/drive/root:{cfg['file_path']}:/content"
    )
    file_resp = requests.get(file_url, headers=headers, timeout=60)
    file_resp.raise_for_status()

    # ── Guardar en archivo temporal ──
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.write(file_resp.content)
    tmp.close()

    cache_key = hashlib.md5(file_resp.content).hexdigest()[:16]
    return Path(tmp.name), cache_key


def _subir_excel_sharepoint(archivo_local_bytes: bytes) -> str:
    """
    Escribe directamente en el Excel de SharePoint usando la API de Excel de
    Microsoft Graph — funciona aunque el archivo esté abierto, y actualiza
    el rango completo de un solo request (sin descargar ni subir el archivo).
    """
    import io, math
    import pandas as pd
    cfg = _get_sharepoint_cfg()

    # ── Autenticación MSAL ──
    app_msal = msal.ConfidentialClientApplication(
        cfg["client_id"],
        authority=f"https://login.microsoftonline.com/{cfg['tenant_id']}",
        client_credential=cfg["client_secret"],
    )
    token = app_msal.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in token:
        raise RuntimeError(f"Error auth: {token.get('error_description', token)}")
    hdrs = {"Authorization": f"Bearer {token['access_token']}", "Content-Type": "application/json"}

    # ── Site ID ──
    parts = cfg["site_url"].rstrip("/").split("/")
    r = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{parts[2]}:/{chr(47).join(parts[3:])}",
        headers=hdrs, timeout=30)
    r.raise_for_status()
    site_id = r.json()["id"]

    # ── Item ID del archivo ──
    file_path_encoded = cfg["file_path"].lstrip("/")
    item_resp = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{file_path_encoded}",
        headers=hdrs, timeout=30)
    item_resp.raise_for_status()
    item_id = item_resp.json()["id"]

    base = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/items/{item_id}/workbook"

    # ── Crear sesión de Excel (persistente para múltiples writes) ──
    sess_resp = requests.post(
        f"{base}/createSession",
        headers=hdrs,
        json={"persistChanges": True},
        timeout=30)
    sess_resp.raise_for_status()
    session_id = sess_resp.json()["id"]
    hdrs_sess = {**hdrs, "workbook-session-id": session_id}

    try:
        # ── Leer Excel local desde fila 26, cols A(1)→AS(45) ──
        COL_INI, COL_FIN, FILA_LOCAL = 1, 45, 26
        filas = []
        try:
            wb_local = openpyxl.load_workbook(io.BytesIO(archivo_local_bytes), data_only=True)
            ws_local = wb_local.active
            for row in ws_local.iter_rows(min_row=FILA_LOCAL, min_col=COL_INI, max_col=COL_FIN, values_only=True):
                if all(v is None for v in row):
                    break
                filas.append(list(row))
        except Exception:
            for engine in ("openpyxl", "xlrd"):
                try:
                    df = pd.read_excel(io.BytesIO(archivo_local_bytes), engine=engine, header=None)
                    df = df.iloc[FILA_LOCAL - 1:, COL_INI - 1:COL_FIN].dropna(how="all")
                    filas = [list(r) for r in df.values.tolist()]
                    break
                except Exception:
                    continue

        if not filas:
            return "⚠️ No se encontraron datos desde la fila 26."

        n_filas = len(filas)
        n_cols  = COL_FIN - COL_INI + 1  # 45

        # ── Convertir valores no serializables (datetime, None→"") ──
        def clean(v):
            if v is None:
                return ""
            if hasattr(v, "isoformat"):
                return v.isoformat()
            return v

        filas_clean = [[clean(v) for v in fila] for fila in filas]

        # Columna AS en notación A1
        def col_letter(n):
            s = ""
            while n > 0:
                n, r = divmod(n - 1, 26)
                s = chr(65 + r) + s
            return s

        ultima_col = col_letter(COL_FIN)  # "AS"

        # ── Helper: escribir un bloque de filas en un rango específico ──
        CHUNK = 200

        def patch_range(start_row, data_rows):
            end_row = start_row + len(data_rows) - 1
            rng = f"Data!A{start_row}:{ultima_col}{end_row}"
            r = requests.patch(
                f"{base}/worksheets/Data/range(address='{rng}')",
                headers=hdrs_sess,
                json={"values": data_rows},
                timeout=120,
            )
            r.raise_for_status()

        # ── PASO 1: Limpiar en chunks (evita 504 en rangos grandes) ──
        used_resp = requests.get(
            f"{base}/worksheets/Data/usedRange",
            headers=hdrs_sess, timeout=30)
        total_filas_sp = used_resp.json().get("rowCount", 10000) if used_resp.ok else 10000

        for start in range(1, total_filas_sp + 1, CHUNK):
            chunk_size = min(CHUNK, total_filas_sp - start + 1)
            patch_range(start, [[""] * n_cols for _ in range(chunk_size)])

        # ── PASO 2: Escribir nuevos datos en chunks ──
        for start in range(0, n_filas, CHUNK):
            patch_range(start + 1, filas_clean[start: start + CHUNK])

    finally:
        # ── Cerrar sesión siempre ──
        requests.post(f"{base}/closeSession", headers=hdrs_sess, timeout=15)

    return f"✅ ¡Listo! {n_filas} filas escritas en SharePoint (hoja Data, A→AS) en un solo request."


st.set_page_config(page_title="Walmex · CFBC", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
.main .block-container { padding: 0 !important; max-width: 100% !important; margin: 0 !important; }
.main { padding: 0 !important; overflow: hidden !important; }
.stApp { margin: 0 !important; }
[data-testid="stHeader"],[data-testid="stSidebar"],[data-testid="stToolbar"],
[data-testid="stDecoration"],[data-testid="stStatusWidget"],
#MainMenu, header, footer {
    display: none !important; visibility: hidden !important; height: 0 !important;
}
.stDeployButton { display: none !important; }
div[style*="bottom: 1.5rem"], div[style*="bottom: 15px"],
div[style*="position: fixed"][style*="bottom"][style*="right"],
iframe[src*="badge"] {
    display: none !important; opacity: 0 !important;
    pointer-events: none !important; visibility: hidden !important;
}
[data-testid='stVerticalBlock'] { gap: 0 !important; padding: 0 !important; }
div[data-testid='stHtml'] { padding: 0 !important; margin: 0 !important; line-height: 0 !important; }
iframe { display: block !important; margin: 0 !important; border: none !important; }
</style>
""", unsafe_allow_html=True)

# Archivo de caché permanente — sobrevive recargas y reinicios
_CACHE_LATEST = _CACHE_DIR / "data_latest.pkl"

def _cache_es_compatible(data) -> bool:
    return isinstance(data, dict) and 'resumen_diario' in data and 'resumen_has_field1' in data

def _cargar_desde_disco():
    """Carga el último caché guardado en disco. Retorna None si no existe."""
    if _CACHE_LATEST.exists():
        try:
            with open(_CACHE_LATEST, "rb") as f:
                data = pickle.load(f)
            if _cache_es_compatible(data):
                return data
        except Exception:
            pass
    return None

@st.cache_data(show_spinner=False)
def cargar_datos(cache_key: str = "") -> dict:
    _CACHE_DIR.mkdir(parents=True, exist_ok=True)

    if cache_key:
        data_cache_file = _CACHE_DIR / f"data_{cache_key}.pkl"
        if data_cache_file.exists():
            try:
                with open(data_cache_file, "rb") as f:
                    data = pickle.load(f)
                if _cache_es_compatible(data):
                    with open(_CACHE_LATEST, "wb") as f:
                        pickle.dump(data, f)
                    return data
            except Exception:
                pass

    # Fallback a último caché local si falla la detección remota
    datos_disco = _cargar_desde_disco()
    if datos_disco is not None and (not cache_key or datos_disco.get('_source_cache_key') == cache_key):
        return datos_disco

    # Solo descargar y reprocesar el Excel cuando cambie de versión o no exista caché local.
    excel_path, cache_key_real = _descargar_excel_sharepoint()
    cache_key = cache_key or cache_key_real
    data_cache_file = _CACHE_DIR / f"data_{cache_key}.pkl"

    if data_cache_file.exists():
        try:
            with open(data_cache_file, "rb") as f:
                data = pickle.load(f)
            if _cache_es_compatible(data):
                with open(_CACHE_LATEST, "wb") as f:
                    pickle.dump(data, f)
                return data
        except Exception:
            pass

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb['Data']

    def sv(v):
        try: return float(v) if v is not None else 0.0
        except: return 0.0

    # Mapear columnas por nombre de encabezado — fila 1
    headers = [str(c).strip() if c else '' for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    def col(name):
        for i, h in enumerate(headers):
            if h == name: return i
        raise ValueError(f'Columna "{name}" no encontrada. Encabezados: {headers}')

    # Log headers para diagnóstico si alguna columna falla
    import sys
    _col_names = [h for h in headers if h]
    
    idx_producto = col('Desc Art 1')
    idx_tienda   = col('Nombre Tienda/Club')
    idx_semana   = col('SEM')
    idx_fecha    = col('Diario')
    idx_ventas   = col('Cnt POS')       # Unidades vendidas (Cnt POS)
    idx_embarque = col('Cntd Embarque') # Unidades embarcadas
    idx_merma_vc = col('Cant VC Tienda') # Merma (Cant VC Tienda)
    
    # Columnas opcionales para Tienda — intentar varios nombres posibles
    idx_venta_cfbc = None
    for _n in ['Venta CFBC / Costo (Facturado)', 'Venta CFBC/Costo (Facturado)',
               'Venta CFBC', 'CFBC']:
        try: idx_venta_cfbc = col(_n); break
        except: pass

    idx_venta_wmx = None
    for _n in ['Venta WMX / Precio Costo (Vendido)', 'Venta WMX/Precio Costo (Vendido)',
               'Venta WMX', 'WMX']:
        try: idx_venta_wmx = col(_n); break
        except: pass

    idx_retail_vc = None
    for _n in ['Costo VC Tienda',
               'Suma de Retail VC Tienda', 'Retail VC Tienda',
               'Suma Retail VC Tienda', 'Retail VC', 'Suma de Retail VC',
               'Suma de Retail VC Tienda ']:  # trailing space variant
        try: idx_retail_vc = col(_n); break
        except: pass

    # Columna de inventario actual
    idx_inventario = None
    for _n in ['Cantidad Actual en Existentes de la tienda', 'Cantidad Actual en Existentes',
               'Cantidad Actual', 'Inventario Actual', 'Existentes']:
        try: idx_inventario = col(_n); break
        except: pass

    idx_venta_pos = None
    for _n in ['Venta POS']:
        try: idx_venta_pos = col(_n); break
        except: pass

    idx_field1 = None
    for _n in ['Field1']:
        try: idx_field1 = col(_n); break
        except: pass

    # Advertir si columnas clave no se encontraron
    if idx_retail_vc is None:
        import streamlit as _st
        _st.warning(
            f"⚠️ No se encontró columna 'Costo VC Tienda'. "
            f"Columnas disponibles: {[h for h in headers if h and ('VC' in h or 'Costo' in h or 'Retail' in h)]}\n"
            f"Todos los encabezados: {[h for h in headers if h]}"
        )
    
    if idx_inventario is None:
        import streamlit as _st
        _st.warning(
            f"⚠️ No se encontró columna 'Cantidad Actual en Existentes de la tienda'. "
            f"Inventario no estará disponible. Columnas disponibles: {[h for h in headers if h]}"
        )

    # Columnas de días — nombres exactos del Excel
    _dias = ['Sab', 'Dom', 'Lun', 'Mar', 'Mie', 'Jue', 'Vie']
    _ctd_nombres = {
        'Sab': 'Cnt Sab',
        'Dom': 'Ctd Dom',
        'Lun': 'Ctd Lun',
        'Mar': 'Ctd Mar',
        'Mie': 'Ctd Mie',
        'Jue': 'Ctd Jue',
        'Vie': 'Ctd Vie',
    }
    idx_ctd  = {}
    idx_vtas = {}
    for d in _dias:
        try: idx_ctd[d] = col(_ctd_nombres[d])
        except: idx_ctd[d] = None
        try: idx_vtas[d] = col(f'Ventas {d}')
        except: idx_vtas[d] = None

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        producto  = str(row[idx_producto]).strip() if row[idx_producto] else None
        tienda    = str(row[idx_tienda]).strip()   if row[idx_tienda]   else None
        # Semana: valor simple como 50, 51 etc
        try:
            semana_num = int(float(row[idx_semana])) if row[idx_semana] is not None else None
        except:
            semana_num = None

        # Fecha: puede venir como datetime o string MM/DD/YYYY
        from datetime import datetime as _dt
        fecha_raw = row[idx_fecha]
        anio = None
        fecha_ymd = ""
        if hasattr(fecha_raw, 'strftime'):
            fecha = fecha_raw.strftime('%d/%m/%Y')
            fecha_ymd = fecha_raw.strftime('%Y/%m/%d')
            anio  = fecha_raw.year
        elif fecha_raw:
            s_fecha = str(fecha_raw).strip()
            for fmt in ('%m/%d/%Y','%d/%m/%Y','%Y-%m-%d'):
                try:
                    dt   = _dt.strptime(s_fecha, fmt)
                    fecha = dt.strftime('%d/%m/%Y')
                    fecha_ymd = dt.strftime('%Y/%m/%d')
                    anio  = dt.year
                    break
                except:
                    continue
            else:
                fecha = s_fecha
                fecha_ymd = s_fecha
        else:
            fecha = ''

        if not producto or not tienda or not semana_num: continue

        records.append({
            'producto':   producto,
            'tienda':     tienda,
            '_semana_num': semana_num,
            '_anio':       anio,
            'semana':     (anio * 100 + semana_num) if anio else semana_num,
            'fecha':      fecha,
            'fecha_ymd':  fecha_ymd,
            'ventas_u':   sv(row[idx_ventas]),
            'embarque_u': sv(row[idx_embarque]),
            'merma_u':    sv(row[idx_merma_vc]),
            'venta_cfbc': sv(row[idx_venta_cfbc]) if idx_venta_cfbc is not None else 0,
            'venta_wmx':  sv(row[idx_venta_wmx]) if idx_venta_wmx is not None else 0,
            'venta_pos':  sv(row[idx_venta_pos]) if idx_venta_pos is not None else 0,
            'field1':     sv(row[idx_field1]) if idx_field1 is not None else 0,
            'retail_vc':  sv(row[idx_retail_vc]) if idx_retail_vc is not None else 0,
            'inventario': sv(row[idx_inventario]) if idx_inventario is not None else 0,
            **{f'ctd_{d.lower()}':  sv(row[idx_ctd[d]])  if idx_ctd[d]  is not None else 0 for d in _dias},
            **{f'vtas_{d.lower()}': sv(row[idx_vtas[d]]) if idx_vtas[d] is not None else 0 for d in _dias},
        })

    # Inferir año para filas sin fecha usando el año más frecuente del mismo número de semana
    from collections import Counter as _Counter
    _sem_anio_votes = _Counter()
    for r in records:
        if r['_anio']:
            _sem_anio_votes[(r['_semana_num'], r['_anio'])] += 1
    _sem_to_anio = {}
    for (wk, yr) in _sem_anio_votes:
        if wk not in _sem_to_anio or _sem_anio_votes[(wk, yr)] > _sem_anio_votes[(_sem_to_anio[wk], yr)]:
            _sem_to_anio[wk] = yr
    for r in records:
        if not r['_anio']:
            yr_inf = _sem_to_anio.get(r['_semana_num'])
            if yr_inf:
                r['semana'] = yr_inf * 100 + r['_semana_num']

    semanas   = sorted(set(r['semana'] for r in records))
    tiendas   = sorted(set(r['tienda']  for r in records))
    productos = sorted(set(r['producto'] for r in records))

    by_stp = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(float))))
    for r in records:
        by_stp[r['semana']][r['tienda']][r['producto']]['ventas_u']   += r['ventas_u']
        by_stp[r['semana']][r['tienda']][r['producto']]['embarque_u'] += r['embarque_u']
        by_stp[r['semana']][r['tienda']][r['producto']]['merma_u']    += r['merma_u']
        by_stp[r['semana']][r['tienda']][r['producto']]['venta_cfbc'] += r['venta_cfbc']
        by_stp[r['semana']][r['tienda']][r['producto']]['venta_wmx']  += r['venta_wmx']
        by_stp[r['semana']][r['tienda']][r['producto']]['retail_vc']  += r['retail_vc']
        for d in ['dom','lun','mar','mie','jue','vie','sab']:
            by_stp[r['semana']][r['tienda']][r['producto']][f'ctd_{d}']  += r.get(f'ctd_{d}', 0)
            by_stp[r['semana']][r['tienda']][r['producto']][f'vtas_{d}'] += r.get(f'vtas_{d}', 0)
        by_stp[r['semana']][r['tienda']][r['producto']]['inventario'] += r['inventario']

    # Fecha real del Excel por semana
    fecha_por_semana = {}
    for r in records:
        if r['fecha']:
            fecha_por_semana[r['semana']] = r['fecha']

    result = {}
    for t in tiendas:
        result[t] = {}
        for s in semanas:
            idx    = semanas.index(s)
            last12 = semanas[max(0, idx-11):idx+1]
            last3  = semanas[max(0, idx-2):idx+1]
            prod_data = {}
            for p in productos:
                v12  = sum(by_stp[sem][t][p]['ventas_u']   for sem in last12)
                v3   = sum(by_stp[sem][t][p]['ventas_u']   for sem in last3)
                emb3 = sum(by_stp[sem][t][p]['embarque_u'] for sem in last3)  # embarque 3 semanas
                m3   = sum(by_stp[sem][t][p]['merma_u']    for sem in last3)  # merma 3 semanas (Cant VC Tienda)
                cfbc3 = sum(by_stp[sem][t][p].get('venta_cfbc', 0) for sem in last3)  # Venta CFBC 3 semanas
                retail3 = sum(by_stp[sem][t][p].get('retail_vc', 0) for sem in last3)  # Retail VC 3 semanas
                avg  = v3 / 3  # Promedio = Ventas 3 semanas / 3
                
                # Proyección = Venta Promedio / (1 - Índice Merma %)
                merma_ratio = m3 / emb3 if emb3 > 0 else 0  # Ratio de merma como decimal
                proj = avg / (1 - merma_ratio) if merma_ratio < 1 else avg  # Evitar división por cero
                
                prod_data[p] = {
                    'v12': round(v12), 'v3': round(v3),
                    'n12': min(s % 100 if s > 9999 else s, 12),
                    'emb': round(emb3), 'm3': round(m3),
                    'avg': round(avg, 1), 'proj': round(proj),
                    'pct_merma': round(m3/emb3*100) if emb3 > 0 else 0,
                    'cfbc': round(cfbc3), 'retail': round(retail3),
                    'wmx': round(sum(by_stp[sem][t][p].get('venta_wmx', 0) for sem in last3))
                }
            result[t][s] = prod_data

    # Totales crudos acumulados por tienda — GLOBAL (todas las fechas, sin ventanas deslizantes)
    totales_tienda = defaultdict(lambda: defaultdict(float))
    # Totales crudos por tienda+semana — para filtrar por semana específica
    raw_semana = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    # Totales crudos por tienda+semana+producto — para cuadrar tablas inferiores con superiores
    totales_prod_tienda = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in records:
        totales_tienda[r['tienda']]['embarque_u'] += r['embarque_u']
        totales_tienda[r['tienda']]['venta_cfbc'] += r['venta_cfbc']
        totales_tienda[r['tienda']]['venta_wmx']  += r['venta_wmx']
        totales_tienda[r['tienda']]['merma_u']    += r['merma_u']
        totales_tienda[r['tienda']]['retail_vc']  += r['retail_vc']
        totales_tienda[r['tienda']]['inventario'] += r['inventario']
        totales_tienda[r['tienda']]['ventas_u']   += r['ventas_u']
        raw_semana[r['tienda']][r['semana']]['embarque_u'] += r['embarque_u']
        raw_semana[r['tienda']][r['semana']]['venta_cfbc'] += r['venta_cfbc']
        raw_semana[r['tienda']][r['semana']]['venta_wmx']  += r['venta_wmx']
        raw_semana[r['tienda']][r['semana']]['merma_u']    += r['merma_u']
        raw_semana[r['tienda']][r['semana']]['retail_vc']  += r['retail_vc']
        raw_semana[r['tienda']][r['semana']]['inventario'] += r['inventario']
        raw_semana[r['tienda']][r['semana']]['ventas_u']   += r['ventas_u']
        totales_prod_tienda[r['tienda']][r['producto']]['embarque_u'] += r['embarque_u']
        totales_prod_tienda[r['tienda']][r['producto']]['venta_cfbc'] += r['venta_cfbc']
        totales_prod_tienda[r['tienda']][r['producto']]['venta_wmx']  += r['venta_wmx']
        totales_prod_tienda[r['tienda']][r['producto']]['merma_u']    += r['merma_u']
        totales_prod_tienda[r['tienda']][r['producto']]['retail_vc']  += r['retail_vc']
        totales_prod_tienda[r['tienda']][r['producto']]['ventas_u']   += r['ventas_u']
        totales_prod_tienda[r['tienda']][r['producto']]['inventario'] += r['inventario']
        for d in ['dom','lun','mar','mie','jue','vie','sab']:
            totales_prod_tienda[r['tienda']][r['producto']][f'ctd_{d}']  += r.get(f'ctd_{d}', 0)
            totales_prod_tienda[r['tienda']][r['producto']][f'vtas_{d}'] += r.get(f'vtas_{d}', 0)

    # ==== DATA DIARIA PARA RESUMEN ====
    resumen_diario = defaultdict(
        lambda: defaultdict(
            lambda: defaultdict(
                lambda: defaultdict(lambda: defaultdict(float))
            )
        )
    )
    for r in records:
        if r['fecha_ymd']:
            resumen_diario[r['tienda']][str(r['semana'])][r['producto']][r['fecha_ymd']]['embarque'] += r['embarque_u']
            resumen_diario[r['tienda']][str(r['semana'])][r['producto']][r['fecha_ymd']]['ventas']   += r['ventas_u']
            resumen_diario[r['tienda']][str(r['semana'])][r['producto']][r['fecha_ymd']]['merma']    += r['merma_u']
            resumen_diario[r['tienda']][str(r['semana'])][r['producto']][r['fecha_ymd']]['venta_pos'] += r['venta_pos']
            resumen_diario[r['tienda']][str(r['semana'])][r['producto']][r['fecha_ymd']]['venta_cfbc'] += r['venta_cfbc']
            resumen_diario[r['tienda']][str(r['semana'])][r['producto']][r['fecha_ymd']]['field1'] += r['field1']
            resumen_diario[r['tienda']][str(r['semana'])][r['producto']][r['fecha_ymd']]['inventario'] += r['inventario']

    # ==== INTEGRACIÓN DE HOJA DETALLE ====
    # La tabla "Ventas por Día" debe respetar exactamente las columnas del Excel Data:
    # Cnt/Ctd Sab..Vie y Ventas Sab..Vie. Por eso no se sobrescriben con la hoja Detalle.

    # 2. Construir mapeo Semana -> Sábado de inicio
    from datetime import datetime as _dt, timedelta as _td
    week_starts_to_semana = {}
    for r in records:
        if r['fecha'] and r['semana']:
            try:
                dt_d = _dt.strptime(r['fecha'], '%d/%m/%Y')
                w_start = dt_d - _td(days=dt_d.weekday())
                week_starts_to_semana[w_start] = r['semana']
            except:
                pass

    # 3. Leer hoja Detalle para el tab INVENTARIO ACTUAL
    detalle_inventario = {'fechas': [], 'fechas_por_semana': {}, 'fecha_to_semana': {}, 'data': {}}
    ws_det = None
    # Buscar la hoja de forma insensible a mayúsculas/minúsculas
    for _sname in wb.sheetnames:
        if _sname.strip().lower() == 'detalle':
            ws_det = wb[_sname]
            break

    if ws_det is None:
        import streamlit as _st2
        _st2.warning(f"⚠️ No se encontró hoja 'Detalle'. Hojas disponibles: {wb.sheetnames}")
    
    if ws_det is not None:
        try:
            det_headers = [str(c).strip() if c else '' for c in next(ws_det.iter_rows(min_row=1, max_row=1, values_only=True))]
            def _det_col(name):
                for i, h in enumerate(det_headers):
                    if h.strip().lower() == name.lower():
                        return i
                return None
            idx_det_tienda     = _det_col('Tienda')
            idx_det_fecha      = _det_col('Fecha')
            idx_det_producto   = _det_col('Producto')
            idx_det_inventario = _det_col('Inventario')

            if not all(v is not None for v in [idx_det_tienda, idx_det_fecha, idx_det_producto, idx_det_inventario]):
                import streamlit as _st2
                _st2.warning(
                    f"⚠️ Hoja Detalle encontrada pero columnas no detectadas. "
                    f"Tienda:{idx_det_tienda} Fecha:{idx_det_fecha} Producto:{idx_det_producto} Inventario:{idx_det_inventario}. "
                    f"Encabezados: {det_headers}"
                )
            else:
                from datetime import datetime as _dt_det, timedelta as _td_det
                det_data = {}   # {tienda: {producto: {fecha_str: valor}}}
                fechas_set = set()
                fecha_to_semana = {}  # {fecha_str: semana_key (YYYYWW)}
                fechas_por_semana = defaultdict(list)  # {semana_key: [fecha_str, ...]}

                for row in ws_det.iter_rows(min_row=2, values_only=True):
                    td = str(row[idx_det_tienda]).strip() if row[idx_det_tienda] else None
                    pd2 = str(row[idx_det_producto]).strip() if row[idx_det_producto] else None
                    inv_v = sv(row[idx_det_inventario])
                    fecha_raw = row[idx_det_fecha]
                    fecha_str = None
                    fecha_dt = None
                    if hasattr(fecha_raw, 'strftime'):
                        fecha_str = fecha_raw.strftime('%m/%d/%y')
                        fecha_dt = fecha_raw
                    elif fecha_raw:
                        fecha_str = str(fecha_raw).strip()
                        for _fmt in ('%m/%d/%y', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d'):
                            try:
                                fecha_dt = _dt_det.strptime(fecha_str[:10], _fmt)
                                break
                            except:
                                pass
                    if td and pd2 and fecha_str:
                        fechas_set.add(fecha_str)
                        if td not in det_data:
                            det_data[td] = {}
                        if pd2 not in det_data[td]:
                            det_data[td][pd2] = {}
                        det_data[td][pd2][fecha_str] = det_data[td][pd2].get(fecha_str, 0) + inv_v

                        # Calcular semana usando el mismo criterio (inicio sábado)
                        if fecha_dt and fecha_str not in fecha_to_semana:
                            # Encontrar el sábado de inicio de esa semana
                            w_start = fecha_dt.date() - _td_det(days=fecha_dt.weekday()) if hasattr(fecha_dt, 'date') else fecha_dt - _td_det(days=fecha_dt.weekday())
                            w_start_dt = _dt_det(w_start.year, w_start.month, w_start.day) if hasattr(w_start, 'year') else w_start
                            sem = week_starts_to_semana.get(w_start_dt)
                            if sem is not None:
                                fecha_to_semana[fecha_str] = sem
                                if fecha_str not in fechas_por_semana[sem]:
                                    fechas_por_semana[sem].append(fecha_str)

                # Ordenar fechas cronológicamente
                def _parse_det_date(s):
                    for fmt in ('%m/%d/%y', '%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d'):
                        try: return _dt_det.strptime(s, fmt)
                        except: pass
                    return _dt_det(2099,1,1)
                sorted_fechas = sorted(fechas_set, key=_parse_det_date)
                # Ordenar también las fechas dentro de cada semana
                for _sem_k in fechas_por_semana:
                    fechas_por_semana[_sem_k] = sorted(fechas_por_semana[_sem_k], key=_parse_det_date)

                detalle_inventario = {
                    'fechas': sorted_fechas,
                    'fechas_por_semana': {str(k): v for k, v in fechas_por_semana.items()},
                    'fecha_to_semana': fecha_to_semana,
                    'data': det_data
                }
        except Exception as _e_det:
            import streamlit as _st2
            _st2.warning(f"⚠️ Error leyendo hoja Detalle: {_e_det}")
    # =====================================


    # raw por tienda+semana+producto (exactamente la semana seleccionada)
    raw_prod_semana = {}
    for t in tiendas:
        raw_prod_semana[t] = {}
        for s in semanas:
            raw_prod_semana[t][str(s)] = {}
            for p in productos:
                d = by_stp[s][t][p]
                if any(d.get(k, 0) for k in ['ventas_u','venta_cfbc','venta_wmx','merma_u','retail_vc','embarque_u','inventario','ctd_sab','ctd_dom','ctd_lun','ctd_mar','ctd_mie','ctd_jue','ctd_vie']):
                    raw_prod_semana[t][str(s)][p] = {
                        'ventas_u':   round(d['ventas_u']),
                        'venta_cfbc': round(d['venta_cfbc']),
                        'venta_wmx':  round(d['venta_wmx']),
                        'merma_u':    round(d['merma_u']),
                        'retail_vc':  round(d['retail_vc']),
                        'embarque_u': round(d['embarque_u']),
                        'inventario': round(d['inventario']),
                        **{f'ctd_{d_str}': round(d[f'ctd_{d_str}']) for d_str in ['dom','lun','mar','mie','jue','vie','sab']},
                        **{f'vtas_{d_str}': round(d[f'vtas_{d_str}']) for d_str in ['dom','lun','mar','mie','jue','vie','sab']}
                    }

    # Agregaciones de inventario por tienda (suma de todos los productos)
    inventario_por_tienda = {}
    for t in tiendas:
        total_inv = sum(totales_prod_tienda[t][p].get('inventario', 0) for p in productos)
        inventario_por_tienda[t] = {
            'total': round(total_inv),
            'productos': {p: round(totales_prod_tienda[t][p].get('inventario', 0)) 
                         for p in productos if totales_prod_tienda[t][p].get('inventario', 0) > 0}
        }
    
    # Agregaciones de inventario por producto (suma de todas las tiendas)
    inventario_por_producto = {}
    for p in productos:
        total_inv = sum(totales_prod_tienda[t][p].get('inventario', 0) for t in tiendas)
        inventario_por_producto[p] = {
            'total': round(total_inv),
            'tiendas': {t: round(totales_prod_tienda[t][p].get('inventario', 0)) 
                       for t in tiendas if totales_prod_tienda[t][p].get('inventario', 0) > 0}
        }

    # ==== MAPEOS PARA VISTA GASTO ====
    PRODUCTO_GASTO = {
        'BQT ALSTROEMERI 8T': 15.00,
        'BQT GIRASOL 6T': 10.00,
        'BQT LILI ASIATIC 6T': 15.00,
        'BQT MINI CLAVEL 8T': 15.00,
        'BQT MIXTO 12T': 23.00,
        'BQT MIXTO 15T': 23.00,
        'BQT MIXTO 18 T': 10.00,
        'BQT MIXTO 9T': 15.00,
        'BQT ROSAS 12T': 20.00,
        'BQT ROSAS 12T BAJA': 20.00,
        'BQT ROSAS 6T': 15.00,
        'BQT SNAPDRAGON 8T': 10.00,
        'BQT ROSAS 6T BAJA': 10.00
    }
    
    TIENDA_RUTA = {
        'SC LOMAS DE SANTA FE': 'Rutas Playas',
        'SC ENSENADA CENTRO': 'ENS',
        'SC ENSENADA': 'ENS',
        'SC ROSARITO': 'Rutas Playas',
        'SC PLAYAS DE TIJUANA': 'Rutas Playas',
        'SC MACROPLAZA INSURGENTES': 'Ruta 2000',
        'SC DIAZ ORDAZ': 'Ruta 2000',
        'SC TIJUANA HIPODROMO': 'Ruta 2000',
        'SC PACIFICO': 'Rutas Playas',
        'SC TIJUANA 2000': 'Ruta 2000',
        'SC MEXICALI NOVENA': 'MXL 1',
        'SC PLAZA SAN PEDRO': 'MXL 1',
        'SC GALERIAS DEL VALLE': 'MXL 1',
        'SC MEXICALI': 'MXL 1',
        'SC TECATE GARITA': 'Ruta 2000',
        'SC NUEVO MEXICALI': 'MXL 1'
    }

    # Procesar datos para vista GASTO: agrupar por ruta/semana/producto
    gasto_data = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    for r in records:
        ruta = TIENDA_RUTA.get(r['tienda'])
        gasto = PRODUCTO_GASTO.get(r['producto'], 0)
        if ruta and gasto > 0:
            gasto_data[ruta][r['semana']][r['producto']] += r['embarque_u']

    # ==== OTROS GASTOS — hoja "Gastos" del mismo SharePoint principal ====
    # Estructura: Fecha | SC | Combustible (Transf.) | Viaticos | Casetas | Mantenimiento | ...
    # Cada columna a partir de la 3ª es un tipo de gasto con su monto.
    # Agrupamos: SC normalizado -> concepto -> semana_key(YYYYWW) -> total $
    gastos_otros = defaultdict(lambda: defaultdict(lambda: defaultdict(float)))
    gastos_sc_order = []  # orden de los SCs para el frontend
    gastos_tipos = []   # orden de los conceptos de gasto para el frontend
    try:
        ws_gastos = None
        for _sname in wb.sheetnames:
            if _sname.strip().lower() == 'gastos':
                ws_gastos = wb[_sname]
                break

        if ws_gastos is not None:
            import datetime as _dtmod

            # Leer encabezados de fila 1
            _gst_hdrs = [str(c).strip() if c else ''
                         for c in next(ws_gastos.iter_rows(min_row=1, max_row=1, values_only=True))]

            # Col 0 = Fecha, Col 1 = SC, Col 2+ = tipos de gasto
            _ig_fecha = 0
            # Tipos de gasto: todas las columnas con encabezado a partir del índice 2
            _gasto_cols = [(i, h) for i, h in enumerate(_gst_hdrs)
                           if i >= 2 and h]

            def _norm_txt(v):
                return ' '.join(str(v or '').strip().upper().split())

            def _norm_sc(v):
                _txt = _norm_txt(v)
                if not _txt:
                    return None
                if 'WMX TJ' in _txt or _txt in ('TJ', 'TIJUANA'):
                    return 'WMX TJ'
                if 'WMX MXL' in _txt or 'WLMX' in _txt or 'MXL' in _txt or 'MEXICALI' in _txt:
                    return 'WMX MXL'
                return None

            def _concepto_gasto(tipo, detalle_sc=None):
                _tipo = ' '.join(str(tipo or '').strip().split())
                _detalle = ' '.join(str(detalle_sc or '').strip().split())
                return f'{_tipo} - {_detalle}' if _detalle else _tipo

            def _parse_fecha_g(v):
                if isinstance(v, _dtmod.datetime):
                    return v
                if isinstance(v, _dtmod.date):
                    return _dtmod.datetime(v.year, v.month, v.day)
                if isinstance(v, (int, float)) and v > 0:
                    return _dtmod.datetime(1899, 12, 30) + _dtmod.timedelta(days=int(v))
                if isinstance(v, str):
                    for _fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                        try:
                            return _dtmod.datetime.strptime(v[:10], _fmt)
                        except Exception:
                            pass
                return None

            for _grow in ws_gastos.iter_rows(min_row=2, values_only=True):
                _fv = _grow[_ig_fecha] if _ig_fecha < len(_grow) else None
                _fdt = _parse_fecha_g(_fv)
                if _fdt is None:
                    continue
                _iso = _fdt.isocalendar()
                _sem_key = int(_iso[0]) * 100 + int(_iso[1])

                # Leer SC (columna 1 = col B del Excel). Si trae texto libre en vez de SC,
                # lo agrupamos en "Gastos Generales" y lo pasamos al concepto para no contaminar nombres.
                _sc_raw = str(_grow[1]).strip() if len(_grow) > 1 and _grow[1] else ''
                _sc = _norm_sc(_sc_raw) or 'Gastos Generales'
                _sc_detalle = '' if _sc != 'Gastos Generales' else _sc_raw
                if _sc and _sc not in gastos_sc_order:
                    gastos_sc_order.append(_sc)

                for _cidx, _ctipo in _gasto_cols:
                    _monto = _grow[_cidx] if _cidx < len(_grow) else None
                    if _monto is None:
                        continue
                    try:
                        _concepto = _concepto_gasto(_ctipo, _sc_detalle)
                        if _concepto not in gastos_tipos:
                            gastos_tipos.append(_concepto)
                        gastos_otros[_sc][_concepto][_sem_key] += float(_monto)
                    except Exception:
                        pass
    except Exception:
        pass

    result_dict = {
        '_source_cache_key': cache_key,
        'semanas':           semanas,
        'tiendas':           tiendas,
        'productos':         productos,
        'fecha_por_semana':  fecha_por_semana,
        'semana_to_week_starts': {s: d.strftime('%Y-%m-%d') for d, s in week_starts_to_semana.items()},
        'data':              {t: {str(s): v for s, v in sv2.items()} for t, sv2 in result.items()},
        'totales_tienda':    {t: dict(v) for t, v in totales_tienda.items()},
        'raw_semana':        {t: {str(s): dict(v) for s, v in sv.items()} for t, sv in raw_semana.items()},
        'raw_prod_semana':   raw_prod_semana,
        'totales_prod_tienda': {t: {p: dict(v) for p, v in pd.items()} for t, pd in totales_prod_tienda.items()},
        'resumen_diario': {t: {sem: {prod: dict(fechas) for prod, fechas in sems.items()} for sem, sems in semanas.items()} for t, semanas in resumen_diario.items()},
        'resumen_has_field1': idx_field1 is not None,
        'inventario_por_tienda': inventario_por_tienda,
        'inventario_por_producto': inventario_por_producto,
        'detalle_inventario': detalle_inventario,
        'producto_gasto': PRODUCTO_GASTO,
        'tienda_ruta': TIENDA_RUTA,
        'gasto_data': {k: {s: dict(p) for s,p in v.items()} for k,v in gasto_data.items()},
        'gastos_otros': {sc: {tipo: dict(sems) for tipo, sems in tipos.items()} for sc, tipos in gastos_otros.items()},
        'gastos_tipos': gastos_tipos,
        'gastos_sc': gastos_sc_order,
    }
    # ── Liberar memoria del workbook antes de guardar caché ──
    try:
        wb.close()
        del wb, ws
    except Exception:
        pass
    import gc as _gc
    _gc.collect()

    # ── Guardar caché en disco (gzip reduce el tamaño ~70%) ──
    import gzip as _gzip
    try:
        _CACHE_DIR.mkdir(parents=True, exist_ok=True)
        with open(data_cache_file, "wb") as f:
            pickle.dump(result_dict, f)
        with open(_CACHE_LATEST, "wb") as f:
            pickle.dump(result_dict, f)
    except Exception:
        pass
    return result_dict

def _encode_share_url_graph(url: str) -> str:
    """
    Codifica una URL de link-compartido al formato que pide el endpoint
    /shares de Microsoft Graph (prefijo 'u!' + base64 url-safe sin padding).
    """
    b64 = base64.b64encode(url.encode("utf-8")).decode("utf-8")
    b64 = b64.replace("/", "_").replace("+", "-").rstrip("=")
    return "u!" + b64


def _descargar_via_graph_shared_link(shared_link: str) -> bytes:
    """
    Descarga el contenido de un archivo a partir de un link "compartido" de
    SharePoint/OneDrive, autenticando con el token de la app (Azure AD)
    en vez de depender de que el link sea público/anónimo.

    Esto es necesario porque al marcar el link como "privado"/restringido,
    una petición sin credenciales (requests.get directo) recibe 403 Forbidden.
    Usando el token de la misma app que ya usamos para el Excel principal,
    Graph resuelve el link aunque ya no acepte acceso anónimo.

    Requisito: la app de Azure AD (CLIENT_ID) debe tener el permiso de
    APLICACIÓN 'Files.Read.All' (o 'Sites.Read.All') consentido por un
    administrador, en el mismo tenant donde vive el archivo
    (pacificafarms-my.sharepoint.com).
    """
    cfg = _get_sharepoint_cfg()
    token = _sharepoint_access_token(cfg["tenant_id"], cfg["client_id"], cfg["client_secret"])
    headers = {"Authorization": f"Bearer {token}"}
    share_id = _encode_share_url_graph(shared_link)
    resp = requests.get(
        f"https://graph.microsoft.com/v1.0/shares/{share_id}/driveItem/content",
        headers=headers,
        timeout=60,
        allow_redirects=True,
    )
    resp.raise_for_status()
    return resp.content


# ══════════════════════════════════════════════════════
# GASOLINA — SharePoint externo (pacificafarms)
# ══════════════════════════════════════════════════════
_GASOLINA_SHARED_LINK = "https://pacificafarms-my.sharepoint.com/:x:/g/personal/anahi_mora_cfbc_co/IQANE_rjNbe-T5n5XZuwt3FwAa9dla1RGnbl1oC9PGuNO-o?e=cqwWsE&download=1"
_VEHICULOS_GASOLINA   = ['FORD / TRANSIT 250 / 2020', 'FORD / TRANSIT / 2019']   

@st.cache_data(ttl=3600, show_spinner=False)
def cargar_gasolina() -> dict:
    """
    Descarga la hoja 'Base datos' desde SharePoint usando link público directo.
    Columnas: F=Fecha (idx 5), L=Total (idx 11), R=Vehículo (idx 17).
    Agrupa por vehículo → semana (YYYYWW) → total $.
    No requiere autenticación.
    """
    import datetime, io, gc

    contenido = _descargar_via_graph_shared_link(_GASOLINA_SHARED_LINK)

    wb_gas = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    if 'Base datos' in wb_gas.sheetnames:
        ws_gas = wb_gas['Base datos']
    elif any(s.lower() == 'base datos' for s in wb_gas.sheetnames):
        ws_gas = wb_gas[[s for s in wb_gas.sheetnames if s.lower() == 'base datos'][0]]
    else:
        ws_gas = wb_gas.active

    values = [list(row) for row in ws_gas.iter_rows(values_only=True)]
    wb_gas.close()

    gc.collect()



    # Índices de columna (0-based): F=5, L=11, R=17
    COL_FECHA = 5
    COL_TOTAL = 11
    COL_VEH   = 17

    def _parse_fecha(v):
        """Acepta serial Excel (int/float), string ISO o datetime."""
        if isinstance(v, (int, float)) and v > 0:
            return datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(v))
        if isinstance(v, str):
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y'):
                try:
                    return datetime.datetime.strptime(v[:10], fmt)
                except ValueError:
                    pass
        if hasattr(v, 'isocalendar'):
            return v
        return None

    # Agrupar: vehiculo → semana (YYYYWW) → total $
    resultado = {v: {} for v in _VEHICULOS_GASOLINA}
    semanas_set = set()

    for row in values[3:]:           # fila 4 en adelante (índice 3)
        if len(row) <= max(COL_FECHA, COL_TOTAL, COL_VEH):
            continue
        vehiculo = str(row[COL_VEH]).strip() if row[COL_VEH] is not None else ''
        if vehiculo not in _VEHICULOS_GASOLINA:
            continue
        fecha = _parse_fecha(row[COL_FECHA])
        if fecha is None:
            continue
        try:
            total = float(row[COL_TOTAL]) if row[COL_TOTAL] not in (None, '', 'None') else 0.0
        except (ValueError, TypeError):
            total = 0.0
        iso  = fecha.isocalendar()
        sem  = int(iso[0]) * 100 + int(iso[1])   # YYYYWW
        resultado[vehiculo][sem] = resultado[vehiculo].get(sem, 0.0) + total
        semanas_set.add(sem)

    return {
        'gasolina_por_vehiculo': resultado,
        'gasolina_semanas':      sorted(semanas_set),
    }

# ══════════════════════════════════════════════════════
# NÓMINA — SharePoint externo (pacificafarms)
# ══════════════════════════════════════════════════════
_NOMINA_SHARED_LINK = "https://pacificafarms-my.sharepoint.com/:x:/g/personal/anahi_mora_cfbc_co/IQAQCb79SzHtRrTQR71pSNQcAT7r1BbxaVtGuiSy1lUzZOY?e=hxAq81&download=1"

@st.cache_data(ttl=3600, show_spinner=False)
def cargar_nomina() -> dict:
    import io, re, gc
    contenido = _descargar_via_graph_shared_link(_NOMINA_SHARED_LINK)

    wb_nomina = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    resultado = {}
    semanas_set = set()

    for sname in wb_nomina.sheetnames:
        sname_clean = sname.strip().upper()
        if sname_clean.startswith('WK'):
            digits = re.findall(r'\d+', sname_clean)
            if digits:
                num_str = "".join(digits)
                if len(num_str) >= 4:
                    yy = int(num_str[:2])
                    ww = int(num_str[2:4])
                    yyyy = 2000 + yy if yy < 100 else yy
                    sem_key = yyyy * 100 + ww
                    
                    ws = wb_nomina[sname]
                    val = ws.cell(row=48, column=6).value
                    try:
                        val_float = float(val) if val is not None else 0.0
                    except (ValueError, TypeError):
                        val_float = 0.0
                        
                    resultado[sem_key] = val_float
                    semanas_set.add(sem_key)

    wb_nomina.close()

    gc.collect()

    return {
        'nomina_data': resultado,
        'nomina_semanas': sorted(semanas_set),
    }

# ══════════════════════════════════════════════════════
# DIAGNÓSTICO TEMPORAL — borrar después de encontrar la ruta
# ══════════════════════════════════════════════════════
def _diagnostico_sharepoint():
    cfg = _get_sharepoint_cfg()
    app_msal = msal.ConfidentialClientApplication(
        cfg["client_id"],
        authority=f"https://login.microsoftonline.com/{cfg['tenant_id']}",
        client_credential=cfg["client_secret"],
    )
    token = app_msal.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in token:
        st.error(f"Error auth: {token.get('error_description')}")
        return
    headers = {"Authorization": f"Bearer {token['access_token']}"}

    # Obtener site_id
    parts    = cfg["site_url"].rstrip("/").split("/")
    hostname = parts[2]
    site_path = "/".join(parts[3:])
    site_resp = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{hostname}:/{site_path}",
        headers=headers, timeout=30
    )
    site_id = site_resp.json()["id"]
    st.info(f"✅ Site ID: `{site_id}`")

    # Listar bibliotecas
    drives = requests.get(
        f"https://graph.microsoft.com/v1.0/sites/{site_id}/drives",
        headers=headers, timeout=30
    ).json()
    st.markdown("### Bibliotecas disponibles:")
    for d in drives.get("value", []):
        st.code(f"Nombre: {d['name']}\nID: {d['id']}")

    # Listar raíz de cada biblioteca
    for d in drives.get("value", []):
        st.markdown(f"### Archivos en raíz de **{d['name']}**:")
        items = requests.get(
            f"https://graph.microsoft.com/v1.0/drives/{d['id']}/root/children",
            headers=headers, timeout=30
        ).json()
        for item in items.get("value", []):
            tipo = "📁" if "folder" in item else "📄"
            st.write(f"{tipo} `{item['name']}`")

if st.sidebar.button("🔍 Diagnóstico SharePoint"):
    _diagnostico_sharepoint()
    st.stop()
# ══════════════════════════════════════════════════════

_excel_meta = {}
try:
    _excel_meta = _obtener_meta_excel_actual()
except Exception:
    _excel_meta = {}

import concurrent.futures

# Functions to wrap loading to return results cleanly for concurrent execution
def _load_main_data():
    try:
        return cargar_datos(_excel_meta.get('cache_key', ''))
    except Exception as e:
        return e

def _load_gas_data():
    try:
        return cargar_gasolina()
    except Exception as e:
        return e

def _load_nom_data():
    try:
        return cargar_nomina()
    except Exception as e:
        return e

# Load data concurrently
with concurrent.futures.ThreadPoolExecutor() as executor:
    future_main = executor.submit(_load_main_data)
    future_gas = executor.submit(_load_gas_data)
    future_nom = executor.submit(_load_nom_data)

    res_main = future_main.result()
    res_gas = future_gas.result()
    res_nom = future_nom.result()

if isinstance(res_main, Exception):
    st.error(f"❌ Error cargando datos: {res_main}")
    st.stop()
else:
    DATA = res_main

# ── Gasolina (fuente separada — SharePoint pacificafarms) ──
if isinstance(res_gas, Exception):
    DATA['gasolina_por_vehiculo'] = {v: {} for v in _VEHICULOS_GASOLINA}
    DATA['gasolina_semanas']      = []
    DATA['vehiculos_gasolina']    = _VEHICULOS_GASOLINA
    st.warning(f"⚠️ No se pudieron cargar datos de gasolina: {res_gas}")
else:
    DATA['gasolina_por_vehiculo'] = res_gas['gasolina_por_vehiculo']
    DATA['gasolina_semanas']      = res_gas['gasolina_semanas']
    DATA['vehiculos_gasolina']    = _VEHICULOS_GASOLINA

# ── Nómina ──
if isinstance(res_nom, Exception):
    DATA['nomina_data'] = {}
    DATA['nomina_semanas'] = []
    st.warning(f"⚠️ No se pudieron cargar datos de nómina: {res_nom}")
else:
    DATA['nomina_data'] = res_nom['nomina_data']
    DATA['nomina_semanas'] = res_nom['nomina_semanas']

HTML = r"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<link href="https://fonts.googleapis.com/icon?family=Material+Icons" rel="stylesheet">
<style>
.material-icons{font-size:15px;vertical-align:middle;line-height:1}
tbody tr.checked { background: #e8f4e8; }
tbody tr.flagged { background: #fde8e8 !important; }
tbody tr.flagged.checked { background: #f5d0d0 !important; }
body.dragging-checks { user-select: none; cursor: grabbing !important; }
.flag-btn {
    background: none; border: 1px solid #ccc; border-radius: 3px; cursor: pointer;
    font-size: 0.7rem; padding: 1px 4px; line-height: 1; color: #999; transition: all 0.15s;
}
.flag-btn:hover { border-color: #e00; color: #e00; }
tbody tr.flagged .flag-btn { background: #e00; border-color: #e00; color: white; }
.nota-input {
    border: none; background: transparent; font-size: 0.68rem; color: #555;
    width: 100%; min-width: 80px; outline: none; font-family: Arial, sans-serif;
}
.nota-input:focus { background: #fffbe6; border-bottom: 1px solid #ccc; }
.nota-input::placeholder { color: #bbb; }
</style>
<style>
*{box-sizing:border-box;margin:0;padding:0}
body{background:#fff;font-family:Arial,sans-serif;font-size:12px;color:#111}
.hdr{display:flex;align-items:center;justify-content:space-between;padding:6px 16px 4px;border-bottom:1px solid #ccc}
.wm-logo{display:flex;align-items:center;gap:4px}
.wm-text{font-size:1.2rem;font-weight:700;color:#0071ce;letter-spacing:-0.5px}
.wm-spark{color:#ffc220;font-size:1.3rem;line-height:1}
.hdr-right{display:flex;align-items:center;gap:12px;font-size:.72rem;color:#333;line-height:1.6}
.btn-print{
  display:inline-flex;align-items:center;gap:5px;
  padding:4px 14px;border-radius:4px;border:1px solid #0071ce;
  background:#fff;color:#0071ce;font-size:.7rem;font-weight:700;
  cursor:pointer;transition:.15s;white-space:nowrap;flex-shrink:0;
}
.btn-print:hover{background:#0071ce;color:#fff}
.btn-reload{
  display:inline-flex;align-items:center;gap:4px;
  padding:4px 8px;border-radius:4px;border:1px solid #ccc;
  background:#fff;color:#888;font-size:.65rem;
  cursor:pointer;transition:.15s;white-space:nowrap;flex-shrink:0;
}
.btn-reload:hover{border-color:#0071ce;color:#0071ce}
.ctrl{display:flex;align-items:center;gap:8px;padding:5px 16px;background:#f5f7fa;border-bottom:1px solid #ddd;flex-wrap:wrap;overflow:visible;position:relative;z-index:5000;isolation:isolate}
.ctrl label{font-size:.7rem;color:#555;font-weight:600}
#semDropWrap,#tiendaDropWrap,#productoDropWrap{position:relative;z-index:5001}
#semDropMenu,#tiendaDropMenu,#productoDropMenu{z-index:5002!important}
select{border:1px solid #bbb;border-radius:4px;padding:3px 7px;font-size:.72rem;cursor:pointer;background:#fff}
.chip-wrap{display:flex;flex-wrap:wrap;gap:4px;flex:1}
.chip{padding:2px 9px;border-radius:12px;font-size:.67rem;cursor:pointer;border:1px solid #bbb;color:#333;background:#fff;transition:.15s}
.chip:hover{border-color:#0071ce;color:#0071ce}
.mode-wrap{display:none;align-items:center;gap:6px}
.mode-btn{padding:4px 10px;border:1px solid #bbb;border-radius:4px;background:#fff;color:#333;font-size:.7rem;font-weight:700;cursor:pointer;transition:.15s}
.mode-btn.active{background:#0071ce;border-color:#0071ce;color:#fff}
.sem-item{display:flex;align-items:center;gap:6px;padding:4px 10px;font-size:.72rem;cursor:pointer;color:#333;white-space:nowrap}
.sem-item:hover{background:#f0f7ff}
.sem-item.on{background:#e8f0fe;color:#0071ce;font-weight:600}
.sem-item input{accent-color:#0071ce;cursor:pointer;width:13px;height:13px;flex-shrink:0}
.grid{display:grid;grid-template-columns:1fr 1fr;padding:8px 16px;gap:8px;width:100%;box-sizing:border-box}
.box{border:1px solid #bbb;border-radius:4px;overflow:visible}
.box-hdr{background:#f0f0f0;border-bottom:1px solid #bbb;padding:4px 10px;text-align:center;font-size:.74rem;font-weight:700;color:#111}
table.t{width:100%;border-collapse:collapse;font-size:.71rem}
table.t th{padding:3px 8px;font-size:.66rem;font-weight:700;color:#333;border-bottom:1px solid #ccc;text-align:right;background:#fafafa}
table.t th:first-child{text-align:left}
table.t td{padding:2px 8px;font-size:.71rem;text-align:right;color:#222;height:20px;line-height:20px;white-space:nowrap}
table.t td:first-child{text-align:left;color:#111}
table.t tr.total td{font-weight:700;border-top:1px solid #ddd;background:#f5f5f5}
.red{color:#d32f2f !important;font-weight:700 !important;background-color:#ffebee !important;}
.bold{font-weight:700}
#viewTienda table.t tr:not(.total):hover td{background:#f0f7ff;cursor:pointer}
#viewInventario table.t tr:hover td{background:#f0f7ff;cursor:pointer}
#viewTienda{overflow:visible}
html,body{height:auto;overflow:auto;margin:0;padding:0}
#app{height:auto;overflow:visible}
@media(max-width:1200px){
  .grid{grid-template-columns:1fr;gap:8px}
}
@media(max-width:768px){
  .grid{gap:6px;padding:6px 12px}
  table.t th,table.t td{padding:1px 6px;font-size:.68rem}
}
#loader{position:fixed;inset:0;background:#fff;display:flex;align-items:center;justify-content:center;z-index:99;flex-direction:column;gap:10px}
.ld-txt{font-size:.85rem;color:#0071ce;font-weight:600}
.ld-bar{width:160px;height:3px;background:#dde;border-radius:2px;overflow:hidden}
.ld-fill{height:100%;background:#0071ce;animation:ld .9s ease-in-out infinite}
@keyframes ld{0%{transform:translateX(-100%)}100%{transform:translateX(200%)}}
/* Estilos para vista GASTO - formato financiero */
#viewGasto{
  overflow:visible!important;height:auto!important;max-height:none!important;
  padding:16px 20px!important;
  background:linear-gradient(180deg,#f7fbff 0%,#f4f8fc 100%)
}
#viewGasto>div{overflow:visible!important;height:auto!important;max-height:none!important}
#gastoGrid{gap:16px!important}
#viewGasto .box{
  overflow:hidden!important;height:auto!important;max-height:none!important;
  border:1px solid #bfd0e3;border-radius:10px;background:#fff;
  box-shadow:0 10px 28px rgba(15,39,72,.07), 0 2px 8px rgba(15,39,72,.04);
  min-width: 0!important;
}
#viewGasto .box-hdr{
  overflow:visible!important;
  background:linear-gradient(180deg,#dcecf9 0%,#c2d9ef 100%);
  border-bottom:1px solid #b3c9de;
  color:#0f2748;padding:10px 14px;text-align:left
}
#viewGasto .gasto-head-kicker{
  font-size:.58rem;font-weight:800;letter-spacing:.12em;text-transform:uppercase;color:#47627d;margin-bottom:3px
}
#viewGasto .gasto-head-title{
  font-size:.9rem;font-weight:800;letter-spacing:.01em;color:#0f2748;line-height:1.15
}
#viewGasto .gasto-head-sub{
  margin-top:4px;font-size:.68rem;color:#4f6882;line-height:1.35
}
#viewGasto .box>div{height:auto!important;max-height:none!important}
#viewGasto .gasto-table-wrap{
  overflow:auto!important;max-height:68vh;
  border-top:1px solid #d8e4f0;background:#fff
}
#viewGasto table{overflow:visible!important}
#viewGasto table.t{
  width:max-content;min-width:100%;border-collapse:separate;border-spacing:0;
  font-size:.7rem;font-variant-numeric:tabular-nums
}
#viewGasto table.t th,
#viewGasto table.t td{
  padding:6px 9px;white-space:nowrap;
  border-bottom:1px solid #e2eaf3
}
#viewGasto table.t thead th{
  position:sticky;top:0;z-index:4;
  background:#d6e6f5;color:#17324d;font-size:.63rem;font-weight:800;
  letter-spacing:.06em;text-transform:uppercase;
  border-bottom:1px solid #b6cadf
}
#viewGasto table.t th:first-child,
#viewGasto table.t td:first-child{
  position:sticky;left:0;z-index:3;
  background:inherit;
  box-shadow:1px 0 0 #d6e2ee;
  text-align:left
}
#viewGasto table.t thead th:first-child{z-index:5;background:#c7dcef}
#viewGasto table.t tbody tr:nth-child(even) td{background:#fbfdff}
#viewGasto table.t tbody tr:hover td{background:#f3f8fd}
#viewGasto table.t tbody tr:hover td:first-child{background:#f3f8fd}
#tGasto{font-size:.66rem}
/* Anular zebra y hover generales para tGasto */
#tGasto tbody tr:nth-child(even) td{background:#fff!important}
#tGasto tbody tr:hover td{background:inherit!important}
#tGasto tbody tr:hover td:first-child{background:inherit!important}
#tGasto .ruta-row td{
  background:#fff!important;color:#111827;font-weight:700;
  border-top:2px solid #e5e7eb;border-bottom:1px solid #e5e7eb;font-size:.72rem
}
#tGasto .ruta-row td:first-child{text-transform:none;letter-spacing:0;color:#111827;font-weight:800}
#tGasto .ruta-row td:not(:first-child){color:#16a34a;font-weight:800}
#tGasto .unidad-row td{font-size:.66rem;color:#374151;background:#fff}
#tGasto .unidad-row td:first-child{padding-left:24px;color:#374151}
#tGasto .unidad-row td:not(:first-child){color:#374151;text-align:right}
#tGasto .totalpiezas-row td{font-size:.66rem;font-weight:700;color:#111827;background:#fff!important;border-top:1px solid #e5e7eb}
#tGasto .totalpiezas-row td:first-child{padding-left:24px}
#tGasto .costo-pieza-row td{font-size:.66rem;font-style:italic;color:#6b7280;background:#fff!important;border-bottom:2px solid #e5e7eb}
#tGasto .costo-pieza-row td:first-child{padding-left:24px}
#tGasto .grand-total td{
  background:#fff!important;color:#111827;font-weight:800;
  border-top:2px solid #e5e7eb;border-bottom:1px solid #e5e7eb;font-size:.69rem
}
#tGasto .grand-total td:first-child{background:#fff!important}
/* Anular zebra y hover para tGasolina */
#tGasolina tbody tr:nth-child(even) td{background:inherit!important}
#tGasolina tbody tr:hover td{background:inherit!important}
#tGasolina tbody tr:hover td:first-child{background:inherit!important}
#tGasolina{font-size:.69rem}
#tGasolina td{color:#1f2937;background:#fff}
/* Sección principal: GASOLINA POR VEHÍCULO, OTROS GASTOS */
#tGasolina .section-row td{
  background:#fff!important;color:#16a34a;font-weight:800;
  font-size:.72rem;letter-spacing:0;text-transform:none;
  border-top:2px solid #e5e7eb;border-bottom:1px solid #e5e7eb;
  padding:6px 9px
}
/* Grupo dentro de sección: WMX TJ, WMX MXL, GASTOS GENERALES */
#tGasolina .group-row td{
  background:#fff!important;color:#16a34a;font-weight:800;
  font-size:.69rem;letter-spacing:0;text-transform:none;
  border-top:1px solid #e5e7eb;border-bottom:1px solid #e5e7eb;
  padding:5px 9px
}
#tGasolina .group-row td:first-child{padding-left:9px}
/* Filas de concepto/item */
#tGasolina .ruta-row td{
  background:#fff!important;color:#374151;font-weight:400;font-size:.69rem
}
#tGasolina .ruta-row td:first-child{padding-left:28px;color:#374151;font-weight:400}
#tGasolina .ruta-row td:not(:first-child){color:#374151}
/* Subtotales: TOTAL WMX TJ, TOTAL WMX MXL, etc */
#tGasolina .subtotal-row td{
  background:#fff!important;color:#111827;font-weight:700;
  border-top:1px solid #e5e7eb;border-bottom:1px solid #e5e7eb;
  font-size:.69rem
}
#tGasolina .subtotal-row td:first-child{padding-left:9px}
/* Grand Total */
#tGasolina .grand-total td{
  background:#fff!important;color:#111827;font-weight:800;
  border-top:2px solid #e5e7eb;border-bottom:1px solid #e5e7eb;font-size:.7rem
}
#tGasolina .grand-total td:first-child{background:#fff!important}
#viewResumen{overflow:visible!important;height:auto!important;max-height:none!important;min-height:400px}
#viewResumen .resumen-layout{
  display:flex;align-items:flex-start;gap:14px;
  overflow:visible;max-width:100%
}
#viewResumen .box{
  overflow:visible!important;height:auto!important;max-height:none!important;
  min-width:0;max-width:100%; width:fit-content;
  border:1px solid #bfc9d4;border-radius:3px;background:#fff;
  box-shadow:0 1px 4px rgba(0,0,0,.08)
}
#viewResumen .resumen-main-box{flex:0 0 auto;width:fit-content!important;min-width:0;max-width:100%}
#viewResumen .resumen-side-box{flex:0 1 auto;width:fit-content!important;max-width:min(74vw, 1380px)}
#viewResumen .box-hdr{
  background:#f6f8fa;
  border-bottom:2px solid #bfc9d4;
  color:#1f2937;font-size:.72rem;font-weight:700;
  padding:6px 10px;text-align:left;letter-spacing:.02em
}
#viewResumen .resumen-table-wrap{
  overflow:auto!important;max-height:calc(100vh - 120px);max-width:100%;
  width:fit-content;background:#fff
}
#viewResumen table.t{
  width:auto;border-collapse:collapse;
  font-size:.58rem;font-variant-numeric:tabular-nums
}
#viewResumen table.t th,
#viewResumen table.t td{
  padding:2px 4px;white-space:nowrap;
  border:none;box-shadow:inset 0 0 0 1px #c8d2e8;text-align:right;color:#000;
  box-sizing:border-box;vertical-align:middle
}
#viewResumen table.t thead th{
  position:sticky;top:0;z-index:4;
  background:#dce6f1;color:#000;
  font-size:.58rem;font-weight:700;
  border-bottom:1px solid #95b3d7;text-align:center
}
#viewResumen table.t thead tr:nth-child(2) th{
  background:#dce6f1;top:20px
}
#viewResumen table.t thead tr:nth-child(3) th{
  background:#dce6f1;top:40px
}
#viewResumen table.t th:first-child,
#viewResumen table.t td:first-child{
  position:sticky;left:0;z-index:3;
  background:#fff;text-align:left;
  border-right:1px solid #e0e0e0;
}
#viewResumen table.t th:nth-child(2),
#viewResumen table.t td:nth-child(2){
  position:sticky;left:0;z-index:2;
  background:#fff;text-align:left;
  border-right:1px solid #e0e0e0;
  vertical-align:top;
}
#viewResumen table.t thead th:first-child{z-index:5;background:#dce6f1;border-right:none}
#viewResumen table.t thead tr:nth-child(2) th:first-child{background:#dce6f1}
#viewResumen table.t thead th:nth-child(2){z-index:4;background:#dce6f1}
#viewResumen table.t tr.pivot-product td{
  background:#fff!important;color:#000;font-weight:700;
  border-top:1px solid #5b9bd5!important;border-bottom:1px solid #5b9bd5!important;
  font-size:.58rem
}
#viewResumen .pivot-product td:first-child{background:#fff!important}
#viewResumen .pivot-metric td{color:#000;background:#fff}
#viewResumen .pivot-metric td:first-child{padding-left:14px;color:#000;background:#fff}
#viewResumen .pivot-total td{
  background:#dce6f1!important;color:#000;font-weight:700;
  border-top:1px solid #95b3d7;font-size:.62rem
}
#viewResumen .pivot-total td:first-child{background:#dce6f1!important}
#viewResumen .hl-yellow{background:#fff176!important;font-weight:700;color:#4a3b00!important}
#viewResumen .hl-green{background:#a9d18e!important;font-weight:700;color:#143d16!important}
#viewResumen .muted{color:#999}
#viewResumen .reabasto-panel{padding:10px;background:#fff;width:fit-content;max-width:100%}
#viewResumen .reabasto-meta{font-size:.62rem;color:#5a6573;line-height:1.4;margin-bottom:10px}
#viewResumen .reabasto-table-wrap{overflow:auto;max-height:58vh;border:1px solid #e3e8ef;border-radius:6px;width:fit-content;max-width:100%}
#viewResumen table.reabasto-table{width:auto;min-width:0;border-collapse:collapse;font-size:.64rem;font-variant-numeric:tabular-nums;table-layout:auto}
#viewResumen table.reabasto-table th,
#viewResumen table.reabasto-table td{padding:5px 6px;border-bottom:1px solid #edf1f5;white-space:normal;word-break:break-word;line-height:1.15;text-align:right;color:#1f2937}
#viewResumen table.reabasto-table th{position:sticky;top:0;z-index:2;background:#eef4fb;font-size:.58rem;text-transform:uppercase;letter-spacing:.03em;padding:6px 6px}
#viewResumen table.reabasto-table th:first-child,
#viewResumen table.reabasto-table td:first-child,
#viewResumen table.reabasto-table th:nth-child(2),
#viewResumen table.reabasto-table td:nth-child(2),
#viewResumen table.reabasto-table th:nth-child(3),
#viewResumen table.reabasto-table td:nth-child(3),
#viewResumen table.reabasto-table th:nth-child(4),
#viewResumen table.reabasto-table td:nth-child(4),
#viewResumen table.reabasto-table th:nth-child(13),
#viewResumen table.reabasto-table td:nth-child(13){text-align:left}
#viewResumen .reabasto-empty{padding:18px 12px;text-align:center;color:#7b8794;font-size:.68rem}
#viewResumen .badge{display:inline-block;padding:3px 7px;border-radius:999px;font-size:.55rem;font-weight:700;letter-spacing:.03em}
#viewResumen .badge-critical{background:#fde8e8;color:#b42318}
#viewResumen .badge-today{background:#fff4db;color:#b54708}
#viewResumen .badge-plan{background:#e8f1ff;color:#175cd3}
#viewResumen .badge-ok{background:#e8f7ed;color:#027a48}
#viewResumen .badge-idle{background:#eef2f6;color:#667085}
#viewResumen .reabasto-order{font-weight:700;background:#eefaf2;color:#0f5132}
#viewResumen .reabasto-muted{color:#98a2b3}
@media(max-width:1200px){
  #viewGasto .gasto-table-wrap{max-height:56vh}
  #viewResumen .resumen-table-wrap{max-height:calc(100vh - 140px)}
  #viewResumen .resumen-layout{flex-direction:column}
  #viewResumen .resumen-side-box{width:100%!important;max-width:100%;flex:1 1 auto}
#viewResumen .reabasto-table-wrap{max-height:42vh}
}
/* Estilos para el markdown del asistente */
#aiChatMessages table { width:100%; border-collapse:collapse; margin-bottom:10px; font-size:0.8rem; background:#fff; }
#aiChatMessages th, #aiChatMessages td { border:1px solid #cbd5e1; padding:6px 8px; text-align:left; }
#aiChatMessages th { background:#f1f5f9; font-weight:600; color:#334155; }
#aiChatMessages p { margin:0 0 8px 0; }
#aiChatMessages p:last-child { margin:0; }
#aiChatMessages ul, #aiChatMessages ol { margin:0 0 8px 0; padding-left:20px; }
#aiChatMessages li { margin-bottom:4px; }
/* Estilos para comentarios dinámicos */
.cell-comment { position: relative !important; cursor: pointer !important; }
.cell-comment::after {
  content: ''; position: absolute; top: 0; right: 0;
  border-left: 7px solid transparent; border-top: 7px solid #e74c3c;
}
.cell-comment:hover { background-color: #e2e8f0 !important; }
#commentModal { display: none; position: fixed; top: 50%; left: 50%; transform: translate(-50%, -50%); background: #fff; border-radius: 8px; box-shadow: 0 10px 30px rgba(0,0,0,0.2); z-index: 10000; width: 320px; padding: 20px; font-family: sans-serif; }
#commentModal h3 { margin: 0 0 10px 0; font-size: 15px; color: #1e293b; }
#commentModal textarea { width: 100%; height: 90px; padding: 8px; box-sizing: border-box; border: 1px solid #cbd5e1; border-radius: 4px; resize: none; font-size: 13px; outline: none; }
#commentModal .actions { display: flex; justify-content: flex-end; gap: 8px; margin-top: 15px; }
#commentModal button { padding: 6px 12px; border: none; border-radius: 4px; cursor: pointer; font-size: 13px; font-weight: bold; }
#btnSaveComment { background: #1565c0; color: #fff; }
#btnDelComment { background: #e74c3c; color: #fff; }
#btnCancelComment { background: #f1f5f9; color: #475569; }
#commentOverlay { display: none; position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(0,0,0,0.3); z-index: 9999; }
</style>
<script src="https://cdn.jsdelivr.net/npm/marked/marked.min.js"></script>
</head>
<body>

<!-- Ventana emergente (Modal) para Comentarios -->
<div id="commentOverlay" onclick="closeCommentModal()"></div>
<div id="commentModal">
  <h3>Comentario</h3>
  <textarea id="txtComment" placeholder="Escribe un comentario o nota..."></textarea>
  <div class="actions">
    <button id="btnDelComment" onclick="deleteComment()">Borrar</button>
    <button id="btnCancelComment" onclick="closeCommentModal()">Cancelar</button>
    <button id="btnSaveComment" onclick="saveComment()">Guardar</button>
  </div>
</div>

<div id="loader">
  <div class="ld-txt">Cargando...</div>
  <div class="ld-bar"><div class="ld-fill"></div></div>
</div>

<div id="app" style="display:none">

  <div class="hdr">
    <div style="margin-top:12px; display:flex; gap:8px;">
      <button onclick="setView('producto')" id="btnProd" style="padding:6px 12px; background:#6A4B59; color:white; border:1px solid #4A3540; border-bottom:3px solid #4A3540; border-radius:4px; cursor:pointer; font-weight:600;">Proyección</button>
      <button onclick="setView('tienda')" id="btnTiend" style="padding:6px 12px; background:#A08A95; color:white; border:1px solid #806A75; border-bottom:3px solid #806A75; border-radius:4px; cursor:pointer; font-weight:600;">Tienda</button>
      <button onclick="setView('comparativo')" id="btnComp" style="padding:6px 12px; background:#A08A95; color:white; border:1px solid #806A75; border-bottom:3px solid #806A75; border-radius:4px; cursor:pointer; font-weight:600;">Comparativo</button>
      <button onclick="setView('inventario')" id="btnInv" style="padding:6px 12px; background:#A08A95; color:white; border:1px solid #806A75; border-bottom:3px solid #806A75; border-radius:4px; cursor:pointer; font-weight:600;">Inventario Actual</button>
      <button onclick="setView('gasto')" id="btnGasto" style="padding:6px 12px; background:#A08A95; color:white; border:1px solid #806A75; border-bottom:3px solid #806A75; border-radius:4px; cursor:pointer; font-weight:600;">Gasto</button>
      <button onclick="setView('resumen')" id="btnResumen" style="padding:6px 12px; background:#A08A95; color:white; border:1px solid #806A75; border-bottom:3px solid #806A75; border-radius:4px; cursor:pointer; font-weight:600;">Resumen</button>
      <div style="position:relative; display:inline-block;" id="choferesDropdownWrap">
        <button onclick="toggleChoferesMenu(event)" id="btnChoferes" style="padding:6px 12px; background:#A08A95; color:white; border:1px solid #806A75; border-bottom:3px solid #806A75; border-radius:4px; cursor:pointer; font-weight:600;">Choferes ▾</button>
        <div id="choferesMenu" style="display:none; position:absolute; top:100%; left:0; background:#fff; border:1px solid #ddd; border-radius:6px; box-shadow:0 4px 12px rgba(0,0,0,0.15); z-index:9999; min-width:160px; padding:4px 0; margin-top:2px;">
          <div onclick="openChoferesSubTab('consolidado')" style="padding:8px 16px; cursor:pointer; font-size:13px; color:#1e293b; font-weight:500;" onmouseover="this.style.background='#f1f5f9'" onmouseout="this.style.background=''">Consolidado</div>
          <div onclick="openChoferesSubTab('devoluciones')" style="padding:8px 16px; cursor:pointer; font-size:13px; color:#1e293b; font-weight:500;" onmouseover="this.style.background='#f1f5f9'" onmouseout="this.style.background=''">Devoluciones</div>
        </div>
      </div>
    </div>
    <div class="hdr-right">
      <div>
        <div id="hdrFecha">—</div>
        <div>Semana&nbsp;&nbsp;<strong id="hdrSem">—</strong></div>
      </div>
      <button class="btn-print" onclick="imprimirReporte()">🖨️ Imprimir</button>
      <button class="btn-reload" onclick="recargarDatos()" title="Actualizar datos">↺</button>
    </div>
  </div>
  <div class="ctrl">
      <div id="globalCtrlContainer" style="display:flex; align-items:center; gap:8px; flex-wrap:wrap;">
    <label>Semana:</label>
    <div style="position:relative;display:inline-block" id="semDropWrap">
      <button id="semDropBtn" onclick="toggleSemDrop()" style="border:1px solid #bbb;border-radius:4px;padding:3px 24px 3px 7px;font-size:.72rem;cursor:pointer;background:#fff;min-width:160px;text-align:left;position:relative">
        <span id="semDropLabel">— Seleccionar semanas —</span>
        <span style="position:absolute;right:6px;top:50%;transform:translateY(-50%);font-size:.6rem">▼</span>
      </button>
      <div id="semDropMenu" style="display:none;position:absolute;top:100%;left:0;z-index:999;background:#fff;border:1px solid #bbb;border-radius:4px;box-shadow:0 3px 10px rgba(0,0,0,.15);min-width:200px;max-height:260px;overflow-y:auto;padding:4px 0"></div>
    </div>
    <label>Ruta:</label>
    <select id="rutaSelect" onchange="onRutaChange(this.value)" style="border:1px solid #bbb;border-radius:4px;padding:3px 8px;font-size:.72rem;background:#fff;margin-right:12px;cursor:pointer;">
      <option value="Todas">Todas las rutas</option>
      <option value="ENS">ENS</option>
      <option value="MXL 1">MXL 1</option>
      <option value="Ruta 2000">Ruta 2000</option>
      <option value="Rutas Playas">Rutas Playas</option>
    </select>
    <label>Tienda:</label>
    <div style="position:relative;display:inline-block" id="tiendaDropWrap">
      <button id="tiendaDropBtn" onclick="toggleTiendaDrop()" style="border:1px solid #bbb;border-radius:4px;padding:3px 24px 3px 7px;font-size:.72rem;cursor:pointer;background:#fff;min-width:160px;text-align:left;position:relative">
        <span id="tiendaDropLabel">— Seleccionar tiendas —</span>
        <span style="position:absolute;right:6px;top:50%;transform:translateY(-50%);font-size:.6rem">▼</span>
      </button>
      <div id="tiendaDropMenu" style="display:none;position:absolute;top:100%;left:0;z-index:999;background:#fff;border:1px solid #bbb;border-radius:4px;box-shadow:0 3px 10px rgba(0,0,0,.15);min-width:200px;max-height:260px;overflow-y:auto;padding:4px 0"></div>
    </div>
    <label>Producto:</label>
    <div style="position:relative;display:inline-block" id="productoDropWrap">
      <button id="productoDropBtn" onclick="toggleProductoDrop()" style="border:1px solid #bbb;border-radius:4px;padding:3px 24px 3px 7px;font-size:.72rem;cursor:pointer;background:#fff;min-width:160px;text-align:left;position:relative">
        <span id="productoDropLabel">— Todos los productos —</span>
        <span style="position:absolute;right:6px;top:50%;transform:translateY(-50%);font-size:.6rem">▼</span>
      </button>
      <div id="productoDropMenu" style="display:none;position:absolute;top:100%;left:0;z-index:999;background:#fff;border:1px solid #bbb;border-radius:4px;box-shadow:0 3px 10px rgba(0,0,0,.15);min-width:200px;max-height:260px;overflow-y:auto;padding:4px 0"></div>
    </div>
    </div>
  </div>

  <div class="grid" id="viewProducto">
    <div class="box">
      <div class="box-hdr">Ventas Históricas</div>
      <table class="t"><thead><tr><th>Producto</th><th>12 Semanas</th><th>3 Semanas</th></tr></thead>
      <tbody id="tHist"></tbody></table>
    </div>
    <div class="box">
      <div class="box-hdr">Índice de Merma por Artículo Últimas 3 Semanas</div>
      <table class="t"><thead><tr><th>Producto</th><th>Embarque</th><th>Merma</th><th>Merma %</th></tr></thead>
      <tbody id="tMerma"></tbody></table>
    </div>
    <div class="box">
      <div class="box-hdr">Venta Promedio Semanal</div>
      <table class="t"><thead><tr><th>Producto</th><th>Prom 12 Sem</th><th>Prom 3 Sem</th></tr></thead>
      <tbody id="tAvg"></tbody></table>
    </div>
    <div class="box">
      <div class="box-hdr" id="projTitle">Proyección Semana Siguiente</div>
      <table class="t"><thead><tr><th>Producto</th><th>Proyección</th></tr></thead>
      <tbody id="tProj"></tbody></table>
    </div>
  </div>

  <div class="grid" id="viewTienda" style="display:none;">
    <div class="box">
      <div class="box-hdr">Top Venta</div>
      <table class="t"><thead><tr><th>Tienda</th><th>UNIDADES</th><th>VENTA CFBC</th><th>VENTA WMX</th><th>%</th></tr></thead>
      <tbody id="tHistT"></tbody></table>
    </div>
    <div class="box">
      <div class="box-hdr">Top Merma</div>
      <table class="t"><thead><tr><th>Tienda</th><th>UNIDADES</th><th>$</th><th>CANTIDAD</th><th>%</th></tr></thead>
      <tbody id="tMermaT"></tbody></table>
    </div>
    <div class="box" id="boxAvgT" style="display:none">
      <div class="box-hdr" id="avgTTitle">Venta Promedio Semanal</div>
      <table class="t"><thead><tr><th>Producto</th><th>Venta CFBC</th><th>Venta WMX</th><th>Unidades</th></tr></thead>
      <tbody id="tAvgT"></tbody></table>
    </div>
    <div id="boxProjWrapper" style="display:none; flex-direction:row; gap:8px; align-items: flex-start;">
      <div class="box" id="boxProjT" style="flex:1; min-width:0; max-height:400px; overflow-y:auto;">
        <div class="box-hdr" id="projTTitle">Comparacion Ultimas 3 Semanas</div>
        <table class="t"><thead><tr><th>Merma Producto</th><th>Unidades</th><th>Cantidad</th></tr></thead>
        <tbody id="tProjT"></tbody></table>
      </div>
      <div class="box" id="boxMermaSem" style="flex:1; min-width:0; max-height:400px; overflow-y:auto;">
        <div class="box-hdr">Merma por Producto/Semana</div>
        <table class="t"><thead><tr><th>Producto</th><th>Merma</th><th>Cantidad</th></tr></thead>
        <tbody id="tMermaSem"></tbody></table>
      </div>
    </div>
    <div class="box" id="boxDiasT" style="display:none; grid-column: 1 / -1; overflow-x: auto;">
      <div class="box-hdr">Ventas por Día</div>
      <table class="t" style="min-width: 1200px;">
        <thead>
          <tr>
            <th>Producto</th>
            <th id="hdrThSab">Cnt Sab</th><th id="hdrThDom">Ctd Dom</th><th id="hdrThLun">Ctd Lun</th><th id="hdrThMar">Ctd Mar</th><th id="hdrThMie">Ctd Mie</th><th id="hdrThJue">Ctd Jue</th><th id="hdrThVie">Ctd Vie</th>
            <th id="hdrThVtaSab">Ventas Sab</th><th id="hdrThVtaDom">Ventas Dom</th><th id="hdrThVtaLun">Ventas Lun</th><th id="hdrThVtaMar">Ventas Mar</th><th id="hdrThVtaMie">Ventas Mie</th><th id="hdrThVtaJue">Ventas Jue</th><th id="hdrThVtaVie">Ventas Vie</th>
          </tr>
        </thead>
        <tbody id="tDiasT"></tbody>
      </table>
    </div>
  </div>

  <!-- Vista COMPARATIVO -->
  <div id="viewComparativo" style="display:none; padding:10px 16px; overflow:visible !important; height:auto !important;">

    <!-- Layout: tabla pivot izquierda | gráfica derecha -->
    <div style="display:flex; gap:10px; align-items:flex-start; flex-wrap:wrap;">

      <!-- Columna izquierda: tabla pivot -->
      <div class="box" style="flex:1 1 420px; overflow:auto; min-width:320px;">
        <div class="box-hdr" id="compTableTitle">Comparativo — ▶ semana → tiendas → productos</div>
        <table class="t" id="tComp">
          <thead id="tCompHead"></thead>
          <tbody id="tCompBody"></tbody>
        </table>
      </div>

      <!-- Columna derecha: gráfica dinámica (oculta) -->
      <div class="box" style="flex:1 1 300px; overflow:visible; min-width:240px; visibility:hidden;">
        <div class="box-hdr" id="compChartTitle">Venta CFBC por Semana</div>
        <div id="compChart" style="padding:8px; overflow-x:auto;"></div>
      </div>
    </div>

    <!-- IDs fantasma para compatibilidad -->
    <div style="display:none">
      <div id="compDrillBox"></div>
      <div id="compDrillTitle"></div>
      <div id="drillTiendaTitle"></div>
      <table><thead id="tDrillTiendaHead"></thead><tbody id="tDrillTiendaBody"></tbody></table>
      <div id="drillChartTienda"></div>
      <div id="drillProdBox"></div>
      <div id="drillProdTitle"></div>
      <table><thead id="tDrillProdHead"></thead><tbody id="tDrillProdBody"></tbody></table>
      <div id="drillChartProd"></div>
      <div id="compChartBox"></div>
      <div id="drillChartTiendaBox"></div>
    </div>
  </div>

  <!-- Vista Inventario Actual -->
  <div class="grid" id="viewInventario" style="display:none">
    <div class="box" style="overflow-x:auto">
      <div class="box-hdr">Inventario Actual — Tiendas</div>
      <table class="t">
        <thead id="tInvTiendaHead"></thead>
        <tbody id="tInvTienda"></tbody>
      </table>
    </div>
    <div class="box" style="overflow-x:auto">
      <div class="box-hdr" id="invProductoTitle">Inventario Actual — Productos</div>
      <table class="t">
        <thead id="tInvProductoHead"></thead>
        <tbody id="tInvProducto"></tbody>
      </table>
    </div>
  </div>

  <!-- Vista GASTO -->
  <div id="viewGasto" style="display:none; padding:12px 20px; overflow:visible !important; height:auto !important;">
    <div id="gastoGrid" style="display:grid; grid-template-columns: 1fr 1fr; gap:12px; overflow:visible !important; height:auto !important;">
      <div class="box" style="overflow:visible !important; height:auto !important;">
        <div class="box-hdr">Presupuesto por Ruta</div>
        <div class="gasto-table-wrap">
          <table class="t" id="tGasto">
            <thead id="tGastoHead"></thead>
            <tbody id="tGastoBody"></tbody>
          </table>
        </div>
      </div>
      
      <!-- Tabla Gasto Gasolina -->
      <div class="box" style="overflow:visible !important; height:auto !important;">
        <div class="box-hdr">Gasto Gasolina y Otros Gastos</div>
        <div class="gasto-table-wrap">
          <table class="t" id="tGasolina">
            <thead id="tGasolinaHead"></thead>
            <tbody id="tGasolinaBody"></tbody>
          </table>
        </div>
      </div>
    </div>
  </div>

  <!-- Vista Choferes -->
  <div id="viewChoferes" style="display:none; padding:12px 20px; font-family:sans-serif;">

    <!-- Sub-panel Consolidado -->
    <div id="subPanelConsolidado" style="display:block;">
    <div style="display:flex; justify-content:space-between; margin-bottom: 20px; gap:20px;">
        <!-- Lado CFBC -->
        <div style="flex:1; background:#fff; border:1px solid #c8d2e8; border-radius:8px; padding:15px; box-shadow:0 1px 3px rgba(0,0,0,0.1);">
            <div style="display:flex; justify-content:space-between; margin-bottom:15px;">
                <div style="border:1px solid #e0e0e0; padding:10px; border-radius:4px; width:48%;">
                    <div style="font-size:10px; color:#666;">UNIDADES CFBC</div>
                    <div id="cfbcUnidades" style="font-size:18px; font-weight:bold;">0</div>
                </div>
                <div style="border:1px solid #e0e0e0; padding:10px; border-radius:4px; width:48%;">
                    <div style="font-size:10px; color:#666;">VENTA CFBC</div>
                    <div id="cfbcVenta" style="font-size:18px; font-weight:bold;">$0.00</div>
                </div>
            </div>
            <h3 style="margin-top:0; font-size:16px; border-bottom:2px solid #2563eb; padding-bottom:5px;">CFBC</h3>
            <div style="margin-bottom:10px; font-size:11px; display:flex; flex-wrap:wrap; gap:10px; align-items:center;">
                <div>Desde: <input type="date" id="cfbcDesde" value="2026-05-01" onchange="renderChoferes()" style="font-size:11px; padding:2px; border:1px solid #ccc;"></div>
                <div>Hasta: <input type="date" id="cfbcHasta" onchange="renderChoferes()" style="font-size:11px; padding:2px; border:1px solid #ccc;"></div>
                <select id="cfbcTienda" onchange="syncFilter(this, 'wmTienda'); renderChoferes()" style="font-size:11px; padding:2px; border:1px solid #ccc;"><option value="ALL">Tienda (Todas)</option></select>
                <select id="cfbcProducto" onchange="syncFilter(this, 'wmProducto'); renderChoferes()" style="font-size:11px; padding:2px; border:1px solid #ccc;"><option value="ALL">Producto (Todos)</option></select>
            </div>
            <div style="height:auto; overflow:visible;">
                <table id="cfbcTable" style="width:100%; border-collapse:collapse; font-size:11px; text-align:left;">
                    <thead style="background:#eef4fb; position:sticky; top:0;">
                        <tr>
                            <th style="padding:5px; border-bottom:1px solid #ddd; width:25px;"></th>
                            <th style="padding:5px; border-bottom:1px solid #ddd; width:25px;"></th>
                            <th style="padding:5px; border-bottom:1px solid #ddd;">Fecha</th>
                            <th style="padding:5px; border-bottom:1px solid #ddd;">Folio</th>
                            <th style="padding:5px; border-bottom:1px solid #ddd;">Tienda</th>
                            <th style="padding:5px; border-bottom:1px solid #ddd;">Producto</th>
                            <th style="padding:5px; text-align:right; border-bottom:1px solid #ddd;">Unidades</th>
                            <th style="padding:5px; text-align:right; border-bottom:1px solid #ddd;">Venta</th>
                            <th style="padding:5px; border-bottom:1px solid #ddd;">Notas</th>
                        </tr>
                    </thead>
                    <tbody></tbody>
                </table>
            </div>
        </div>
        
        <!-- Lado WALMART -->
        <div style="flex:1; background:#fff; border:1px solid #c8d2e8; border-radius:8px; padding:15px; box-shadow:0 1px 3px rgba(0,0,0,0.1);">
            <div style="display:flex; justify-content:space-between; margin-bottom:15px;">
                <div style="border:1px solid #e0e0e0; padding:10px; border-radius:4px; width:48%;">
                    <div style="font-size:10px; color:#666;">UNIDADES WALMART</div>
                    <div id="walmartUnidades" style="font-size:18px; font-weight:bold;">0</div>
                </div>
                <div style="border:1px solid #e0e0e0; padding:10px; border-radius:4px; width:48%;">
                    <div style="font-size:10px; color:#666;">VENTA WALMART</div>
                    <div id="walmartVenta" style="font-size:18px; font-weight:bold;">$0.00</div>
                </div>
            </div>
            <h3 style="margin-top:0; font-size:16px; border-bottom:2px solid #10b981; padding-bottom:5px;">Walmart</h3>
            <div style="margin-bottom:10px; font-size:11px; display:flex; flex-wrap:wrap; gap:10px; align-items:center;">
                <div>Desde: <input type="date" id="wmDesde" value="2026-05-01" onchange="renderChoferes()" style="font-size:11px; padding:2px; border:1px solid #ccc;"></div>
                <div>Hasta: <input type="date" id="wmHasta" onchange="renderChoferes()" style="font-size:11px; padding:2px; border:1px solid #ccc;"></div>
                <select id="wmTienda" onchange="syncFilter(this, 'cfbcTienda'); renderChoferes()" style="font-size:11px; padding:2px; border:1px solid #ccc;"><option value="ALL">Tienda (Todas)</option></select>
                <select id="wmProducto" onchange="syncFilter(this, 'cfbcProducto'); renderChoferes()" style="font-size:11px; padding:2px; border:1px solid #ccc;"><option value="ALL">Producto (Todos)</option></select>
            </div>
            <div style="height:auto; overflow:visible;">
                <table id="walmartTable" style="width:100%; border-collapse:collapse; font-size:11px; text-align:left;">
                    <thead style="background:#eef4fb; position:sticky; top:0;">
                        <tr>
                            <th style="padding:5px; border-bottom:1px solid #ddd; width:25px;"></th>
                            <th style="padding:5px; border-bottom:1px solid #ddd; width:25px;"></th>
                            <th style="padding:5px; border-bottom:1px solid #ddd;">Fecha</th>
                            <th style="padding:5px; border-bottom:1px solid #ddd;">Tienda</th>
                            <th style="padding:5px; border-bottom:1px solid #ddd;">Producto</th>
                            <th style="padding:5px; text-align:right; border-bottom:1px solid #ddd;">Unidades</th>
                            <th style="padding:5px; text-align:right; border-bottom:1px solid #ddd;">Venta</th>
                            <th style="padding:5px; border-bottom:1px solid #ddd;">Notas</th>
                        </tr>
                    </thead>
                    <tbody></tbody>
                </table>
            </div>
        </div>
    </div>
    </div><!-- end subPanelConsolidado -->

    <!-- Sub-panel Devoluciones -->
    <div id="subPanelDevoluciones" style="display:none;">
      <div style="background:#fff; border:1px solid #c8d2e8; border-radius:8px; padding:18px; box-shadow:0 1px 3px rgba(0,0,0,0.1);">
        <!-- Filtros -->
        <div style="margin-bottom:12px; font-size:11px; display:flex; flex-wrap:wrap; gap:10px; align-items:center;">
          <div>Desde: <input type="date" id="devDesde" onchange="renderDevoluciones()" style="font-size:11px; padding:2px; border:1px solid #ccc;"></div>
          <div>Hasta: <input type="date" id="devHasta" onchange="renderDevoluciones()" style="font-size:11px; padding:2px; border:1px solid #ccc;"></div>
          <div>Folio: <input type="text" id="devFolioFilter" oninput="renderDevoluciones()" placeholder="Buscar folio..." style="font-size:11px; padding:2px; border:1px solid #ccc; width:100px;"></div>
          <div>Producto: <input type="text" id="devProductoFilter" oninput="renderDevoluciones()" placeholder="Buscar producto..." style="font-size:11px; padding:2px; border:1px solid #ccc; width:130px;"></div>
        </div>
        <!-- Tabla -->
        <div style="overflow-x:auto;">
          <table id="devolucionesTable" style="width:100%; border-collapse:collapse; font-size:12px; text-align:left;">
            <thead style="background:#fff;">
              <tr>
                <th style="padding:8px 6px; border-bottom:2px solid #dc2626; color:#dc2626;">Fecha</th>
                <th style="padding:8px 6px; border-bottom:2px solid #dc2626; color:#dc2626;">Folio</th>
                <th style="padding:8px 6px; border-bottom:2px solid #dc2626; color:#dc2626;">Serie</th>
                <th style="padding:8px 6px; border-bottom:2px solid #dc2626; color:#dc2626;">Producto</th>
                <th style="padding:8px 6px; border-bottom:2px solid #dc2626; color:#dc2626; text-align:right;">Cant. Devuelta</th>
                <th style="padding:8px 6px; border-bottom:2px solid #dc2626; color:#dc2626; text-align:right;">Precio Unitario</th>
                <th style="padding:8px 6px; border-bottom:2px solid #dc2626; color:#dc2626; text-align:right;">Total Devolucion</th>
                <th style="padding:8px 6px; border-bottom:2px solid #dc2626; color:#dc2626; text-align:center;">Motivo</th>
              </tr>
            </thead>
            <tbody id="devolucionesTbody"></tbody>
          </table>
        </div>
      </div>
    </div><!-- end subPanelDevoluciones -->

  </div><!-- end viewChoferes -->

  
<!-- Vista Resumen -->
  <div id="viewResumen" style="display:none; padding:12px 20px; overflow:visible; height:auto !important;">
    <div class="resumen-layout">
      <div class="box resumen-main-box" style="height:auto !important; min-width:0; overflow:visible;">
        <div class="box-hdr" id="resumenTitle" style="display:flex; justify-content:space-between; align-items:center;">
          <span id="resumenTitleText" style="font-size:1.6em; font-weight:bold; color:#1a3a5c;">Resumen Diario</span>
          <div class="mode-btn-group" style="display:flex; gap:5px; align-items:center;">
            <button id="btnAddCaptureSem" onclick="window.addCaptureRowAll()" title="Agregar semana de captura" style="display:inline-flex;align-items:center;gap:4px;font-size:12px;font-weight:600;background:#1565c0;color:#fff;border:none;border-radius:6px;padding:4px 11px;cursor:pointer;letter-spacing:.02em;"><span class="material-icons" style="font-size:16px">add_circle</span> SEM</button>
            <button id="btnRemoveCaptureSem" onclick="window.removeCaptureRowAll()" title="Eliminar última semana de captura" style="display:inline-flex;align-items:center;gap:4px;font-size:12px;font-weight:600;background:#c62828;color:#fff;border:none;border-radius:6px;padding:4px 11px;cursor:pointer;opacity:0.45;letter-spacing:.02em;" disabled><span class="material-icons" style="font-size:16px">remove_circle</span> SEM</button>
            <button id="btnSaveCaptureAll" onclick="window.saveAllCaptureRows()" title="Guardar todas las capturas visibles" style="display:inline-flex;align-items:center;justify-content:center;font-size:12px;font-weight:600;background:#2e7d32;color:#fff;border:none;border-radius:6px;padding:4px 7px;cursor:pointer;letter-spacing:.02em;"><span class="material-icons" style="font-size:16px">save</span></button>
            <button id="btnClearCapture" onclick="_showDeleteWeekDialog()" title="Borrar capturas guardadas (afecta a todos los usuarios)" style="display:inline-flex;align-items:center;justify-content:center;font-size:12px;font-weight:600;background:#fff;color:#c62828;border:1px solid #c62828;border-radius:6px;padding:4px 7px;cursor:pointer;letter-spacing:.02em;"><span class="material-icons" style="font-size:16px">delete_sweep</span></button>
            <button id="btnToggleTotalPos" onclick="if(typeof window.toggleTotalPosition === 'function') window.toggleTotalPosition();" title="Mover Total General Arriba/Abajo" style="display:inline-flex;align-items:center;gap:4px;font-size:12px;font-weight:600;background:#eef2fa;color:#1a3a5c;border:1px solid #c8d2e8;border-radius:6px;padding:4px 11px;cursor:pointer;letter-spacing:.02em;"><span class="material-icons" style="font-size:16px">swap_vert</span> Total</button>
            <div id="metricFilters" style="display:inline-flex; align-items:center; gap:8px; background:#eef2fa; padding:4px 8px; border-radius:6px; border:1px solid #c8d2e8; font-size:12px; font-weight:600; color:#1a3a5c;">
              <label style="cursor:pointer; display:flex; align-items:center; gap:2px;" title="Embarque"><input type="checkbox" value="embarque" checked onclick="if(document.querySelectorAll('#metricFilters input:checked').length===0) return false; if(typeof renderResumen==='function') renderResumen()"> Emb</label>
              <label style="cursor:pointer; display:flex; align-items:center; gap:2px;" title="Ventas POS"><input type="checkbox" value="ventas" checked onclick="if(document.querySelectorAll('#metricFilters input:checked').length===0) return false; if(typeof renderResumen==='function') renderResumen()"> POS</label>
              <label style="cursor:pointer; display:flex; align-items:center; gap:2px;" title="Sell Thru"><input type="checkbox" value="sell_thru" checked onclick="if(document.querySelectorAll('#metricFilters input:checked').length===0) return false; if(typeof renderResumen==='function') renderResumen()"> Thru</label>
            </div>
            <button id="btnExportCapture" onclick="window.exportCaptureData()" title="Descargar captura a CSV" style="display:inline-flex;align-items:center;gap:4px;font-size:12px;font-weight:600;background:#107c41;color:#fff;border:none;border-radius:6px;padding:4px 11px;cursor:pointer;letter-spacing:.02em;"><span class="material-icons" style="font-size:16px">download</span> Excel</button>
            <button class="mode-btn active" id="btnPivotProducto" onclick="setResumenPivot('producto')">Producto</button>
            <button class="mode-btn" id="btnPivotTienda" onclick="setResumenPivot('tienda')">Tienda</button>
          </div>
        </div>
        <div class="resumen-table-wrap">
          <table class="t" id="tResumen">
            <thead id="tResumenHead"></thead>
            <tbody id="tResumenBody"></tbody>
          </table>
        </div>
      </div>
      <div class="box resumen-side-box" style="display:none;">
        <div class="box-hdr" style="display:flex; justify-content:space-between; align-items:center;">
          <span>Ordenes de Reabasto</span>
          <div style="display:flex; gap:10px; align-items:center;">
            <select id="reabastoWindowSelect" onchange="setReabastoWindow(this.value)" style="font-size:0.75rem; padding:2px 5px; border-radius:3px; border:1px solid #ccc; background:#fff;">
              <option value="1">1 Semana</option>
              <option value="2">2 Semanas</option>
              <option value="3" selected>3 Semanas</option>
            </select>
            <button onclick="exportReabastoExcel()" style="font-size:0.75rem; padding:3px 8px; background:#107c41; color:white; border:none; border-radius:3px; cursor:pointer; font-weight:bold;">📥 Excel</button>
          </div>
        </div>
        <div class="reabasto-panel">
          <div class="reabasto-meta" id="reabastoMeta">Proyección de reabasto dinámico basado en las semanas seleccionadas.</div>
          <div class="reabasto-table-wrap">
            <table class="reabasto-table">
              <thead>
                <tr>
                  <th>Prioridad</th>
                  <th>Ruta</th>
                  <th>Tienda</th>
                  <th>Producto</th>
                  <th>Inv.</th>
                  <th>Prom/dia</th>
                  <th>Cob.</th>
                  <th>Entrega</th>
                  <th>Dias ent.</th>
                  <th>Seg.</th>
                  <th>Mult.</th>
                  <th>Orden</th>
                  <th>Motivo</th>
                </tr>
              </thead>
              <tbody id="tReabastoBody"></tbody>
            </table>
          </div>
        </div>
      </div>
      <!-- Chat IA Box -->
      <div class="box ai-chat-box" style="position:sticky; top:20px; flex:0 0 350px; width:350px; height:600px !important; max-height:600px !important; display:flex; flex-direction:column; border-radius:8px; box-shadow:0 4px 12px rgba(0,0,0,0.15); background:#fff; margin-left:auto; overflow:hidden;">
        <div class="box-hdr" style="display:flex; justify-content:space-between; align-items:center; background:#1a3a5c; color:#fff; border-bottom:none; padding:10px 14px;">
          <div style="display:flex; align-items:center; gap:8px;">
            <span class="material-icons" style="font-size:18px;">auto_awesome</span>
            <span style="font-size:0.9rem; font-weight:700;">Asistente Walmex</span>
          </div>
          <button title="Cambiar API Key" onclick="localStorage.removeItem('GROQ_API_KEY'); alert('API Key borrada. Envia un mensaje para ingresar una nueva.');" style="background:transparent; border:none; color:#fff; cursor:pointer;"><span class="material-icons" style="font-size:16px;">key</span></button>
        </div>
        <div id="aiChatMessages" style="flex:1; padding:14px; overflow-y:auto; background:#f8fafc; display:flex; flex-direction:column; gap:12px; font-size:0.8rem;">
          <div style="display:flex; flex-direction:column; gap:4px; max-width:85%; align-self:flex-start;">
            <div style="background:#e2e8f0; color:#334155; padding:8px 12px; border-radius:12px; border-top-left-radius:2px;">
              ¡Hola! Soy tu asistente de datos. Pregúntame sobre la <strong>Tabla Resumen</strong> que estás viendo. (Ej. "¿Cuál es la mayor venta?" o "¿Qué tienda vendió más?").
            </div>
          </div>
        </div>
        <div style="padding:10px; background:#fff; border-top:1px solid #e2e8f0; display:flex; gap:8px; align-items:center;">
          <input type="text" id="aiChatInput" placeholder="Haz una pregunta..." style="flex:1; padding:8px 12px; border:1px solid #cbd5e1; border-radius:20px; outline:none; font-size:0.8rem; background:#fff; color:#334155;" onkeypress="if(event.key === 'Enter') sendChatMessage()">
          <button onclick="sendChatMessage()" style="background:#1565c0; color:#fff; border:none; width:32px; height:32px; border-radius:50%; display:flex; align-items:center; justify-content:center; cursor:pointer; flex-shrink:0;">
            <span class="material-icons" style="font-size:16px;">send</span>
          </button>
        </div>
      </div>
    </div>
  </div>
</div>

<script>
window.syncFilter = function(el, targetId) {
    var targetEl = document.getElementById(targetId);
    if (targetEl && targetEl.value !== el.value) {
        targetEl.value = el.value;
    }
};
var checkedCFBC = new Set();
var checkedWalmart = new Set();
var flaggedCFBC = new Set();
var flaggedWalmart = new Set();
var notesCFBC = {};
var notesWalmart = {};
try {
    if(localStorage.getItem('checkedCFBC')) checkedCFBC = new Set(JSON.parse(localStorage.getItem('checkedCFBC')));
    if(localStorage.getItem('checkedWalmart')) checkedWalmart = new Set(JSON.parse(localStorage.getItem('checkedWalmart')));
    if(localStorage.getItem('flaggedCFBC')) flaggedCFBC = new Set(JSON.parse(localStorage.getItem('flaggedCFBC')));
    if(localStorage.getItem('flaggedWalmart')) flaggedWalmart = new Set(JSON.parse(localStorage.getItem('flaggedWalmart')));
    if(localStorage.getItem('notesCFBC')) notesCFBC = JSON.parse(localStorage.getItem('notesCFBC'));
    if(localStorage.getItem('notesWalmart')) notesWalmart = JSON.parse(localStorage.getItem('notesWalmart'));
} catch(e) {}

var SUPABASE_UI_ROW_ID = 'choferes_ui_state';

window.saveChecks = function() {
    try {
        localStorage.setItem('checkedCFBC', JSON.stringify([...checkedCFBC]));
        localStorage.setItem('checkedWalmart', JSON.stringify([...checkedWalmart]));
        localStorage.setItem('flaggedCFBC', JSON.stringify([...flaggedCFBC]));
        localStorage.setItem('flaggedWalmart', JSON.stringify([...flaggedWalmart]));
        localStorage.setItem('notesCFBC', JSON.stringify(notesCFBC));
        localStorage.setItem('notesWalmart', JSON.stringify(notesWalmart));
    } catch (e) {}
    
    // Sync to Supabase
    if (typeof SUPABASE_URL !== 'undefined' && SUPABASE_URL) {
        var payload = {
            checkedCFBC: [...checkedCFBC],
            checkedWalmart: [...checkedWalmart],
            flaggedCFBC: [...flaggedCFBC],
            flaggedWalmart: [...flaggedWalmart],
            notesCFBC: notesCFBC,
            notesWalmart: notesWalmart
        };
        fetch(SUPABASE_URL + '/rest/v1/walmex_resumen_captura?on_conflict=id', {
            method: 'POST',
            headers: {
                'apikey': SUPABASE_KEY,
                'Authorization': 'Bearer ' + SUPABASE_KEY,
                'Content-Type': 'application/json',
                'Prefer': 'resolution=merge-duplicates,return=minimal'
            },
            body: JSON.stringify({
                id: SUPABASE_UI_ROW_ID,
                data: payload,
                updated_at: new Date().toISOString()
            })
        }).catch(function(e) { console.error("Error saving UI state to Supabase", e); });
    }
};

window.loadUIStateFromSupabase = function() {
    if (typeof SUPABASE_URL !== 'undefined' && SUPABASE_URL) {
        fetch(SUPABASE_URL + '/rest/v1/walmex_resumen_captura?id=eq.' + encodeURIComponent(SUPABASE_UI_ROW_ID) + '&select=data', {
            headers: {
                'apikey': SUPABASE_KEY,
                'Authorization': 'Bearer ' + SUPABASE_KEY
            }
        })
        .then(function(res) { return res.ok ? res.json() : []; })
        .then(function(rows) {
            if (rows && rows[0] && rows[0].data) {
                var d = rows[0].data;
                if (d.checkedCFBC) checkedCFBC = new Set(d.checkedCFBC);
                if (d.checkedWalmart) checkedWalmart = new Set(d.checkedWalmart);
                if (d.flaggedCFBC) flaggedCFBC = new Set(d.flaggedCFBC);
                if (d.flaggedWalmart) flaggedWalmart = new Set(d.flaggedWalmart);
                if (d.notesCFBC) notesCFBC = d.notesCFBC;
                if (d.notesWalmart) notesWalmart = d.notesWalmart;
                
                // Save locally to keep it in sync
                try {
                    localStorage.setItem('checkedCFBC', JSON.stringify([...checkedCFBC]));
                    localStorage.setItem('checkedWalmart', JSON.stringify([...checkedWalmart]));
                    localStorage.setItem('flaggedCFBC', JSON.stringify([...flaggedCFBC]));
                    localStorage.setItem('flaggedWalmart', JSON.stringify([...flaggedWalmart]));
                    localStorage.setItem('notesCFBC', JSON.stringify(notesCFBC));
                    localStorage.setItem('notesWalmart', JSON.stringify(notesWalmart));
                } catch(e) {}
                
                // Re-render if view is currently Choferes
                if (state.view === 'choferes' && typeof renderChoferes === 'function') {
                    renderChoferes();
                }
            }
        })
        .catch(function(e) { console.error("Error loading UI state from Supabase", e); });
    }
};

// Initiate load right away
setTimeout(window.loadUIStateFromSupabase, 500);

window.toggleFlag = function(btn, key, source) {
    var row = btn.closest('tr');
    var set = source === 'cfbc' ? flaggedCFBC : flaggedWalmart;
    if (set.has(key)) {
        set.delete(key);
        row.classList.remove('flagged');
    } else {
        set.add(key);
        row.classList.add('flagged');
    }
    window.saveChecks();
};

window.applyCheck = function(row, value, key, source) {
    if (!key) key = row.dataset.key;
    if (!source) source = row.dataset.source;
    var cb = row.querySelector('input[type="checkbox"]');
    if (cb) cb.checked = value;
    if (value) {
        row.classList.add('checked');
        // Quitar estado 'Procesando' visual
        row.classList.remove('flagged');
        var lbl = row.querySelector('.procesando-lbl');
        if(lbl) lbl.style.display = 'none';
        
        if (source === 'cfbc') checkedCFBC.add(key);
        else checkedWalmart.add(key);
    } else {
        row.classList.remove('checked');
        // Restaurar estado 'Procesando' visual
        row.classList.add('flagged');
        var lbl = row.querySelector('.procesando-lbl');
        if(lbl) lbl.style.display = 'block';
        
        if (source === 'cfbc') checkedCFBC.delete(key);
        else checkedWalmart.delete(key);
    }
};

var isDragging = false;
var dragSource = null;
var dragValue = false;

document.addEventListener('mousedown', function(e) {
    var cb = e.target.closest('input[type="checkbox"]');
    if (!cb) return;
    var row = cb.closest('tr');
    if (!row || !row.dataset.source) return;
    isDragging = true;
    dragSource = row.dataset.source;
    dragValue = !cb.checked;
    document.body.classList.add('dragging-checks');
    window.applyCheck(row, dragValue);
    window.saveChecks();
    e.preventDefault();
});

document.addEventListener('mousemove', function(e) {
    if (!isDragging) return;
    var el = document.elementFromPoint(e.clientX, e.clientY);
    if (!el) return;
    var row = el.closest('tr');
    if (!row || row.dataset.source !== dragSource) return;
    window.applyCheck(row, dragValue);
});

document.addEventListener('mouseup', function() {
    if (isDragging) {
        isDragging = false;
        dragSource = null;
        document.body.classList.remove('dragging-checks');
        window.saveChecks();
    }
});

window.toggleRowCheck = function(checkbox, key, source) {
    var row = checkbox.closest('tr');
    window.applyCheck(row, checkbox.checked, key, source);
    window.saveChecks();
};

window.saveNota = function(key, value, source) {
    var map = source === 'cfbc' ? notesCFBC : notesWalmart;
    if (value.trim()) {
        map[key] = value.trim();
    } else {
        delete map[key];
    }
    window.saveChecks();
};

var DATA = JSON.parse(atob('__DATA_JSON__'));
var SUPABASE_DATA = null;
try {
    SUPABASE_DATA = JSON.parse(atob('__SUPABASE_JSON__'));
} catch(e) { console.error("Error parsing supabase JSON:", e); }

var DEVOLUCIONES_DATA = [];
try {
    DEVOLUCIONES_DATA = JSON.parse(atob('__DEVOLUCIONES_JSON__'));
} catch(e) { console.error("Error parsing devoluciones JSON:", e); }

// ── Sub-tabs Choferes ──
function toggleChoferesMenu(e) {
    e.stopPropagation();
    var menu = document.getElementById('choferesMenu');
    menu.style.display = (menu.style.display === 'none' || menu.style.display === '') ? 'block' : 'none';
}

// Cerrar el menú al hacer clic fuera
document.addEventListener('click', function() {
    var menu = document.getElementById('choferesMenu');
    if (menu) menu.style.display = 'none';
});

function openChoferesSubTab(tab) {
    // Cerrar el menú
    var menu = document.getElementById('choferesMenu');
    if (menu) menu.style.display = 'none';

    // Activar la vista choferes
    setView('choferes');

    // Mostrar el sub-panel correcto
    var panelCons = document.getElementById('subPanelConsolidado');
    var panelDev  = document.getElementById('subPanelDevoluciones');
    if (tab === 'consolidado') {
        if (panelCons) panelCons.style.display = 'block';
        if (panelDev)  panelDev.style.display  = 'none';
    } else {
        if (panelCons) panelCons.style.display = 'none';
        if (panelDev)  panelDev.style.display  = 'block';
        renderDevoluciones();
    }
}

function renderDevoluciones() {
    var tbody = document.getElementById('devolucionesTbody');
    if (!tbody) return;
    var data = DEVOLUCIONES_DATA || [];
    var desde   = (document.getElementById('devDesde')   || {}).value || '';
    var hasta   = (document.getElementById('devHasta')   || {}).value || '';
    var folio   = ((document.getElementById('devFolioFilter')   || {}).value || '').toLowerCase().trim();
    var prod    = ((document.getElementById('devProductoFilter') || {}).value || '').toLowerCase().trim();

    var totalCount = 0, totalUnid = 0, totalMonto = 0;
    var rows = '';
    data.forEach(function(r) {
        var fechaRaw = (r.created_at || '').substring(0, 10);
        if (desde && fechaRaw < desde) return;
        if (hasta && fechaRaw > hasta) return;
        if (folio && (r.folio || '').toLowerCase().indexOf(folio) === -1) return;
        if (prod  && (r.producto || '').toLowerCase().indexOf(prod) === -1) return;
        var cant  = parseFloat(r.cantidad_devuelta || 0);
        var precio = parseFloat(r.precio_unidad || 0);
        var total = parseFloat(r.total_devolucion || 0);
        totalCount++;
        totalUnid  += cant;
        totalMonto += total;
        var fechaDisp = fechaRaw || 'N/A';
        rows += '<tr style="border-bottom:1px solid #fee2e2;">' +
            '<td style="padding:6px;">'+fechaDisp+'</td>' +
            '<td style="padding:6px; font-weight:600;">'+( r.folio || '-')+'</td>' +
            '<td style="padding:6px;">'+( r.serie || '-')+'</td>' +
            '<td style="padding:6px;">'+( r.producto || '-')+'</td>' +
            '<td style="padding:6px; text-align:right; color:#d97706; font-weight:700;">'+cant.toLocaleString('en-US')+'</td>' +
            '<td style="padding:6px; text-align:right;">$'+precio.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2})+'</td>' +
            '<td style="padding:6px; text-align:right; color:#dc2626; font-weight:700;">$'+total.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2})+'</td>' +
            '<td style="padding:6px; text-align:center; font-style:italic;">'+(r.razon_devolucion || '-')+'</td>' +
            '</tr>';
    });
    tbody.innerHTML = rows || '<tr><td colspan="8" style="text-align:center; padding:20px; color:#aaa;">Sin devoluciones en el rango seleccionado</td></tr>';
    var el; 
    el = document.getElementById('devTotalCount');   if(el) el.textContent = totalCount;
    el = document.getElementById('devTotalUnidades'); if(el) el.textContent = totalUnid.toLocaleString('en-US');
    el = document.getElementById('devTotalMonto');   if(el) el.textContent = '$'+totalMonto.toLocaleString('en-US',{minimumFractionDigits:2,maximumFractionDigits:2});
}
var state = { semana: null, semanas_sel: null, tienda: null, tiendas_sel: null, producto: null, productos_sel: null, view: 'producto', tiendaT: null, invMode: null, invSelected: null, compMode: 'semanas', drillSem: null, drillTienda: null, resumenMode: 'semanas', resumenPivot: 'tienda', reabastoWindow: 3 };
var DIAS  = ['domingo','lunes','martes','miércoles','jueves','viernes','sábado'];
var MESES = ['enero','febrero','marzo','abril','mayo','junio','julio','agosto','septiembre','octubre','noviembre','diciembre'];

// ====== SUPABASE (captura compartida "Programado") ======
// Solo se usa la publishable_key (clave publica), nunca una clave secreta.
var SUPABASE_URL = '__SUPABASE_URL__';
var SUPABASE_KEY = '__SUPABASE_PUBLISHABLE_KEY__';
var SUPABASE_CAPTURE_TABLE = 'walmex_resumen_captura_v2';
var SUPABASE_CAPTURE_ROW_ID = 'captura_global';

function _supabaseHeaders(extra) {
  var h = {
    'apikey': SUPABASE_KEY,
    'Authorization': 'Bearer ' + SUPABASE_KEY,
    'Content-Type': 'application/json'
  };
  if (extra) { for (var k in extra) { h[k] = extra[k]; } }
  return h;
}

function _supabaseConfigured() {
  return !!(SUPABASE_URL && SUPABASE_KEY && SUPABASE_URL.indexOf('http') === 0);
}


// ====== COMENTARIOS DINAMICOS ======
var walmexComments = JSON.parse(localStorage.getItem('WALMEX_COMMENTS') || '{}');
var currentCommentCellId = null;

function openCommentModal(cellId) {
  currentCommentCellId = cellId;
  var comm = walmexComments[cellId];
  document.getElementById('txtComment').value = comm ? (typeof comm === 'string' ? comm : comm.text) : '';
  document.getElementById('commentOverlay').style.display = 'block';
  document.getElementById('commentModal').style.display = 'block';
  document.getElementById('txtComment').focus();
}

function closeCommentModal() {
  document.getElementById('commentOverlay').style.display = 'none';
  document.getElementById('commentModal').style.display = 'none';
  currentCommentCellId = null;
}

function saveComment() {
  if(!currentCommentCellId) return;
  var text = document.getElementById('txtComment').value.trim();
  if(text) {
    var now = new Date();
    var dStr = now.toLocaleDateString('es-MX');
    var tStr = now.toLocaleTimeString('es-MX', {hour: '2-digit', minute:'2-digit'});
    walmexComments[currentCommentCellId] = { text: text, date: dStr, time: tStr };
  } else {
    delete walmexComments[currentCommentCellId];
  }
  localStorage.setItem('WALMEX_COMMENTS', JSON.stringify(walmexComments));
  closeCommentModal();
  renderResumen(); // re-render para mostrar/ocultar el indicador
}

function deleteComment() {
  if(!currentCommentCellId) return;
  delete walmexComments[currentCommentCellId];
  localStorage.setItem('WALMEX_COMMENTS', JSON.stringify(walmexComments));
  closeCommentModal();
  renderResumen();
}

function getCommentAttr(cId) {
  var comm = walmexComments[cId];
  var onClickStr = ' onclick="openCommentModal(\'' + cId + '\')"';
  if(!comm) return { cls: '', str: onClickStr };
  var cText = typeof comm === 'string' ? comm : comm.text;
  var cDate = typeof comm === 'string' ? '' : ' (' + comm.date + ' ' + comm.time + ')';
  var titleText = cText.replace(/"/g, '&quot;') + cDate;
  return { cls: ' cell-comment', str: onClickStr + ' title="' + titleText + '"' };
}
// ===================================

function fmt(v){ return Math.round(v||0).toLocaleString('es-MX'); }

function toggleTodasSemanas(){
  var chks = document.querySelectorAll('#semDropMenu input[type=checkbox].sem-chk');
  var chkAll = document.getElementById('chkTodasSem');
  var allChecked = chkAll.checked;
  chks.forEach(function(c){
    c.checked = allChecked;
    var s = parseInt(c.value);
    var row = document.getElementById('sem-row-'+s);
    if(row) row.className = 'sem-item' + (allChecked ? ' on' : '');
  });
  onSemChk();
}

function syncChkTodas(){
  var chks = document.querySelectorAll('#semDropMenu input[type=checkbox].sem-chk');
  var chkAll = document.getElementById('chkTodasSem');
  if(!chkAll) return;
  var total = chks.length, checked = 0;
  chks.forEach(function(c){ if(c.checked) checked++; });
  chkAll.checked = (checked === total);
  chkAll.indeterminate = (checked > 0 && checked < total);
}

function init(){
  /* ── Restaurar capturas guardadas antes de cualquier render ── */
  /* Se dispara en paralelo (no bloquea el resto de init); cuando llegue la
     respuesta de Supabase, si el usuario ya está en Resumen, se re-renderiza. */
  _loadPersistedCapture();
  /* Refresco periódico de la captura compartida mientras se está en Resumen,
     para que los cambios guardados por otras personas aparezcan sin recargar. */
  setInterval(function () {
    if (typeof state !== 'undefined' && state.view === 'resumen') {
      _loadPersistedCapture();
    }
  }, 30000);
  window.onerror = function(m,s,l){
    var msg = String(m || '');
    if(msg.indexOf('ResizeObserver loop completed with undelivered notifications') >= 0 ||
       msg.indexOf('ResizeObserver loop limit exceeded') >= 0){
      return true;
    }
    document.body.innerHTML='<p style="padding:20px;color:red">Error: '+m+' (línea '+l+')</p>';
  };
  var menu = document.getElementById('semDropMenu');

  // ── Opción "Seleccionar todas" ──
  var rowAll = document.createElement('label');
  rowAll.id = 'sem-row-all';
  rowAll.style.cssText = 'display:flex;align-items:center;gap:6px;padding:5px 10px;cursor:pointer;font-weight:700;border-bottom:1px solid #ddd;background:#f5f5f5;font-size:.72rem';
  var chkAll = document.createElement('input');
  chkAll.type = 'checkbox';
  chkAll.id = 'chkTodasSem';
  chkAll.onchange = function(){ toggleTodasSemanas(); };
  rowAll.appendChild(chkAll);
  rowAll.appendChild(document.createTextNode('Seleccionar todas'));
  menu.appendChild(rowAll);

  // Filtrar semanas sin año si ya existe la versión con año del mismo número de semana
  var semanasConAnio = DATA.semanas.filter(function(s){ return s > 9999; });
  var numsSemConAnio = semanasConAnio.map(function(s){ return s % 100; });
  var semanasRender  = DATA.semanas.filter(function(s){
    if(s > 9999) return true;                       // siempre incluir las que tienen año
    return numsSemConAnio.indexOf(s) === -1;         // bare solo si no hay duplicado con año
  });

  semanasRender.forEach(function(s){
    var yr = Math.floor(s/100), wk = s%100;
    var labelTxt = (yr >= 2000) ? yr+' · Semana '+String(wk).padStart(2,'0') : 'Semana '+String(s).padStart(2,'0');
    var isLast = (s === DATA.semanas[DATA.semanas.length-1]);
    var row = document.createElement('label');
    row.className = 'sem-item' + (isLast ? ' on' : '');
    row.id = 'sem-row-'+s;
    var chk = document.createElement('input');
    chk.type = 'checkbox';
    chk.className = 'sem-chk';
    chk.value = s;
    chk.checked = isLast;
    chk.onchange = function(){ onSemChk(); };
    row.appendChild(chk);
    row.appendChild(document.createTextNode(labelTxt));
    menu.appendChild(row);
  });
  // Cerrar dropdown al clicar fuera
  document.addEventListener('click', function(e){
    var wrap = document.getElementById('semDropWrap');
    if(wrap && !wrap.contains(e.target)) closeSemDrop();
    var wrapT = document.getElementById('tiendaDropWrap');
    if(wrapT && !wrapT.contains(e.target)) closeTiendaDrop();
    var wrapP = document.getElementById('productoDropWrap');
    if(wrapP && !wrapP.contains(e.target)) closeProductoDrop();
  });
  
  // ── Crear dropdown de tiendas ──
  var menuT = document.getElementById('tiendaDropMenu');
  
  // Opción "Seleccionar todas" para tiendas
  var rowAllT = document.createElement('label');
  rowAllT.id = 'tienda-row-all';
  rowAllT.style.cssText = 'display:flex;align-items:center;gap:6px;padding:5px 10px;cursor:pointer;font-weight:700;border-bottom:1px solid #ddd;background:#f5f5f5;font-size:.72rem';
  var chkAllT = document.createElement('input');
  chkAllT.type = 'checkbox';
  chkAllT.id = 'chkTodasTienda';
  chkAllT.onchange = function(){ toggleTodasTiendas(); };
  rowAllT.appendChild(chkAllT);
  rowAllT.appendChild(document.createTextNode('Seleccionar todas'));
  menuT.appendChild(rowAllT);
  
  DATA.tiendas.forEach(function(t){
    var labelTxt = t.replace('SC ','');
    var isFirst = (t === DATA.tiendas[0]);
    var rowT = document.createElement('label');
    rowT.className = 'sem-item' + (isFirst ? ' on' : '');
    rowT.id = 'tienda-row-'+t;
    var chkT = document.createElement('input');
    chkT.type = 'checkbox';
    chkT.className = 'tienda-chk';
    chkT.value = t;
    chkT.checked = isFirst;
    chkT.onchange = function(){ onTiendaChk(); };
    rowT.appendChild(chkT);
    rowT.appendChild(document.createTextNode(labelTxt));
    menuT.appendChild(rowT);
  });
  
  // ── Crear dropdown de productos ──
  var menuP = document.getElementById('productoDropMenu');
  
  // Opción "Seleccionar todos" para productos
  var rowAllP = document.createElement('label');
  rowAllP.id = 'producto-row-all';
  rowAllP.style.cssText = 'display:flex;align-items:center;gap:6px;padding:5px 10px;cursor:pointer;font-weight:700;border-bottom:1px solid #ddd;background:#f5f5f5;font-size:.72rem';
  var chkAllP = document.createElement('input');
  chkAllP.type = 'checkbox';
  chkAllP.id = 'chkTodosProducto';
  chkAllP.onchange = function(){ toggleTodosProductos(); };
  rowAllP.appendChild(chkAllP);
  rowAllP.appendChild(document.createTextNode('Seleccionar todos'));
  menuP.appendChild(rowAllP);
  
  DATA.productos.forEach(function(p){
    var labelTxt = p.replace('BQT ','');
    var rowP = document.createElement('label');
    rowP.className = 'sem-item on';
    rowP.id = 'producto-row-'+p.replace(/\s+/g, '');
    var chkP = document.createElement('input');
    chkP.type = 'checkbox';
    chkP.className = 'producto-chk';
    chkP.value = p;
    chkP.checked = true;
    chkP.onchange = function(){ onProductoChk(); };
    rowP.appendChild(chkP);
    rowP.appendChild(document.createTextNode(labelTxt));
    menuP.appendChild(rowP);
  });
  state.productos_sel = DATA.productos.slice();

  state.semana = DATA.semanas[DATA.semanas.length-1];
  state.semanas_sel = [state.semana];
  state.tienda = DATA.tiendas[0];
  state.tiendas_sel = [DATA.tiendas[0]];
  updateResumenModeButtons();
  updateHeader(); updateSemLabel(); updateTiendaLabel(); render();
  document.getElementById('loader').style.display = 'none';
  document.getElementById('app').style.display    = 'block';
}

function toggleSemDrop(){
  toggleFloatingMenu('semDropMenu', 'semDropBtn');
}
function closeSemDrop(){
  closeFloatingMenu('semDropMenu');
}

function toggleTiendaDrop(){
  toggleFloatingMenu('tiendaDropMenu', 'tiendaDropBtn');
}
function closeTiendaDrop(){
  closeFloatingMenu('tiendaDropMenu');
}

function areAllTiendasSelected(sel){
  return Array.isArray(sel) && sel.length === DATA.tiendas.length;
}

function applyTiendaSelection(selected){
  state.tiendas_sel = Array.isArray(selected) ? selected.slice() : [];
  state.tienda = state.tiendas_sel.length > 0 ? state.tiendas_sel[0] : DATA.tiendas[0];
  state.tiendaT = null;
  updateTiendaLabel();
  updateHeader();
  syncChkTodasTiendas();
  if(state.view==='producto') render();
  else if(state.view==='tienda') renderTienda();
  else if(state.view==='inventario') renderInventario();
  else if(state.view==='comparativo') renderComparativo();
  else if(state.view==='gasto') renderGasto();
  else if(state.view==='resumen') renderResumen();
  else if(state.view==='choferes' && typeof renderChoferes === 'function') renderChoferes();
}

function updateTiendaLabel(){
  var sel = state.tiendas_sel;
  var lbl = document.getElementById('tiendaDropLabel');
  if(areAllTiendasSelected(sel)){
    lbl.textContent = '— Todas las tiendas —';
  } else if(!sel || sel.length === 0){
    lbl.textContent = '— Sin tiendas —';
  } else if(sel.length === 1){
    lbl.textContent = sel[0].replace('SC ','');
  } else {
    lbl.textContent = sel.length+' tiendas seleccionadas';
  }
}
function onRutaChange(ruta) {
  var chks = document.querySelectorAll('#tiendaDropMenu input[type=checkbox].tienda-chk');
  chks.forEach(function(chk) {
    var t = chk.value;
    if (ruta === 'Todas') {
      chk.checked = true;
    } else {
      var tiendaRuta = (typeof DATA !== 'undefined' && DATA.tienda_ruta && DATA.tienda_ruta[t]) || 'Sin ruta';
      chk.checked = (tiendaRuta === ruta);
    }
  });
  onTiendaChk();
}

function onTiendaChk(){
  var chks = document.querySelectorAll('#tiendaDropMenu input[type=checkbox].tienda-chk');
  var selected = [];
  chks.forEach(function(c){
    var t = c.value;
    var row = document.getElementById('tienda-row-'+t);
    if(c.checked){
      selected.push(t);
      if(row) row.className = 'sem-item on';
    } else {
      if(row) row.className = 'sem-item';
    }
  });
  applyTiendaSelection(selected);
}

function toggleTodasTiendas(){
  var chkAll = document.getElementById('chkTodasTienda');
  var allChecked = chkAll.checked;
  var chks = document.querySelectorAll('#tiendaDropMenu input[type=checkbox].tienda-chk');
  var selected = allChecked ? DATA.tiendas.slice() : [];
  chks.forEach(function(c){
    c.checked = allChecked;
    var t = c.value;
    var row = document.getElementById('tienda-row-'+t);
    if(row) row.className = 'sem-item' + (allChecked ? ' on' : '');
  });
  applyTiendaSelection(selected);
}

function syncChkTodasTiendas(){
  var chks = document.querySelectorAll('#tiendaDropMenu input[type=checkbox].tienda-chk');
  var chkAll = document.getElementById('chkTodasTienda');
  if(!chkAll) return;
  var total = chks.length, checked = 0;
  chks.forEach(function(c){ if(c.checked) checked++; });
  chkAll.checked = (checked === total);
  chkAll.indeterminate = (checked > 0 && checked < total);
}

function toggleProductoDrop(){
  toggleFloatingMenu('productoDropMenu', 'productoDropBtn');
}
function closeProductoDrop(){
  closeFloatingMenu('productoDropMenu');
}
function closeFloatingMenu(menuId){
  var menu = document.getElementById(menuId);
  if(!menu) return;
  menu.style.display = 'none';
  menu.style.position = '';
  menu.style.top = '';
  menu.style.left = '';
  menu.style.visibility = '';
  menu.style.minWidth = '';
  menu.style.maxHeight = '260px';
}
function closeAllFloatingMenus(){
  ['semDropMenu','tiendaDropMenu','productoDropMenu'].forEach(closeFloatingMenu);
}
function openFloatingMenu(menuId, btnId){
  var menu = document.getElementById(menuId);
  var btn = document.getElementById(btnId);
  if(!menu || !btn) return;

  menu.style.display = 'block';
  menu.style.position = 'fixed';
  menu.style.top = '-9999px';
  menu.style.left = '-9999px';
  menu.style.visibility = 'hidden';

  var rect = btn.getBoundingClientRect();
  var measuredWidth = Math.max(menu.offsetWidth || 0, rect.width, 200);
  var desiredHeight = Math.min(menu.scrollHeight || 260, 320);
  var spaceBelow = window.innerHeight - rect.bottom - 12;
  var spaceAbove = rect.top - 12;
  var openUpwards = spaceBelow < Math.min(desiredHeight, 220) && spaceAbove > spaceBelow;
  var available = openUpwards ? spaceAbove : spaceBelow;
  var maxHeight = Math.max(140, Math.min(desiredHeight, available - 6));

  menu.style.minWidth = measuredWidth + 'px';
  menu.style.maxHeight = maxHeight + 'px';

  var left = Math.max(12, Math.min(rect.left, window.innerWidth - measuredWidth - 12));
  var top = openUpwards
    ? Math.max(12, rect.top - maxHeight - 6)
    : Math.max(12, Math.min(rect.bottom + 6, window.innerHeight - maxHeight - 12));

  menu.style.left = left + 'px';
  menu.style.top = top + 'px';
  menu.style.visibility = 'visible';
}
function toggleFloatingMenu(menuId, btnId){
  var menu = document.getElementById(menuId);
  if(!menu) return;
  var isOpen = window.getComputedStyle(menu).display !== 'none';
  closeAllFloatingMenus();
  if(!isOpen) openFloatingMenu(menuId, btnId);
}
window.addEventListener('resize', closeAllFloatingMenus);
window.addEventListener('scroll', closeAllFloatingMenus);
function areAllProductosSelected(sel){
  return Array.isArray(sel) && sel.length === DATA.productos.length;
}
function getProductosActivos(){
  if(!state.productos_sel || state.productos_sel.length === 0) return [];
  if(areAllProductosSelected(state.productos_sel)) return DATA.productos;
  return state.productos_sel;
}
function applyProductoSelection(selected){
  state.productos_sel = Array.isArray(selected) ? selected.slice() : [];
  state.producto = state.productos_sel.length > 0 ? state.productos_sel[0] : DATA.productos[0];
  updateProductoLabel();
  syncChkTodosProductos();
  if(state.view==='producto') render();
  else if(state.view==='tienda') renderTienda();
  else if(state.view==='inventario') renderInventario();
  else if(state.view==='comparativo') renderComparativo();
  else if(state.view==='gasto') renderGasto();
  else if(state.view==='resumen') renderResumen();
  else if(state.view==='choferes' && typeof renderChoferes === 'function') renderChoferes();
}
function updateProductoLabel(){
  var sel = state.productos_sel;
  var lbl = document.getElementById('productoDropLabel');
  if(areAllProductosSelected(sel)){
    lbl.textContent = '— Todos los productos —';
  } else if(!sel || sel.length === 0){
    lbl.textContent = '— Sin productos —';
  } else if(sel.length === 1){
    lbl.textContent = sel[0].replace('BQT ','');
  } else {
    lbl.textContent = sel.length+' productos seleccionados';
  }
}
function onProductoChk(){
  var chks = document.querySelectorAll('#productoDropMenu input[type=checkbox].producto-chk');
  var selected = [];
  chks.forEach(function(c){
    var p = c.value;
    var row = document.getElementById('producto-row-'+p.replace(/\s+/g, ''));
    if(c.checked){
      selected.push(p);
      if(row) row.className = 'sem-item on';
    } else {
      if(row) row.className = 'sem-item';
    }
  });
  applyProductoSelection(selected);
}
function toggleTodosProductos(){
  var chkAll = document.getElementById('chkTodosProducto');
  var allChecked = chkAll.checked;
  var chks = document.querySelectorAll('#productoDropMenu input[type=checkbox].producto-chk');
  var selected = allChecked ? DATA.productos.slice() : [];
  chks.forEach(function(c){
    c.checked = allChecked;
    var p = c.value;
    var row = document.getElementById('producto-row-'+p.replace(/\s+/g, ''));
    if(row) row.className = 'sem-item' + (allChecked ? ' on' : '');
  });
  applyProductoSelection(selected);
}
function syncChkTodosProductos(){
  var chks = document.querySelectorAll('#productoDropMenu input[type=checkbox].producto-chk');
  var chkAll = document.getElementById('chkTodosProducto');
  if(!chkAll) return;
  var total = chks.length, checked = 0;
  chks.forEach(function(c){ if(c.checked) checked++; });
  chkAll.checked = (checked === total);
  chkAll.indeterminate = (checked > 0 && checked < total);
}

function updateSemLabel(){
  var sel = state.semanas_sel;
  var lbl = document.getElementById('semDropLabel');
  if(!sel || sel.length === 0){
    lbl.textContent = '— Todas las semanas —';
  } else if(sel.length === 1){
    var s = sel[0], yr = Math.floor(s/100), wk = s%100;
    lbl.textContent = (yr >= 2000) ? yr+' · Semana '+String(wk).padStart(2,'0') : 'Semana '+String(s).padStart(2,'0');
  } else {
    lbl.textContent = sel.length+' semanas seleccionadas';
  }
}

function onSemChk(){
  // Solo .sem-chk para excluir el checkbox "Seleccionar todas" (sin value → NaN)
  var chks = document.querySelectorAll('#semDropMenu input[type=checkbox].sem-chk');
  var selected = [];
  chks.forEach(function(c){
    var s = parseInt(c.value);
    var row = document.getElementById('sem-row-'+s);
    if(c.checked){
      selected.push(s);
      if(row) row.className = 'sem-item on';
    } else {
      if(row) row.className = 'sem-item';
    }
  });
  
  // LÍMITE DE 2 SEMANAS SOLO EN VISTA GASTO
  if(state.view === 'gasto' && selected.length > 2){
    // Mantener solo las últimas 2 seleccionadas
    var keep = selected.slice(-2);
    chks.forEach(function(c){
      var s = parseInt(c.value);
      if(keep.indexOf(s) === -1){
        c.checked = false;
        var row = document.getElementById('sem-row-'+s);
        if(row) row.className = 'sem-item';
      }
    });
    selected = keep;
  }
  
  // Si están TODAS seleccionadas → tratar como Global (semanas_sel vacío)
  var esGlobal = (selected.length === DATA.semanas.length);
  state.semanas_sel = esGlobal ? [] : selected;
  state.semana = selected.length > 0 ? selected[selected.length-1] : 'all';
  state.tiendaT = null;
  updateSemLabel();
  updateHeader();
  syncChkTodas();
  if(state.view==='producto') render();
  else if(state.view==='tienda') renderTienda();
  else if(state.view==='inventario') renderInventario();
  else if(state.view==='comparativo') renderComparativo();
  else if(state.view==='gasto') renderGasto();
  else if(state.view==='resumen') renderResumen();
}

function onSem(sel){ onSemChk(); }

function updateHeader(){
  var sems = getSemanasActivas();
  var tiendas = getTiendasActivas();
  var isAll = (!state.semanas_sel || state.semanas_sel.length === 0);
  
  if(isAll){
    var s0 = DATA.semanas[0], sN = DATA.semanas[DATA.semanas.length-1];
    var f0 = (DATA.fecha_por_semana && (DATA.fecha_por_semana[String(s0)] || DATA.fecha_por_semana[s0])) || '—';
    var fN = (DATA.fecha_por_semana && (DATA.fecha_por_semana[String(sN)] || DATA.fecha_por_semana[sN])) || '—';
    document.getElementById('hdrFecha').textContent  = f0 + ' — ' + fN;
    document.getElementById('hdrSem').textContent    = 'Global';
    document.getElementById('projTitle').textContent = 'Proyección';
    return;
  }
  if(sems.length > 1){
    // Incluir año cuando hay semanas de distintos años para evitar "1, 1" duplicados
    var primerAnio = sems.find(function(s){ return s > 9999 ? Math.floor(s/100) : null; });
    primerAnio = primerAnio ? Math.floor(primerAnio/100) : null;
    var multiAnio = sems.some(function(s){ return s > 9999 && Math.floor(s/100) !== primerAnio; });
    var semNums = sems.map(function(s){
      var yr = s > 9999 ? Math.floor(s/100) : null;
      var wk = s > 9999 ? s%100 : s;
      return (multiAnio && yr) ? yr+'\xb7S'+String(wk).padStart(2,'0') : String(wk).padStart(2,'0');
    });
    document.getElementById('hdrFecha').textContent  = sems.length + ' semanas seleccionadas';
    document.getElementById('hdrSem').textContent    = 'Sem ' + semNums.join(', ');
    document.getElementById('projTitle').textContent = 'Proyección';
    return;
  }
  var semKey = String(state.semana);
  var fecha = DATA.fecha_por_semana && DATA.fecha_por_semana[semKey]
    ? DATA.fecha_por_semana[semKey]
    : DATA.fecha_por_semana && DATA.fecha_por_semana[state.semana]
    ? DATA.fecha_por_semana[state.semana]
    : '—';
  document.getElementById('hdrFecha').textContent   = fecha;
  var semNum = state.semana > 9999 ? state.semana%100 : state.semana;
  var semAnio = state.semana > 9999 ? Math.floor(state.semana/100) : '';
  document.getElementById('hdrSem').textContent     = (semAnio ? semAnio+' · ' : '')+'Semana '+String(semNum).padStart(2,'0');
  document.getElementById('projTitle').textContent  = 'Proyección Semana '+(semNum+1);
}

function getSemanasActivas(){
  if(!state.semanas_sel || state.semanas_sel.length === 0) return DATA.semanas;
  return state.semanas_sel;
}

function getTiendasActivas(){
  if(areAllTiendasSelected(state.tiendas_sel)) return DATA.tiendas;
  if(!state.tiendas_sel || state.tiendas_sel.length === 0) return [];
  return state.tiendas_sel;
}

function updateResumenModeButtons(){
  var titleText = document.getElementById('resumenTitleText');
  if(titleText) titleText.textContent = 'Resumen Semanal';
}

function setResumenMode(mode){
  state.resumenMode = 'semanas';
  updateResumenModeButtons();
  if(state.view === 'resumen') renderResumen();
}

function getD(){
  var sems = getSemanasActivas();
  var tiendas = getTiendasActivas();
  var prods = getProductosActivos();
  var merged = {};
  prods.forEach(function(p){
    var v12=0,v3=0,emb=0,m3=0,cfbc=0,retail=0,n12=0;
    tiendas.forEach(function(t){
      sems.forEach(function(s){
        var key = String(s);
        var d = (DATA.data[t]&&DATA.data[t][key]&&DATA.data[t][key][p]) || {};
        v12   += d.v12   || 0;
        v3    += d.v3    || 0;
        emb   += d.emb   || 0;
        m3    += d.m3    || 0;
        cfbc  += d.cfbc  || 0;
        retail+= d.retail|| 0;
        if((d.n12||0) > n12) n12 = d.n12;
      });
    });
    if(n12 < 1) n12 = 1;
    var avg = sems.length > 0 ? v3 / sems.length : 0;
    var merma_ratio = emb > 0 ? m3/emb : 0;
    var proj = merma_ratio < 1 ? avg/(1-merma_ratio) : avg;
    merged[p] = {
      v12: v12, v3: v3, n12: n12, emb: emb, m3: m3,
      avg: avg, proj: proj,
      pct_merma: emb > 0 ? Math.round(m3/emb*100) : 0,
      cfbc: cfbc, retail: retail
    };
  });
  return merged;
}

function render(){
  var d = getD(), prods = getProductosActivos();
  var totV12=0,totV3=0,totEmb=0,totM3=0,totAvg=0,totProj=0,totEmb2=0;
  var histRows='',mermaRows='',avgRows='',projRows='';

  // Construir array y ordenar cada tabla de mayor a menor
  var prodArr = prods.map(function(p){ return {p:p, v:d[p]||{v12:0,v3:0,emb:0,m3:0,avg:0,proj:0,pct_merma:0}}; });

  prodArr.forEach(function(o){ var v=o.v; totV12+=v.v12; totV3+=v.v3; totEmb+=v.emb; totM3+=v.m3; totAvg+=v.avg; totProj+=v.proj; totEmb2+=v.emb; });

  prodArr.slice().sort(function(a,b){ return b.v.v12-a.v.v12; }).forEach(function(o){
    var name=o.p.replace('BQT ',''), v=o.v;
    histRows += '<tr><td>'+name+'</td><td>'+fmt(v.v12)+'</td><td>'+fmt(v.v3)+'</td></tr>';
  });
  prodArr.slice().sort(function(a,b){ return b.v.m3-a.v.m3; }).forEach(function(o){
    var name=o.p.replace('BQT ',''), v=o.v;
    mermaRows += '<tr><td>'+name+'</td><td>'+fmt(v.emb)+'</td><td class="'+(v.m3>0?'red':'')+'">'+fmt(v.m3)+'</td><td class="'+(v.pct_merma>0?'red':'')+'">'+v.pct_merma+'%</td></tr>';
  });
  prodArr.slice().sort(function(a,b){ return b.v.v12-a.v.v12; }).forEach(function(o){
    var name=o.p.replace('BQT ',''), v=o.v;
    var div12 = (v.n12 && v.n12 > 0) ? v.n12 : 1;
    avgRows += '<tr><td>'+name+'</td><td>'+parseFloat((v.v12/div12).toFixed(3))+'</td><td>'+Math.round(v.v3/3)+'</td></tr>';
  });
  prodArr.slice().sort(function(a,b){ return b.v.proj-a.v.proj; }).forEach(function(o){
    var name=o.p.replace('BQT ',''), v=o.v;
    projRows += '<tr><td>'+name+'</td><td class="bold">'+fmt(v.proj)+'</td></tr>';
  });

  histRows  += '<tr class="total"><td>Total</td><td>'+fmt(totV12)+'</td><td>'+fmt(totV3)+'</td></tr>';
  var pct_merma_total = totEmb2 > 0 ? Math.round(totM3/totEmb2*100) : 0;
  mermaRows += '<tr class="total"><td>Total</td><td>'+fmt(totEmb)+'</td><td class="red">'+fmt(totM3)+'</td><td class="red">'+pct_merma_total+'%</td></tr>';
  var totDiv12 = 1;
  prodArr.forEach(function(o){ if((o.v.n12||0) > totDiv12) totDiv12 = o.v.n12; });
  avgRows   += '<tr class="total"><td>Total</td><td>'+parseFloat((totV12/totDiv12).toFixed(3))+'</td><td>'+Math.round(totV3/3)+'</td></tr>';
  projRows  += '<tr class="total"><td>Total</td><td>'+fmt(totProj)+'</td></tr>';
  document.getElementById('tHist').innerHTML  = histRows;
  document.getElementById('tMerma').innerHTML = mermaRows;
  document.getElementById('tAvg').innerHTML   = avgRows;
  document.getElementById('tProj').innerHTML  = projRows;
}

function renderChoferes() {
    var cfbcTbody = document.querySelector('#cfbcTable tbody');
    var cfbcHTML = '';
    var cfbcU = 0;
    var cfbcV = 0;

    var selCfbcTienda = document.getElementById('cfbcTienda');
    var selCfbcProd = document.getElementById('cfbcProducto');
    var selWmTienda = document.getElementById('wmTienda');
    var selWmProd = document.getElementById('wmProducto');
    
    if (selCfbcTienda && selCfbcTienda.options.length <= 1) {
        DATA.tiendas.forEach(function(t) {
            selCfbcTienda.add(new Option(t, t));
            if(selWmTienda) selWmTienda.add(new Option(t, t));
        });
        DATA.productos.forEach(function(p) {
            selCfbcProd.add(new Option(p, p));
            if(selWmProd) selWmProd.add(new Option(p, p));
        });
    }

    var cfbcD = document.getElementById('cfbcDesde') ? document.getElementById('cfbcDesde').value : '';
    var cfbcH = document.getElementById('cfbcHasta') ? document.getElementById('cfbcHasta').value : '';
    var cfbcT = selCfbcTienda ? selCfbcTienda.value : 'ALL';
    var cfbcP = selCfbcProd ? selCfbcProd.value : 'ALL';

    var wmD = document.getElementById('wmDesde') ? document.getElementById('wmDesde').value : '';
    var wmH = document.getElementById('wmHasta') ? document.getElementById('wmHasta').value : '';
    var wmT = selWmTienda ? selWmTienda.value : 'ALL';
    var wmP = selWmProd ? selWmProd.value : 'ALL';
    
    var cfbcCount = 0;
    if (SUPABASE_DATA && SUPABASE_DATA.length > 0) {
        var cfbcRowsArr = [];
        SUPABASE_DATA.forEach(function(row) {
            var rowFecha = row.diario || row.fecha || '';
            var normRowFecha = rowFecha.replace(/\//g, '-');
            
            if (cfbcT !== 'ALL' && row.tienda !== cfbcT) return;
            if (cfbcP !== 'ALL' && row.producto !== cfbcP) return;
            if (cfbcD && normRowFecha < cfbcD) return;
            if (cfbcH && normRowFecha > cfbcH) return;
            
            var rowVenta = parseFloat(row.venta_total || row.venta || 0);
            cfbcRowsArr.push({
                  fecha: rowFecha,
                  folio: row.folio || '',
                  tienda: row.tienda || '',
                  producto: row.producto || '',
                  unidades: parseInt(row.unidades || 0),
                  venta: rowVenta,
                  notas: row.notas || '',
                  url_factura: row.url_factura || '',
                  url_acuse: row.url_acuse || '',
                  razon_sin_acuse: row.razon_sin_acuse || ''
              });
        });
        
        cfbcCount = cfbcRowsArr.length;
        if (cfbcCount > 0) {
            cfbcRowsArr.sort(function(a, b){ 
                if(a.fecha !== b.fecha) return a.fecha < b.fecha ? -1 : 1;
                return a.folio < b.folio ? -1 : (a.folio > b.folio ? 1 : 0); 
            });
            var currentGroup = null;
              var subU = 0, subV = 0;
              var curFactura = '', curAcuse = '', curRazon = '';
              
              function emitCfbcSub() {
                  if (currentGroup !== null) {
                      var iconsHtml = '<div style="display:inline-flex; align-items:center; float:left; padding-left:10px; gap:8px;">';
                      if (curFactura) iconsHtml += '<img src="'+curFactura+'" onclick="openImageViewer(this.src)" style="height:25px; width:25px; object-fit:cover; border-radius:4px; cursor:pointer; border:1px solid #ccc; box-shadow:0 2px 4px rgba(0,0,0,0.1);" title="Ver Factura">';
                      if (curAcuse) iconsHtml += '<img src="'+curAcuse+'" onclick="openImageViewer(this.src)" style="height:25px; width:25px; object-fit:cover; border-radius:4px; cursor:pointer; border:1px solid #ccc; box-shadow:0 2px 4px rgba(0,0,0,0.1);" title="Ver Acuse">';
                      if (curRazon) iconsHtml += '<span style="color:#d97706; font-size:12px; cursor:help; vertical-align:middle; display:flex; align-items:center; gap:4px;" title="Sin Folio: '+curRazon+'">⚠️ Sin Folio</span>';
                      iconsHtml += '</div>';

                      cfbcHTML += '<tr class="subtotal-row" style="background:#e2e8f0; font-weight:bold;">' +
                                  '<td colspan="6" style="text-align:right; font-size:11px; padding:4px 5px;">' + iconsHtml + '<span style="line-height:25px;">Total ' + currentGroup + ':</span></td>' +
                                  '<td class="sub-u" style="text-align:right; font-size:11px; padding:4px 5px; vertical-align:middle;">' + subU.toLocaleString('en-US') + '</td>' +
                                  '<td class="sub-v" style="text-align:right; font-size:11px; padding:4px 5px; vertical-align:middle;">$' + subV.toLocaleString('en-US', {minimumFractionDigits:2, maximumFractionDigits:2}) + '</td>' +
                                  '<td></td></tr>';
                  }
              }
              
              cfbcRowsArr.forEach(function(r) {
                  var groupKey = r.folio || r.fecha;
                  if (groupKey !== currentGroup) {
                      emitCfbcSub();
                      currentGroup = groupKey;
                      subU = 0;
                      subV = 0;
                      curFactura = r.url_factura;
                      curAcuse = r.url_acuse;
                      curRazon = r.razon_sin_acuse;
                  } else {
                      if (!curFactura && r.url_factura) curFactura = r.url_factura;
                      if (!curAcuse && r.url_acuse) curAcuse = r.url_acuse;
                      if (!curRazon && r.razon_sin_acuse) curRazon = r.razon_sin_acuse;
                  }
                subU += r.unidades;
                subV += r.venta;
                cfbcU += r.unidades;
                cfbcV += r.venta;
                var rKey = r.fecha + '|' + r.tienda + '|' + r.producto;
                
                // Auto-verificar si el chofer ya procesó (subió acuse o puso razón)
                if (r.url_acuse || r.razon_sin_acuse) {
                    checkedCFBC.add(rKey);
                    // Opcional: si queremos limpiar la bandera manual que tuviera antes
                    // flaggedCFBC.delete(rKey);
                }
                
                var chk = checkedCFBC.has(rKey) ? 'checked' : '';
                var flg = flaggedCFBC.has(rKey) ? 'flagged' : '';
                var cClass = '';
                if(chk) cClass += ' checked';
                if(!chk || flg) cClass += ' flagged';
                var sNota = notesCFBC[rKey] || r.notas || '';
                var flagLabel = '<div class="procesando-lbl" style="color:#e00; font-size:9px; font-weight:bold; margin-top:2px; display:' + (chk ? 'none' : 'block') + ';">Procesando</div>';
                cfbcHTML += '<tr class="'+cClass+'" data-source="cfbc" data-key="'+rKey+'" style="border-bottom:1px solid #eee;">' +
                            '<td style="padding:4px 5px; text-align:center;"><input type="checkbox" '+(chk?'checked':'')+' onchange="toggleRowCheck(this, \''+rKey+'\', \'cfbc\')"></td>' +
                            '<td style="padding:4px 5px; text-align:center;"><button class="flag-btn" title="Marcar diferencia" onclick="toggleFlag(this, \''+rKey+'\', \'cfbc\')">&#9873;</button>'+flagLabel+'</td>' +
                            '<td style="padding:4px 5px;">'+r.fecha+'</td>' +
                            '<td style="padding:4px 5px; font-weight:bold;">'+r.folio+'</td>' +
                            '<td style="padding:4px 5px;">'+r.tienda+'</td>' +
                            '<td style="padding:4px 5px;">'+r.producto+'</td>' +
                            '<td style="padding:4px 5px; text-align:right;">'+r.unidades.toLocaleString('en-US')+'</td>' +
                            '<td style="padding:4px 5px; text-align:right;">$'+r.venta.toLocaleString('en-US', {minimumFractionDigits:2, maximumFractionDigits:2})+'</td>' +
                            '<td style="padding:4px 5px;"><input type="text" class="nota-input" placeholder="Agregar nota..." value="'+sNota+'" onchange="saveNota(\''+rKey+'\', this.value, \'cfbc\')"></td>' +
                            '</tr>';
            });
            emitCfbcSub();
        }
    }
    if (cfbcCount === 0) {
        cfbcHTML = '<tr><td colspan="7" style="text-align:center; padding:10px;">No hay datos en CFBC para estos filtros</td></tr>';
    }
    if(cfbcTbody) cfbcTbody.innerHTML = cfbcHTML;
    var elmCU = document.getElementById('cfbcUnidades');
    if(elmCU) elmCU.innerText = cfbcU.toLocaleString('en-US');
    var elmCV = document.getElementById('cfbcVenta');
    if(elmCV) elmCV.innerText = '$' + cfbcV.toLocaleString('en-US', {minimumFractionDigits:2, maximumFractionDigits:2});

    var wmRows = [];
    var wmU = 0;
    var wmV = 0;
    if (DATA && DATA.resumen_diario) {
        Object.keys(DATA.resumen_diario).forEach(function(tienda) {
            if (wmT !== 'ALL' && tienda !== wmT) return;
            Object.keys(DATA.resumen_diario[tienda]).forEach(function(sem) {
                Object.keys(DATA.resumen_diario[tienda][sem]).forEach(function(prod) {
                    if (wmP !== 'ALL' && prod !== wmP) return;
                    Object.keys(DATA.resumen_diario[tienda][sem][prod]).forEach(function(fecha) {
                        var normFecha = fecha.replace(/\//g, '-');
                        if (wmD && normFecha < wmD) return;
                        if (wmH && normFecha > wmH) return;
                        var metricas = DATA.resumen_diario[tienda][sem][prod][fecha];
                        var emb = metricas.embarque || 0;
                        if (emb > 0) {
                            var vCfbc = parseFloat(metricas.venta_cfbc || 0);
                            wmRows.push({
                                fecha: fecha,
                                tienda: tienda,
                                producto: prod,
                                unidades: emb,
                                venta: vCfbc
                            });
                        }
                    });
                });
            });
        });
    }
    
    var wmHTML = '';
    if (wmRows.length > 0) {
        wmRows.sort(function(a, b){ return a.fecha < b.fecha ? -1 : (a.fecha > b.fecha ? 1 : 0); });
        var currentGroupWm = null;
        var subUWm = 0, subVWm = 0;
        
        function emitWmSub() {
            if (currentGroupWm !== null) {
                wmHTML += '<tr class="subtotal-row" style="background:#e2e8f0; font-weight:bold;">' +
                          '<td colspan="5" style="text-align:right; font-size:11px; padding:4px 5px;">Total ' + currentGroupWm + ':</td>' +
                          '<td class="sub-u" style="text-align:right; font-size:11px; padding:4px 5px;">' + subUWm.toLocaleString('en-US') + '</td>' +
                          '<td class="sub-v" style="text-align:right; font-size:11px; padding:4px 5px;">$' + subVWm.toLocaleString('en-US', {minimumFractionDigits:2, maximumFractionDigits:2}) + '</td>' +
                          '<td></td></tr>';
            }
        }
        
        wmRows.forEach(function(row) {
            var groupKey = row.fecha;
            if (groupKey !== currentGroupWm) {
                emitWmSub();
                currentGroupWm = groupKey;
                subUWm = 0;
                subVWm = 0;
            }
            subUWm += row.unidades;
            subVWm += row.venta;
            var rKey = row.fecha + '|' + row.tienda + '|' + row.producto;
            var chk = checkedWalmart.has(rKey) ? 'checked' : '';
            var flg = flaggedWalmart.has(rKey) ? 'flagged' : '';
            var cClass = '';
            if(chk) cClass += ' checked';
            if(!chk || flg) cClass += ' flagged';
            var sNota = notesWalmart[rKey] || '';
            var flagLabel = '<div class="procesando-lbl" style="color:#e00; font-size:9px; font-weight:bold; margin-top:2px; display:' + (chk ? 'none' : 'block') + ';">Procesando</div>';
            wmHTML += '<tr class="'+cClass+'" data-source="walmart" data-key="'+rKey+'" style="border-bottom:1px solid #eee;">' +
                '<td style="padding:4px 5px; text-align:center;"><input type="checkbox" '+(chk?'checked':'')+' onchange="toggleRowCheck(this, \''+rKey+'\', \'walmart\')"></td>' +
                '<td style="padding:4px 5px; text-align:center;"><button class="flag-btn" title="Marcar diferencia" onclick="toggleFlag(this, \''+rKey+'\', \'walmart\')">&#9873;</button>'+flagLabel+'</td>' +
                '<td style="padding:4px; border-bottom:1px solid #eee;">' + row.fecha + '</td>' +
                '<td style="padding:4px; border-bottom:1px solid #eee;">' + row.tienda + '</td>' +
                '<td style="padding:4px; border-bottom:1px solid #eee;">' + row.producto + '</td>' +
                '<td style="padding:4px; text-align:right; border-bottom:1px solid #eee;">' + row.unidades + '</td>' +
                '<td style="padding:4px; text-align:right; border-bottom:1px solid #eee;">$' + (row.venta.toLocaleString('en-US', {minimumFractionDigits:2})) + '</td>' +
                '<td style="padding:4px; border-bottom:1px solid #eee;"><input type="text" class="nota-input" placeholder="Agregar nota..." value="'+sNota+'" onchange="saveNota(\''+rKey+'\', this.value, \'walmart\')"></td>' +
            '</tr>';
            wmU += row.unidades;
            wmV += row.venta;
        });
        emitWmSub();
    } else {
        wmHTML = '<tr><td colspan="6" style="text-align:center; padding:10px;">No hay datos de Walmart para estos filtros</td></tr>';
    }
    var wmTbody = document.querySelector('#walmartTable tbody');
    if(wmTbody) wmTbody.innerHTML = wmHTML;
    var elmWU = document.getElementById('walmartUnidades');
    if(elmWU) elmWU.innerText = wmU.toLocaleString('en-US');
    var elmWV = document.getElementById('walmartVenta');
    if(elmWV) elmWV.innerText = '$' + wmV.toLocaleString('en-US', {minimumFractionDigits:2, maximumFractionDigits:2});
}

function setView(v){
  var ctrlCont = document.getElementById('globalCtrlContainer');
  if(ctrlCont) ctrlCont.style.display = (v === 'choferes') ? 'none' : 'flex';
  if(v==='choferes') renderChoferes();
  if(state.view === 'resumen' && v !== 'resumen') saveCaptureProjectionsFromDom();
  state.view = v;
  updateHeader();
  updateResumenModeButtons();
  document.getElementById('btnProd').style.background = v==='producto' ? '#6A4B59' : '#A08A95';
  document.getElementById('btnProd').style.color = v==='producto' ? 'white' : 'white';
  document.getElementById('btnTiend').style.background = v==='tienda' ? '#6A4B59' : '#A08A95';
  document.getElementById('btnTiend').style.color = v==='tienda' ? 'white' : 'white';
  document.getElementById('btnComp').style.background = v==='comparativo' ? '#6A4B59' : '#A08A95';
  document.getElementById('btnComp').style.color = v==='comparativo' ? 'white' : 'white';
  document.getElementById('btnInv').style.background = v==='inventario' ? '#6A4B59' : '#A08A95';
  document.getElementById('btnInv').style.color = v==='inventario' ? 'white' : 'white';
  document.getElementById('btnGasto').style.background = v==='gasto' ? '#6A4B59' : '#A08A95';
  document.getElementById('btnGasto').style.color = v==='gasto' ? 'white' : 'white';
  document.getElementById('btnResumen').style.background = v==='resumen' ? '#6A4B59' : '#A08A95';
  document.getElementById('btnResumen').style.color = v==='resumen' ? 'white' : 'white';
  var bc = document.getElementById('btnChoferes'); if(bc) { bc.style.background = v==='choferes' ? '#6A4B59' : '#A08A95'; bc.style.color = v==='choferes' ? 'white' : 'white'; }
  document.getElementById('viewProducto').style.display = v==='producto' ? 'grid' : 'none';
  document.getElementById('viewTienda').style.display = v==='tienda' ? 'grid' : 'none';
  document.getElementById('viewComparativo').style.display = v==='comparativo' ? 'block' : 'none';
  document.getElementById('viewInventario').style.display = v==='inventario' ? 'grid' : 'none';
  document.getElementById('viewGasto').style.display = v==='gasto' ? 'block' : 'none';
  document.getElementById('viewResumen').style.display = v==='resumen' ? 'block' : 'none';
  var vc = document.getElementById('viewChoferes'); if(vc) vc.style.display = v==='choferes' ? 'block' : 'none';
  
  // Ocultar filtros de tienda solo en vistas que no los usan
  var tiendaDropWrap = document.getElementById('tiendaDropWrap');
  var tiendaLabel = Array.from(document.querySelectorAll('.ctrl label')).find(el => el.textContent === 'Tienda:');
  if(v==='tienda' || v==='inventario' || v==='gasto' || v==='comparativo'){
    if(tiendaDropWrap) tiendaDropWrap.style.display = 'none';
    if(tiendaLabel) tiendaLabel.style.display = 'none';
  } else {
    if(tiendaDropWrap) tiendaDropWrap.style.display = 'inline-block';
    if(tiendaLabel) tiendaLabel.style.display = 'block';
  }
  
  // Ocultar filtro de ruta en comparativo y tienda
  var rutaSelect = document.getElementById('rutaSelect');
  var rutaLabel = Array.from(document.querySelectorAll('.ctrl label')).find(el => el.textContent === 'Ruta:');
  if(v==='tienda' || v==='comparativo'){
    if(rutaSelect) rutaSelect.style.display = 'none';
    if(rutaLabel) rutaLabel.style.display = 'none';
  } else {
    if(rutaSelect) rutaSelect.style.display = 'inline-block';
    if(rutaLabel) rutaLabel.style.display = 'block';
  }
  
  if(v==='tienda'){ state.tiendaT = null; renderTienda(); }
  else if(v==='comparativo'){ if(!state.compMode) state.compMode='semanas'; renderComparativo(); }
  else if(v==='inventario'){ state.invMode = null; state.invSelected = null; renderInventario(); }
  else if(v==='gasto'){ renderGasto(); }
  else if(v==='resumen'){ renderResumen(); }
  else render();
}

function selTiendaT(t){
  // Toggle: si ya está seleccionada, deseleccionar
  state.tiendaT = (state.tiendaT === t) ? null : t;
  renderTienda();
}

function renderTienda(){
  var tiendas = DATA.tiendas;
  var sems = getSemanasActivas();
  var isAll = (!state.semanas_sel || state.semanas_sel.length === 0);

  var hdrSab='Cnt Sab', hdrDom='Ctd Dom', hdrLun='Ctd Lun', hdrMar='Ctd Mar', hdrMie='Ctd Mie', hdrJue='Ctd Jue', hdrVie='Ctd Vie';
  var hdrVtaSab='Ventas Sab', hdrVtaDom='Ventas Dom', hdrVtaLun='Ventas Lun', hdrVtaMar='Ventas Mar', hdrVtaMie='Ventas Mie', hdrVtaJue='Ventas Jue', hdrVtaVie='Ventas Vie';
  if(document.getElementById('hdrThSab')){
    document.getElementById('hdrThSab').innerHTML = hdrSab;
    document.getElementById('hdrThDom').innerHTML = hdrDom;
    document.getElementById('hdrThLun').innerHTML = hdrLun;
    document.getElementById('hdrThMar').innerHTML = hdrMar;
    document.getElementById('hdrThMie').innerHTML = hdrMie;
    document.getElementById('hdrThJue').innerHTML = hdrJue;
    document.getElementById('hdrThVie').innerHTML = hdrVie;
    document.getElementById('hdrThVtaSab').innerHTML = hdrVtaSab;
    document.getElementById('hdrThVtaDom').innerHTML = hdrVtaDom;
    document.getElementById('hdrThVtaLun').innerHTML = hdrVtaLun;
    document.getElementById('hdrThVtaMar').innerHTML = hdrVtaMar;
    document.getElementById('hdrThVtaMie').innerHTML = hdrVtaMie;
    document.getElementById('hdrThVtaJue').innerHTML = hdrVtaJue;
    document.getElementById('hdrThVtaVie').innerHTML = hdrVtaVie;
  }

  // ── Obtener totales por tienda según semanas activas ──
  var totEmb=0, totCfbc=0, totWmx=0, totMerma=0, totRetail=0, totVentasU=0;
  var tiendaData = [];

  tiendas.forEach(function(tienda){
    var emb=0, cfbc=0, wmx=0, merma=0, retail=0, ventas_u=0;
    if(isAll){
      var tot = (DATA.totales_tienda && DATA.totales_tienda[tienda]) || {};
      emb    = tot.embarque_u || 0;
      cfbc   = tot.venta_cfbc || 0;
      wmx    = tot.venta_wmx  || 0;
      merma  = tot.merma_u    || 0;
      retail = tot.retail_vc  || 0;
      ventas_u = tot.ventas_u || 0;
    } else {
      sems.forEach(function(s){
        var raw = (DATA.raw_semana && DATA.raw_semana[tienda] && DATA.raw_semana[tienda][String(s)]) || {};
        emb    += raw.embarque_u || 0;
        cfbc   += raw.venta_cfbc || 0;
        wmx    += raw.venta_wmx  || 0;
        merma  += raw.merma_u    || 0;
        retail += raw.retail_vc  || 0;
        ventas_u += raw.ventas_u || 0;
      });
    }
    totEmb+=emb; totCfbc+=cfbc; totWmx+=wmx; totMerma+=merma; totRetail+=retail; totVentasU+=ventas_u;
    tiendaData.push({tienda:tienda, emb:emb, cfbc:cfbc, wmx:wmx, merma:merma, retail:retail, ventas_u:ventas_u});
  });

  // ── TOP VENTA: ordenar por VENTA (cfbc) de mayor a menor ──
  var histRows='';
  tiendaData.slice().sort(function(a,b){ return b.cfbc-a.cfbc; }).forEach(function(t){
    var pct = totCfbc > 0 ? Math.round(t.cfbc/totCfbc*100) : 0;
    var sel = state.tiendaT === t.tienda;
    var style = sel ? 'style="background:#e8f0fe;font-weight:700;cursor:pointer"' : 'style="cursor:pointer"';
    histRows += '<tr '+style+' onclick="selTiendaT(\''+t.tienda.replace(/'/g,"\\'")+'\')">'
      +'<td>'+t.tienda+'</td><td>'+fmt(t.ventas_u)+'</td><td>$'+fmt(t.cfbc)+'</td><td>$'+fmt(t.wmx)+'</td><td>'+pct+'%</td></tr>';
  });
  histRows += '<tr class="total"><td>Total</td><td>'+fmt(totVentasU)+'</td><td>$'+fmt(totCfbc)+'</td><td>$'+fmt(totWmx)+'</td><td>100%</td></tr>';

  // ── TOP MERMA: ordenar por RETAIL (cantidad $) de mayor a menor ──
  var mermaRows='';
  tiendaData.slice().sort(function(a,b){ return b.retail-a.retail; }).forEach(function(t){
    var pct_retail = totRetail > 0 ? Math.round(t.retail/totRetail*100) : 0;
    var sel = state.tiendaT === t.tienda;
    var style = sel ? 'style="background:#fff0f0;font-weight:700;cursor:pointer"' : 'style="cursor:pointer"';
    mermaRows += '<tr '+style+' onclick="selTiendaT(\''+t.tienda.replace(/'/g,"\\'")+'\')">'
      +'<td>'+t.tienda+'</td>'
      +'<td class="'+(t.merma>0?'red':'')+'">'+fmt(t.merma)+'</td>'
      +'<td>$</td>'
      +'<td class="'+(t.retail>0?'red':'')+'">'+fmt(t.retail)+'</td>'
      +'<td class="'+(pct_retail>0?'red':'')+'">'+pct_retail+'%</td></tr>';
  });
  mermaRows += '<tr class="total"><td>Total</td><td class="red">'+fmt(totMerma)+'</td><td>$</td><td class="red">'+fmt(totRetail)+'</td><td class="red">100%</td></tr>';

  // ── Semana clave para productos (última activa) ──
  var semKeyProd = sems.length > 0 ? String(sems[sems.length-1]) : String(DATA.semanas[DATA.semanas.length-1]);
  var prods = getProductosActivos();

  // Determinar qué tiendas y productos mostrar en las tablas inferiores
  var avgRows='', projRows='', diasRows='';

  if(state.tiendaT){
    var tSel = state.tiendaT;
    var tName = tSel.replace('SC ','');
    document.getElementById('avgTTitle').textContent  = 'Venta — '+tName;
    document.getElementById('projTTitle').textContent = 'Merma — '+tName;

    var totVenta=0, totWmx=0, totUnid=0, totMermaU=0, totMermaR=0;
    var tDias = {cs:0,cd:0,cl:0,cma:0,cmi:0,cj:0,cv:0,
                 vs:0,vd:0,vl:0,vma:0,vmi:0,vj:0,vv:0};

    // Construir array con datos por producto
    var prodItems = prods.map(function(p){
      if(isAll){
        var d = (DATA.totales_prod_tienda && DATA.totales_prod_tienda[tSel] && DATA.totales_prod_tienda[tSel][p]) || {};
        return { p:p, venta:d.venta_cfbc||0, wmx:d.venta_wmx||0, unid:d.ventas_u||0, mermaU:d.merma_u||0, mermaR:d.retail_vc||0, dias:d };
      } else {
        var venta=0, wmx=0, unid=0, mermaU=0, mermaR=0;
        var r = { ctd_sab:0, ctd_dom:0, ctd_lun:0, ctd_mar:0, ctd_mie:0, ctd_jue:0, ctd_vie:0,
                  vtas_sab:0, vtas_dom:0, vtas_lun:0, vtas_mar:0, vtas_mie:0, vtas_jue:0, vtas_vie:0 };
        sems.forEach(function(s){
          var dr = (DATA.raw_prod_semana && DATA.raw_prod_semana[tSel] && DATA.raw_prod_semana[tSel][String(s)] && DATA.raw_prod_semana[tSel][String(s)][p]) || {};
          venta  += dr.venta_cfbc||0;
          wmx    += dr.venta_wmx||0;
          unid   += dr.ventas_u||0;
          mermaU += dr.merma_u||0;
          mermaR += dr.retail_vc||0;
          r.ctd_sab += dr.ctd_sab||0; r.ctd_dom += dr.ctd_dom||0; r.ctd_lun += dr.ctd_lun||0;
          r.ctd_mar += dr.ctd_mar||0; r.ctd_mie += dr.ctd_mie||0; r.ctd_jue += dr.ctd_jue||0; r.ctd_vie += dr.ctd_vie||0;
          r.vtas_sab += dr.vtas_sab||0; r.vtas_dom += dr.vtas_dom||0; r.vtas_lun += dr.vtas_lun||0;
          r.vtas_mar += dr.vtas_mar||0; r.vtas_mie += dr.vtas_mie||0; r.vtas_jue += dr.vtas_jue||0; r.vtas_vie += dr.vtas_vie||0;
        });
        return { p:p, venta:venta, wmx:wmx, unid:unid, mermaU:mermaU, mermaR:mermaR, dias:r };
      }
    });
    prodItems.forEach(function(o){ totVenta+=o.venta; totWmx+=(o.wmx||0); totUnid+=o.unid; totMermaU+=o.mermaU; totMermaR+=o.mermaR; });
    // Venta: ordenar por venta desc
    prodItems.slice().sort(function(a,b){ return b.venta-a.venta; }).forEach(function(o){
      var pname = o.p.replace('BQT ','');
      avgRows += '<tr><td>'+pname+'</td><td>$'+fmt(o.venta)+'</td><td>$'+fmt(o.wmx||0)+'</td><td>'+fmt(o.unid)+'</td></tr>';
    });
    // Merma: ordenar por mermaR desc
    prodItems.slice().sort(function(a,b){ return b.mermaR-a.mermaR; }).forEach(function(o){
      var pname = o.p.replace('BQT ','');
      projRows += '<tr><td>'+pname+'</td>'
        +'<td class="'+(o.mermaU>0?'red':'')+'">'+fmt(o.mermaU)+'</td>'
        +'<td class="'+(o.mermaR>0?'red':'')+'">$'+fmt(o.mermaR)+'</td></tr>';
    });
    
    // Dias: mostrar productos y sus totales de Sab a Vie
    prodItems.forEach(function(o){
        var d = o.dias || {};
        var totU = (d.ctd_sab||0)+(d.ctd_dom||0)+(d.ctd_lun||0)+(d.ctd_mar||0)+(d.ctd_mie||0)+(d.ctd_jue||0)+(d.ctd_vie||0);
        o.totUDias = totU;
    });
    prodItems.slice().sort(function(a,b){ return b.venta-a.venta; }).forEach(function(o){
      var pname = o.p.replace('BQT ','');
      var d = o.dias || {};
      tDias.cs += d.ctd_sab||0; tDias.cd += d.ctd_dom||0; tDias.cl += d.ctd_lun||0; tDias.cma += d.ctd_mar||0; tDias.cmi += d.ctd_mie||0; tDias.cj += d.ctd_jue||0; tDias.cv += d.ctd_vie||0;
      tDias.vs += d.vtas_sab||0; tDias.vd += d.vtas_dom||0; tDias.vl += d.vtas_lun||0; tDias.vma += d.vtas_mar||0; tDias.vmi += d.vtas_mie||0; tDias.vj += d.vtas_jue||0; tDias.vv += d.vtas_vie||0;

      var q = function(v){ return '<td class="'+(v>0?'red':'')+'">'+fmt(v)+'</td>'; };
      var qv = function(v){ return '<td class="'+(v>0?'red':'')+'">$'+fmt(v)+'</td>'; };
      diasRows += '<tr><td>'+pname+'</td>' +
        q(d.ctd_sab||0)+q(d.ctd_dom||0)+q(d.ctd_lun||0)+q(d.ctd_mar||0)+q(d.ctd_mie||0)+q(d.ctd_jue||0)+q(d.ctd_vie||0) +
        qv(d.vtas_sab||0)+qv(d.vtas_dom||0)+qv(d.vtas_lun||0)+qv(d.vtas_mar||0)+qv(d.vtas_mie||0)+qv(d.vtas_jue||0)+qv(d.vtas_vie||0) + '</tr>';
    });

    avgRows  += '<tr class="total"><td>Total</td><td>$'+fmt(totVenta)+'</td><td>$'+fmt(totWmx)+'</td><td>'+fmt(totUnid)+'</td></tr>';
    projRows += '<tr class="total"><td>Total</td><td class="red">'+fmt(totMermaU)+'</td><td class="red">$'+fmt(totMermaR)+'</td></tr>';
    var qT = function(v){ return '<td class="'+(v>0?'red':'')+'">'+fmt(v)+'</td>'; };
    var qvT = function(v){ return '<td class="'+(v>0?'red':'')+'">$'+fmt(v)+'</td>'; };
    diasRows += '<tr class="total"><td>Total</td>' +
      qT(tDias.cs)+qT(tDias.cd)+qT(tDias.cl)+qT(tDias.cma)+qT(tDias.cmi)+qT(tDias.cj)+qT(tDias.cv) +
      qvT(tDias.vs)+qvT(tDias.vd)+qvT(tDias.vl)+qvT(tDias.vma)+qvT(tDias.vmi)+qvT(tDias.vj)+qvT(tDias.vv) + '</tr>';

  } else {
    document.getElementById('avgTTitle').textContent  = 'Venta Promedio Semanal';
    document.getElementById('projTTitle').textContent = 'Comparacion Ultimas 3 Semanas';

    var totVenta=0, totWmx=0, totUnid=0, totMermaU=0, totMermaR=0;
    // Construir array sumando todas las tiendas
    var prodItems = prods.map(function(p){
      var ventaSum=0, wmxSum=0, unidSum=0, mermaUSum=0, mermaRSum=0;
      tiendas.forEach(function(t){
        var d;
        if(isAll){
          d = (DATA.totales_prod_tienda && DATA.totales_prod_tienda[t] && DATA.totales_prod_tienda[t][p]) || {};
          ventaSum  += d.venta_cfbc || 0;
          wmxSum    += d.venta_wmx  || 0;
          unidSum   += d.ventas_u || 0;
          mermaUSum += d.merma_u    || 0;
          mermaRSum += d.retail_vc  || 0;
        } else {
          sems.forEach(function(s){
            var dr = (DATA.raw_prod_semana && DATA.raw_prod_semana[t] && DATA.raw_prod_semana[t][String(s)] && DATA.raw_prod_semana[t][String(s)][p]) || {};
            ventaSum  += dr.venta_cfbc || 0;
            wmxSum    += dr.venta_wmx  || 0;
            unidSum   += dr.ventas_u || 0;
            mermaUSum += dr.merma_u    || 0;
            mermaRSum += dr.retail_vc  || 0;
          });
        }
      });
      return { p:p, venta:ventaSum, wmx:wmxSum, unid:unidSum, mermaU:mermaUSum, mermaR:mermaRSum };
    }).filter(function(o){ return o.venta||o.unid||o.mermaU||o.mermaR; });

    prodItems.forEach(function(o){ totVenta+=o.venta; totWmx+=(o.wmx||0); totUnid+=o.unid; totMermaU+=o.mermaU; totMermaR+=o.mermaR; });
    // Venta: ordenar por venta desc
    prodItems.slice().sort(function(a,b){ return b.venta-a.venta; }).forEach(function(o){
      var pname = o.p.replace('BQT ','');
      avgRows += '<tr><td>'+pname+'</td><td>$'+fmt(o.venta)+'</td><td>$'+fmt(o.wmx||0)+'</td><td>'+fmt(o.unid)+'</td></tr>';
    });
    // Merma: ordenar por mermaR desc
    prodItems.slice().sort(function(a,b){ return b.mermaR-a.mermaR; }).forEach(function(o){
      var pname = o.p.replace('BQT ','');
      projRows += '<tr><td>'+pname+'</td>'
        +'<td class="'+(o.mermaU>0?'red':'')+'">'+fmt(o.mermaU)+'</td>'
        +'<td class="'+(o.mermaR>0?'red':'')+'">$'+fmt(o.mermaR)+'</td></tr>';
    });
    avgRows  += '<tr class="total"><td>Total</td><td class="red">$'+fmt(totVenta)+'</td><td class="red">$'+fmt(totWmx)+'</td><td class="red">'+fmt(totUnid)+'</td></tr>';
    projRows += '<tr class="total"><td>Total</td><td class="red">'+fmt(totMermaU)+'</td><td class="red">$'+fmt(totMermaR)+'</td></tr>';
  }

  document.getElementById('tHistT').innerHTML  = histRows;
  document.getElementById('tMermaT').innerHTML = mermaRows;
  document.getElementById('tAvgT').innerHTML   = avgRows;
  document.getElementById('tProjT').innerHTML  = projRows;
  if(document.getElementById('tDiasT')) document.getElementById('tDiasT').innerHTML  = diasRows;

  // ── Merma por Producto (total todas las tiendas, semanas activas) ──
  var mermaSemsRows = '';
  var semsToShow2 = (state.semanas_sel && state.semanas_sel.length > 0) ? state.semanas_sel : (DATA.semanas || []);
  var prodsMerma = getProductosActivos() || [];
  var totMermaSemTotal = 0, totMermaSemRetail = 0;
  prodsMerma.forEach(function(p){
    var mu = 0, mr = 0;
    DATA.tiendas.forEach(function(t){
      semsToShow2.forEach(function(s){
        var dr = (DATA.raw_prod_semana && DATA.raw_prod_semana[t] && DATA.raw_prod_semana[t][String(s)] && DATA.raw_prod_semana[t][String(s)][p]) || {};
        mu += dr.merma_u  || 0;
        mr += dr.retail_vc || 0;
      });
    });
    if(mu > 0 || mr > 0){
      totMermaSemTotal  += mu;
      totMermaSemRetail += mr;
      var pname = p.replace('BQT ','');
      mermaSemsRows += '<tr><td>'+pname+'</td><td class="red">'+fmt(mu)+'</td><td class="red">$'+fmt(mr)+'</td></tr>';
    }
  });
  mermaSemsRows += '<tr class="total"><td>Total</td><td class="red">'+fmt(totMermaSemTotal)+'</td><td class="red">$'+fmt(totMermaSemRetail)+'</td></tr>';
  if(document.getElementById('tMermaSem')) document.getElementById('tMermaSem').innerHTML = mermaSemsRows;

  // Mostrar/ocultar tablas inferiores
  var showBottom = state.tiendaT ? 'block' : 'none';
  var showFlex   = state.tiendaT ? 'flex'  : 'none';
  document.getElementById('boxAvgT').style.display      = showBottom;
  document.getElementById('boxProjWrapper').style.display = showFlex;
  if(document.getElementById('boxDiasT')) document.getElementById('boxDiasT').style.display = showBottom;
}

// ─── INVENTARIO ─────────────────────────────────────────────────────────────
function renderInventario(){
  document.getElementById('viewInventario').style.display = 'grid';

  var det = DATA.detalle_inventario;
  if(!det || !det.fechas || det.fechas.length === 0){
    document.getElementById('tInvTiendaHead').innerHTML = '<tr><th colspan="2">Sin datos en hoja Detalle</th></tr>';
    document.getElementById('tInvTienda').innerHTML = '';
    document.getElementById('tInvProductoHead').innerHTML = '<tr><th colspan="2">Sin datos en hoja Detalle</th></tr>';
    document.getElementById('tInvProducto').innerHTML = '';
    document.getElementById('invProductoTitle').innerHTML = 'Inventario Actual — Productos';
    return;
  }

  var sems    = getSemanasActivas();
  var detData = det.data;
  var selTiendas = getTiendasActivas();
  var tiendas = Object.keys(detData).filter(function(t) { return selTiendas.indexOf(t) !== -1; }).sort();

  // ── Determinar fechas a mostrar según semanas activas ──
  var fechas;
  if(sems.length === 0){
    // Sin filtro de semana: mostrar todas las fechas
    fechas = det.fechas || [];
  } else {
    // Filtrar solo las fechas que pertenecen a las semanas seleccionadas
    var fechasSet = {};
    sems.forEach(function(s){
      var fs = (det.fechas_por_semana && det.fechas_por_semana[String(s)]) || [];
      fs.forEach(function(f){ fechasSet[f] = true; });
    });
    // Mantener el orden cronológico original
    fechas = (det.fechas || []).filter(function(f){ return fechasSet[f]; });
  }

  if(fechas.length === 0){
    document.getElementById('tInvTiendaHead').innerHTML = '<tr><th colspan="2">Sin datos para las semanas seleccionadas</th></tr>';
    document.getElementById('tInvTienda').innerHTML = '';
    document.getElementById('tInvProductoHead').innerHTML = '<tr><th colspan="2">Sin datos para las semanas seleccionadas</th></tr>';
    document.getElementById('tInvProducto').innerHTML = '';
    document.getElementById('invProductoTitle').innerHTML = 'Inventario Actual — Productos';
    return;
  }

  // Recolectar todos los productos de todas las tiendas
  var prodSet = {};
  tiendas.forEach(function(t){
    Object.keys(detData[t]).forEach(function(p){ prodSet[p] = true; });
  });
  var prods = Object.keys(prodSet).sort();

  // ── Cabeceras dinámicas ──
  function buildHead(firstColLabel){
    var h = '<tr><th>'+firstColLabel+'</th>';
    fechas.forEach(function(f){ h += '<th>'+f+'</th>'; });
    h += '<th>Total</th>';
    return h + '</tr>';
  }
  document.getElementById('tInvTiendaHead').innerHTML   = buildHead('Tienda');
  document.getElementById('tInvProductoHead').innerHTML = buildHead('Producto');

  var q = function(v){ return '<td class="'+(v>0?'red':'')+'">' + fmt(v) + '</td>'; };

  // ── Tabla izquierda: totales por tienda ──
  var rowsTienda = '';
  var totT = fechas.map(function(){ return 0; });
  var grandTotT = 0;
  tiendas.forEach(function(t){
    var vals = fechas.map(function(f){
      var total = 0;
      Object.keys(detData[t]).forEach(function(p){
        total += (detData[t][p][f] || 0);
      });
      return total;
    });
    var rowTotal = vals.reduce(function(s,v){ return s+v; }, 0);
    if(rowTotal > 0){
      vals.forEach(function(v,i){ totT[i] += v; });
      grandTotT += rowTotal;
      var sel = (state.invMode==='tienda' && state.invSelected===t);
      var st2 = sel ? ' style="background:#e8f0fe;font-weight:700;cursor:pointer"' : ' style="cursor:pointer"';
      rowsTienda += '<tr'+st2+' onclick="selInvTienda(\''+t.replace(/'/g,"\\'")+'\')"><td>'+t+'</td>';
      vals.forEach(function(v){ rowsTienda += q(v); });
      rowsTienda += q(rowTotal);
      rowsTienda += '</tr>';
    }
  });
  rowsTienda += '<tr class="total"><td>Total</td>';
  totT.forEach(function(v){ rowsTienda += q(v); });
  rowsTienda += q(grandTotT);
  rowsTienda += '</tr>';
  document.getElementById('tInvTienda').innerHTML = rowsTienda;

  // ── Tabla derecha: por producto (filtrada si hay tienda seleccionada) ──
  var rowsProducto = '';
  var totP = fechas.map(function(){ return 0; });
  var grandTotP = 0;
  var title = 'Inventario Actual — Productos';
  var tFiltro = (state.invMode==='tienda' && state.invSelected) ? [state.invSelected] : tiendas;

  if(state.invMode==='tienda' && state.invSelected){
    title = 'Inventario Actual — '+state.invSelected.replace('SC ','')+
      ' <button onclick="limpiarInvFiltro()" style="margin-left:8px;padding:2px 6px;background:#999;color:white;border:none;border-radius:3px;cursor:pointer;font-size:.65rem">✕</button>';
  }

  prods.forEach(function(p){
    var vals = fechas.map(function(f){
      var total = 0;
      tFiltro.forEach(function(t){
        if(detData[t] && detData[t][p]){
          total += (detData[t][p][f] || 0);
        }
      });
      return total;
    });
    var rowTotal = vals.reduce(function(s,v){ return s+v; }, 0);
    if(rowTotal > 0){
      vals.forEach(function(v,i){ totP[i] += v; });
      grandTotP += rowTotal;
      rowsProducto += '<tr><td>'+p.replace('BQT ','')+'</td>';
      vals.forEach(function(v){ rowsProducto += q(v); });
      rowsProducto += q(rowTotal);
      rowsProducto += '</tr>';
    }
  });
  rowsProducto += '<tr class="total"><td>Total</td>';
  totP.forEach(function(v){ rowsProducto += q(v); });
  rowsProducto += q(grandTotP);
  rowsProducto += '</tr>';

  document.getElementById('invProductoTitle').innerHTML = title;
  document.getElementById('tInvProducto').innerHTML = rowsProducto;
}

function renderGasto(){
  var sems = getSemanasActivas();
  var tiendas = getTiendasActivas();
  var gData = DATA.gasto_data;
  var pGasto = DATA.producto_gasto;
  
  var rutasActivasSet = {};
  tiendas.forEach(function(t) {
    if(typeof DATA !== 'undefined' && DATA.tienda_ruta && DATA.tienda_ruta[t]) {
      rutasActivasSet[DATA.tienda_ruta[t]] = true;
    }
  });
  
  // Rutas ordenadas y filtradas
  var rutasTodas = ['ENS', 'MXL 1', 'Ruta 2000', 'Rutas Playas'];
  var rutas = rutasTodas.filter(function(r) { return rutasActivasSet[r] || areAllTiendasSelected(state.tiendas_sel); });
  if(!rutas.length) rutas = rutasTodas;
  
  // Construir headers de tabla
  var headHTML = '<tr><th>RUTA / PRODUCTO</th>';
  sems.forEach(function(s){
    var yr = Math.floor(s/100), wk = s%100;
    var label = (yr >= 2000) ? yr+'-'+String(wk).padStart(2,'0') : String(s).padStart(2,'0');
    headHTML += '<th>'+label+'</th>';
  });
  headHTML += '<th>Grand Total</th></tr>';
  document.getElementById('tGastoHead').innerHTML = headHTML;
  
  // Construir filas de datos
  var bodyHTML = '';
  var grandTotals = {};
  var grandTotal = 0;
  
  sems.forEach(function(s){ grandTotals[s] = 0; });
  
  rutas.forEach(function(ruta){
    var rutaData = gData[ruta] || {};
    
    // Recopilar todos los productos que aparecen en esta ruta
    var productosSet = new Set();
    sems.forEach(function(s){
      var semData = rutaData[s] || {};
      for(var prod in semData){
        if(semData[prod] > 0) productosSet.add(prod);
      }
    });
    var productos = Array.from(productosSet).sort();
    
    if(productos.length === 0) return;
    
    // Fila de ruta con GASTO total
    var rutaTotal = 0;
    var rutaGastoSem = {};
    var rutaRow = '<tr class="ruta-row"><td><strong>'+ruta+'</strong></td>';
    sems.forEach(function(s){
      var semData = rutaData[s] || {};
      var gasto = 0;
      for(var prod in semData){
        var unidades = semData[prod];
        var gastoUnit = pGasto[prod] || 0;
        gasto += unidades * gastoUnit;
      }
      rutaGastoSem[s] = gasto;
      rutaTotal += gasto;
      grandTotals[s] += gasto;
      rutaRow += '<td>$'+fmt(gasto)+'</td>';
    });
    rutaRow += '<td>$'+fmt(rutaTotal)+'</td></tr>';
    grandTotal += rutaTotal;
    bodyHTML += rutaRow;
    
    // Filas de productos con UNIDADES embarcadas
    var rutaPiezasSem = {};
    sems.forEach(function(s){ rutaPiezasSem[s] = 0; });
    var rutaPiezasTotal = 0;

    productos.forEach(function(prod){
      var prodName = prod.replace('BQT ','');
      var prodRow = '<tr class="unidad-row"><td>&nbsp;&nbsp;'+prodName+'</td>';
      var prodTotal = 0;
      sems.forEach(function(s){
        var semData = rutaData[s] || {};
        var unidades = semData[prod] || 0;
        prodTotal += unidades;
        rutaPiezasSem[s] += unidades;
        prodRow += '<td>'+(unidades > 0 ? Math.round(unidades) : '')+'</td>';
      });
      rutaPiezasTotal += prodTotal;
      prodRow += '<td>'+(prodTotal > 0 ? Math.round(prodTotal) : '')+'</td></tr>';
      bodyHTML += prodRow;
    });

    var piezasRow = '<tr class="totalpiezas-row"><td>Total Piezas</td>';
    sems.forEach(function(s){
      var p = rutaPiezasSem[s];
      piezasRow += '<td>'+(p > 0 ? Math.round(p) : '')+'</td>';
    });
    piezasRow += '<td>'+(rutaPiezasTotal > 0 ? Math.round(rutaPiezasTotal) : '')+'</td></tr>';
    bodyHTML += piezasRow;

    var cppRow = '<tr class="costo-pieza-row"><td>Costo / Pieza</td>';
    sems.forEach(function(s){
      var p = rutaPiezasSem[s];
      var g = rutaGastoSem[s];
      cppRow += '<td>'+(p > 0 ? '$'+(g/p).toFixed(2) : '')+'</td>';
    });
    cppRow += '<td>'+(rutaPiezasTotal > 0 ? '$'+(rutaTotal/rutaPiezasTotal).toFixed(2) : '')+'</td></tr>';
    bodyHTML += cppRow;
  });
  
  // Fila de Grand Total
  var gtRow = '<tr class="grand-total"><td><strong>Grand Total</strong></td>';
  sems.forEach(function(s){
    gtRow += '<td><strong>$'+fmt(grandTotals[s])+'</strong></td>';
  });
  gtRow += '<td><strong>$'+fmt(grandTotal)+'</strong></td></tr>';
  bodyHTML += gtRow;
  
  document.getElementById('tGastoBody').innerHTML = bodyHTML;
  renderGasolina();
}

function renderGasolina(){
  var sems      = getSemanasActivas();
  var gData     = DATA.gasolina_por_vehiculo || {};
  var vehiculos = DATA.vehiculos_gasolina    || [];
  var gOtros    = DATA.gastos_otros          || {};
  var gTipos    = DATA.gastos_tipos          || Object.keys(gOtros);

  // ── Cabecera ──
  var headHTML = '<tr><th>Concepto</th>';
  sems.forEach(function(s){
    var yr = Math.floor(s/100), wk = s%100;
    var label = (yr >= 2000) ? yr+'-'+String(wk).padStart(2,'0') : String(s).padStart(2,'0');
    headHTML += '<th>'+label+'</th>';
  });
  headHTML += '<th>Grand Total</th></tr>';
  document.getElementById('tGasolinaHead').innerHTML = headHTML;

  var bodyHTML    = '';
  var grandTotals = {};
  var grandTotal  = 0;
  sems.forEach(function(s){ grandTotals[s] = 0; });

  // ── Sección: Gasolina por Vehículo ──
  if(vehiculos.length > 0){
    var vehRowsHTML = '';
    vehiculos.forEach(function(veh){
      var vehData  = gData[veh] || {};
      var vehTotal = 0;
      var vehRow   = '<tr class="ruta-row"><td>'+veh+'</td>';
      sems.forEach(function(s){
        var total = vehData[s] || 0;
        vehTotal += total;
        vehRow += '<td>'+(total !== 0 ? '$'+fmt(total) : '—')+'</td>';
      });
      vehRow += '<td>$'+fmt(vehTotal)+'</td></tr>';
      if(vehTotal !== 0){
        sems.forEach(function(s){ grandTotals[s] += (vehData[s] || 0); });
        grandTotal += vehTotal;
        vehRowsHTML += vehRow;
      }
    });
    if(vehRowsHTML !== ''){
      bodyHTML += '<tr class="section-row"><td colspan="'+(sems.length+2)+'">Gasolina por Vehículo</td></tr>';
      bodyHTML += vehRowsHTML;
    }
  }

  // ── Sección: Nómina ──
  var nomData = DATA.nomina_data || {};
  if(Object.keys(nomData).length > 0){
    var nomRow = '<tr class="ruta-row"><td>GASTO NOMINA</td>';
    var nomTotal = 0;
    sems.forEach(function(s){
      var total = nomData[s] || 0;
      nomTotal += total;
      nomRow += '<td>'+(total !== 0 ? '$'+fmt(total) : '—')+'</td>';
    });
    nomRow += '<td>$'+fmt(nomTotal)+'</td></tr>';
    if(nomTotal !== 0){
      sems.forEach(function(s){ grandTotals[s] += (nomData[s] || 0); });
      grandTotal += nomTotal;
      bodyHTML += '<tr class="section-row"><td colspan="'+(sems.length+2)+'">Nómina</td></tr>';
      bodyHTML += nomRow;
    }
  }

  // ── Sección: Otros Gastos (hoja Gastos del SharePoint) agrupada por SC ──
  var gSCs = DATA.gastos_sc || Object.keys(gOtros);
  var scConDatos = gSCs.filter(function(sc){ return gOtros[sc]; });
  if(scConDatos.length > 0){
    var scSectionHTML = '';
    scConDatos.forEach(function(sc){
      var scData = gOtros[sc] || {};
      var scTotals = {};
      var scGrandTotal = 0;
      sems.forEach(function(s){ scTotals[s] = 0; });
      
      var tiposEnSC = gTipos.filter(function(t){ return scData[t]; });
      var scRowsHTML = '';
      
      tiposEnSC.forEach(function(tipo){
        var tipoData  = scData[tipo] || {};
        var tipoTotal = 0;
        var tipoRow   = '<tr class="ruta-row"><td>'+tipo+'</td>';
        sems.forEach(function(s){
          var total = tipoData[s] || 0;
          tipoTotal += total;
          tipoRow += '<td>'+(total !== 0 ? '$'+fmt(total) : '—')+'</td>';
        });
        tipoRow += '<td>$'+fmt(tipoTotal)+'</td></tr>';
        
        if(tipoTotal !== 0){
          sems.forEach(function(s){ 
            var total = tipoData[s] || 0;
            scTotals[s] += total;
            grandTotals[s] += total;
          });
          scGrandTotal += tipoTotal;
          grandTotal += tipoTotal;
          scRowsHTML += tipoRow;
        }
      });
      
      if(scGrandTotal !== 0){
        scSectionHTML += '<tr class="group-row"><td colspan="'+(sems.length+2)+'">'+sc+'</td></tr>';
        scSectionHTML += scRowsHTML;
        var subtotalRow = '<tr class="subtotal-row"><td>Total '+sc+'</td>';
        sems.forEach(function(s){
          subtotalRow += '<td>'+(scTotals[s] !== 0 ? '$'+fmt(scTotals[s]) : '—')+'</td>';
        });
        subtotalRow += '<td>$'+fmt(scGrandTotal)+'</td></tr>';
        scSectionHTML += subtotalRow;
      }
    });
    
    if(scSectionHTML !== ''){
      bodyHTML += '<tr class="section-row"><td colspan="'+(sems.length+2)+'">Otros Gastos</td></tr>';
      bodyHTML += scSectionHTML;
    }
  }

  // ── Grand Total general ──
  var gtRow = '<tr class="grand-total"><td><strong>Grand Total</strong></td>';
  sems.forEach(function(s){
    gtRow += '<td><strong>$'+fmt(grandTotals[s])+'</strong></td>';
  });
  gtRow += '<td><strong>$'+fmt(grandTotal)+'</strong></td></tr>';
  bodyHTML += gtRow;

  document.getElementById('tGasolinaBody').innerHTML = bodyHTML;
}

function fmtMoney(v){ return '$'+fmt(v); }

function resumenSemLabel(s){
  var n = parseInt(s, 10);
  if(!n) return String(s);
  var yr = Math.floor(n/100), wk = n%100;
  return String(yr >= 2000 ? wk : n);
}

function parseResumenDate(s){
  if(!s) return null;
  var m = /^(\d{4})[\/\-](\d{2})[\/\-](\d{2})$/.exec(String(s).trim());
  if(!m) return null;
  return new Date(Number(m[1]), Number(m[2]) - 1, Number(m[3]));
}

function addDays(baseDate, days){
  if(!(baseDate instanceof Date) || isNaN(baseDate.getTime())) return null;
  var dt = new Date(baseDate.getTime());
  dt.setDate(dt.getDate() + days);
  return dt;
}

function fmtDecimal(v, digits){
  if(v === null || v === undefined || !isFinite(v)) return 'N/A';
  return Number(v).toLocaleString('es-MX', {minimumFractionDigits: digits, maximumFractionDigits: digits});
}

function fmtShortDate(dt){
  if(!(dt instanceof Date) || isNaN(dt.getTime())) return 'N/A';
  return String(dt.getDate()).padStart(2,'0') + '/' + String(dt.getMonth()+1).padStart(2,'0');
}

function fmtDayDate(dt){
  if(!(dt instanceof Date) || isNaN(dt.getTime())) return 'N/A';
  var dia = DIAS[dt.getDay()] || '';
  return dia.charAt(0).toUpperCase() + dia.slice(1) + ' ' + fmtShortDate(dt);
}

function daysBetweenDates(baseDate, targetDate){
  if(!(baseDate instanceof Date) || isNaN(baseDate.getTime())) return null;
  if(!(targetDate instanceof Date) || isNaN(targetDate.getTime())) return null;
  var ms = targetDate.getTime() - baseDate.getTime();
  return Math.max(0, Math.round(ms / 86400000));
}

var _routeSchedulesCache = null;
var _storeSchedulesCache = null;

function getStoreSchedules(){
  if(_storeSchedulesCache) return _storeSchedulesCache;

  var resumenSrc = DATA.resumen_diario || {};
  _storeSchedulesCache = {};

  Object.keys(resumenSrc).forEach(function(tienda){
    var tiendaData = resumenSrc[tienda] || {};
    var embarqueDates = {};

    Object.keys(tiendaData).forEach(function(sem){
      var semanaData = tiendaData[sem] || {};
      Object.keys(semanaData).forEach(function(prod){
        Object.keys(semanaData[prod] || {}).forEach(function(fecha){
          var emb = ((semanaData[prod][fecha] || {}).embarque || 0);
          if(emb > 0) embarqueDates[fecha] = true;
        });
      });
    });

    var recentDates = Object.keys(embarqueDates)
      .map(parseResumenDate)
      .filter(function(dt){ return dt instanceof Date && !isNaN(dt.getTime()); })
      .sort(function(a, b){ return a - b; })
      .slice(-12);

    var counts = [0,0,0,0,0,0,0];
    recentDates.forEach(function(dt){ counts[dt.getDay()] += 1; });

    var maxCount = Math.max.apply(null, counts.concat([0]));
    var dias = [];
    counts.forEach(function(count, dow){
      if(count >= 2 && count >= maxCount * 0.6){
        dias.push(dow);
      }
    });

    if(!dias.length && maxCount > 0){
      counts.forEach(function(count, dow){ if(count === maxCount) dias.push(dow); });
    }

    _storeSchedulesCache[tienda] = dias.sort(function(a, b){ return a - b; });
  });

  return _storeSchedulesCache;
}

function getRouteSchedules(){
  if(_routeSchedulesCache) return _routeSchedulesCache;

  var resumenSrc = DATA.resumen_diario || {};
  var tiendaRuta = DATA.tienda_ruta || {};
  var routeDates = {};

  Object.keys(resumenSrc).forEach(function(tienda){
    var ruta = tiendaRuta[tienda] || 'Sin ruta';
    if(!routeDates[ruta]) routeDates[ruta] = {};
    var tiendaData = resumenSrc[tienda] || {};

    Object.keys(tiendaData).forEach(function(sem){
      var semanaData = tiendaData[sem] || {};
      Object.keys(semanaData).forEach(function(prod){
        Object.keys(semanaData[prod] || {}).forEach(function(fecha){
          var emb = ((semanaData[prod][fecha] || {}).embarque || 0);
          if(emb <= 0) return;
          routeDates[ruta][fecha] = (routeDates[ruta][fecha] || 0) + emb;
        });
      });
    });
  });

  _routeSchedulesCache = {};
  Object.keys(routeDates).forEach(function(ruta){
    var counts = [0,0,0,0,0,0,0];
    var weights = [0,0,0,0,0,0,0];

    Object.keys(routeDates[ruta]).forEach(function(fecha){
      var dt = parseResumenDate(fecha);
      if(!dt) return;
      var dow = dt.getDay();
      counts[dow] += 1;
      weights[dow] += routeDates[ruta][fecha] || 0;
    });

    var maxCount = Math.max.apply(null, counts.concat([0]));
    var maxWeight = Math.max.apply(null, weights.concat([0]));
    var dias = [];

    counts.forEach(function(count, dow){
      if(count <= 0) return;
      var weight = weights[dow] || 0;
      if(count >= maxCount * 0.6 || weight >= maxWeight * 0.6){
        dias.push(dow);
      }
    });

    if(!dias.length && maxCount > 0){
      counts.forEach(function(count, dow){ if(count === maxCount) dias.push(dow); });
    }

    _routeSchedulesCache[ruta] = dias.sort(function(a,b){ return a - b; });
  });

  return _routeSchedulesCache;
}

function getUpcomingRouteDates(routeDays, lastDate, count){
  if(!(lastDate instanceof Date) || isNaN(lastDate.getTime()) || !routeDays || !routeDays.length) return [];
  var dates = [];
  for(var offset = 1; offset <= 28 && dates.length < count; offset++){
    var dt = addDays(lastDate, offset);
    if(routeDays.indexOf(dt.getDay()) !== -1) dates.push(dt);
  }
  return dates;
}

function getSafetyDays(avgDiaria, ventasPorFecha){
  if(!(avgDiaria > 0)) return 0;
  var vals = Object.keys(ventasPorFecha || {}).map(function(fecha){
    return Number((ventasPorFecha || {})[fecha] || 0);
  });
  if(!vals.length) return 1;
  var peak = 0;
  var sumSq = 0;
  vals.forEach(function(v){
    if(v > peak) peak = v;
    var diff = v - avgDiaria;
    sumSq += diff * diff;
  });
  var stdev = Math.sqrt(sumSq / vals.length);
  var cv = avgDiaria > 0 ? (stdev / avgDiaria) : 0;
  return (vals.length >= 4 && (cv >= 0.45 || peak >= avgDiaria * 1.75)) ? 2 : 1;
}

function getOrderMultiple(producto){
  var fuentes = [DATA.producto_multiplo, DATA.producto_empaque, DATA.producto_lote];
  for(var i = 0; i < fuentes.length; i++){
    var src = fuentes[i] || {};
    var raw = src[producto];
    var val = Number(raw);
    if(isFinite(val) && val > 0) return val;
  }
  return 1;
}

function roundUpToMultiple(qty, multiple){
  if(!(qty > 0)) return 0;
  var mult = (isFinite(multiple) && multiple > 0) ? multiple : 1;
  if(mult <= 1) return Math.ceil(qty);
  return Math.ceil(qty / mult) * mult;
}

function getPriorityLabel(avgDiaria, diasCobertura, daysToNextRoute, orderFinal, routeDetected){
  if(!(avgDiaria > 0)) return {key:'idle', label:'Sin mov.', reason:'Sin venta reciente'};
  if(!routeDetected){
    if(!(orderFinal > 0)) return {key:'ok', label:'OK', reason:'Inventario suficiente'};
    if(diasCobertura !== null && diasCobertura <= 2) return {key:'today', label:'Pedir hoy', reason:'Ruta sin patron; revisar manual'};
    return {key:'plan', label:'Programar', reason:'Ruta sin patron; planear manualmente'};
  }
  if(diasCobertura === null || diasCobertura < daysToNextRoute){
    return {key:'critical', label:'Critico', reason:'Quiebre antes de proxima entrega'};
  }
  if(!(orderFinal > 0)){
    return {key:'ok', label:'OK', reason:'Inventario suficiente hasta siguiente ciclo'};
  }
  if(diasCobertura <= daysToNextRoute + 1){
    return {key:'today', label:'Pedir hoy', reason:'Cobertura justa para llegar a la ruta'};
  }
  return {key:'plan', label:'Programar', reason:'Reponer ciclo y stock de seguridad'};
}

function buildReabastoRows(){
  // Productos excluidos de reabasto (órdenes especiales que inflan la proyección)
  var PRODUCTOS_EXCLUIDOS = [
    'BQT 18 ROSAS',
    'BQT MDAY PREMIUM',
    'ROSAS 12 MDAY',
    'JARRON MDAY',
    'BQT MDAY M'
  ];

  var sems = getSemanasActivas().map(function(s){ return String(s); });
  var tiendas = getTiendasActivas();
  var resumenSrc = DATA.resumen_diario || {};
  var rows = [];

  tiendas.forEach(function(t){
    var tiendaData = resumenSrc[t] || {};
    var aggByProd = {};
    var latestByProd = {};

    sems.forEach(function(s){
      var semanaData = tiendaData[s] || {};
      Object.keys(semanaData).forEach(function(p){
        if(PRODUCTOS_EXCLUIDOS.indexOf(p.trim().toUpperCase()) !== -1) return;
        if(!aggByProd[p]){
          aggByProd[p] = {ventas:0, embarque:0, merma:0, fechas:{}, dias:0, ventasPorFecha:{}};
        }
        Object.keys(semanaData[p] || {}).forEach(function(fecha){
          var cell = semanaData[p][fecha] || {};
          aggByProd[p].ventas += cell.ventas || 0;
          aggByProd[p].embarque += cell.embarque || 0;
          aggByProd[p].merma += cell.merma || 0;
          aggByProd[p].ventasPorFecha[fecha] = (aggByProd[p].ventasPorFecha[fecha] || 0) + (cell.ventas || 0);
          if(!aggByProd[p].fechas[fecha]){
            aggByProd[p].fechas[fecha] = true;
            aggByProd[p].dias += 1;
          }

          var dt = parseResumenDate(fecha);
          var prev = latestByProd[p];
          if(!prev || (dt && prev.dt && dt > prev.dt) || (dt && !prev.dt)){
            latestByProd[p] = {
              fecha: fecha,
              dt: dt,
              inventario: cell.inventario || 0,
            };
          }
        });
      });
    });

    Object.keys(aggByProd).forEach(function(p){
      if(PRODUCTOS_EXCLUIDOS.indexOf(p.trim().toUpperCase()) !== -1) return;
      var agg = aggByProd[p];
      var last = latestByProd[p] || {inventario:0, dt:null};
      var avgDiaria = agg.dias > 0 ? agg.ventas / agg.dias : 0;
      var inventario = last.inventario || 0;
      var diasCobertura = avgDiaria > 0 ? inventario / avgDiaria : null;
      var mermaPct = agg.embarque > 0 ? (agg.merma / agg.embarque) * 100 : 0;
      var ruta = (DATA.tienda_ruta && DATA.tienda_ruta[t]) || 'Sin ruta';
      var routeDays = getStoreSchedules()[t] || [];
      if(!routeDays.length) routeDays = getRouteSchedules()[ruta] || [];
      var nextDates = getUpcomingRouteDates(routeDays, last.dt, 2);
      var nextRouteDate = nextDates[0] || null;
      var followingRouteDate = nextDates[1] || null;
      var routeDetected = !!nextRouteDate;
      var daysToNextRoute = routeDetected ? daysBetweenDates(last.dt, nextRouteDate) : 1;
      var routeCycleDays = routeDetected
        ? (followingRouteDate ? daysBetweenDates(nextRouteDate, followingRouteDate) : 7)
        : 3;
      var safetyDays = getSafetyDays(avgDiaria, agg.ventasPorFecha);
      var targetDays = (daysToNextRoute || 0) + (routeCycleDays || 0) + safetyDays;
      var orderRaw = avgDiaria > 0 ? Math.max(0, (avgDiaria * targetDays) - inventario) : 0;
      var multiple = getOrderMultiple(p);
      var orderFinal = roundUpToMultiple(orderRaw, multiple);
      var prioridad = getPriorityLabel(avgDiaria, diasCobertura, daysToNextRoute || 0, orderFinal, routeDetected);

      if(avgDiaria <= 0 && inventario <= 0 && agg.ventas <= 0) return;

      rows.push({
        ruta: ruta,
        tienda: t,
        producto: p,
        inventario: inventario,
        avgDiaria: avgDiaria,
        diasCobertura: diasCobertura,
        entrega: nextRouteDate ? fmtDayDate(nextRouteDate) : 'Sin patron',
        diasEntrega: daysToNextRoute,
        safetyDays: safetyDays,
        multiple: multiple,
        ordenRaw: orderRaw,
        ordenFinal: orderFinal,
        routeCycleDays: routeCycleDays,
        mermaPct: mermaPct,
        ventas: agg.ventas,
        baseDate: last.dt,
        prioridad: prioridad,
      });
    });
  });

  rows.sort(function(a, b){
    var order = {critical:0, today:1, plan:2, ok:3, idle:4};
    var oa = order[a.prioridad.key] !== undefined ? order[a.prioridad.key] : 9;
    var ob = order[b.prioridad.key] !== undefined ? order[b.prioridad.key] : 9;
    if(oa !== ob) return oa - ob;
    var da = a.diasEntrega === null ? 999999 : a.diasEntrega;
    var db = b.diasEntrega === null ? 999999 : b.diasEntrega;
    if(da !== db) return da - db;
    if(b.ordenFinal !== a.ordenFinal) return b.ordenFinal - a.ordenFinal;
    return a.producto.localeCompare(b.producto);
  });

  return rows;
}

function getReabastoBaseWeek(){
  if(state.semana && state.semana !== 'all') return state.semana;
  return (DATA.semanas && DATA.semanas.length) ? DATA.semanas[DATA.semanas.length-1] : null;
}

function getReabastoWindowWeeks(){
  var baseWeek = getReabastoBaseWeek();
  var semanas = DATA.semanas || [];
  var idx = semanas.indexOf(baseWeek);
  if(idx < 0) idx = semanas.length - 1;
  var win = parseInt(state.reabastoWindow) || 3;
  return semanas.slice(Math.max(0, idx-(win-1)), idx+1).map(function(s){ return String(s); });
}

function setReabastoWindow(val){
  state.reabastoWindow = parseInt(val) || 3;
  renderReabastoPanel();
}

function exportReabastoExcel(){
  var table = document.querySelector('.reabasto-table');
  if(!table) return;
  var html = '<html xmlns:x="urn:schemas-microsoft-com:office:excel"><meta charset="utf-8"><head><!--[if gte mso 9]><xml><x:ExcelWorkbook><x:ExcelWorksheets><x:ExcelWorksheet><x:Name>Reabasto</x:Name><x:WorksheetOptions><x:DisplayGridlines/></x:WorksheetOptions></x:ExcelWorksheet></x:ExcelWorksheets></x:ExcelWorkbook></xml><![endif]--></head><body>';
  html += '<table border="1">' + table.innerHTML + '</table></body></html>';
  var blob = new Blob([html], {type: 'application/vnd.ms-excel'});
  var url = URL.createObjectURL(blob);
  var a = document.createElement('a');
  a.href = url;
  a.download = 'Orden_Reabasto_Sem' + getReabastoBaseWeek() + '.xls';
  a.click();
  URL.revokeObjectURL(url);
}

function getStoreDeliveryDaysForWeeks(tienda, sems){
  var resumenSrc = DATA.resumen_diario || {};
  var tiendaData = resumenSrc[tienda] || {};
  var embarqueDates = {};

  (sems || []).forEach(function(s){
    var semanaData = tiendaData[String(s)] || {};
    Object.keys(semanaData).forEach(function(prod){
      Object.keys(semanaData[prod] || {}).forEach(function(fecha){
        var emb = ((semanaData[prod][fecha] || {}).embarque || 0);
        if(emb > 0) embarqueDates[fecha] = true;
      });
    });
  });

  var counts = [0,0,0,0,0,0,0];
  Object.keys(embarqueDates).forEach(function(fecha){
    var dt = parseResumenDate(fecha);
    if(!dt) return;
    var dow = dt.getDay() - 1;
    if(dow < 0) dow = 6;
    counts[dow] += 1;
  });

  var maxCount = Math.max.apply(null, counts.concat([0]));
  var dias = [];
  counts.forEach(function(count, dow){
    if(count >= 2 && count >= maxCount * 0.6) dias.push(dow);
  });
  if(!dias.length && maxCount > 0){
    counts.forEach(function(count, dow){ if(count === maxCount) dias.push(dow); });
  }
  return dias.sort(function(a, b){ return a - b; });
}

function renderReabastoPanel(){
  var head = document.querySelector('.reabasto-table thead') || document.querySelector('#tReabastoBody').parentElement.querySelector('thead');
  var body = document.getElementById('tReabastoBody');
  var meta = document.getElementById('reabastoMeta');
  var PRODUCTOS_EXCLUIDOS_PANEL = [
    'BQT 18 ROSAS',
    'BQT MDAY PREMIUM',
    'ROSAS 12 MDAY',
    'JARRON MDAY',
    'BQT MDAY M'
  ];
  var baseWeek = getReabastoBaseWeek();
  var semsAct = getReabastoWindowWeeks();
  var baseWeekKey = String(baseWeek || '');
  var tiendas = getTiendasActivas();
  var productos = (getProductosActivos() || []).filter(function(p){
    return PRODUCTOS_EXCLUIDOS_PANEL.indexOf(p.trim().toUpperCase()) === -1;
  });
  var tiendaRuta = DATA.tienda_ruta || {};

  if(meta){
    meta.textContent = 'Proyección de reabasto calculada dinámicamente según la ventana de tiempo seleccionada (' + state.reabastoWindow + ' semana(s)).';
  }

  if(head){
    var thr = '<tr><th style="min-width:140px; text-align:left; background:#eef4fb; color:#1f2937;">Tienda / Día</th>';
    productos.forEach(function(p){
      thr += '<th style="text-align:center; background:#eef4fb; color:#1f2937;">'+p.replace('BQT ','')+'</th>';
    });
    thr += '<th style="text-align:center; background:#eef4fb; color:#1f2937; font-weight:bold;">TOTAL</th></tr>';
    head.innerHTML = thr;
  }

  var storeProjs = {};
  tiendas.forEach(function(t){
    storeProjs[t] = {};
    var myDows = getStoreDeliveryDaysForWeeks(t, semsAct);
    if(!myDows.length){
      var ruta = tiendaRuta[t] || 'Sin ruta';
      myDows = (getRouteSchedules()[ruta] || []).map(function(dow){
        var mondayDow = dow - 1;
        return mondayDow < 0 ? 6 : mondayDow;
      }).sort(function(a, b){ return a - b; });
    }

    productos.forEach(function(p){
      var salesDow = [0,0,0,0,0,0,0];
      var sum_vtas = 0, sum_merma = 0, sum_emb = 0;

      semsAct.forEach(function(s){
        var raw = (DATA.raw_prod_semana && DATA.raw_prod_semana[t] && DATA.raw_prod_semana[t][String(s)] && DATA.raw_prod_semana[t][String(s)][p]) || {};
        salesDow[0] += raw.ctd_lun || 0;
        salesDow[1] += raw.ctd_mar || 0;
        salesDow[2] += raw.ctd_mie || 0;
        salesDow[3] += raw.ctd_jue || 0;
        salesDow[4] += raw.ctd_vie || 0;
        salesDow[5] += raw.ctd_sab || 0;
        salesDow[6] += raw.ctd_dom || 0;
        
        sum_vtas += raw.ventas_u || 0;
        sum_merma += raw.merma_u || 0;
        sum_emb += raw.embarque_u || 0;
      });
      
      var avg = semsAct.length > 0 ? sum_vtas / semsAct.length : 0;
      var merma_ratio = sum_emb > 0 ? sum_merma / sum_emb : 0;
      var storeProj = merma_ratio < 1 ? avg / (1 - merma_ratio) : avg;

      var alloc = {};
      if(myDows.length > 0){
        var bucketSales = {};
        myDows.forEach(function(dow){ bucketSales[dow] = 0; });
        var totalSales = 0;
        for(var dow = 0; dow < 7; dow++){
          totalSales += salesDow[dow];
          var chosenBucket = myDows[0];
          for(var i = myDows.length - 1; i >= 0; i--){
            if(myDows[i] <= dow){
              chosenBucket = myDows[i];
              break;
            }
          }
          if(dow < myDows[0]) chosenBucket = myDows[myDows.length-1];
          bucketSales[chosenBucket] += salesDow[dow];
        }
        myDows.forEach(function(dow){
          alloc[dow] = totalSales > 0 ? storeProj * (bucketSales[dow] / totalSales) : (storeProj / myDows.length);
        });
      }

      storeProjs[t][p] = { proj: storeProj, alloc: alloc, dows: myDows };
    });
  });

  var html = '';
  var dowNames = ['Lunes','Martes','Miercoles','Jueves','Viernes','Sabado','Domingo'];
  tiendas.sort().forEach(function(t){
    var myDows = [];
    productos.forEach(function(p){
      var dows = (storeProjs[t][p] && storeProjs[t][p].dows) || [];
      dows.forEach(function(dow){ if(myDows.indexOf(dow) === -1) myDows.push(dow); });
    });
    myDows.sort(function(a, b){ return a - b; });

    html += '<tr style="background:#f9fafb; border-top:2px solid #e5e7eb;">' +
            '<td class="bold" style="text-align:left; padding-top:10px; color:#111827;">'+t+'</td>' +
            '<td colspan="'+(productos.length+1)+'"></td></tr>';

    if(myDows.length === 0){
      var tds = '<td style="padding-left:20px; text-align:left;">Semanal</td>';
      var gTotal = 0;
      productos.forEach(function(p){
        var val = Math.round((storeProjs[t][p] && storeProjs[t][p].proj) || 0);
        gTotal += val;
        tds += '<td style="text-align:center;">'+(val||'-')+'</td>';
      });
      tds += '<td class="bold" style="text-align:center; color:#374151;">'+gTotal+'</td>';
      html += '<tr>'+tds+'</tr>';
    } else {
      myDows.forEach(function(dow){
        var dRow = '<td style="padding-left:20px; text-align:left; color:#4b5563;">'+dowNames[dow]+'</td>';
        var dTotal = 0;
        productos.forEach(function(p){
          var val = Math.round(((storeProjs[t][p] && storeProjs[t][p].alloc && storeProjs[t][p].alloc[dow]) || 0));
          dTotal += val;
          dRow += '<td style="text-align:center;">'+(val||'-')+'</td>';
        });
        dRow += '<td class="bold" style="text-align:center; color:#374151;">'+dTotal+'</td>';
        html += '<tr>'+dRow+'</tr>';
      });
    }

    var totRow = '<td class="bold" style="text-align:right; color:#0071ce; padding-bottom:10px;">TOTAL</td>';
    var finalTotal = 0;
    productos.forEach(function(p){
      var val = Math.round((storeProjs[t][p] && storeProjs[t][p].proj) || 0);
      finalTotal += val;
      totRow += '<td class="bold" style="color:#0071ce; text-align:center; padding-bottom:10px;">'+(val||'-')+'</td>';
    });
    totRow += '<td class="bold" style="color:#0071ce; text-align:center; padding-bottom:10px;">'+finalTotal+'</td>';
    html += '<tr style="border-bottom:2px solid #e5e7eb;">'+totRow+'</tr>';
  });

  if(!tiendas.length){
    html = '<tr><td colspan="'+(productos.length+2)+'" style="text-align:center; padding:20px;">Selecciona tiendas para calcular la proyección.</td></tr>';
  }

  body.innerHTML = html;
}

function getResumenMetricDefs(){
  var defs = [
    {key:'embarque', label:'Sum of Cntd Embarque', type:'num', bgEven:'#eef6fc', bgOdd:'#eef6fc'},
    {key:'ventas',   label:'Sum of Cnt POS',        type:'num', bgEven:'#ebfceb', bgOdd:'#ebfceb'},
    {key:'sell_thru',label:'Sum of Sell Thru',      type:'pct', bgEven:'#fcf8e3', bgOdd:'#fcf8e3'}
  ];
  if(DATA.resumen_has_field1) defs.push({key:'field1', label:'Sum of Field1', type:'num', bgEven:'#ffffff', bgOdd:'#f4f6fa'});
  return defs;
}

/* ── Helper: obtener filas programado guardadas para un (prod, tienda) ── */
function _getSavedProgramadoRows(r1Key, r2Key) {
  if(!window._captureProjections) return [];
  var modeKey = _captureProjModeKey();
  var store = window._captureProjections[modeKey];
  if(!store) return [];
  var pivotMode = state.resumenPivot || 'producto';
  var prod = pivotMode === 'producto' ? r1Key : r2Key;
  var tienda = pivotMode === 'producto' ? r2Key : r1Key;
  var key = _captureProjStoreKey(prod, tienda);
  var entry = store[key];
  if(!entry || !entry.rows || !entry.rows.length) return [];
  /* Filtrar solo filas con datos y que estén marcadas como guardadas */
  return entry.rows.filter(function(r){
    return r.saved && (r.sem || (r.values||[]).some(function(v){ return v !== '' && parseFloat(v) > 0; }));
  });
}

/* ── Persistencia de capturas: Supabase (compartido entre todos los usuarios) ── */
/* Antes vivía solo en localStorage (solo el navegador de cada persona la veía).
   Ahora se guarda en una tabla de Supabase para que todos vean lo mismo.
   Se mantiene una copia en localStorage como respaldo/caché por si se pierde
   la conexión, pero la fuente de verdad es Supabase. */
var _CAPTURE_LS_KEY = 'walmex_cfbc_capture_v1';
var _captureLoaded = false;     // true cuando ya llegó la primera respuesta de Supabase
var _capturePersistInFlight = null;

/* Las llaves internas usan un byte nulo ('\x00') para combinar producto+tienda
   (ver _captureProjStoreKey). Postgres/Supabase NO acepta ese byte dentro de
   texto/JSON ("\u0000 cannot be converted to text"), así que antes de mandar
   o recibir datos de Supabase lo intercambiamos por un símbolo seguro. */
var _NULL_SEP = '\x00';
var _SAFE_SEP = '\u241F'; // símbolo imprimible, no es un carácter de control

function _swapKeySeparators(obj, fromSep, toSep) {
  if (Array.isArray(obj)) {
    return obj.map(function (item) { return _swapKeySeparators(item, fromSep, toSep); });
  }
  if (obj && typeof obj === 'object') {
    var out = {};
    Object.keys(obj).forEach(function (k) {
      var newKey = k.indexOf(fromSep) >= 0 ? k.split(fromSep).join(toSep) : k;
      out[newKey] = _swapKeySeparators(obj[k], fromSep, toSep);
    });
    return out;
  }
  return obj;
}

function _captureForSupabase(data) {
  return _swapKeySeparators(data, _NULL_SEP, _SAFE_SEP);
}

function _captureFromSupabase(data) {
  return _swapKeySeparators(data, _SAFE_SEP, _NULL_SEP);
}

/* Llaves (producto\x00tienda) borradas localmente desde el último _persistCapture
   exitoso. Se usan para que el merge contra Supabase respete las eliminaciones
   en vez de "revivirlas" porque seguían existiendo del lado remoto. */
window._captureRemovedKeys = window._captureRemovedKeys || { sem: {}, norm: {} };

function _markCaptureKeyRemoved(modeKey, sk) {
  if (!window._captureRemovedKeys) window._captureRemovedKeys = { sem: {}, norm: {} };
  if (!window._captureRemovedKeys[modeKey]) window._captureRemovedKeys[modeKey] = {};
  window._captureRemovedKeys[modeKey][sk] = true;
}

/* Combina lo que ya estaba guardado en Supabase con los cambios hechos en este
   navegador, en vez de reemplazar el registro completo. Por cada llave
   (producto+tienda) se usa la versión local si existe (es el cambio más
   reciente de esta sesión); si una llave solo existe en remoto, se conserva
   (son capturas guardadas por otra persona, o por esta misma sesión antes de
   que terminara de cargar). Las llaves marcadas como eliminadas localmente se
   quitan del resultado final aunque sigan en remoto. */
function _mergeCaptureStores(remote, local, removedKeys) {
  var remoteObj = (remote && typeof remote === 'object') ? remote : { sem: {}, norm: {}, _meta: {} };
  var localObj = (local && typeof local === 'object') ? local : { sem: {}, norm: {}, _meta: {} };
  var out = { sem: {}, norm: {}, _meta: localObj._meta || remoteObj._meta || {} };
  ['sem', 'norm'].forEach(function (modeKey) {
    var remoteStore = (remoteObj[modeKey] && typeof remoteObj[modeKey] === 'object') ? remoteObj[modeKey] : {};
    var localStore = (localObj[modeKey] && typeof localObj[modeKey] === 'object') ? localObj[modeKey] : {};
    var merged = {};
    Object.keys(remoteStore).forEach(function (k) { merged[k] = remoteStore[k]; });
    Object.keys(localStore).forEach(function (k) { merged[k] = localStore[k]; });
    var removed = (removedKeys && removedKeys[modeKey]) || {};
    Object.keys(removed).forEach(function (k) { delete merged[k]; });
    out[modeKey] = merged;
  });
  return out;
}

function _showCaptureSaveError(msg) {
  try {
    var el = document.getElementById('captureSaveError');
    if (!el) {
      el = document.createElement('div');
      el.id = 'captureSaveError';
      el.style.cssText = 'position:fixed;bottom:16px;right:16px;background:#fff5f5;color:#c62828;border:1px solid #c62828;border-radius:8px;padding:10px 14px;font-size:13px;z-index:99999;max-width:320px;box-shadow:0 2px 8px rgba(0,0,0,.15);font-family:"Segoe UI",sans-serif;';
      document.body.appendChild(el);
    }
    el.textContent = '⚠️ ' + msg;
    el.style.display = 'block';
    clearTimeout(el._hideTimer);
    el._hideTimer = setTimeout(function () { el.style.display = 'none'; }, 7000);
  } catch (e) {}
}

function _persistCapture() {
  // Obsoleto: ahora se guarda fila por fila
}

function _saveRowToSupabase(mode, prod, tienda, sem, vals) {
  if (!_supabaseConfigured()) return;
  var payload = { mode: mode, producto: prod, tienda: tienda, semana: sem, valores: vals };
  fetch(SUPABASE_URL + '/rest/v1/' + SUPABASE_CAPTURE_TABLE + '?on_conflict=mode,producto,tienda,semana', {
    method: 'POST',
    headers: _supabaseHeaders({ 'Prefer': 'resolution=merge-duplicates,return=minimal' }),
    body: JSON.stringify(payload)
  }).catch(function(e){ console.error(e); });
}

function _deleteRowFromSupabase(mode, prod, tienda, sem) {
  if (!_supabaseConfigured()) return;
  var url = SUPABASE_URL + '/rest/v1/' + SUPABASE_CAPTURE_TABLE + '?mode=eq.' + encodeURIComponent(mode) + '&producto=eq.' + encodeURIComponent(prod) + '&tienda=eq.' + encodeURIComponent(tienda) + '&semana=eq.' + encodeURIComponent(sem);
  fetch(url, { method: 'DELETE', headers: _supabaseHeaders() }).catch(function(e){ console.error(e); });
}

function _loadPersistedCapture() {
  if (!_supabaseConfigured()) {
    _captureLoaded = true;
    return Promise.resolve();
  }
  var url = SUPABASE_URL + '/rest/v1/' + SUPABASE_CAPTURE_TABLE + '?select=*';
  return fetch(url, { headers: _supabaseHeaders() })
    .then(function (resp) {
      if (!resp.ok) return [];
      return resp.json();
    })
    .then(function (rows) {
      var store = { sem: {}, norm: {}, _meta: { visibleInitial: 0 } };
      rows.forEach(function(r) {
        var sk = _captureProjStoreKey(r.producto, r.tienda);
        if(!store[r.mode]) store[r.mode] = {};
        if(!store[r.mode][sk]) store[r.mode][sk] = { visibleInitial: 0, rows: [] };
        
        var hasVals = (r.valores || []).some(function (v) {
           return v !== '' && parseFloat(v) > 0;
        });
        
        store[r.mode][sk].rows.push({
           sem: r.semana,
           values: r.valores,
           hidden: false,
           saved: hasVals
        });
      });
      window._captureProjections = store;
      window._captureVisibleInitial = 0;
      _captureLoaded = true;
      if (typeof state !== 'undefined' && state.view === 'resumen' && typeof renderResumen === 'function') {
        renderResumen();
      }
    })
    .catch(function (e) {
      console.warn("Load failed", e);
      _captureLoaded = true;
    });
}

function _clearPersistedCapture() {
  try { localStorage.removeItem(_CAPTURE_LS_KEY); } catch (e) {}
  window._captureProjections = { sem: {}, norm: {}, _meta: {} };
  window._captureRemovedKeys = { sem: {}, norm: {} };
  window._captureVisibleInitial = 0;
  window._nextSemanaProyeccion = null;
  var tbody = document.getElementById('tResumenBody');
  if (tbody) tbody.innerHTML = '';

  if (_supabaseConfigured()) {
    fetch(SUPABASE_URL + '/rest/v1/' + SUPABASE_CAPTURE_TABLE + '?id=gt.0', {
      method: 'DELETE',
      headers: _supabaseHeaders()
    }).catch(function (err) {
      console.warn('No se pudo borrar la captura en Supabase:', err);
    });
  }
}

/* ── Filtra una semana de un objeto captureProjections ── */
function _filterWeekFromStore(proj, targetSem) {
  var t = String(targetSem).trim();
  ['sem', 'norm'].forEach(function(modeKey) {
    var store = proj[modeKey];
    if (!store) return;
    Object.keys(store).forEach(function(sk) {
      var entry = store[sk];
      if (!entry || !entry.rows) return;
      entry.rows = entry.rows.filter(function(r) {
        return String(r.sem || '').trim() !== t;
      });
    });
  });
  return proj;
}

/* ── Borrar UNA semana específica de las capturas guardadas ── */
function _clearWeekFromCapture(targetSem) {
  // 0) Save any current unsaved data in DOM first!
  if (typeof saveCaptureProjectionsFromDom === 'function') {
    saveCaptureProjectionsFromDom();
  }
  
  // 1) Aplicar filtro local inmediatamente
  if (!window._captureProjections) window._captureProjections = { sem: {}, norm: {}, _meta: {} };
  _filterWeekFromStore(window._captureProjections, targetSem);
  try { localStorage.setItem(_CAPTURE_LS_KEY, JSON.stringify(window._captureProjections)); } catch(e) {}
  window._skipCaptureSaveOnce = true;
  if (typeof renderResumen === 'function') renderResumen();

  if (!_supabaseConfigured()) return;

  // 2) Borrar en Supabase
  var url = SUPABASE_URL + '/rest/v1/' + SUPABASE_CAPTURE_TABLE + '?semana=eq.' + encodeURIComponent(String(targetSem).trim());
  fetch(url, {
    method: 'DELETE',
    headers: _supabaseHeaders()
  }).then(function(res) {
    if (!res.ok) {
      console.warn('Supabase DELETE failed:', res.status, res.statusText);
      alert('Error al borrar en Supabase (Código ' + res.status + '). Puede que falten permisos (RLS) para DELETE.');
    } else {
      console.log('Semana ' + targetSem + ' borrada de Supabase.');
    }
  }).catch(function(err) {
    console.warn('Error de red al borrar la semana en Supabase:', err);
  });
}

/* ── Modal para seleccionar qué semana borrar ── */
function _showDeleteWeekDialog() {
  // Recopilar semanas únicas de ambos modos
  var semsSet = {};
  var proj = window._captureProjections || {};
  ['sem', 'norm'].forEach(function(modeKey) {
    var store = proj[modeKey] || {};
    Object.keys(store).forEach(function(sk) {
      var entry = store[sk];
      if (!entry || !entry.rows) return;
      entry.rows.forEach(function(r) {
        var s = String(r.sem || '').trim();
        if (s) semsSet[s] = true;
      });
    });
  });
  var sems = Object.keys(semsSet).sort(function(a,b){ return parseInt(a)-parseInt(b); });

  // Eliminar modal previo si existe
  var old = document.getElementById('_deleteWeekModal');
  if (old) old.remove();

  // Construir opciones
  var optionsHtml = '';
  if (sems.length === 0) {
    optionsHtml = '<p style="color:#888;font-size:13px;margin:12px 0;">No hay semanas guardadas actualmente.</p>';
  } else {
    sems.forEach(function(s) {
      optionsHtml += '<label style="display:flex;align-items:center;gap:10px;padding:8px 10px;border-radius:6px;cursor:pointer;transition:background .15s;" onmouseover="this.style.background=\'#f0f4fa\'" onmouseout="this.style.background=\'\'">'
        + '<input type="radio" name="_delWeekOpt" value="' + s + '" style="accent-color:#c62828;width:16px;height:16px;">'
        + '<span style="font-size:14px;font-weight:600;color:#1a3a5c;">Semana ' + s + '</span>'
        + '</label>';
    });
    optionsHtml += '<hr style="border:none;border-top:1px solid #e0e0e0;margin:10px 0;">'
      + '<label style="display:flex;align-items:center;gap:10px;padding:8px 10px;border-radius:6px;cursor:pointer;transition:background .15s;" onmouseover="this.style.background=\'#fff0f0\'" onmouseout="this.style.background=\'\'">'
      + '<input type="radio" name="_delWeekOpt" value="__ALL__" style="accent-color:#c62828;width:16px;height:16px;">'
      + '<span style="font-size:14px;font-weight:600;color:#c62828;">⚠ Todas las semanas</span>'
      + '</label>';
  }

  var modal = document.createElement('div');
  modal.id = '_deleteWeekModal';
  modal.style.cssText = 'position:fixed;inset:0;z-index:99999;display:flex;align-items:center;justify-content:center;background:rgba(10,20,50,0.45);backdrop-filter:blur(3px);';
  modal.innerHTML =
    '<div style="background:#fff;border-radius:12px;box-shadow:0 8px 40px rgba(0,0,0,0.22);width:340px;max-width:94vw;padding:0;overflow:hidden;">'
    + '<div style="background:#c62828;padding:16px 20px;display:flex;align-items:center;gap:10px;">'
    + '<span class="material-icons" style="color:#fff;font-size:22px;">delete_sweep</span>'
    + '<span style="color:#fff;font-size:15px;font-weight:700;letter-spacing:.02em;">Borrar captura guardada</span>'
    + '</div>'
    + '<div style="padding:18px 20px;">'
    + '<p style="font-size:12px;color:#666;margin:0 0 14px 0;">Selecciona la semana que deseas eliminar de las capturas guardadas.<br><span style="color:#c62828;font-weight:600;">Esto afecta a todos los usuarios.</span></p>'
    + '<div id="_delWeekOptions" style="max-height:260px;overflow-y:auto;">' + optionsHtml + '</div>'
    + '</div>'
    + '<div style="padding:12px 20px 18px;display:flex;gap:10px;justify-content:flex-end;border-top:1px solid #eee;">'
    + '<button onclick="document.getElementById(\'_deleteWeekModal\').remove()" style="background:#f0f0f0;color:#444;border:none;border-radius:6px;padding:8px 18px;font-size:13px;font-weight:600;cursor:pointer;">Cancelar</button>'
    + '<button id="_delWeekConfirmBtn" onclick="_confirmDeleteWeek()" style="background:#c62828;color:#fff;border:none;border-radius:6px;padding:8px 18px;font-size:13px;font-weight:600;cursor:pointer;">Borrar</button>'
    + '</div>'
    + '</div>';
  document.body.appendChild(modal);
  // Cerrar al clic en el fondo
  modal.addEventListener('click', function(e){ if(e.target === modal) modal.remove(); });
}

function _confirmDeleteWeek() {
  var sel = document.querySelector('input[name="_delWeekOpt"]:checked');
  if (!sel) { alert('Selecciona una semana primero.'); return; }
  var val = sel.value;
  document.getElementById('_deleteWeekModal').remove();
  if (val === '__ALL__') {
    if (prompt('Esto borra las capturas de TODAS las semanas para todos los usuarios.\n\nEscribe BORRAR para confirmar:') === 'BORRAR') {
      _clearPersistedCapture();
      window._captureProjections = { sem: {}, norm: {}, _meta: {} };
      window._skipCaptureSaveOnce = true;
      if (typeof renderResumen === 'function') renderResumen();
    }
  } else {
    if (confirm('¿Borrar la captura de la Semana ' + val + ' para todos los usuarios?')) {
      _clearWeekFromCapture(val);
    }
  }
}

/* ── Helper: guardar captura actual como programado ── */
function _guardarCaptura(r1i, r2i, rowIndex) {
  var blocks = window._resumenCaptureBlocks || [];
  var block = null;
  for(var b=0; b<blocks.length; b++){
    if(blocks[b].i1 === r1i && blocks[b].r2i === r2i){ block = blocks[b]; break; }
  }
  if(!block || !block.prod || !block.tienda) return;
  if(!window._captureProjections) window._captureProjections = { sem: {}, norm: {}, _meta: {} };
  var modeKey = _captureProjModeKey();
  if(!window._captureProjections[modeKey]) window._captureProjections[modeKey] = {};
  var sk = _captureProjStoreKey(block.prod, block.tienda);
  if(!window._captureProjections[modeKey][sk]) {
    window._captureProjections[modeKey][sk] = { visibleInitial: 1, rows: [] };
  }
  var entry = window._captureProjections[modeKey][sk];
  /* Colectar datos de la fila solicitada */
  var nCols = modeKey === 'sem' ? (block.ngr || 1) : (block.ncols || 1);
  var wi = 0;
  while(document.getElementById((modeKey === 'sem' ? 'ct_' : 'ctn_') + r1i + '_' + r2i + '_' + wi)) wi++;
  if(wi === 0) return;
  var targetWi = parseInt(rowIndex, 10);
  if(isNaN(targetWi) || targetWi < 0 || targetWi >= wi) targetWi = wi - 1;
  var semEl = document.getElementById('csem_' + r1i + '_' + r2i + '_' + targetWi);
  var semVal = semEl ? String(semEl.value || '').trim() : '';
  var values = [];
  for(var ci=0; ci<nCols; ci++){
    var inp = document.getElementById((modeKey === 'sem' ? 'ci_' : 'cin_') + r1i + '_' + r2i + '_' + targetWi + '_' + ci);
    values.push(inp ? String(inp.value || '') : '');
  }
  if(!entry.rows) entry.rows = [];
  
  // Separar guardados y reconstruir no guardados desde el DOM para no perder otras semanas abiertas.
  var savedRows = entry.rows.filter(function(r){ return r.saved; });
  var domRows = _collectCaptureRowsFromBlock(block, modeKey);
  var unsavedRows = domRows.filter(function(row, idx){ return idx !== targetWi; });
  
  // No sobrescribir programados existentes desde Guardar; para editar se usa el lapiz.
  if(semVal) {
    var alreadySaved = savedRows.some(function(row){
      return String(row.sem || '').trim() === semVal;
    });
    if(alreadySaved) {
      alert('La semana ' + semVal + ' ya esta programada. Usa el lapiz para editarla.');
      return;
    }
  }
  savedRows.push({ sem: semVal, values: values.slice(), hidden: false, saved: true });
  
  entry.rows = savedRows.concat(unsavedRows);
  
  // Decrementar visibilidad para ocultar los inputs que acabamos de guardar
  if(entry.visibleInitial > 0) {
    entry.visibleInitial = entry.visibleInitial - 1;
  }
  if(window._captureVisibleInitial > 0) {
    window._captureVisibleInitial = window._captureVisibleInitial - 1;
  }
  
  _saveRowToSupabase(modeKey, block.prod, block.tienda, semVal, values);
  window._skipCaptureSaveOnce = true;
  if(typeof renderResumen === 'function') renderResumen();
  /* Feedback visual */
  var btn = document.getElementById('btnGuardar_' + r1i + '_' + r2i + '_' + targetWi) || document.getElementById('btnGuardar_' + r1i + '_' + r2i);
  if(btn) {
    btn.textContent = '✓';
    btn.style.background = '#2e7d32';
    setTimeout(function(){ if(btn) { btn.textContent = '💾'; btn.style.background = '#1565c0'; } }, 1200);
  }
}

function _captureRowHasPositiveValues(row) {
  return (row.values || []).some(function(v){ return parseFloat(String(v || '').replace(/,/g,'')) > 0; });
}

function saveAllCaptureRows() {
  var blocks = window._resumenCaptureBlocks || [];
  if(!blocks.length) return;
  if(!window._captureProjections) window._captureProjections = { sem: {}, norm: {}, _meta: {} };
  var modeKey = _captureProjModeKey();
  if(!window._captureProjections[modeKey]) window._captureProjections[modeKey] = {};
  var store = window._captureProjections[modeKey];
  var savedCount = 0;
  var duplicateCount = 0;

  blocks.forEach(function(block){
    if(!block.prod || !block.tienda || !_captureBlockDomExists(block, modeKey)) return;
    var sk = _captureProjStoreKey(block.prod, block.tienda);
    var entry = store[sk] || { visibleInitial: window._captureVisibleInitial || 0, rows: [] };
    var savedRows = (entry.rows || []).filter(function(r){ return r.saved; });
    var savedBySem = {};
    savedRows.forEach(function(r){
      var k = String(r.sem || '').trim();
      if(k) savedBySem[k] = true;
    });

    var remainingRows = [];
    _collectCaptureRowsFromBlock(block, modeKey).forEach(function(row){
      var semVal = String(row.sem || '').trim();
      var shouldSave = !row.hidden && semVal && _captureRowHasPositiveValues(row);
      if(shouldSave) {
        if(savedBySem[semVal]) {
          duplicateCount++;
          return;
        }
        savedRows.push({ sem: semVal, values: (row.values || []).slice(), hidden: false, saved: true });
        _saveRowToSupabase(modeKey, block.prod, block.tienda, semVal, (row.values || []).slice());
        savedBySem[semVal] = true;
        savedCount++;
        return;
      }
      if(row.hidden && (row.sem || _captureRowHasPositiveValues(row))) remainingRows.push(row);
    });

    store[sk] = { visibleInitial: 0, rows: savedRows.concat(remainingRows) };
  });

  if(savedCount > 0) {
    window._captureVisibleInitial = 0;
    if(window._captureProjections._meta) window._captureProjections._meta.visibleInitial = 0;
    window._skipCaptureSaveOnce = true;
    if(typeof renderResumen === 'function') renderResumen();
    return;
  }
  if(duplicateCount > 0) {
    alert('Las semanas seleccionadas ya estan programadas. Usa el lapiz para editarlas.');
    return;
  }
  alert('No hay capturas con datos para guardar.');
}

/* ── Helper: editar programado (cambia todos los valores de una fila) ── */
function _editProgramadoRow(r1Key, r2Key, rowIdx, colIdx, modeKey) {
  var store = window._captureProjections && window._captureProjections[modeKey];
  if(!store) return;
  var pivotMode = state.resumenPivot || 'producto';
  var prod = pivotMode === 'producto' ? r1Key : r2Key;
  var tienda = pivotMode === 'producto' ? r2Key : r1Key;
  var sk = _captureProjStoreKey(prod, tienda);
  var entry = store[sk];
  if(!entry || !entry.rows) return;
  var savedRows = entry.rows.filter(function(r){ return r.saved; });
  var row = savedRows[rowIdx];
  if(!row) return;
  var nCols = row.values.length;
  if(colIdx >= 0 && colIdx < nCols){
    var curVal = parseFloat(row.values[colIdx] || '0');
    var newVal = prompt('Editar Programado - Valor:', curVal || '');
    if(newVal === null) return;
    row.values[colIdx] = String(Math.round(parseFloat(String(newVal).replace(/,/g,'')) || 0));
  } else {
    var newSem = prompt('Editar Programado - Semana:', row.sem || '');
    if(newSem === null) return;
    var oldSem = row.sem;
    row.sem = newSem;
    if (oldSem !== newSem) _deleteRowFromSupabase(modeKey, prod, tienda, oldSem);
  }
  _saveRowToSupabase(modeKey, prod, tienda, row.sem, row.values);
  if(typeof renderResumen === 'function') renderResumen();
}

/* ── Helper: eliminar una fila programado ── */
function _deleteProgramadoRow(r1Key, r2Key, rowIdx, modeKey, semVal) {
  if(!confirm('¿Eliminar esta fila programado?')) return;
  var store = window._captureProjections && window._captureProjections[modeKey];
  if(!store) return;
  var pivotMode = state.resumenPivot || 'producto';
  var prod = pivotMode === 'producto' ? r1Key : r2Key;
  var tienda = pivotMode === 'producto' ? r2Key : r1Key;
  var sk = _captureProjStoreKey(prod, tienda);
  var entry = store[sk];
  if(!entry || !entry.rows) return;
  var savedRows = entry.rows.filter(function(r){ return r.saved; });
  var unsavedRows = entry.rows.filter(function(r){ return !r.saved; });
  var deleteIdx = -1;
  var semKey = String(semVal || '').trim();
  if(semKey) {
    for(var i=0; i<savedRows.length; i++){
      if(String(savedRows[i].sem || '').trim() === semKey){ deleteIdx = i; break; }
    }
  }
  if(deleteIdx < 0) deleteIdx = rowIdx;
  if(deleteIdx < 0 || deleteIdx >= savedRows.length) return;
  var deletedRow = savedRows.splice(deleteIdx, 1)[0];
  entry.rows = savedRows.concat(unsavedRows);
  if(entry.rows.length === 0) {
    delete store[sk];
    _markCaptureKeyRemoved(modeKey, sk);
  }
  if(deletedRow) _deleteRowFromSupabase(modeKey, prod, tienda, deletedRow.sem);
  if(typeof renderResumen === 'function') renderResumen();
}

/* Proyecciones Capture SEM: persisten en localStorage entre sesiones */
function _captureProjStoreKey(prod, tienda) {
  return prod + '\x00' + tienda;
}

function _captureProjModeKey() {
  return state.resumenMode === 'semanas' ? 'sem' : 'norm';
}

function _captureBlockDomExists(b, modeKey) {
  if(modeKey === 'sem') return !!document.getElementById('ct_'+b.i1+'_'+b.r2i+'_0');
  return !!document.getElementById('ctn_'+b.i1+'_'+b.r2i+'_0');
}

function _collectCaptureRowsFromBlock(b, modeKey) {
  var rows = [];
  if(modeKey === 'sem') {
    var wi = 0;
    while(document.getElementById('ct_'+b.i1+'_'+b.r2i+'_'+wi)) {
      var totEl = document.getElementById('ct_'+b.i1+'_'+b.r2i+'_'+wi);
      var tr = totEl ? totEl.closest('tr') : null;
      var hidden = !!(tr && tr.style.display === 'none');
      var semEl = document.getElementById('csem_'+b.i1+'_'+b.r2i+'_'+wi);
      var vals = [];
      for(var gi=0; gi<(b.ngr||0); gi++) {
        var inp = document.getElementById('ci_'+b.i1+'_'+b.r2i+'_'+wi+'_'+gi);
        vals.push(inp ? String(inp.value||'') : '');
      }
      rows.push({ sem: semEl ? String(semEl.value||'') : '', values: vals, hidden: hidden, saved: false });
      wi++;
    }
  } else {
    var ri = 0;
    while(document.getElementById('ctn_'+b.i1+'_'+b.r2i+'_'+ri)) {
      var totEl2 = document.getElementById('ctn_'+b.i1+'_'+b.r2i+'_'+ri);
      var tr2 = totEl2 ? totEl2.closest('tr') : null;
      var hidden2 = !!(tr2 && tr2.style.display === 'none');
      var semEl2 = document.getElementById('csem_'+b.i1+'_'+b.r2i+'_'+ri);
      var vals2 = [];
      for(var ci=0; ci<(b.ncols||0); ci++) {
        var inp2 = document.getElementById('cin_'+b.i1+'_'+b.r2i+'_'+ri+'_'+ci);
        vals2.push(inp2 ? String(inp2.value||'') : '');
      }
      rows.push({ sem: semEl2 ? String(semEl2.value||'') : '', values: vals2, hidden: hidden2, saved: false });
      ri++;
    }
  }
  return rows;
}

function _captureRowsHaveData(rows) {
  return rows.some(function(r) {
    return r.sem || (r.values||[]).some(function(v) { return v !== '' && v !== '0'; });
  });
}

function _writeCaptureProjEntry(store, key, domRows) {
  var existingRows = (store[key] && store[key].rows) || [];
  var savedRows = existingRows.filter(function(r){ return r.saved; });
  store[key] = {
    visibleInitial: window._captureVisibleInitial||0,
    rows: savedRows.concat(domRows)
  };
}

/* ── Export de Captura a CSV: formato "Cronograma de entregas" ──
   Ruta | Tienda | Dia/Periodo | <una columna por producto>
   Agrupado en bloques separados por semana (CSV no soporta hojas). */
function _exportCaptureToCSV(modeKey, subColLabels) {
  var store = (window._captureProjections && window._captureProjections[modeKey]) || {};
  var tiendaRuta = (typeof DATA !== 'undefined' && DATA.tienda_ruta) || {};

  /* bySemana[semana][tienda][rowLabel][producto] = valor */
  var bySemana = {};
  var productosOrden = [];
  var productosSet = {};
  var hasAnyData = false;

  /* Recorremos TODO lo capturado en memoria (window._captureProjections),
     no solo los bloques de la tienda/filtro actualmente visible. Así se
     incluye lo capturado en tiendas que ya no están filtradas/visibles. */
  Object.keys(store).forEach(function(key) {
    var sepIdx = key.indexOf('\x00');
    if (sepIdx === -1) return;
    var prod = key.slice(0, sepIdx);
    var tienda = key.slice(sepIdx + 1);
    if (!prod || !tienda) return;

    var entry = store[key];
    if (!entry || !entry.rows) return;
    if (!productosSet[prod]) { productosSet[prod] = true; productosOrden.push(prod); }

    entry.rows.forEach(function(row) {
      var values = row.values || [];
      var rowHasData = values.some(function(v) { return v !== '' && parseFloat(v) > 0; });
      if (!rowHasData) return;
      hasAnyData = true;
      var semKey = (row.sem || '').trim() || 'Sin semana';
      if (!bySemana[semKey]) bySemana[semKey] = {};
      if (!bySemana[semKey][tienda]) bySemana[semKey][tienda] = {};

      values.forEach(function(v, idx) {
        var rowLabel = subColLabels[idx] || ('Col' + idx);
        var num = parseFloat(v) || 0;
        if (num <= 0) return;
        if (!bySemana[semKey][tienda][rowLabel]) bySemana[semKey][tienda][rowLabel] = {};
        bySemana[semKey][tienda][rowLabel][prod] = num;
      });
    });
  });

  if (!hasAnyData) {
    alert('No hay datos capturados para exportar.');
    return;
  }

  function csvRow(arr) {
    return arr.map(function(c) {
      return '"' + String(c == null ? '' : c).replace(/"/g, '""') + '"';
    }).join(',');
  }

  var semanas = Object.keys(bySemana).sort();
  var lines = [];

  semanas.forEach(function(sem, si) {
    if (si > 0) lines.push('');
    lines.push(csvRow(['Cronograma entregas semana ' + sem]));
    lines.push(csvRow(['Ruta', 'Tienda', 'Dia/Periodo'].concat(productosOrden)));

    var tiendas = Object.keys(bySemana[sem]).sort();
    tiendas.forEach(function(tienda) {
      var ruta = tiendaRuta[tienda] || 'Sin ruta';
      subColLabels.forEach(function(rowLabel, rowIdx) {
        var rowData = (bySemana[sem][tienda] && bySemana[sem][tienda][rowLabel]) || {};
        var rowVals = productosOrden.map(function(p) {
          return rowData[p] != null ? rowData[p] : '';
        });
        /* Solo poner Ruta y Tienda en la primera fila de cada tienda */
        var rutaCol   = rowIdx === 0 ? ruta   : '';
        var tiendaCol = rowIdx === 0 ? tienda : '';
        lines.push(csvRow([rutaCol, tiendaCol, rowLabel].concat(rowVals)));
      });
    });
  });

  var csvContent = lines.join('\n');
  var blob = new Blob([csvContent], { type: 'text/csv;charset=utf-8;' });
  var url = URL.createObjectURL(blob);
  var a = document.createElement('a');
  a.href = url;
  a.download = 'Cronograma_Entregas_Captura.csv';
  a.click();
  URL.revokeObjectURL(url);
}

function syncCaptureProjFromBlock(i1, r2i) {
  var blocks = window._resumenCaptureBlocks;
  if(!blocks || !blocks.length) return;
  var b = null;
  for(var i=0; i<blocks.length; i++){
    if(blocks[i].i1 === i1 && blocks[i].r2i === r2i){ b = blocks[i]; break; }
  }
  if(!b || !b.prod || !b.tienda) return;
  if(!window._captureProjections) window._captureProjections = { sem: {}, norm: {}, _meta: {} };
  var modeKey = _captureProjModeKey();
  if(!window._captureProjections[modeKey]) window._captureProjections[modeKey] = {};
  if(!window._captureProjections._meta) window._captureProjections._meta = {};
  if(!_captureBlockDomExists(b, modeKey)) return;
  var rows = _collectCaptureRowsFromBlock(b, modeKey);
  /* Siempre sobreescribir para reflejar borrados */
  _writeCaptureProjEntry(window._captureProjections[modeKey], _captureProjStoreKey(b.prod, b.tienda), rows);
  window._captureProjections._meta.visibleInitial = window._captureVisibleInitial||0;
  if(window._nextSemanaProyeccion) window._captureProjections._meta.nextSem = window._nextSemanaProyeccion;
}

function saveCaptureProjectionsFromDom() {
  var blocks = window._resumenCaptureBlocks;
  if(!window._captureProjections) window._captureProjections = { sem: {}, norm: {}, _meta: {} };
  if(!window._captureProjections._meta) window._captureProjections._meta = {};
  var modeKey = _captureProjModeKey();
  if(!window._captureProjections[modeKey]) window._captureProjections[modeKey] = {};
  var store = window._captureProjections[modeKey];
  var anyDom = false;
  /* Usar el primer bloque disponible para capturar filas globales (sem y valores son compartidos entre bloques) */
  var globalRowsSaved = false;
  if(blocks && blocks.length){
    blocks.forEach(function(b) {
      if(!b.prod || !b.tienda) return;
      if(!_captureBlockDomExists(b, modeKey)) return;
      anyDom = true;
      var rows = _collectCaptureRowsFromBlock(b, modeKey);
      /* Siempre sobreescribir — aunque rows esté vacío, para limpiar datos borrados */
      _writeCaptureProjEntry(store, _captureProjStoreKey(b.prod, b.tienda), rows);
      /* Guardar también en meta.globalRows para sobrevivir cambios de filtro */
      if(!globalRowsSaved && _captureRowsHaveData(rows.filter(function(r){ return !r.hidden; }))){
        window._captureProjections._meta.globalRows = rows;
        window._captureProjections._meta.globalMode = modeKey;
        globalRowsSaved = true;
      }
    });
  }
  if(anyDom) {
    window._captureProjections._meta.visibleInitial = window._captureVisibleInitial||0;
    if(window._nextSemanaProyeccion) window._captureProjections._meta.nextSem = window._nextSemanaProyeccion;
    _persistCapture();
  }
}

function restoreCaptureProjections() {
  var blocks = window._resumenCaptureBlocks;
  if(!blocks || !blocks.length) return;
  var modeKey = _captureProjModeKey();
  var store = (window._captureProjections && window._captureProjections[modeKey]) || {};
  var meta = (window._captureProjections && window._captureProjections._meta) || {};
  var maxVis = meta.visibleInitial || window._captureVisibleInitial || 0;
  /* No restaurar nextSem: debe recalcularse desde la ultima semana real de DATA. */
  /* BUG FIX: Eliminar el fallback a globalRows.
     Antes, cuando no había datos para la nueva tienda (clave prod+tienda no encontrada),
     se usaban las filas de globalRows que pertenecían a OTRA tienda, mezclando datos.
     Ahora: si no hay datos para esta combinación prod+tienda, simplemente no se restaura nada. */
  var globalRows = null; // fallback deshabilitado intencionalmente

  blocks.forEach(function(b) {
    if(!b.prod || !b.tienda) return;
    /* Buscar datos por storeKey exacto; si no, usar globalRows */
    var saved = store[_captureProjStoreKey(b.prod, b.tienda)];
    if((!saved || !saved.rows || !saved.rows.length) && globalRows && globalRows.length) {
      saved = { visibleInitial: meta.visibleInitial||0, rows: globalRows };
    }
    if(!saved || !saved.rows) return;
    if((saved.visibleInitial||0) > maxVis) maxVis = saved.visibleInitial||0;

    var allRows = saved.rows.filter(function(r){ return !r.saved; });
    /* Filas visibles guardadas (las que el usuario había revelado) */
    var visibleRows = allRows.filter(function(r){ return !r.hidden; });
    /* Filas dinámicas: las que superan el número de filas iniciales del DOM */
    var initCap = modeKey === 'sem' ? (window._captureMinRowsSem||1) : (window._captureMinRowsNorm||1);
    var extraNeeded = visibleRows.length - initCap;

    if(modeKey === 'sem') {
      /* 1. Agregar filas dinámicas extras si hacen falta */
      for(var ei=0; ei<extraNeeded; ei++){
        window.addCaptureSemRow(b.i1, b.r2i, b.id1, b.bg, b.ngr);
      }
      /* 2. Poblar valores en TODAS las filas (estén visibles u ocultas en el DOM) */
      allRows.forEach(function(row, ri) {
        var semEl = document.getElementById('csem_'+b.i1+'_'+b.r2i+'_'+ri);
        if(semEl) semEl.value = row.sem || '';
        for(var gi=0; gi<(b.ngr||0); gi++){
          var inp = document.getElementById('ci_'+b.i1+'_'+b.r2i+'_'+ri+'_'+gi);
          if(!inp) continue;
          var v = (row.values||[])[gi];
          inp.value = (v !== undefined && v !== null && v !== '') ? v : '';
        }
      });
      window.recalcCaptureSemTotals(b.i1, b.r2i, b.ngr);
      /* 3. Revelar las filas iniciales que estaban visibles — usando revealCaptureSemRow
            para que bumpResumenRowspans ajuste los rowspans correctamente */
      var revealCount = Math.min(visibleRows.length, initCap);
      for(var vi=0; vi<revealCount; vi++){
        window.revealCaptureSemRow(b.i1, b.r2i, vi);
      }
    } else {
      for(var ej=0; ej<extraNeeded; ej++){
        window.addCaptureNormalRow(b.i1, b.r2i, b.ncols, b.bgEven, b.bgOdd);
      }
      allRows.forEach(function(row, ri) {
        var semEl = document.getElementById('csem_'+b.i1+'_'+b.r2i+'_'+ri);
        if(semEl) semEl.value = row.sem || '';
        for(var ci=0; ci<(b.ncols||0); ci++){
          var inp = document.getElementById('cin_'+b.i1+'_'+b.r2i+'_'+ri+'_'+ci);
          if(!inp) continue;
          var v = (row.values||[])[ci];
          inp.value = (v !== undefined && v !== null && v !== '') ? v : '';
        }
        window.updateCaptureNormalTotal(b.i1, b.r2i, ri, b.ncols);
      });
      var revealCountN = Math.min(visibleRows.length, initCap);
      for(var vj=0; vj<revealCountN; vj++){
        window.revealCaptureNormalRow(b.i1, b.r2i, vj);
      }
    }
  });

  /* Calcular _captureVisibleInitial desde las filas iniciales visibles */
  var refRows = globalRows || (blocks[0] && store[_captureProjStoreKey(blocks[0].prod, blocks[0].tienda)] && store[_captureProjStoreKey(blocks[0].prod, blocks[0].tienda)].rows) || [];
  refRows = refRows.filter(function(r){ return !r.saved; });
  var initCapRef = modeKey === 'sem' ? (window._captureMinRowsSem||1) : (window._captureMinRowsNorm||1);
  var visCount = 0;
  refRows.slice(0, initCapRef).forEach(function(r){ if(!r.hidden) visCount++; });
  maxVis = visCount || maxVis;
  var capMax = modeKey === 'sem' ? (window._captureMinRowsSem||1) : (window._captureMinRowsNorm||1);
  if(maxVis > capMax) maxVis = capMax;
  window._captureVisibleInitial = maxVis;
  if(window._captureProjections && window._captureProjections._meta) {
    window._captureProjections._meta.visibleInitial = maxVis;
  }
  window.updateRemoveCaptureBtn();
}

function setResumenPivot(mode) {
  state.resumenPivot = mode;
  document.getElementById('btnPivotProducto').className = mode === 'producto' ? 'mode-btn active' : 'mode-btn';
  document.getElementById('btnPivotTienda').className = mode === 'tienda' ? 'mode-btn active' : 'mode-btn';
  renderResumen();
}

function collectResumenPivot(mode){
  var sems = getSemanasActivas().map(function(s){ return String(s); });
  var tiendas = getTiendasActivas();
  var resumenSrc = DATA.resumen_diario || {};

  var rowsData = {};
  var totalsByColumn = {};
  var totalsGrand = {embarque:0, ventas:0, merma:0, venta_pos:0, field1:0};
  var hasData = false;
  var totalsByDay = {}; // To store which days exist in which week

  var pivotMode = state.resumenPivot || 'producto';
  var prods = getProductosActivos();

  tiendas.forEach(function(t){
    var tiendaData = resumenSrc[t] || {};
    sems.forEach(function(s){
      var semanaData = tiendaData[s] || {};
      Object.keys(semanaData).forEach(function(p){
        if(prods.indexOf(p) === -1) return;
        var r1 = pivotMode === 'producto' ? p : t;
        var r2 = pivotMode === 'producto' ? t : p;

        if(!rowsData[r1]) rowsData[r1] = { _total:{}, subRows:{} };
        if(!rowsData[r1].subRows[r2]) rowsData[r1].subRows[r2] = { _total:{} };

        var wKey = 'W' + s;

        if(!rowsData[r1]._total[wKey]) rowsData[r1]._total[wKey] = {embarque:0,ventas:0,merma:0,venta_pos:0,field1:0};
        if(!rowsData[r1].subRows[r2]._total[wKey]) rowsData[r1].subRows[r2]._total[wKey] = {embarque:0,ventas:0,merma:0,venta_pos:0,field1:0};
        if(!totalsByColumn[wKey]) totalsByColumn[wKey] = {embarque:0,ventas:0,merma:0,venta_pos:0,field1:0};

        Object.keys(semanaData[p] || {}).forEach(function(fecha){
          hasData = true;
          var dKey = 'D' + fecha;
          
          if(!totalsByDay[s]) totalsByDay[s] = {};
          totalsByDay[s][fecha] = true;

          if(!rowsData[r1]._total[dKey]) rowsData[r1]._total[dKey] = {embarque:0,ventas:0,merma:0,venta_pos:0,field1:0};
          if(!rowsData[r1].subRows[r2]._total[dKey]) rowsData[r1].subRows[r2]._total[dKey] = {embarque:0,ventas:0,merma:0,venta_pos:0,field1:0};
          if(!totalsByColumn[dKey]) totalsByColumn[dKey] = {embarque:0,ventas:0,merma:0,venta_pos:0,field1:0};

          var src = semanaData[p][fecha] || {};
          ['embarque','ventas','merma','venta_pos','field1'].forEach(function(k){
            var val = src[k] || 0;
            // Add to week total
            rowsData[r1]._total[wKey][k] += val;
            rowsData[r1].subRows[r2]._total[wKey][k] += val;
            totalsByColumn[wKey][k] += val;
            // Add to day total
            rowsData[r1]._total[dKey][k] += val;
            rowsData[r1].subRows[r2]._total[dKey][k] += val;
            totalsByColumn[dKey][k] += val;
            // Grand total
            totalsGrand[k] += val;
          });
        });
      });
    });
  });

  window.expandedWeeks = window.expandedWeeks || {};
  var columns = [];
  var DIAS_ABREV = ['Dom','Lun','Mar','Mié','Jue','Vie','Sáb'];

  sems.forEach(function(s){
    // always add week column
    var shortWeek = s.slice(-2); // converts "202621" to "21"
    columns.push({ key: 'W'+s, label: shortWeek, type: 'week', rawWeek: s });
    
    // add days if expanded
    if(window.expandedWeeks[s]){
      var days = Object.keys(totalsByDay[s] || {}).sort();
      days.forEach(function(d){
        var dt = parseResumenDate(d);
        var lbl = dt ? DIAS_ABREV[dt.getDay()] + ' ' + dt.getDate() : d;
        columns.push({ key: 'D'+d, label: lbl, type: 'day' });
      });
    }
  });

  var primaryKeys = [];
  if (pivotMode === 'producto') {
    primaryKeys = getProductosActivos().filter(function(p){ return !!rowsData[p]; });
    Object.keys(rowsData).sort().forEach(function(p){
      if(primaryKeys.indexOf(p) === -1) primaryKeys.push(p);
    });
  } else {
    primaryKeys = Object.keys(rowsData).sort();
  }

  return {
    hasData: hasData,
    columns: columns,
    primaryKeys: primaryKeys,
    rowsData: rowsData,
    totalsByColumn: totalsByColumn,
    totalsGrand: totalsGrand,
    totalsByDay: totalsByDay,
  };
}
function _getMaxGlobalWeek() {
  var globalMaxSem = DATA.semanas && DATA.semanas.length ? Math.max.apply(null, DATA.semanas.map(function(s){ return parseInt(s, 10); })) : 0;
  return globalMaxSem > 9999 ? globalMaxSem % 100 : globalMaxSem;
}

function _getBaseNextSemanaProyeccion() {
  var maxW = _getMaxGlobalWeek();
  return maxW > 0 ? maxW + 1 : 1;
}

function _ensureNextSemanaProyeccion() {
  var baseNext = _getBaseNextSemanaProyeccion();
  var curNext = parseInt(window._nextSemanaProyeccion, 10);
  if(isNaN(curNext) || curNext < baseNext) curNext = baseNext;
  window._nextSemanaProyeccion = curNext;
  return curNext;
}

function renderResumen(){
  if(window._skipCaptureSaveOnce) {
    window._skipCaptureSaveOnce = false;
  } else {
    saveCaptureProjectionsFromDom();
  }
  /* BUG FIX: Siempre recalcular _nextSemanaProyeccion desde DATA al re-renderizar.
     Antes se guardaba el valor ya-incrementado en meta.nextSem y se restauraba tal cual,
     causando que al cambiar tienda la semana apareciera como 28 en vez de 26.
     La semana base siempre viene de DATA; los incrementos temporales no deben persistir. */
  /* [FIX REVERTIDO] No reiniciar window._nextSemanaProyeccion. 
     El código anterior propagaba a las tiendas invisibles y por tanto el contador no debe reiniciarse. */
  var pivot    = collectResumenPivot('semanas');
  var defsAll  = getResumenMetricDefs();
  var metricFiltersBox = document.getElementById('metricFilters');
  var defs = defsAll;
  if (metricFiltersBox) {
    var checkedInputs = metricFiltersBox.querySelectorAll('input[type="checkbox"]:checked');
    if (checkedInputs.length > 0) {
      var selectedKeys = [];
      for(var i=0; i<checkedInputs.length; i++) selectedKeys.push(checkedInputs[i].value);
      defs = defsAll.filter(function(d){ return selectedKeys.indexOf(d.key) !== -1; });
    }
  }
  var pivotMode = state.resumenPivot || 'producto';

  window.updateRemoveCaptureBtn = function(){
    var btn = document.getElementById('btnRemoveCaptureSem');
    if(!btn) return;
    var canRemove = false;
    var blocks = window._resumenCaptureBlocks || [];
    var visibleInitial = window._captureVisibleInitial || 0;
    if(visibleInitial > 0) canRemove = true;
    if(state.resumenMode === 'semanas'){
      var minSem = window._captureMinRowsSem || 1;
      blocks.forEach(function(b){
        var wi = 0;
        while(document.getElementById('ct_'+b.i1+'_'+b.r2i+'_'+wi)) wi++;
        if(wi > minSem) canRemove = true;
      });
    } else {
      var minNorm = window._captureMinRowsNorm || 1;
      blocks.forEach(function(b){
        var ri = 0;
        while(document.getElementById('ctn_'+b.i1+'_'+b.r2i+'_'+ri)) ri++;
        if(ri > minNorm) canRemove = true;
      });
    }
    btn.disabled = !canRemove;
    btn.style.opacity = canRemove ? '1' : '0.45';
    btn.style.cursor = canRemove ? 'pointer' : 'not-allowed';
  };

  if(typeof window._captureVisibleInitial !== 'number') window._captureVisibleInitial = 0;

  var label1 = pivotMode === 'producto' ? 'Desc Art 1'         : 'Nombre Tienda/Club';
  var label2 = pivotMode === 'producto' ? 'Nombre Tienda/Club' : 'Desc Art 1';

  if(!pivot.hasData || !pivot.primaryKeys.length){
    document.getElementById('tResumenHead').innerHTML =
      '<tr style="background:#1a3a5c;color:#fff">' +
      '<th style="font-size:15px;text-align:left;vertical-align:bottom;min-width:130px;padding:6px 8px;border-right:1px solid #3a5a8c">'+label1+'</th>' +
      '<th style="font-size:15px;text-align:left;vertical-align:bottom;min-width:130px;padding:6px 8px;border-right:1px solid #3a5a8c">'+label2+'</th>' +
      '<th style="font-size:15px;text-align:left;vertical-align:bottom;min-width:145px;padding:6px 8px;border-right:1px solid #3a5a8c">Values</th>' +
      '<th style="font-size:15px;text-align:left;vertical-align:bottom;padding:6px 8px;">Estado</th></tr>';
    document.getElementById('tResumenBody').innerHTML =
      '<tr><td colspan="4" style="text-align:center; padding:30px; font-weight:bold; color:#6b7280; background:#f9fafb;">Sin datos para la selección actual</td></tr>';
    var btnAddCapEmpty = document.getElementById('btnAddCaptureSem');
    if(btnAddCapEmpty) btnAddCapEmpty.style.display = 'none';
    var btnRemCapEmpty = document.getElementById('btnRemoveCaptureSem');
    if(btnRemCapEmpty) btnRemCapEmpty.style.display = 'none';
    var btnSaveCapEmpty = document.getElementById('btnSaveCaptureAll');
    if(btnSaveCapEmpty) btnSaveCapEmpty.style.display = 'none';
    window._resumenCaptureBlocks = [];
    renderReabastoPanel();
    return;
  }

  window.bumpResumenRowspans = function(r1i, r2i){
    var r2Cell = document.querySelector('[data-r2-rowspan="'+r1i+'_'+r2i+'"]');
    if(r2Cell) r2Cell.rowSpan = (parseInt(r2Cell.getAttribute('rowspan')||r2Cell.rowSpan,10)||0) + 1;
    var r1Cell = document.querySelector('[data-r1-rowspan="'+r1i+'"]');
    if(r1Cell) r1Cell.rowSpan = (parseInt(r1Cell.getAttribute('rowspan')||r1Cell.rowSpan,10)||0) + 1;
  };

  window.decreaseResumenRowspans = function(r1i, r2i){
    var r2Cell = document.querySelector('[data-r2-rowspan="'+r1i+'_'+r2i+'"]');
    if(r2Cell) r2Cell.rowSpan = Math.max(1, (parseInt(r2Cell.getAttribute('rowspan')||r2Cell.rowSpan,10)||1) - 1);
    var r1Cell = document.querySelector('[data-r1-rowspan="'+r1i+'"]');
    if(r1Cell) r1Cell.rowSpan = Math.max(1, (parseInt(r1Cell.getAttribute('rowspan')||r1Cell.rowSpan,10)||1) - 1);
  };

  if(state.resumenMode === 'semanas'){
    var semsResumen = getSemanasActivas().map(function(s){ return String(s); });
    if(!semsResumen.length){
      document.getElementById('tResumenHead').innerHTML =
        '<tr style="background:#1a3a5c;color:#fff">' +
        '<th style="font-size:15px;text-align:left;vertical-align:bottom;min-width:130px;padding:6px 8px;border-right:1px solid #3a5a8c">'+label1+'</th>' +
        '<th style="font-size:15px;text-align:left;vertical-align:bottom;min-width:130px;padding:6px 8px;border-right:1px solid #3a5a8c">'+label2+'</th>' +
        '<th style="font-size:15px;text-align:left;vertical-align:bottom;min-width:145px;padding:6px 8px;border-right:1px solid #3a5a8c">Values</th>' +
        '<th style="font-size:15px;text-align:left;vertical-align:bottom;padding:6px 8px;">Estado</th></tr>';
      document.getElementById('tResumenBody').innerHTML =
        '<tr><td colspan="4" style="text-align:center; padding:30px; font-weight:bold; color:#6b7280; background:#f9fafb;">Selecciona al menos una semana</td></tr>';
      var btnAddCapNoSem = document.getElementById('btnAddCaptureSem');
      if(btnAddCapNoSem) btnAddCapNoSem.style.display = 'none';
      var btnRemCapNoSem = document.getElementById('btnRemoveCaptureSem');
      if(btnRemCapNoSem) btnRemCapNoSem.style.display = 'none';
      var btnSaveCapNoSem = document.getElementById('btnSaveCaptureAll');
      if(btnSaveCapNoSem) btnSaveCapNoSem.style.display = 'none';
      window._resumenCaptureBlocks = [];
      renderReabastoPanel();
      return;
    }
    window._captureMinRowsSem = semsResumen.length;
    
    /* Calcular la siguiente semana para autocompletar SEM — siempre desde la última sem GLOBAL */
    if(!window._nextSemanaProyeccion){
      window._nextSemanaProyeccion = _getBaseNextSemanaProyeccion();
    }
    var diasOrden = [1,2,3,4,5,6,0];
    var diasAbrev = ['Dom','Lun','Mar','Mié','Jue','Vie','Sáb'];
    var totalsByDay = pivot.totalsByDay || {};
    var weekCols = pivot.columns.filter(function(col){ return col.type === 'week'; });
    var dayGroups = [];

    diasOrden.forEach(function(dow){
      var items = [];
      semsResumen.forEach(function(s){
        var fechas = Object.keys(totalsByDay[s] || {}).sort();
        var fechaDia = null;
        fechas.some(function(fecha){
          var dt = parseResumenDate(fecha);
          if(dt && dt.getDay() === dow){ fechaDia = fecha; return true; }
          return false;
        });
        if(fechaDia){
          items.push({week:s, weekLabel:resumenSemLabel(s), key:'D'+fechaDia});
        }
      });
      if(items.length){
        dayGroups.push({dow:dow, label:diasAbrev[dow], items:items});
      }
    });

    if(!dayGroups.length){
      dayGroups = [{dow:null, label:'SEM', items:weekCols.map(function(col){ return {week:col.rawWeek, weekLabel:col.label, key:col.key}; })}];
    }
    window._captureDayGroups = dayGroups;

    var headSemanal =
      '<tr style="background:#1a3a5c;color:#fff;font-size:14px;">' +
      '<th style="font-size:15px;text-align:left;vertical-align:bottom;min-width:130px;padding:10px 8px;border-right:1px solid #3a5a8c">'+label1+'</th>' +
      '<th style="font-size:15px;text-align:left;vertical-align:bottom;min-width:130px;padding:10px 8px;border-right:1px solid #3a5a8c">'+label2+'</th>' +
      '<th style="font-size:15px;text-align:left;vertical-align:bottom;min-width:145px;padding:10px 8px;border-right:1px solid #3a5a8c">Values</th>';

    dayGroups.forEach(function(group){
      headSemanal += '<th style="font-size:15px;text-align:right;width:68px;min-width:68px;max-width:68px;padding:10px 8px;font-weight:bold">'+group.label+'</th>';
    });
    headSemanal += '<th style="font-size:15px;text-align:right;width:82px;min-width:82px;vertical-align:bottom;padding:10px 8px;border-left:2px solid #4a6a9c">Total general</th></tr>';
    document.getElementById('tResumenHead').innerHTML = headSemanal;

    window.toggleResumenLevel = function(id){
      var rows = document.querySelectorAll('[data-group="'+id+'"]');
      if(!rows.length) return;
      var isHidden = rows[0].style.display === 'none';
      rows.forEach(function(r){ r.style.display = isHidden ? '' : 'none'; });
      var btn = document.querySelector('[data-togbtn="'+id+'"]');
      if(btn) btn.innerText = isHidden ? '−' : '+';
    };

    var TOG_STYLE_SEM =
      'display:inline-block;width:13px;height:13px;line-height:11px;text-align:center;' +
      'border:1px solid #888;margin-right:5px;font-size:11px;font-weight:bold;' +
      'color:#333;background:#fff;cursor:pointer;vertical-align:middle;flex-shrink:0;';

    var bodySemanal = [];
    window._resumenCaptureBlocks = [];

    /* ── Función para actualizar totales de Capture (modo semanas) ── */
    window.updateCaptureRowTotal = function(r1i, r2i, wi, ngr){
      /* Suma inputs de esta fila */
      var sum = 0;
      for(var gi=0; gi<ngr; gi++){
        var inp = document.getElementById('ci_'+r1i+'_'+r2i+'_'+wi+'_'+gi);
        if(inp) sum += parseFloat(inp.value)||0;
      }
      var tot = document.getElementById('ct_'+r1i+'_'+r2i+'_'+wi);
      if(tot) tot.textContent = sum ? fmt(sum) : '0';
      /* Contar todas las filas de capture existentes (dinámicas incluidas) */
      var semsLen = 0;
      while(document.getElementById('ct_'+r1i+'_'+r2i+'_'+semsLen)) semsLen++;
      /* Grand total */
      var grand = 0;
      for(var wi2=0; wi2<semsLen; wi2++){
        var tEl = document.getElementById('ct_'+r1i+'_'+r2i+'_'+wi2);
        if(tEl) grand += parseFloat(tEl.textContent.replace(/,/g,''))||0;
      }
      var gEl = document.getElementById('ctgrand_'+r1i+'_'+r2i);
      if(gEl) gEl.textContent = grand ? fmt(grand) : '0';
      /* Totales por columna */
      for(var gi2=0; gi2<ngr; gi2++){
        var colSum = 0;
        for(var wi3=0; wi3<semsLen; wi3++){
          var inp3 = document.getElementById('ci_'+r1i+'_'+r2i+'_'+wi3+'_'+gi2);
          if(inp3) colSum += parseFloat(inp3.value)||0;
        }
        var cEl = document.getElementById('ctcol_'+r1i+'_'+r2i+'_'+gi2);
        if(cEl) cEl.textContent = colSum ? fmt(colSum) : '';
      }
      /* Actualizar celdas Programado */
      if(typeof _updateProgramadoDisplay === 'function') _updateProgramadoDisplay(r1i, r2i, ngr);
    };

    /* ── Agrega fila Capture dinámica (modo semanas) ── */
    window.recalcCaptureSemTotals = function(r1i, r2i, ngr){
      var semsLen = 0;
      while(document.getElementById('ct_'+r1i+'_'+r2i+'_'+semsLen)) semsLen++;
      var grand = 0;
      for(var wi2=0; wi2<semsLen; wi2++){
        var sum = 0;
        for(var gi=0; gi<ngr; gi++){
          var inp = document.getElementById('ci_'+r1i+'_'+r2i+'_'+wi2+'_'+gi);
          if(inp) sum += parseFloat(inp.value)||0;
        }
        var tot = document.getElementById('ct_'+r1i+'_'+r2i+'_'+wi2);
        if(tot) tot.textContent = sum ? fmt(sum) : '0';
        grand += sum;
      }
      var gEl = document.getElementById('ctgrand_'+r1i+'_'+r2i);
      if(gEl) gEl.textContent = grand ? fmt(grand) : '0';
      for(var gi2=0; gi2<ngr; gi2++){
        var colSum = 0;
        for(var wi3=0; wi3<semsLen; wi3++){
          var inp3 = document.getElementById('ci_'+r1i+'_'+r2i+'_'+wi3+'_'+gi2);
          if(inp3) colSum += parseFloat(inp3.value)||0;
        }
        var cEl = document.getElementById('ctcol_'+r1i+'_'+r2i+'_'+gi2);
        if(cEl) cEl.textContent = colSum ? fmt(colSum) : '';
      }
      if(typeof _updateProgramadoDisplay === 'function') _updateProgramadoDisplay(r1i, r2i, ngr);
    };

    window.removeCaptureSemRow = function(r1i, r2i, wi, ngr){
      var minRows = window._captureMinRowsSem || 1;
      if(wi < minRows) return;
      var totEl = document.getElementById('ct_'+r1i+'_'+r2i+'_'+wi);
      if(!totEl) return;
      var tr = totEl.closest('tr');
      if(tr) tr.remove();
      /* Actualizar store: quitar la fila wi de todas las entradas prod+tienda de este bloque */
      var modeKey = _captureProjModeKey();
      if(window._captureProjections && window._captureProjections[modeKey]){
        var blocks = window._resumenCaptureBlocks || [];
        blocks.forEach(function(b){
          if(b.i1 !== r1i || b.r2i !== r2i) return;
          if(!b.prod || !b.tienda) return;
          var storeKey = _captureProjStoreKey(b.prod, b.tienda);
          var entry = window._captureProjections[modeKey][storeKey];
          if(entry && entry.rows){
            entry.rows.splice(wi, 1);
          }
        });
      }
      window.decreaseResumenRowspans(r1i, r2i);
      window.recalcCaptureSemTotals(r1i, r2i, ngr);
      window.updateRemoveCaptureBtn();
    };

    window.removeLastCaptureSemRow = function(r1i, r2i, ngr){
      var minRows = window._captureMinRowsSem || 1;
      var lastWi = -1, wi = 0;
      while(document.getElementById('ct_'+r1i+'_'+r2i+'_'+wi)){ lastWi = wi; wi++; }
      if(lastWi < minRows) return;
      window.removeCaptureSemRow(r1i, r2i, lastWi, ngr);
    };

    window.revealCaptureSemRow = function(r1i, r2i, wi){
      var totEl = document.getElementById('ct_'+r1i+'_'+r2i+'_'+wi);
      if(!totEl) return;
      var tr = totEl.closest('tr');
      if(!tr || tr.style.display !== 'none') return;
      /* Poner semana automática si el input está vacío */
      var semEl = document.getElementById('csem_'+r1i+'_'+r2i+'_'+wi);
      if(semEl && !semEl.value && window._nextSemanaProyeccion){
        semEl.value = window._nextSemanaProyeccion;
      }
      tr.style.display = '';
      window.bumpResumenRowspans(r1i, r2i);
    };

    window.hideCaptureSemRow = function(r1i, r2i, wi, ngr){
      var totEl = document.getElementById('ct_'+r1i+'_'+r2i+'_'+wi);
      if(!totEl) return;
      var tr = totEl.closest('tr');
      if(!tr || tr.style.display === 'none') return;
      /* Limpiar inputs de semana y valores antes de ocultar */
      var semEl = document.getElementById('csem_'+r1i+'_'+r2i+'_'+wi);
      if(semEl) semEl.value = '';
      for(var gi=0; gi<(ngr||0); gi++){
        var inp = document.getElementById('ci_'+r1i+'_'+r2i+'_'+wi+'_'+gi);
        if(inp) inp.value = '';
      }
      totEl.textContent = '0';
      tr.style.display = 'none';
      window.decreaseResumenRowspans(r1i, r2i);
      window.recalcCaptureSemTotals(r1i, r2i, ngr);
    };

    window.addCaptureSemRow = function(r1i, r2i, id1, bg, ngr){
      var dayGroups = window._captureDayGroups || [];
      var wi = 0;
      var lastCaptureRow = null;
      while(document.getElementById('ct_'+r1i+'_'+r2i+'_'+wi)){
        var totEl = document.getElementById('ct_'+r1i+'_'+r2i+'_'+wi);
        if(totEl) lastCaptureRow = totEl.closest('tr');
        wi++;
      }
      if(!lastCaptureRow) return;
      var tr = document.createElement('tr');
      tr.setAttribute('data-group', id1);
      tr.style.cssText = 'background:'+bg+';border-top:1px dashed #64a8d8';
      /* Celda Values con input de semana */
      var td1 = document.createElement('td');
      td1.setAttribute('data-group', id1);
      td1.style.cssText = 'position:static;background:'+bg+';padding:3px 6px;white-space:nowrap;font-size:14px;font-weight:600;color:#1565c0;border-right:1px solid #e0e6f0;text-align:left;vertical-align:middle';
      td1.innerHTML = '<span style="font-size:12px;color:#1565c0">Capture SEM</span><div style="font-size:12px;font-weight:bold;color:#1a3a5c;margin-top:2px;background:#ddeeff;display:inline-block;border-radius:3px;padding:1px 6px"><input type="text" id="csem_'+r1i+'_'+r2i+'_'+wi+'" placeholder="SEM" value="'+(window._nextSemanaProyeccion||'')+'" oninput="window.syncCaptureProjFromBlock('+r1i+','+r2i+')" style="width:50px;border:2px solid #2980b9;border-radius:4px;padding:2px 4px;font-size:13px;text-align:center;background:#eef6fc;color:#0a2a4a;font-weight:bold;outline:none"></div><button type="button" title="Eliminar fila" onclick="window.removeCaptureSemRow('+r1i+','+r2i+','+wi+','+ngr+')" style="margin-left:3px;font-size:11px;font-weight:bold;background:#c62828;color:#fff;border:none;border-radius:3px;padding:1px 5px;cursor:pointer;vertical-align:middle">×</button>';
      tr.appendChild(td1);
      /* Inputs por día */
      dayGroups.forEach(function(group, gi){
        var td = document.createElement('td');
        td.setAttribute('data-group', id1);
        td.style.cssText = 'background:'+bg+';padding:2px 4px;text-align:right;width:68px;min-width:68px;max-width:68px;vertical-align:middle';
        td.innerHTML = '<input type="number" id="ci_'+r1i+'_'+r2i+'_'+wi+'_'+gi+'" min="0" placeholder="0" oninput="window.updateCaptureRowTotal('+r1i+','+r2i+','+wi+','+ngr+');window.syncCaptureProjFromBlock('+r1i+','+r2i+')" style="width:72px;border:2px solid #2980b9;border-radius:6px;padding:4px 6px;font-size:14px;font-weight:bold;text-align:right;background:#eef6fc;color:#0a2a4a;outline:none;box-shadow:inset 0 1px 3px rgba(0,0,0,0.1)">';
        tr.appendChild(td);
      });
      /* Total fila */
      var tdTot = document.createElement('td');
      tdTot.id = 'ct_'+r1i+'_'+r2i+'_'+wi;
      tdTot.setAttribute('data-group', id1);
      tdTot.style.cssText = 'font-size:14px;font-weight:bold;color:#0d47a1;background:'+bg+';text-align:right;width:82px;min-width:82px;padding:3px 8px;border-left:2px solid #4a6a9c';
      tdTot.textContent = '0';
      tr.appendChild(tdTot);
      window.bumpResumenRowspans(r1i, r2i);
      if(lastCaptureRow.nextSibling){
        lastCaptureRow.parentNode.insertBefore(tr, lastCaptureRow.nextSibling);
      } else {
        lastCaptureRow.parentNode.appendChild(tr);
      }
      window.updateRemoveCaptureBtn();
    };

    /* ── Helper: propaga el estado de captura a TODAS las tiendas/productos
       (no solo los bloques del DOM visible). Así, cuando se cambia el filtro
       de tienda, las tiendas que no estaban en pantalla también reciben el
       mismo número de semanas de captura. ── */
    function _propagateCaptureToAllStores(modeKey, newVisibleInitial, newNextSem) {
      if(!window._captureProjections) window._captureProjections = { sem: {}, norm: {}, _meta: {} };
      if(!window._captureProjections[modeKey]) window._captureProjections[modeKey] = {};
      var store = window._captureProjections[modeKey];
      var allTiendas = (typeof DATA !== 'undefined' && DATA.tiendas) || [];
      var allProds   = (typeof DATA !== 'undefined' && DATA.productos) || [];
      var pivotMode  = state.resumenPivot || 'producto';
      /* Obtener filas de referencia de algún bloque visible */
      var blocks = window._resumenCaptureBlocks || [];
      var refRows = null;
      if(blocks.length) {
        var refB = blocks[0];
        if(_captureBlockDomExists(refB, modeKey)) {
          refRows = _collectCaptureRowsFromBlock(refB, modeKey);
        }
      }
      /* Para cada combinación prod+tienda que NO esté ya en el DOM visible,
         crear o actualizar la entrada en el store con el mismo estado. */
      var visibleKeys = {};
      blocks.forEach(function(b){ visibleKeys[_captureProjStoreKey(b.prod, b.tienda)] = true; });
      allTiendas.forEach(function(t){
        allProds.forEach(function(p){
          var sk = _captureProjStoreKey(p, t);
          if(visibleKeys[sk]) return; /* ya fue manejado por saveCaptureProjectionsFromDom */
          var existing = store[sk];
          /* Conservar SIEMPRE las filas ya guardadas (Programado) de esta tienda/producto,
             aunque no esté visible en este momento — antes se perdían porque se
             reemplazaba todo el registro con solo la plantilla de captura vacía. */
          var existingSaved = (existing && existing.rows) ? existing.rows.filter(function(r){ return r.saved; }) : [];
          if(refRows) {
            /* Copiar las mismas filas de semana (sin guardar) pero con valores vacíos
               o conservar los valores si ya existían para no perder lo capturado sin guardar. */
            var existingUnsaved = (existing && existing.rows) ? existing.rows.filter(function(r){ return !r.saved; }) : [];
            var mergedRows = refRows.map(function(r, i){
              var oldRow = existingUnsaved[i];
              var vals = oldRow ? oldRow.values : r.values.map(function(){ return ''; });
              var semVal = (oldRow && oldRow.sem) ? oldRow.sem : r.sem;
              return { sem: semVal, values: vals, hidden: r.hidden, saved: false };
            });
            store[sk] = { visibleInitial: newVisibleInitial, rows: existingSaved.concat(mergedRows) };
          } else if(existing) {
            /* Solo actualizar el visibleInitial para sincronizar */
            existing.visibleInitial = newVisibleInitial;
          }
        });
      });
      window._captureProjections._meta.visibleInitial = newVisibleInitial;
      if(newNextSem) window._captureProjections._meta.nextSem = newNextSem;
    }

    window.addCaptureRowAll = function(){
      var blocks = window._resumenCaptureBlocks || [];
      var minRows = window._captureMinRowsSem || 1;
      var visibleInitial = window._captureVisibleInitial || 0;
      _ensureNextSemanaProyeccion();
      if(visibleInitial < minRows){
        window._captureVisibleInitial = visibleInitial + 1;
        blocks.forEach(function(b){
          var semEl = document.getElementById('csem_'+b.i1+'_'+b.r2i+'_'+visibleInitial);
          if(semEl) semEl.value = window._nextSemanaProyeccion || '';
          window.revealCaptureSemRow(b.i1, b.r2i, visibleInitial);
        });
        /* Incrementar semana para la siguiente */
        if(window._nextSemanaProyeccion){
          window._nextSemanaProyeccion = parseInt(window._nextSemanaProyeccion, 10) + 1;
        }
        window.updateRemoveCaptureBtn();
        saveCaptureProjectionsFromDom();
        /* Propagar a todas las tiendas no visibles */
        _propagateCaptureToAllStores(_captureProjModeKey(), window._captureVisibleInitial, window._nextSemanaProyeccion);
        return;
      }
      /* Agregar nueva fila con semana auto-incremental */
      blocks.forEach(function(b){
        window.addCaptureSemRow(b.i1, b.r2i, b.id1, b.bg, b.ngr);
        /* La nueva fila ya tiene _nextSemanaProyeccion — solo ajustar el input recién creado */
        var wi = 0;
        while(document.getElementById('csem_'+b.i1+'_'+b.r2i+'_'+wi)) wi++;
        var lastWi = wi - 1;
        var semInput = document.getElementById('csem_'+b.i1+'_'+b.r2i+'_'+lastWi);
        if(semInput) semInput.value = window._nextSemanaProyeccion || '';
      });
      /* Incrementar para la siguiente */
      if(window._nextSemanaProyeccion){
        window._nextSemanaProyeccion = parseInt(window._nextSemanaProyeccion, 10) + 1;
      }
      saveCaptureProjectionsFromDom();
      /* Propagar a todas las tiendas no visibles */
      _propagateCaptureToAllStores(_captureProjModeKey(), window._captureVisibleInitial, window._nextSemanaProyeccion);
    };

    window.removeCaptureRowAll = function(){
      if(!confirm('¿Estás seguro que deseas eliminar?')) return;
      var blocks = window._resumenCaptureBlocks || [];
      var minRows = window._captureMinRowsSem || 1;
      var modeKey = _captureProjModeKey();
      var hasDynamic = false;
      blocks.forEach(function(b){
        var wi = 0;
        while(document.getElementById('ct_'+b.i1+'_'+b.r2i+'_'+wi)) wi++;
        if(wi > minRows) hasDynamic = true;
      });
      if(hasDynamic){
        /* Semana objetivo a quitar = la última agregada con +SEM (nextSemanaProyeccion - 1).
           Así solo se borra esa semana puntual en cada tienda/producto, en vez de
           quitar a ciegas "la última fila" (lo cual borraba semanas distintas,
           incluso ya guardadas, en tiendas que no tenían la semana nueva capturada). */
        var targetSemRemoveAll = window._nextSemanaProyeccion
          ? String(parseInt(window._nextSemanaProyeccion, 10) - 1)
          : null;
        /* Limpiar store ANTES de borrar del DOM para que quede vacío */
        if(window._captureProjections && window._captureProjections[modeKey]){
          var store = window._captureProjections[modeKey];
          /* Borrar SOLO la fila de la semana objetivo en cada key del store (visibles y no visibles) */
          Object.keys(store).forEach(function(sk){
            var entry = store[sk];
            if(entry && entry.rows && entry.rows.length > 0){
              var visRows = entry.rows.filter(function(r){
                return r.sem || (r.values||[]).some(function(v){ return v !== '' && parseFloat(v) > 0; });
              });
              var idxRemove = -1;
              if(targetSemRemoveAll !== null){
                for(var iR = visRows.length - 1; iR >= 0; iR--){
                  if(String(visRows[iR].sem || '').trim() === targetSemRemoveAll){ idxRemove = iR; break; }
                }
              } else if(visRows.length > 0){
                idxRemove = visRows.length - 1;
              }
              if(idxRemove >= 0) visRows.splice(idxRemove, 1);
              store[sk] = { visibleInitial: window._captureVisibleInitial||0, rows: visRows };
            }
          });
        }
        blocks.forEach(function(b){
          window.removeLastCaptureSemRow(b.i1, b.r2i, b.ngr);
        });
        /* Decrementar semana al borrar fila dinámica */
        if(window._nextSemanaProyeccion){
          window._nextSemanaProyeccion = Math.max(_getBaseNextSemanaProyeccion(), parseInt(window._nextSemanaProyeccion, 10) - 1);
        }
        if(window._captureProjections && window._captureProjections._meta){
          window._captureProjections._meta.nextSem = window._nextSemanaProyeccion;
        }
        return;
      }
      var visibleInitial = window._captureVisibleInitial || 0;
      if(visibleInitial > 0){
        var hideWi = visibleInitial - 1;
        window._captureVisibleInitial = hideWi;
        blocks.forEach(function(b){
          window.hideCaptureSemRow(b.i1, b.r2i, hideWi, b.ngr);
        });
        /* Decrementar semana para que el siguiente +SEM sea correcto */
        if(window._nextSemanaProyeccion){
          window._nextSemanaProyeccion = Math.max(_getBaseNextSemanaProyeccion(), parseInt(window._nextSemanaProyeccion, 10) - 1);
        }
        window.updateRemoveCaptureBtn();
        saveCaptureProjectionsFromDom();
        /* Propagar a todas las tiendas no visibles */
        _propagateCaptureToAllStores(modeKey, window._captureVisibleInitial, window._nextSemanaProyeccion);
      }
    };

    pivot.primaryKeys.forEach(function(r1Key, i1){
      var r1Data = pivot.rowsData[r1Key];
      var r2Keys = Object.keys(r1Data.subRows).sort();
      var hasSellThru = defs.some(function(d){ return d.key === 'sell_thru'; });
      var hasCumulative = (hasSellThru && semsResumen.length >= 2) ? 1 : 0;
      var baseR2RowCount = defs.length * semsResumen.length + hasCumulative;
      /* Precomputar filas programado por r2 */
      var r2ProgRowsMap = {};
      var totalProgRows = 0;
      r2Keys.forEach(function(r2k){
        var pr = _getSavedProgramadoRows(r1Key, r2k);
        r2ProgRowsMap[r2k] = pr;
        totalProgRows += pr.length;
      });
      var r1Span = r2Keys.length * baseR2RowCount + totalProgRows;
      var id1 = 'grp-'+i1;

      r2Keys.forEach(function(r2Key, r2i){
        var r2Data = r1Data.subRows[r2Key];
        var isFirstR2 = (r2i === 0);
        var borderTop = isFirstR2 ? '2px solid #b0bcd8' : '1px solid #e4e8f0';
        var r2Bg = r2i % 2 === 0 ? '#ffffff' : '#f4f6fa';
        var savedProgRows = r2ProgRowsMap[r2Key] || [];
        var r2RowCount = baseR2RowCount + savedProgRows.length;

        /* ── Filas Programado (arriba de embarque) ── */
        var hasProgRows = savedProgRows.length > 0;
        savedProgRows.forEach(function(progRow, progIdx){
          var isFirstProg = (progIdx === 0);
          var progBg = r2i % 2 === 0 ? '#fff8e1' : '#fff3c3';
          var pRow = '<tr style="background:'+progBg+';border-top:'+(isFirstProg ? borderTop : '1px solid #edf0f6')+'">';

          if(isFirstR2 && isFirstProg){
            pRow +=
              '<td rowspan="'+r1Span+'" data-r1-rowspan="'+i1+'" style="font-size:15px;font-weight:bold;vertical-align:top;'+
              'white-space:nowrap;padding:5px 8px;border-right:1px solid #c8d2e8;'+
              'border-top:2px solid #b0bcd8;background:#eef2fa">'+
              '<span data-togbtn="'+id1+'" style="'+TOG_STYLE_SEM+'" onclick="toggleResumenLevel(\''+id1+'\')">−</span>'+r1Key+'</td>';
          }

          if(isFirstProg){
            pRow +=
              '<td rowspan="'+r2RowCount+'" data-r2-rowspan="'+i1+'_'+r2i+'" data-group="'+id1+'" style="vertical-align:top;white-space:nowrap;padding:4px 6px;'+
              'border-top:'+borderTop+';border-right:1px solid #d0d8ea;font-size:14px;color:#2980B9;font-weight:600;background:'+r2Bg+'">'+
              r2Key+'</td>';
          }

          pRow += '<td data-group="'+id1+'" style="position:static;background:'+progBg+';padding:3px 6px;white-space:nowrap;'+
            'font-size:15px;font-weight:600;color:#e65100;border-right:1px solid #e0e6f0;text-align:left">Programado '+
            '<span style="font-size:13px;font-weight:bold;color:#333;background:#ffe082;border-radius:3px;padding:1px 5px">'+(progRow.sem || '')+'</span>'+
            '<button type="button" onclick="_deleteProgramadoRow(\''+r1Key+'\',\''+r2Key+'\','+progIdx+',\'sem\',\''+(progRow.sem || '')+'\')" style="margin-left:6px;font-size:11px;background:transparent;border:none;cursor:pointer;color:#c62828">🗑️</button>'+
            '</td>';

          dayGroups.forEach(function(group, gi){
            var v = progRow.values[gi] !== undefined ? progRow.values[gi] : '';
            var fv = v !== '' ? fmt(parseFloat(v)||0) : '';
            pRow += '<td data-group="'+id1+'" style="font-size:14px;color:#e65100;background:'+progBg+';text-align:right;width:68px;min-width:68px;max-width:68px;padding:3px 8px;vertical-align:middle">'+
              (fv||'') +
              '<button type="button" onclick="_editProgramadoRow(\''+r1Key+'\',\''+r2Key+'\','+progIdx+','+gi+',\'sem\')" style="margin-left:2px;font-size:10px;background:transparent;border:none;cursor:pointer;color:#1565c0">✏️</button>' +
              '</td>';
          });

          var progRowTotal = progRow.values.reduce(function(s, x){ return s + (parseFloat(x)||0); }, 0);
          pRow += '<td data-group="'+id1+'" style="font-size:14px;font-weight:bold;color:#bf360c;background:'+progBg+';text-align:right;width:82px;min-width:82px;padding:3px 8px;border-left:2px solid #e0a030">'+
            fmt(progRowTotal)+'</td>';

          pRow += '</tr>';
          bodySemanal.push(pRow);
        });

        defs.forEach(function(def, di){
          var rowBg = r2i % 2 === 0 ? (def.bgEven || '#ffffff') : (def.bgOdd || '#f4f6fa');
          var isFirstDef = (di === 0);
          semsResumen.forEach(function(semActual, wi){
          var isFirstWeek = (wi === 0);
          var isVeryFirst = (isFirstR2 && di === 0 && wi === 0);
          var rowHTML = '<tr style="background:'+rowBg+';border-top:'+(isFirstDef && isFirstWeek ? borderTop : '1px solid #edf0f6')+'">';

          if(!hasProgRows && isVeryFirst){
            rowHTML +=
              '<td rowspan="'+r1Span+'" data-r1-rowspan="'+i1+'" style="font-size:15px;font-weight:bold;vertical-align:top;'+
              'white-space:nowrap;padding:5px 8px;border-right:1px solid #c8d2e8;'+
              'border-top:2px solid #b0bcd8;background:#eef2fa">'+
              '<span data-togbtn="'+id1+'" style="'+TOG_STYLE_SEM+'" onclick="toggleResumenLevel(\''+id1+'\')">−</span>'+r1Key+'</td>';
          }

          if(!hasProgRows && isFirstDef && isFirstWeek){
            rowHTML +=
              '<td rowspan="'+r2RowCount+'" data-r2-rowspan="'+i1+'_'+r2i+'" data-group="'+id1+'" style="vertical-align:top;white-space:nowrap;padding:4px 6px;'+
              'border-top:'+borderTop+';border-right:1px solid #d0d8ea;font-size:14px;color:#2980B9;font-weight:600;background:'+r2Bg+';">'+
              r2Key+'</td>';
          }

          var cIdLabel = 'sem_lbl_' + r1Key + '_' + r2Key + '_' + def.key + '_' + semActual;
          var cAttrLabel = getCommentAttr(cIdLabel);
          rowHTML +=
            '<td data-group="'+id1+'" class="'+cAttrLabel.cls+'"'+cAttrLabel.str+' style="position:static;background:'+rowBg+';padding:3px 6px;white-space:nowrap;'+
            'font-size:15px;font-weight:normal;color:#333;border-right:1px solid #e0e6f0;text-align:left">'+
            (isFirstWeek ? def.label : '')+
            '<div style="font-size:14px;color:#1a3a5c;font-weight:bold;margin-top:2px">'+resumenSemLabel(semActual)+'</div></td>';

          dayGroups.forEach(function(group){
            var item = group.items.find(function(x){ return x.week === semActual; });
            var isDaily = group.dow !== null;
            var v = item ? cellVal(def, (r2Data._total||{})[item.key], isDaily) : '';
            var hasVal = v && v !== '0';
            var cId = 'sem_' + r1Key + '_' + r2Key + '_' + def.key + '_' + semActual + '_' + (group.dow||'T');
            var cAttr = getCommentAttr(cId);
            rowHTML +=
              '<td data-group="'+id1+'" class="'+cAttr.cls+'"'+cAttr.str+' style="font-size:14px;position:static;background:'+rowBg+';text-align:right;width:68px;min-width:68px;max-width:68px;padding:3px 8px;vertical-align:middle;'+
              (hasVal ? 'color:#1a3a5c' : 'color:#bbb')+'">'+(v||'')+'</td>';
            });

          var tot = cellVal(def, (r2Data._total||{})['W'+semActual]);
          if(def.key === 'sell_thru' && !tot) tot = '0.0%';
          var cellBg = rowBg;
          var cellColor = (def.key === 'sell_thru' ? '#000' : '');
          if (def.key === 'sell_thru' && tot) {
            var valPct = parseFloat(tot);
            if (valPct <= 79) cellBg = '#ffcdd2';
            else if (valPct < 90) cellBg = '#fff9c4';
            else cellBg = '#c8e6c9';
          }
          var colorStyle = cellColor ? 'color:'+cellColor+';' : '';
          var cIdTot = 'sem_' + r1Key + '_' + r2Key + '_' + def.key + '_' + semActual + '_TOTAL';
          var cAttrTot = getCommentAttr(cIdTot);
          rowHTML +=
            '<td data-group="'+id1+'" class="'+cAttrTot.cls+'"'+cAttrTot.str+' style="font-size:14px;'+colorStyle+'position:static;background:'+cellBg+';text-align:right;width:82px;min-width:82px;padding:3px 8px;vertical-align:middle;'+
            'font-weight:bold;border-left:1px solid #c0cce0">'+(tot||'0')+'</td>';

          rowHTML += '</tr>';
          bodySemanal.push(rowHTML);
          });

          /* ── Fila Acumulado para Sell Thru ── */
          if(def.key === 'sell_thru' && semsResumen.length >= 2){
            var acumBg = r2i % 2 === 0 ? (def.bgEven || '#fcf8e3') : (def.bgOdd || '#fcf8e3');
            var acumRow = '<tr style="background:'+acumBg+';border-top:1px solid #edf0f6">';
            acumRow += '<td data-group="'+id1+'" style="position:static;background:'+acumBg+';padding:3px 6px;white-space:nowrap;font-size:15px;font-weight:normal;color:#333;border-right:1px solid #e0e6f0;text-align:left"><div style="font-size:14px;color:#1a3a5c;font-weight:bold;margin-top:2px">Accumulated</div></td>';
            dayGroups.forEach(function(group){
              acumRow += '<td data-group="'+id1+'" style="font-size:14px;position:static;background:'+acumBg+';text-align:right;width:68px;min-width:68px;max-width:68px;padding:3px 8px;vertical-align:middle;color:#bbb"></td>';
            });
            var acumV = 0, acumE = 0;
            semsResumen.forEach(function(sem){
              var wData = (r2Data._total || {})['W'+sem] || {};
              acumV += wData.ventas || 0;
              acumE += wData.embarque || 0;
            });
            var acumPct = acumE > 0 ? ((acumV/acumE)*100).toFixed(1)+'%' : '0.0%';
            var acumCellBg = acumBg;
            if(acumPct){
              var acumVal = parseFloat(acumPct);
              if(acumVal <= 79) acumCellBg = '#ffcdd2';
              else if(acumVal < 90) acumCellBg = '#fff9c4';
              else acumCellBg = '#c8e6c9';
            }
            acumRow += '<td data-group="'+id1+'" style="font-size:14px;color:#000;position:static;background:'+acumCellBg+';text-align:right;width:82px;min-width:82px;padding:3px 8px;vertical-align:middle;font-weight:bold;border-left:1px solid #c0cce0">'+acumPct+'</td>';
            acumRow += '</tr>';
            bodySemanal.push(acumRow);
          }

          /* ── Filas Capture ── */
          if(di === defs.length - 1){
            var ngrCap = dayGroups.length;
            semsResumen.forEach(function(semActual, wi){
              var capBg = r2i % 2 === 0 ? '#e3f2fd' : '#d6ebf9';
              var capRow = '<tr data-group="'+id1+'" data-capture-initial="1" style="display:none;background:'+capBg+';border-top:1px dashed #64a8d8">';
              /* Values cell: Capture SEM en todas las filas */
              capRow += '<td data-group="'+id1+'" style="position:static;background:'+capBg+';padding:3px 6px;white-space:nowrap;font-size:14px;font-weight:600;color:#1565c0;border-right:1px solid #e0e6f0;text-align:left;vertical-align:middle">';
              capRow += '<span style="font-size:12px;color:#1565c0">Capture SEM</span>';
              capRow += '<div style="font-size:12px;font-weight:bold;color:#1a3a5c;margin-top:2px;background:#ddeeff;display:inline-block;border-radius:3px;padding:1px 6px"><input type="text" id="csem_'+i1+'_'+r2i+'_'+wi+'" placeholder="SEM" value="'+(window._nextSemanaProyeccion||'')+'" oninput="window.syncCaptureProjFromBlock('+i1+','+r2i+')" style="width:50px;border:2px solid #2980b9;border-radius:4px;padding:2px 4px;font-size:13px;text-align:center;background:#eef6fc;color:#0a2a4a;font-weight:bold;outline:none"></div>';
              capRow += '</td>';
              dayGroups.forEach(function(group, gi){
                /* Siempre mostrar input, independientemente de si hay dato */
                var inputId = 'ci_'+i1+'_'+r2i+'_'+wi+'_'+gi;
                capRow += '<td data-group="'+id1+'" style="background:'+capBg+';padding:2px 4px;text-align:right;width:68px;min-width:68px;max-width:68px;vertical-align:middle"><input type="number" id="'+inputId+'" min="0" placeholder="0" oninput="window.updateCaptureRowTotal('+i1+','+r2i+','+wi+','+ngrCap+');window.syncCaptureProjFromBlock('+i1+','+r2i+')" style="width:72px;border:2px solid #2980b9;border-radius:6px;padding:4px 6px;font-size:14px;font-weight:bold;text-align:right;background:#eef6fc;color:#0a2a4a;outline:none;box-shadow:inset 0 1px 3px rgba(0,0,0,0.1)"></td>';
              });
              var totId = 'ct_'+i1+'_'+r2i+'_'+wi;
              capRow += '<td id="'+totId+'" data-group="'+id1+'" style="font-size:14px;font-weight:bold;color:#0d47a1;background:'+capBg+';text-align:right;width:82px;min-width:82px;padding:3px 8px;border-left:2px solid #4a6a9c">0</td>';
              capRow += '</tr>';
              bodySemanal.push(capRow);
            });
            window._resumenCaptureBlocks.push({
              i1: i1,
              r2i: r2i,
              id1: id1,
              bg: r2i % 2 === 0 ? '#e3f2fd' : '#d6ebf9',
              ngr: ngrCap,
              minRows: semsResumen.length,
              prod: pivotMode === 'producto' ? r1Key : r2Key,
              tienda: pivotMode === 'producto' ? r2Key : r1Key
            });
          }
        });
      });
    });

    if(typeof window._totalTop === 'undefined') window._totalTop = true;
    window.toggleTotalPosition = function() {
      window._totalTop = !window._totalTop;
      if(typeof renderResumen === 'function') renderResumen();
    };

    var totalRowsGroup = [];
    defsAll.forEach(function(def, di){
      semsResumen.forEach(function(semActual, wi){
      var totalRow = '<tr class="pivot-total" style="background:#1a3a5c;color:#fff">';
      if(di === 0 && wi === 0){
        totalRow +=
          '<td colspan="2" rowspan="'+(defsAll.length * semsResumen.length)+'" style="font-size:16px;vertical-align:middle;padding:5px 8px;font-weight:bold;border-right:1px solid #3a5a8c">'+
          'Total general</td>';
      }
      totalRow += '<td style="position:static;background:#1a3a5c;padding:3px 8px;font-size:15px;font-weight:normal;color:#000;">'+(wi === 0 ? def.label : '')+'<div style="font-size:15px;color:#000;font-weight:bold;margin-top:2px">'+resumenSemLabel(semActual)+'</div></td>';
      dayGroups.forEach(function(group){
        var item = group.items.find(function(x){ return x.week === semActual; });
        var isDaily = group.dow !== null;
        var v = '';
        if(def.key === 'sell_thru' && isDaily && item){
          var dayData = pivot.totalsByColumn[item.key] || {};
          var weekData = pivot.totalsByColumn['W'+semActual] || {};
          var dayPOS = dayData.ventas || 0;
          var weekPOS = weekData.ventas || 0;
          v = weekPOS > 0 ? ((dayPOS/weekPOS)*100).toFixed(1)+'%' : '';
        } else {
          v = item ? cellVal(def, pivot.totalsByColumn[item.key], isDaily) : '';
        }
        totalRow += '<td style="font-size:14px;position:static;background:#1a3a5c;color:#000;text-align:right;width:68px;min-width:68px;max-width:68px;padding:3px 8px;vertical-align:middle">'+(v||'')+'</td>';
      });
      var gt = cellVal(def, pivot.totalsByColumn['W'+semActual]);
      var cellBg = '#1a3a5c';
      var cellColor = '#000';
      if (def.key === 'sell_thru' && gt) {
        var valPct = parseFloat(gt);
        if (valPct <= 79) cellBg = '#ffcdd2';
        else if (valPct < 90) cellBg = '#fff9c4';
        else cellBg = '#c8e6c9';
      }
      totalRow += '<td style="font-size:14px;position:static;background:'+cellBg+';color:'+cellColor+';text-align:right;width:82px;min-width:82px;padding:3px 8px;vertical-align:middle;font-weight:bold;border-left:2px solid #4a6a9c">'+gt+'</td>';
      totalRow += '</tr>';
      totalRowsGroup.push(totalRow);
      });
    });

    if (window._totalTop !== false) {
      bodySemanal.unshift.apply(bodySemanal, totalRowsGroup);
    } else {
      bodySemanal.push.apply(bodySemanal, totalRowsGroup);
    }

    document.getElementById('tResumenBody').innerHTML = bodySemanal.join('');
    restoreCaptureProjections();
    var btnAddCap = document.getElementById('btnAddCaptureSem');
    if(btnAddCap) btnAddCap.style.display = (window._resumenCaptureBlocks.length ? 'inline-block' : 'none');
    var btnRemCap = document.getElementById('btnRemoveCaptureSem');
    if(btnRemCap) btnRemCap.style.display = (window._resumenCaptureBlocks.length ? 'inline-block' : 'none');
    var btnSaveCap = document.getElementById('btnSaveCaptureAll');
    if(btnSaveCap) btnSaveCap.style.display = (window._resumenCaptureBlocks.length ? 'inline-flex' : 'none');
    window.updateRemoveCaptureBtn();

    /* ── Función global para agregar sem activa desde el botón + SEM ── */
    window.agregarSemResumen = function(){
      var semInput = prompt('Ingresa el número de semana a agregar (ej: 26):');
      if(!semInput) return;
      var semNum = parseInt(semInput.trim(), 10);
      if(isNaN(semNum)) return;
      var semStr = String(semNum);
      var current = state.semanas_sel && state.semanas_sel.length ? state.semanas_sel.slice() : (DATA.semanas ? DATA.semanas.slice() : []);
      if(current.indexOf(semNum) === -1 && current.indexOf(semStr) === -1){
        current.push(semNum);
        state.semanas_sel = current;
      }
      renderResumen();
    };

    /* ── Función para exportar datos capturados a Excel/CSV ── */
    window.exportCaptureData = function(){
      /* Modo SEM: las sub-columnas de cada fila capturada son días (Lunes…Domingo) */
      saveCaptureProjectionsFromDom();
      var dowFull = {0:'Domingo',1:'Lunes',2:'Martes',3:'Miercoles',4:'Jueves',5:'Viernes',6:'Sabado'};
      var dayLabels = diasOrden.map(function(d){ return dowFull[d]; });
      _exportCaptureToCSV('sem', dayLabels);
    };

    renderReabastoPanel();
    if(typeof resizeAfterChange === 'function') resizeAfterChange();
    return;
  }

  /* ── HEADER ─────────────────────────────────────────────────────── */
  var ncols = pivot.columns.length;
  
  /* Calcular la siguiente semana para autocompletar SEM en modo normal — siempre desde la última sem GLOBAL */
  if(!window._nextSemanaProyeccion){
    window._nextSemanaProyeccion = _getBaseNextSemanaProyeccion();
  }
  var headHTML =
    '<tr style="background:#1a3a5c;color:#fff;font-size:14px;">' +
    '<th rowspan="2" style="font-size:15px;text-align:left;vertical-align:bottom;min-width:130px;padding:10px 8px;border-right:1px solid #3a5a8c">'+label1+'</th>' +
    '<th rowspan="2" style="font-size:15px;text-align:left;vertical-align:bottom;min-width:130px;padding:10px 8px;border-right:1px solid #3a5a8c">'+label2+'</th>' +
    '<th rowspan="2" style="font-size:15px;text-align:left;vertical-align:bottom;min-width:145px;padding:10px 8px;border-right:1px solid #3a5a8c">Values</th>' +
    '<th colspan="'+ncols+'" style="font-size:15px;text-align:center;padding:10px 8px">SEM</th>' +
    '<th rowspan="2" style="font-size:15px;text-align:right;vertical-align:bottom;padding:10px 8px;border-left:2px solid #4a6a9c">Total general</th>' +
    '</tr><tr style="background:#223f6e;color:#fff;font-size:14px;">';
  
  pivot.columns.forEach(function(col){
    if(col.type === 'week'){
      var isExpanded = window.expandedWeeks[col.rawWeek];
      var icon = isExpanded ? '−' : '+';
      headHTML += '<th style="font-size:15px;text-align:right;min-width:68px;padding:8px;font-weight:bold;cursor:pointer" '+
                  'onclick="toggleWeekColumn(\''+col.rawWeek+'\')">'+col.label+' <span style="font-size:10px;color:#a0b8e0;vertical-align:middle">['+icon+']</span></th>';
    } else {
      headHTML += '<th style="font-size:15px;text-align:right;min-width:68px;padding:8px;font-weight:normal;background:#142d4c;color:#c0d4ea;border-left:1px solid #2a4a7c">'+col.label+'</th>';
    }
  });
  headHTML += '</tr>';
  document.getElementById('tResumenHead').innerHTML = headHTML;

  /* ── TOGGLE ──────────────────────────────────────────────────────── */
  window.toggleResumenLevel = function(id){
    var rows = document.querySelectorAll('[data-group="'+id+'"]');
    if(!rows.length) return;
    var isHidden = rows[0].style.display === 'none';
    rows.forEach(function(r){ r.style.display = isHidden ? '' : 'none'; });
    var btn = document.querySelector('[data-togbtn="'+id+'"]');
    if(btn) btn.innerText = isHidden ? '−' : '+';
  };

  window.toggleWeekColumn = function(s){
    window.expandedWeeks = window.expandedWeeks || {};
    window.expandedWeeks[s] = !window.expandedWeeks[s];
    renderResumen();
  };

  /* ── HELPERS ─────────────────────────────────────────────────────── */
  var TOG_STYLE =
    'display:inline-block;width:13px;height:13px;line-height:11px;text-align:center;' +
    'border:1px solid #888;margin-right:5px;font-size:11px;font-weight:bold;' +
    'color:#333;background:#fff;cursor:pointer;vertical-align:middle;flex-shrink:0;';

  function cellVal(def, colData, isDaily){
    var d = colData || {};
    if(def.key === 'sell_thru'){
      if(isDaily) return '';
      var n = d.ventas||0, e = d.embarque||0;
      return e > 0 ? ((n/e)*100).toFixed(1)+'%' : '';
    }
    var v = d[def.key]||0;
    return v ? (def.type === 'money' ? fmtMoney(v) : fmt(v)) : '';
  }

  function rowTotalVal(def, byCol){
    var totV=0, totE=0, totX=0;
    pivot.columns.forEach(function(col){
      if(col.type === 'week'){ // Only sum weeks to avoid double counting days!
        var d = byCol[col.key]||{};
        totV += d.ventas||0;
        totE += d.embarque||0;
        totX += d[def.key]||0;
      }
    });
    if(def.key === 'sell_thru') return totE > 0 ? ((totV/totE)*100).toFixed(1)+'%' : '';
    return totX ? (def.type === 'money' ? fmtMoney(totX) : fmt(totX)) : '';
  }

  /* ── Función para actualizar totales de Capture (modo normal) ── */
  window.updateCaptureNormalTotal = function(r1i, r2i, rowi, ncols){
    var sum = 0;
    for(var ci=0; ci<ncols; ci++){
      var inp = document.getElementById('cin_'+r1i+'_'+r2i+'_'+rowi+'_'+ci);
      if(inp) sum += parseFloat(inp.value)||0;
    }
    var tot = document.getElementById('ctn_'+r1i+'_'+r2i+'_'+rowi);
    if(tot) tot.textContent = sum ? fmt(sum) : '0';
    /* Actualizar totales de columna y gran total */
    window.updateCaptureNormalGrand(r1i, r2i, ncols);
  };

  /* ── Actualiza Total Capture (modo normal): columnas + gran total ── */
  window.updateCaptureNormalGrand = function(r1i, r2i, ncols){
    var grand = 0;
    for(var ci=0; ci<ncols; ci++){
      var colSum = 0;
      for(var ri=0; ; ri++){
        var inp = document.getElementById('cin_'+r1i+'_'+r2i+'_'+ri+'_'+ci);
        if(!inp) break;
        colSum += parseFloat(inp.value)||0;
      }
      var colEl = document.getElementById('ctn_col_'+r1i+'_'+r2i+'_'+ci);
      if(colEl) colEl.textContent = colSum ? fmt(colSum) : '';
      grand += colSum;
    }
    var grandEl = document.getElementById('ctn_grand_'+r1i+'_'+r2i);
    if(grandEl) grandEl.textContent = grand ? fmt(grand) : '0';
    if(typeof _updateProgramadoDisplay === 'function') _updateProgramadoDisplay(r1i, r2i, ncols);
  };

  /* ── Agregar fila Capture adicional (modo normal) ── */
  window._captureMinRowsNorm = 1;

  window.removeCaptureNormalRow = function(r1i, r2i, rowi, ncols){
    var minRows = window._captureMinRowsNorm || 1;
    if(rowi < minRows) return;
    var totEl = document.getElementById('ctn_'+r1i+'_'+r2i+'_'+rowi);
    if(!totEl) return;
    var tr = totEl.closest('tr');
    if(tr) tr.remove();
    window.decreaseResumenRowspans(r1i, r2i);
    window.updateCaptureNormalGrand(r1i, r2i, ncols);
    window.updateRemoveCaptureBtn();
  };

  window.removeLastCaptureNormalRow = function(r1i, r2i, ncols){
    var minRows = window._captureMinRowsNorm || 1;
    var lastRi = -1, ri = 0;
    while(document.getElementById('ctn_'+r1i+'_'+r2i+'_'+ri)){ lastRi = ri; ri++; }
    if(lastRi < minRows) return;
    window.removeCaptureNormalRow(r1i, r2i, lastRi, ncols);
  };

  window.revealCaptureNormalRow = function(r1i, r2i, rowi){
    var totEl = document.getElementById('ctn_'+r1i+'_'+r2i+'_'+rowi);
    if(!totEl) return;
    var tr = totEl.closest('tr');
    if(!tr || tr.style.display !== 'none') return;
    tr.style.display = '';
    window.bumpResumenRowspans(r1i, r2i);
  };

  window.hideCaptureNormalRow = function(r1i, r2i, rowi, ncols){
    var totEl = document.getElementById('ctn_'+r1i+'_'+r2i+'_'+rowi);
    if(!totEl) return;
    var tr = totEl.closest('tr');
    if(!tr || tr.style.display === 'none') return;
    /* Limpiar inputs antes de ocultar */
    var semEl = document.getElementById('csem_'+r1i+'_'+r2i+'_'+rowi);
    if(semEl) semEl.value = '';
    for(var ci=0; ci<(ncols||0); ci++){
      var inp = document.getElementById('cin_'+r1i+'_'+r2i+'_'+rowi+'_'+ci);
      if(inp) inp.value = '';
    }
    totEl.textContent = '0';
    tr.style.display = 'none';
    window.decreaseResumenRowspans(r1i, r2i);
    window.updateCaptureNormalGrand(r1i, r2i, ncols);
  };

  window.addCaptureNormalRow = function(r1i, r2i, ncols, bgEven, bgOdd){
    var rowi = 0;
    var lastCaptureRow = null;
    while(document.getElementById('ctn_'+r1i+'_'+r2i+'_'+rowi)){
      var totEl = document.getElementById('ctn_'+r1i+'_'+r2i+'_'+rowi);
      if(totEl) lastCaptureRow = totEl.closest('tr');
      rowi++;
    }
    if(!lastCaptureRow) return;
    var bg = (r2i % 2 === 0 ? (bgEven||'#e3f2fd') : (bgOdd||'#d6ebf9'));
    var tr = document.createElement('tr');
    tr.setAttribute('data-group', lastCaptureRow.getAttribute('data-group')||'');
    tr.style.cssText = 'background:'+bg+';border-top:1px dashed #64a8d8';
    /* Values cell con sem input */
    var td1 = document.createElement('td');
    td1.setAttribute('data-group', tr.getAttribute('data-group'));
    td1.style.cssText = 'background:'+bg+';padding:3px 6px;white-space:nowrap;font-size:14px;font-weight:600;color:#1565c0;border-right:1px solid #e0e6f0;text-align:left;vertical-align:middle';
    td1.innerHTML = '<span style="font-size:12px;color:#1565c0">Capture SEM</span><div style="font-size:12px;font-weight:bold;color:#1a3a5c;margin-top:2px;background:#ddeeff;display:inline-block;border-radius:3px;padding:1px 6px"><input type="text" id="csem_'+r1i+'_'+r2i+'_'+rowi+'" placeholder="SEM" value="'+(window._nextSemanaProyeccion||'')+'" oninput="window.syncCaptureProjFromBlock('+r1i+','+r2i+')" style="width:50px;border:2px solid #2980b9;border-radius:4px;padding:2px 4px;font-size:13px;text-align:center;background:#eef6fc;color:#0a2a4a;font-weight:bold;outline:none"></div><button type="button" title="Eliminar fila" onclick="window.removeCaptureNormalRow('+r1i+','+r2i+','+rowi+','+ncols+')" style="margin-left:3px;font-size:11px;font-weight:bold;background:#c62828;color:#fff;border:none;border-radius:3px;padding:1px 5px;cursor:pointer;vertical-align:middle">×</button>';
    tr.appendChild(td1);
    for(var ci=0; ci<ncols; ci++){
      var td = document.createElement('td');
      td.setAttribute('data-group', tr.getAttribute('data-group'));
      td.style.cssText = 'background:'+bg+';padding:2px 4px;text-align:right;vertical-align:middle;width:68px;min-width:68px;max-width:68px';
      td.innerHTML = '<input type="number" id="cin_'+r1i+'_'+r2i+'_'+rowi+'_'+ci+'" min="0" placeholder="0" oninput="window.updateCaptureNormalTotal('+r1i+','+r2i+','+rowi+','+ncols+');window.syncCaptureProjFromBlock('+r1i+','+r2i+')" style="width:72px;border:2px solid #2980b9;border-radius:6px;padding:4px 6px;font-size:14px;font-weight:bold;text-align:right;background:#eef6fc;color:#0a2a4a;outline:none;box-shadow:inset 0 1px 3px rgba(0,0,0,0.1)">';
      tr.appendChild(td);
    }
    var tdTot = document.createElement('td');
    tdTot.id = 'ctn_'+r1i+'_'+r2i+'_'+rowi;
    tdTot.setAttribute('data-group', tr.getAttribute('data-group'));
    tdTot.style.cssText = 'font-size:14px;font-weight:bold;color:#0d47a1;background:'+bg+';text-align:right;width:82px;min-width:82px;padding:3px 8px;border-left:2px solid #4a6a9c';
    tdTot.textContent = '0';
    tr.appendChild(tdTot);
    window.bumpResumenRowspans(r1i, r2i);
    if(lastCaptureRow.nextSibling){
      lastCaptureRow.parentNode.insertBefore(tr, lastCaptureRow.nextSibling);
    } else {
      lastCaptureRow.parentNode.appendChild(tr);
    }
    window.updateRemoveCaptureBtn();
  };

  window.addCaptureRowAll = function(){
    var blocks = window._resumenCaptureBlocks || [];
    var minRows = window._captureMinRowsNorm || 1;
    var visibleInitial = window._captureVisibleInitial || 0;
    _ensureNextSemanaProyeccion();
    if(visibleInitial < minRows){
      window._captureVisibleInitial = visibleInitial + 1;
      blocks.forEach(function(b){
        var semEl = document.getElementById('csem_'+b.i1+'_'+b.r2i+'_'+visibleInitial);
        if(semEl) semEl.value = window._nextSemanaProyeccion || '';
        window.revealCaptureNormalRow(b.i1, b.r2i, visibleInitial);
      });
      /* Incrementar semana para la siguiente */
      if(window._nextSemanaProyeccion){
        window._nextSemanaProyeccion = parseInt(window._nextSemanaProyeccion, 10) + 1;
      }
      window.updateRemoveCaptureBtn();
      saveCaptureProjectionsFromDom();
      _propagateCaptureToAllStores(_captureProjModeKey(), window._captureVisibleInitial, window._nextSemanaProyeccion);
      return;
    }
    /* Agregar nueva fila con semana incrementada */
    blocks.forEach(function(b){
      window.addCaptureNormalRow(b.i1, b.r2i, b.ncols, b.bgEven, b.bgOdd);
      /* Incrementar el número de semana en la nueva fila */
      var ri = 0;
      while(document.getElementById('csem_'+b.i1+'_'+b.r2i+'_'+ri)){
        ri++;
      }
      var lastRi = ri - 1;
      var semInput = document.getElementById('csem_'+b.i1+'_'+b.r2i+'_'+lastRi);
      if(semInput){
        var currentVal = parseInt(semInput.value, 10);
        if(!isNaN(currentVal)){
          semInput.value = currentVal + 1;
        }
      }
    });
    saveCaptureProjectionsFromDom();
    _propagateCaptureToAllStores(_captureProjModeKey(), window._captureVisibleInitial, window._nextSemanaProyeccion);
  };

  window.removeCaptureRowAll = function(){
    if(!confirm('¿Estás seguro que deseas eliminar?')) return;
    var blocks = window._resumenCaptureBlocks || [];
    var minRows = window._captureMinRowsNorm || 1;
    var modeKeyN = _captureProjModeKey();
    var hasDynamic = false;
    blocks.forEach(function(b){
      var ri = 0;
      while(document.getElementById('ctn_'+b.i1+'_'+b.r2i+'_'+ri)) ri++;
      if(ri > minRows) hasDynamic = true;
    });
    if(hasDynamic){
      /* Semana objetivo a quitar = la última agregada con +SEM (nextSemanaProyeccion - 1).
         Así solo se borra esa semana puntual en cada tienda/producto, en vez de
         quitar a ciegas "la última fila" (lo cual borraba semanas distintas,
         incluso ya guardadas, en tiendas que no tenían la semana nueva capturada). */
      var targetSemRemoveAllN = window._nextSemanaProyeccion
        ? String(parseInt(window._nextSemanaProyeccion, 10) - 1)
        : null;
      if(window._captureProjections && window._captureProjections[modeKeyN]){
        var storeN = window._captureProjections[modeKeyN];
        /* Borrar SOLO la fila de la semana objetivo en cada key (visibles y no visibles) */
        Object.keys(storeN).forEach(function(sk){
          var entry = storeN[sk];
          if(entry && entry.rows && entry.rows.length > 0){
            var visRows = entry.rows.filter(function(r){
              return r.sem || (r.values||[]).some(function(v){ return v !== '' && parseFloat(v) > 0; });
            });
            var idxRemoveN = -1;
            if(targetSemRemoveAllN !== null){
              for(var iRN = visRows.length - 1; iRN >= 0; iRN--){
                if(String(visRows[iRN].sem || '').trim() === targetSemRemoveAllN){ idxRemoveN = iRN; break; }
              }
            } else if(visRows.length > 0){
              idxRemoveN = visRows.length - 1;
            }
            if(idxRemoveN >= 0) visRows.splice(idxRemoveN, 1);
            storeN[sk] = { visibleInitial: window._captureVisibleInitial||0, rows: visRows };
          }
        });
      }
      blocks.forEach(function(b){
        window.removeLastCaptureNormalRow(b.i1, b.r2i, b.ncols);
      });
      return;
    }
    var visibleInitial = window._captureVisibleInitial || 0;
    if(visibleInitial > 0){
      var hideRi = visibleInitial - 1;
      window._captureVisibleInitial = hideRi;
      blocks.forEach(function(b){
        window.hideCaptureNormalRow(b.i1, b.r2i, hideRi, b.ncols);
      });
      window.updateRemoveCaptureBtn();
      saveCaptureProjectionsFromDom();
      _propagateCaptureToAllStores(modeKeyN, window._captureVisibleInitial, window._nextSemanaProyeccion);
    }
  };

  /* ── BUILD ROWS ──────────────────────────────────────────────────── */
  var bodyRows = [];
  window._resumenCaptureBlocks = [];

  pivot.primaryKeys.forEach(function(r1Key, i1){
    var r1Data = pivot.rowsData[r1Key];
    var r2Keys = Object.keys(r1Data.subRows).sort();
    /* normal mode: defs rows (capture oculta al inicio) */
    var baseR2RowCountN = defs.length;
    /* Precomputar filas programado por r2 */
    var r2ProgRowsMapN = {};
    var totalProgRowsN = 0;
    r2Keys.forEach(function(r2k){
      var pr = _getSavedProgramadoRows(r1Key, r2k);
      r2ProgRowsMapN[r2k] = pr;
      totalProgRowsN += pr.length;
    });
    var r1Span = r2Keys.length * baseR2RowCountN + totalProgRowsN;
    var id1    = 'grp-'+i1;

    r2Keys.forEach(function(r2Key, r2i){
      var r2Data = r1Data.subRows[r2Key];
      var isFirstR2 = (r2i === 0);
      var borderTop = isFirstR2 ? '2px solid #b0bcd8' : '1px solid #e4e8f0';
      var r2Bg = r2i % 2 === 0 ? '#ffffff' : '#f4f6fa';
      var savedProgRowsN = r2ProgRowsMapN[r2Key] || [];
      var hasProgRowsN = savedProgRowsN.length > 0;
      var r2RowCountNorm = baseR2RowCountN + savedProgRowsN.length;

      /* ── Filas Programado (modo normal) ── */
      savedProgRowsN.forEach(function(progRow, progIdx){
        var isFirstProg = (progIdx === 0);
        var progBg = r2i % 2 === 0 ? '#fff8e1' : '#fff3c3';
        var pRow = '<tr style="background:'+progBg+';border-top:'+(isFirstProg ? borderTop : '1px solid #edf0f6')+'">';

        if(isFirstR2 && isFirstProg){
          pRow +=
            '<td rowspan="'+r1Span+'" data-r1-rowspan="'+i1+'" style="font-size:15px;font-weight:bold;vertical-align:top;'+
            'white-space:nowrap;padding:5px 8px;border-right:1px solid #c8d2e8;'+
            'border-top:2px solid #b0bcd8;background:#eef2fa">'+
            '<span data-togbtn="'+id1+'" style="'+TOG_STYLE+'" onclick="toggleResumenLevel(\''+id1+'\')">−</span>'+
            r1Key+'</td>';
        }

        if(isFirstProg){
          pRow +=
            '<td rowspan="'+r2RowCountNorm+'" data-r2-rowspan="'+i1+'_'+r2i+'" data-group="'+id1+'" style="vertical-align:top;white-space:nowrap;padding:4px 6px;'+
            'border-top:'+borderTop+';border-right:1px solid #d0d8ea;font-size:14px;color:#2980B9;font-weight:600;background:'+r2Bg+'">'+
            r2Key+'</td>';
        }

        pRow += '<td data-group="'+id1+'" style="position:static;background:'+progBg+';padding:3px 6px;white-space:nowrap;'+
          'font-size:15px;font-weight:600;color:#e65100;border-right:1px solid #e0e6f0;text-align:left">Programado '+
          '<span style="font-size:13px;font-weight:bold;color:#333;background:#ffe082;border-radius:3px;padding:1px 5px">'+(progRow.sem || '')+'</span>'+
          '<button type="button" onclick="_deleteProgramadoRow(\''+r1Key+'\',\''+r2Key+'\','+progIdx+',\'norm\',\''+(progRow.sem || '')+'\')" style="margin-left:6px;font-size:11px;background:transparent;border:none;cursor:pointer;color:#c62828">🗑️</button>'+
          '</td>';

        pivot.columns.forEach(function(col, ci){
          var v = progRow.values[ci] !== undefined ? progRow.values[ci] : '';
          var fv = v !== '' ? fmt(parseFloat(v)||0) : '';
          pRow += '<td data-group="'+id1+'" style="font-size:14px;color:#e65100;background:'+progBg+';text-align:right;vertical-align:middle;width:68px;min-width:68px;max-width:68px;padding:3px 8px">'+
            (fv||'') +
            '<button type="button" onclick="_editProgramadoRow(\''+r1Key+'\',\''+r2Key+'\','+progIdx+','+ci+',\'norm\')" style="margin-left:2px;font-size:10px;background:transparent;border:none;cursor:pointer;color:#1565c0">✏️</button>' +
            '</td>';
        });

        var progRowTotal = progRow.values.reduce(function(s, x){ return s + (parseFloat(x)||0); }, 0);
        pRow += '<td data-group="'+id1+'" style="font-size:14px;font-weight:bold;color:#bf360c;background:'+progBg+';text-align:right;padding:3px 8px;width:82px;min-width:82px;border-left:2px solid #e0a030">'+
          fmt(progRowTotal)+'</td>';

        pRow += '</tr>';
        bodyRows.push(pRow);
      });

      defs.forEach(function(def, di){
        var rowBg = r2i % 2 === 0 ? (def.bgEven || '#ffffff') : (def.bgOdd || '#f4f6fa');
        var isFirstDef = (di === 0);
        var isVeryFirst = (isFirstR2 && di === 0);   // very first row of r1 block
        var rowHTML = '<tr style="background:'+rowBg+';border-top:'+(isFirstDef ? borderTop : '1px solid #edf0f6')+'">';

        /* Col 1: r1 label — spans ALL rows of this r1 block, only on very first row */
        if(!hasProgRowsN && isVeryFirst){
          rowHTML +=
            '<td rowspan="'+r1Span+'" data-r1-rowspan="'+i1+'" style="font-size:15px;font-weight:bold;vertical-align:top;'+
            'white-space:nowrap;padding:5px 8px;border-right:1px solid #c8d2e8;'+
            'border-top:2px solid #b0bcd8;background:#eef2fa">'+
            '<span data-togbtn="'+id1+'" style="'+TOG_STYLE+'" onclick="toggleResumenLevel(\''+id1+'\')">−</span>'+
            r1Key+
            '</td>';
        }

        /* Col 2: r2 label — spans defs rows, only on first def of each r2 */
        if(!hasProgRowsN && isFirstDef){
          rowHTML +=
            '<td rowspan="'+r2RowCountNorm+'" data-r2-rowspan="'+i1+'_'+r2i+'" data-group="'+id1+'" '+
            'style="vertical-align:top;white-space:nowrap;padding:4px 6px;'+
            'border-top:'+borderTop+';border-right:1px solid #d0d8ea;font-size:14px;color:#2980B9;font-weight:600;background:'+r2Bg+';">'+
            r2Key+'</td>';
        }

        /* Col 3: metric label */
        rowHTML +=
          '<td data-group="'+id1+'" style="background:'+rowBg+';padding:3px 6px;white-space:nowrap;'+
          'font-size:15px;font-weight:normal;color:#333;border-right:1px solid #e0e6f0;text-align:left">'+
          def.label+'</td>';

        /* Semana data columns */
        pivot.columns.forEach(function(col){
          var isDaily = col.type === 'day';
          var v = cellVal(def, (r2Data._total||{})[col.key], isDaily);
          var hasVal = v && v !== '0';
          var cId = 'norm_' + r1Key + '_' + r2Key + '_' + def.key + '_' + col.key;
          var cAttr = getCommentAttr(cId);
          rowHTML +=
            '<td data-group="'+id1+'" class="'+cAttr.cls+'"'+cAttr.str+' style="font-size:14px;position:static;background:'+rowBg+';text-align:right;padding:3px 8px;'+
            (hasVal ? 'color:#1a3a5c' : 'color:#bbb')+'">'+(v||'0')+'</td>';
        });

        /* Row total */
        var tot = rowTotalVal(def, r2Data._total||{});
        if(def.key === 'sell_thru' && !tot) tot = '0.0%';
        var cellBg = rowBg;
        var cellColor = (def.key === 'sell_thru' ? '#000' : '');
        if (def.key === 'sell_thru' && tot) {
          var valPct = parseFloat(tot);
          if (valPct <= 79) cellBg = '#ffcdd2';
          else if (valPct < 90) cellBg = '#fff9c4';
          else cellBg = '#c8e6c9';
        }
        var colorStyle = cellColor ? 'color:'+cellColor+';' : '';
        var cIdTot = 'norm_' + r1Key + '_' + r2Key + '_' + def.key + '_TOTAL';
        var cAttrTot = getCommentAttr(cIdTot);
        rowHTML +=
          '<td data-group="'+id1+'" class="'+cAttrTot.cls+'"'+cAttrTot.str+' style="font-size:14px;'+colorStyle+'position:static;background:'+cellBg+';text-align:right;padding:3px 8px;'+
          'font-weight:bold;border-left:2px solid #c0cce0">'+(tot||'0')+'</td>';

        rowHTML += '</tr>';
        bodyRows.push(rowHTML);

        /* ── Fila Capture (modo normal) ── */
        if(di === defs.length - 1){
          var ncapBg = r2i % 2 === 0 ? '#e3f2fd' : '#d6ebf9';
          var nCols = pivot.columns.length;
          /* Primera fila de capture: Capture SEM + input de semana + inputs por columna + total */
          var capRow2 = '<tr data-group="'+id1+'" data-capture-initial="1" style="display:none;background:'+ncapBg+';border-top:1px dashed #64a8d8">';
          capRow2 += '<td data-group="'+id1+'" style="background:'+ncapBg+';padding:3px 6px;white-space:nowrap;font-size:14px;font-weight:600;color:#1565c0;border-right:1px solid #e0e6f0;text-align:left;vertical-align:middle">';
          capRow2 += '<span style="font-size:12px;color:#1565c0">Capture SEM</span>';
          capRow2 += '<div style="font-size:12px;font-weight:bold;color:#1a3a5c;margin-top:3px;background:#ddeeff;display:inline-block;border-radius:3px;padding:1px 6px"><input type="text" id="csem_'+i1+'_'+r2i+'_0" placeholder="SEM" value="'+(window._nextSemanaProyeccion||'')+'" oninput="window.syncCaptureProjFromBlock('+i1+','+r2i+')" style="width:50px;border:2px solid #2980b9;border-radius:4px;padding:2px 4px;font-size:13px;text-align:center;background:#eef6fc;color:#0a2a4a;font-weight:bold;outline:none"></div>';
          capRow2 += '</td>';
          pivot.columns.forEach(function(col, ci){
            var inputId2 = 'cin_'+i1+'_'+r2i+'_0_'+ci;
            capRow2 += '<td data-group="'+id1+'" style="background:'+ncapBg+';padding:2px 4px;text-align:right;vertical-align:middle;width:68px;min-width:68px;max-width:68px"><input type="number" id="'+inputId2+'" min="0" placeholder="0" oninput="window.updateCaptureNormalTotal('+i1+','+r2i+',0,'+nCols+');window.syncCaptureProjFromBlock('+i1+','+r2i+')" style="width:72px;border:2px solid #2980b9;border-radius:6px;padding:4px 6px;font-size:14px;font-weight:bold;text-align:right;background:#eef6fc;color:#0a2a4a;outline:none;box-shadow:inset 0 1px 3px rgba(0,0,0,0.1)"></td>';
          });
          var nTotId = 'ctn_'+i1+'_'+r2i+'_0';
          capRow2 += '<td id="'+nTotId+'" data-group="'+id1+'" style="font-size:14px;font-weight:bold;color:#0d47a1;background:'+ncapBg+';text-align:right;padding:3px 8px;width:82px;min-width:82px;border-left:2px solid #4a6a9c">0</td>';
          capRow2 += '</tr>';
          bodyRows.push(capRow2);
          window._resumenCaptureBlocks.push({
            i1: i1,
            r2i: r2i,
            ncols: nCols,
            bgEven: ncapBg,
            bgOdd: ncapBg,
            minRows: 1,
            prod: pivotMode === 'producto' ? r1Key : r2Key,
            tienda: pivotMode === 'producto' ? r2Key : r1Key
          });
        }
      });
    });
  });

  /* ── GRAND TOTAL ROWS (one per metric) ───────────────────────────── */
  var totalRowsGroupNorm = [];
  defsAll.forEach(function(def, di){
    var rowHTML = '<tr class="pivot-total" style="background:#1a3a5c;color:#fff">';
    if(di === 0){
      rowHTML +=
        '<td colspan="2" rowspan="'+defsAll.length+'" '+
        'style="font-size:16px;vertical-align:middle;padding:5px 8px;font-weight:bold;border-right:1px solid #3a5a8c">'+
        'Total general</td>';
    }
    rowHTML += '<td style="position:static;background:#1a3a5c;padding:3px 8px;font-size:15px;font-weight:bold;">'+def.label+'</td>';
    pivot.columns.forEach(function(col){
      var isDaily = col.type === 'day';
      var v = cellVal(def, pivot.totalsByColumn[col.key], isDaily);
      rowHTML += '<td style="font-size:14px;position:static;background:#1a3a5c;color:#000;text-align:right;padding:3px 8px">'+(v||(def.key==='sell_thru'?'0.0%':'0'))+'</td>';
    });
    /* Grand total for this metric across all columns */
    var gV=0, gE=0, gX=0;
    pivot.columns.forEach(function(col){
      if(col.type === 'week'){
        var d = pivot.totalsByColumn[col.key]||{};
        gV += d.ventas||0; gE += d.embarque||0; gX += d[def.key]||0;
      }
    });
    var gt = def.key === 'sell_thru'
      ? (gE > 0 ? ((gV/gE)*100).toFixed(1)+'%' : '')
      : (gX ? (def.type === 'money' ? fmtMoney(gX) : fmt(gX)) : '0');
    var cellBg = '#1a3a5c';
    var cellColor = '#000';
    if (def.key === 'sell_thru' && gt) {
      var valPct = parseFloat(gt);
      if (valPct <= 79) cellBg = '#ffcdd2';
      else if (valPct < 90) cellBg = '#fff9c4';
      else cellBg = '#c8e6c9';
    }
    rowHTML += '<td style="font-size:14px;position:static;background:'+cellBg+';color:'+cellColor+';text-align:right;padding:3px 8px;font-weight:bold;border-left:2px solid #4a6a9c">'+gt+'</td>';
    rowHTML += '</tr>';
    totalRowsGroupNorm.push(rowHTML);
  });

  if (window._totalTop !== false) {
    bodyRows.unshift.apply(bodyRows, totalRowsGroupNorm);
  } else {
    bodyRows.push.apply(bodyRows, totalRowsGroupNorm);
  }

  document.getElementById('tResumenBody').innerHTML = bodyRows.join('');
  restoreCaptureProjections();
  var btnAddCapNorm = document.getElementById('btnAddCaptureSem');
  if(btnAddCapNorm) btnAddCapNorm.style.display = (window._resumenCaptureBlocks.length ? 'inline-block' : 'none');
  var btnRemCapNorm = document.getElementById('btnRemoveCaptureSem');
  if(btnRemCapNorm) btnRemCapNorm.style.display = (window._resumenCaptureBlocks.length ? 'inline-block' : 'none');
  var btnSaveCapNorm = document.getElementById('btnSaveCaptureAll');
  if(btnSaveCapNorm) btnSaveCapNorm.style.display = (window._resumenCaptureBlocks.length ? 'inline-flex' : 'none');
  window.updateRemoveCaptureBtn();
  
  /* ── Función para exportar datos capturados a Excel/CSV (modo normal) ── */
  window.exportCaptureData = function(){
    /* Modo normal: las sub-columnas de cada fila capturada son las columnas del pivot (semanas/fechas) */
    saveCaptureProjectionsFromDom();
    var subLabels = pivot.columns.map(function(c){ return c.label || c.key; });
    _exportCaptureToCSV('norm', subLabels);
  };
  
  renderReabastoPanel();
  if(typeof resizeAfterChange === 'function') resizeAfterChange();
}



// ─── HELPERS GLOBALES COMPARATIVO ────────────────────────────────────────────
function semLabel(s){
  var yr = Math.floor(s/100), wk = s%100;
  return (yr >= 2000) ? String(yr).slice(-2)+String(wk).padStart(2,'0') : 'Sem '+String(s).padStart(2,'0');
}
function pctSpan(a, b){
  if(!b) return '—';
  var v = (a - b) / b * 100;
  var cls = v >= 0 ? 'color:#007700' : 'color:#c00';
  return '<span style="'+cls+'">'+(v>=0?'+':'')+v.toFixed(1)+'%</span>';
}
function sumMetrics(semSet, tiendaSet, prodSet){
  var r = {cfbc:0, wmx:0, unid:0, mermaU:0, mermaR:0, emb:0};
  tiendaSet.forEach(function(t){
    semSet.forEach(function(s){
      prodSet.forEach(function(p){
        var d = (DATA.raw_prod_semana&&DATA.raw_prod_semana[t]&&DATA.raw_prod_semana[t][String(s)]&&DATA.raw_prod_semana[t][String(s)][p])||{};
        r.cfbc+=d.venta_cfbc||0; r.wmx+=d.venta_wmx||0; r.unid+=d.ventas_u||0;
        r.mermaU+=d.merma_u||0; r.mermaR+=d.retail_vc||0; r.emb+=d.embarque_u||0;
      });
    });
  });
  return r;
}

// ── mini gráfica de barras compacta (W fijo, H pequeño) ─────────────────────
function buildBarChart(labels, vals1, vals2, lbl1, lbl2, W, H){
  W = W||340; H = H||160;
  var maxV = Math.max.apply(null, vals1.concat(vals2||[]).concat([1]));
  var n = labels.length;
  var slotW = (W-50)/n;
  var barW = Math.max(6, Math.min(20, slotW/2 - 3));
  var html = '<svg width="'+W+'" height="'+(H+52)+'" font-family="Arial" font-size="9" style="display:block">';
  html += '<line x1="42" y1="8" x2="42" y2="'+H+'" stroke="#ccc"/>';
  html += '<line x1="42" y1="'+H+'" x2="'+W+'" y2="'+H+'" stroke="#ccc"/>';
  for(var i=0;i<=3;i++){
    var yv = Math.round(maxV*i/3);
    var yp = H - Math.round((yv/maxV)*H*0.88) - 4;
    var ylab = yv>=1000000 ? (yv/1000000).toFixed(1)+'M' : yv>=1000 ? Math.round(yv/1000)+'k' : yv;
    html += '<text x="40" y="'+(yp+4)+'" text-anchor="end" fill="#999">'+ylab+'</text>';
    if(i>0) html += '<line x1="42" y1="'+yp+'" x2="'+W+'" y2="'+yp+'" stroke="#eee" stroke-dasharray="3"/>';
  }
  labels.forEach(function(lbl,i){
    var cx = 48 + i*slotW + slotW/2;
    var h1 = Math.max(1, Math.round((vals1[i]||0)/maxV*H*0.88));
    html += '<rect x="'+(cx - barW - 1)+'" y="'+(H-h1)+'" width="'+barW+'" height="'+h1+'" fill="#0071ce" rx="1"/>';
    if(vals2 && vals2[i]!==undefined){
      var h2 = Math.max(1, Math.round((vals2[i]||0)/maxV*H*0.88));
      html += '<rect x="'+(cx+1)+'" y="'+(H-h2)+'" width="'+barW+'" height="'+h2+'" fill="#ffc220" rx="1"/>';
    }
    var shortLbl = lbl.length>8 ? lbl.slice(0,8)+'…' : lbl;
    html += '<text x="'+cx+'" y="'+(H+12)+'" text-anchor="middle" fill="#555" transform="rotate(-30 '+cx+' '+(H+12)+')" font-size="8">'+shortLbl+'</text>';
  });
  html += '<rect x="48" y="10" width="9" height="7" fill="#0071ce" rx="1"/>';
  html += '<text x="60" y="17" fill="#444">'+lbl1+'</text>';
  if(vals2&&vals2.length){
    html += '<rect x="130" y="10" width="9" height="7" fill="#ffc220" rx="1"/>';
    html += '<text x="142" y="17" fill="#444">'+(lbl2||'')+'</text>';
  }
  html += '</svg>';
  return html;
}

// ─── RENDER COMPARATIVO (solo Semana vs Semana) ───────────────────────────────
function renderComparativo(){
  var sems = getSemanasActivas();
  var tiendas = (typeof DATA !== 'undefined' && DATA.tiendas) ? DATA.tiendas : [];
  var prods = getProductosActivos();
  var semsAct = sems.length ? sems : DATA.semanas.slice(-4);

  var headHTML = '<tr>'
    +'<th style="text-align:left;min-width:160px">Nivel</th>'
    +'<th>Unidades</th><th>Venta CFBC</th><th>Venta WMX</th>'
    +'<th>Merma U</th><th>Merma $</th><th>Embarque</th>'
    +'<th>% Merma</th><th>vs Ant.</th></tr>';

  var bodyRows = [];
  var prevCFBC = null;

  // Detectar cuántas tiendas en total están abiertas (productos expandidos)
  // ocultarTiendas = true solo cuando todas las semanas abiertas tienen el MISMO numero de tiendas (1-1, 2-2, 3-3...)
  var ocultarTiendas = false;
  if(state.openSemanas && state.openSemanas.length > 0 && state.openTiendas){
    var counts = state.openSemanas.map(function(sem){
      return (state.openTiendas[sem] && state.openTiendas[sem].length) || 0;
    });
    var allSame = counts.every(function(c){ return c === counts[0]; });
    ocultarTiendas = allSame && counts[0] >= 1;
  }

  semsAct.forEach(function(s){
    var m = sumMetrics([s], tiendas, prods);
    var pm = m.emb>0?(m.mermaU/m.emb*100).toFixed(1)+'%':'\u2014';
    var vs = prevCFBC!==null ? pctSpan(m.cfbc,prevCFBC) : '\u2014';
    var semOpen = (state.openSemanas && state.openSemanas.indexOf(s) >= 0);

    bodyRows.push('<tr class="pivot-row-sem" style="cursor:pointer;background:'+(semOpen?'#cce5ff':'#f5f7fa')+';font-weight:700;" onclick="drillSemana('+s+')">'
      +'<td style="padding-left:6px;color:#0071ce;">'
      +'<span style="font-size:.8rem;margin-right:4px;">'+(semOpen?'\u25bc':'\u25b6')+'</span>'
      +semLabel(s)+'</td>'
      +'<td>'+fmt(m.unid)+'</td>'
      +'<td style="font-weight:700">$'+fmt(m.cfbc)+'</td>'
      +'<td>$'+fmt(m.wmx)+'</td>'
      +'<td class="'+(m.mermaU>0?'red':'')+'">'+fmt(m.mermaU)+'</td>'
      +'<td class="'+(m.mermaR>0?'red':'')+'">$'+fmt(m.mermaR)+'</td>'
      +'<td>'+fmt(m.emb)+'</td>'
      +'<td class="'+(parseFloat(pm)>10?'red':'')+'">'+pm+'</td>'
      +'<td>'+vs+'</td></tr>');

    prevCFBC = m.cfbc;

    if(semOpen){
      var tiendaItems = tiendas.map(function(t){
        var mt = sumMetrics([s],[t],prods);
        return {t:t,m:mt};
      }).filter(function(x){return x.m.cfbc||x.m.unid||x.m.mermaU;})
        .sort(function(a,b){return b.m.cfbc-a.m.cfbc;});

      var totT = tiendaItems.reduce(function(acc,x){
        return {cfbc:acc.cfbc+x.m.cfbc,wmx:acc.wmx+x.m.wmx,unid:acc.unid+x.m.unid,
                mermaU:acc.mermaU+x.m.mermaU,mermaR:acc.mermaR+x.m.mermaR,emb:acc.emb+x.m.emb};
      },{cfbc:0,wmx:0,unid:0,mermaU:0,mermaR:0,emb:0});

      // Solo renderizar tiendas abiertas cuando ocultarTiendas=true
      var tiendaItemsRender = ocultarTiendas
        ? tiendaItems.filter(function(x){
            return state.openTiendas && state.openTiendas[s] && state.openTiendas[s].indexOf(x.t) >= 0;
          })
        : tiendaItems;

      tiendaItemsRender.forEach(function(x){
        var mt=x.m;
        var pm2=mt.emb>0?(mt.mermaU/mt.emb*100).toFixed(1)+'%':'\u2014';
        var share=totT.unid>0?(mt.unid/totT.unid*100).toFixed(1):'0.0';
        var tOpen=(state.openTiendas && state.openTiendas[s] && state.openTiendas[s].indexOf(x.t) >= 0);

        // Solo mostrar fila de tienda si NO estamos ocultando tiendas
        if(!ocultarTiendas){
          bodyRows.push('<tr class="pivot-row-tienda" style="cursor:pointer;background:'+(tOpen?'#e8f4fd':'#fff')+';border-left:3px solid #0071ce;" onclick="drillTienda('+s+',\''+(x.t).replace(/\'/g,"\\\'")+'\')">'
            +'<td style="padding-left:22px;color:#0071ce;">'
            +'<span style="font-size:.75rem;margin-right:4px;">'+(tOpen?'\u25bc':'\u25b6')+'</span>'
            +x.t.replace('SC ','')
            +'<span style="margin-left:5px;font-size:.65rem;color:#888;">'+share+'%</span>'
            +'</td>'
            +'<td>'+fmt(mt.unid)+'</td>'
            +'<td style="font-weight:600">$'+fmt(mt.cfbc)+'</td>'
            +'<td>$'+fmt(mt.wmx)+'</td>'
            +'<td class="'+(mt.mermaU>0?'red':'')+'">'+fmt(mt.mermaU)+'</td>'
            +'<td class="'+(mt.mermaR>0?'red':'')+'">$'+fmt(mt.mermaR)+'</td>'
            +'<td>'+fmt(mt.emb)+'</td>'
            +'<td class="'+(parseFloat(pm2)>10?'red':'')+'">'+pm2+'</td>'
            +'<td></td></tr>');
        }

        if(tOpen){
          var prodItems = prods.map(function(p){
            var d=(DATA.raw_prod_semana&&DATA.raw_prod_semana[x.t]&&DATA.raw_prod_semana[x.t][String(s)]&&DATA.raw_prod_semana[x.t][String(s)][p])||{};
            return {p:p,cfbc:d.venta_cfbc||0,wmx:d.venta_wmx||0,unid:d.ventas_u||0,emb:d.embarque_u||0,mermaU:d.merma_u||0,mermaR:d.retail_vc||0};
          }).filter(function(o){return o.cfbc||o.unid||o.mermaU;})
            .sort(function(a,b){return b.cfbc-a.cfbc;});

          var totP=prodItems.reduce(function(acc,o){
            return {cfbc:acc.cfbc+o.cfbc,wmx:acc.wmx+o.wmx,unid:acc.unid+o.unid,emb:acc.emb+o.emb,mermaU:acc.mermaU+o.mermaU,mermaR:acc.mermaR+o.mermaR};
          },{cfbc:0,wmx:0,unid:0,emb:0,mermaU:0,mermaR:0});

          // Cuando se ocultan tiendas: mostrar encabezado de tienda clickeable (para cerrar) arriba de sus productos
          if(ocultarTiendas){
            var pm2=mt.emb>0?(mt.mermaU/mt.emb*100).toFixed(1)+'%':'\u2014';
            var share=totT.unid>0?(mt.unid/totT.unid*100).toFixed(1):'0.0';
            bodyRows.push('<tr class="pivot-row-tienda" style="cursor:pointer;background:#e8f4fd;border-left:3px solid #0071ce;" onclick="drillTienda('+s+',\''+(x.t).replace(/\'/g,"\\\'")+'\')">'
              +'<td style="padding-left:22px;color:#0071ce;">'
              +'<span style="font-size:.75rem;margin-right:4px;">\u25bc</span>'
              +x.t.replace('SC ','')
              +'<span style="margin-left:5px;font-size:.65rem;color:#888;">'+share+'%</span>'
              +'<span style="margin-left:8px;font-size:.62rem;color:#888;font-style:italic;">clic para cerrar</span>'
              +'</td>'
              +'<td>'+fmt(mt.unid)+'</td>'
              +'<td style="font-weight:600">$'+fmt(mt.cfbc)+'</td>'
              +'<td>$'+fmt(mt.wmx)+'</td>'
              +'<td class="'+(mt.mermaU>0?'red':'')+'">'+fmt(mt.mermaU)+'</td>'
              +'<td class="'+(mt.mermaR>0?'red':'')+'">$'+fmt(mt.mermaR)+'</td>'
              +'<td>'+fmt(mt.emb)+'</td>'
              +'<td class="'+(parseFloat(pm2)>10?'red':'')+'">'+pm2+'</td>'
              +'<td></td></tr>');
          }

          prodItems.forEach(function(o){
            var sharep=totP.unid>0?(o.unid/totP.unid*100).toFixed(1):'0.0';
            bodyRows.push('<tr class="pivot-row-prod" style="background:#f0f8ff;border-left:6px solid #5bc0de;">'
              +'<td style="padding-left:40px;font-size:.68rem;color:#333;">'+o.p.replace('BQT ','')
              +'<span style="margin-left:4px;font-size:.62rem;color:#999;">'+sharep+'%</span></td>'
              +'<td style="font-size:.68rem">'+fmt(o.unid)+'</td>'
              +'<td style="font-weight:600;font-size:.68rem">$'+fmt(o.cfbc)+'</td>'
              +'<td style="font-size:.68rem">$'+fmt(o.wmx)+'</td>'
              +'<td class="'+(o.mermaU>0?'red':'')+'" style="font-size:.68rem">'+fmt(o.mermaU)+'</td>'
              +'<td class="'+(o.mermaR>0?'red':'')+'" style="font-size:.68rem">$'+fmt(o.mermaR)+'</td>'
              +'<td style="font-size:.68rem">'+fmt(o.emb)+'</td>'
              +'<td></td><td></td></tr>');
          });

          // Subtotal de tienda: solo cuando NO se ocultan tiendas
          if(!ocultarTiendas){
            bodyRows.push('<tr style="background:#daeaf5;border-left:6px solid #5bc0de;font-weight:700;font-size:.68rem;">'
              +'<td style="padding-left:40px;color:#0071ce;">Subtotal '+x.t.replace('SC ','')+'</td>'
              +'<td>'+fmt(totP.unid)+'</td><td>$'+fmt(totP.cfbc)+'</td><td>$'+fmt(totP.wmx)+'</td>'
              +'<td class="red">'+fmt(totP.mermaU)+'</td><td class="red">$'+fmt(totP.mermaR)+'</td>'
              +'<td>'+fmt(totP.emb)+'</td><td></td><td></td></tr>');
          }
        }
      });

      var totAll2=sumMetrics([s],tiendas,prods);
      bodyRows.push('<tr style="background:#b8d4f0;font-weight:700;border-left:3px solid #0071ce;">'
        +'<td style="padding-left:22px;color:#0071ce;">Subtotal Sem '+semLabel(s)+'</td>'
        +'<td>'+fmt(totAll2.unid)+'</td><td style="font-weight:700">$'+fmt(totAll2.cfbc)+'</td>'
        +'<td>$'+fmt(totAll2.wmx)+'</td>'
        +'<td class="red">'+fmt(totAll2.mermaU)+'</td><td class="red">$'+fmt(totAll2.mermaR)+'</td>'
        +'<td>'+fmt(totAll2.emb)+'</td><td></td><td></td></tr>');
    }
  });

  var totAll = sumMetrics(semsAct, tiendas, prods);
  bodyRows.push('<tr class="total"><td>TOTAL GENERAL</td>'
    +'<td>'+fmt(totAll.unid)+'</td><td>$'+fmt(totAll.cfbc)+'</td><td>$'+fmt(totAll.wmx)+'</td>'
    +'<td class="red">'+fmt(totAll.mermaU)+'</td><td class="red">$'+fmt(totAll.mermaR)+'</td>'
    +'<td>'+fmt(totAll.emb)+'</td><td></td><td></td></tr>');

  document.getElementById('tCompHead').innerHTML = headHTML;
  document.getElementById('tCompBody').innerHTML = bodyRows.join('');

  // ── Gráfica lateral: actualiza según nivel activo ──────────────────────────
  var chartEl   = document.getElementById('compChart');
  var chartTitle = document.getElementById('compChartTitle');
  var tiendas = (typeof DATA !== 'undefined' && DATA.tiendas) ? DATA.tiendas : [];
  var prods = getProductosActivos();

  // Obtener la primera tienda abierta (si hay alguna)
  var firstOpenSem = null;
  var firstOpenTienda = null;
  if(state.openTiendas){
    for(var sem in state.openTiendas){
      if(state.openTiendas[sem] && state.openTiendas[sem].length > 0){
        firstOpenSem = parseInt(sem);
        firstOpenTienda = state.openTiendas[sem][0];
        break;
      }
    }
  }

  if(firstOpenTienda && firstOpenSem){
    // Nivel productos: gráfica de productos de la primera tienda abierta
    var s2=firstOpenSem; var t2=firstOpenTienda;
    var pi=prods.map(function(p){
      var d=(DATA.raw_prod_semana&&DATA.raw_prod_semana[t2]&&DATA.raw_prod_semana[t2][String(s2)]&&DATA.raw_prod_semana[t2][String(s2)][p])||{};
      return {p:p,cfbc:d.venta_cfbc||0,mermaR:d.retail_vc||0};
    }).filter(function(o){return o.cfbc;}).sort(function(a,b){return b.cfbc-a.cfbc;});
    chartTitle.textContent='CFBC por Producto — '+t2.replace('SC ','');
    chartEl.innerHTML=buildBarChart(pi.map(function(o){return o.p.replace('BQT ','');}),pi.map(function(o){return o.cfbc;}),pi.map(function(o){return o.mermaR;}),'Venta CFBC','Merma $',320,150);

  } else if(state.openSemanas && state.openSemanas.length > 0){
    // Nivel tiendas: gráfica de tiendas de la primera semana abierta
    var s3=state.openSemanas[0];
    var ti=tiendas.map(function(t){var mt=sumMetrics([s3],[t],prods);return {t:t,cfbc:mt.cfbc,mermaR:mt.mermaR};})
      .filter(function(x){return x.cfbc;}).sort(function(a,b){return b.cfbc-a.cfbc;});
    chartTitle.textContent='CFBC por Tienda — Sem '+semLabel(s3);
    chartEl.innerHTML=buildBarChart(ti.map(function(x){return x.t.replace('SC ','');}),ti.map(function(x){return x.cfbc;}),ti.map(function(x){return x.mermaR;}),'Venta CFBC','Merma $',320,150);

  } else {
    // Nivel semanas: gráfica de todas las semanas activas
    var chartLabels=[],chartCFBC=[],chartMerma=[];
    semsAct.forEach(function(s){var m=sumMetrics([s],tiendas,prods);chartLabels.push(semLabel(s));chartCFBC.push(m.cfbc);chartMerma.push(m.mermaR);});
    chartTitle.textContent='Venta CFBC por Semana';
    chartEl.innerHTML=buildBarChart(chartLabels,chartCFBC,chartMerma,'Venta CFBC','Merma $',320,160);
  }
}

function drillSemana(s){
  // Cambiar a array para permitir múltiples semanas abiertas
  if(!state.openSemanas) state.openSemanas = [];
  var idx = state.openSemanas.indexOf(s);
  if(idx >= 0){
    state.openSemanas.splice(idx, 1); // Cerrar esta semana
  } else {
    state.openSemanas.push(s); // Abrir esta semana
  }
  renderComparativo();
}

function drillTienda(s,t){
  if(!state.openTiendas) state.openTiendas = {};
  if(!state.openTiendas[s]) state.openTiendas[s] = [];

  var idx = state.openTiendas[s].indexOf(t);
  if(idx >= 0){
    state.openTiendas[s].splice(idx, 1);
    if(state.openTiendas[s].length === 0) delete state.openTiendas[s];
  } else {
    state.openTiendas[s].push(t);
  }

  renderComparativo();
}

function cerrarDrill(){ 
  state.openSemanas = []; 
  state.openTiendas = {}; 
  renderComparativo(); 
}
function _renderDrillTiendas(s){ /* integrado en renderComparativo */ }
function _renderDrillProductos(s,t){ /* integrado en renderComparativo */ }

function setCompMode(m){ renderComparativo(); }


function selInvTienda(t){
  // Si ya está seleccionada, deseleccionar
  if(state.invMode === 'tienda' && state.invSelected === t){
    state.invMode = null;
    state.invSelected = null;
  } else {
    state.invMode = 'tienda';
    state.invSelected = t;
  }
  renderInventario();
}

function selInvProducto(p){
  // Si ya está seleccionado, deseleccionar
  if(state.invMode === 'producto' && state.invSelected === p){
    state.invMode = null;
    state.invSelected = null;
  } else {
    state.invMode = 'producto';
    state.invSelected = p;
  }
  renderInventario();
}

function limpiarInvFiltro(){
  state.invMode = null;
  state.invSelected = null;
  renderInventario();
}

// ─── IMPRIMIR ───────────────────────────────────────────────────────────────
// Construye un HTML completo en memoria y lo abre en una pestaña nueva.
// onafterprint cierra la pestaña para que no quede about:blank.
// No hay footer con fecha — la fecha solo está en el encabezado.
// ────────────────────────────────────────────────────────────────────────────
function recargarDatos(){
  try{
    var url = window.parent.location.href;
    if(url.indexOf('reload=1') === -1){
      window.parent.location.href = url + (url.indexOf('?') === -1 ? '?' : '&') + 'reload=1';
    } else {
      window.parent.location.reload(true);
    }
  }catch(e){
    window.location.reload(true);
  }
}

function imprimirReporte(){
  // ── Mapeo de cada pestaña (state.view) a su contenedor visible en el DOM ──
  var viewMap = {
    producto:     { id: 'viewProducto',     titulo: 'Producto' },
    tienda:       { id: 'viewTienda',       titulo: 'Por Tienda' },
    comparativo:  { id: 'viewComparativo',  titulo: 'Comparativo' },
    inventario:   { id: 'viewInventario',   titulo: 'Inventario' },
    gasto:        { id: 'viewGasto',        titulo: 'Gasto' },
    resumen:      { id: 'viewResumen',      titulo: 'Resumen Ejecutivo' }
  };
  var actual = viewMap[state.view] || viewMap.producto;
  var contEl = document.getElementById(actual.id);
  if(!contEl){ return; }

  // Si estamos en "resumen", aseguramos guardar las proyecciones capturadas antes de imprimir
  if(state.view === 'resumen' && typeof saveCaptureProjectionsFromDom === 'function'){
    saveCaptureProjectionsFromDom();
  }

  var semana  = document.getElementById('hdrSem').textContent;
  var fecha   = document.getElementById('hdrFecha').textContent;
  var tienda  = document.getElementById('tiendaDropLabel') ? document.getElementById('tiendaDropLabel').textContent : (state.tienda || '');

  // Clonamos el contenedor de la pestaña activa tal cual está en pantalla
  var clone = contEl.cloneNode(true);
  clone.style.display = 'block';
  clone.removeAttribute('style');
  clone.classList.add('print-area');

  // Quitamos del clon cualquier control interactivo que no tenga sentido imprimir
  Array.prototype.forEach.call(clone.querySelectorAll('button, select, input, .no-print'), function(el){
    el.remove();
  });

  // Reutilizamos TODAS las hojas de estilo ya cargadas en el dashboard,
  // así cada pestaña conserva exactamente su propio diseño sin tener que
  // reconstruir el CSS a mano para cada vista.
  var stylesHTML = Array.prototype.map.call(document.querySelectorAll('style'), function(s){
    return s.outerHTML;
  }).join('');

  var printCSS = '<style>'
    + '@media print{ @page{margin:10mm} body{padding:0 !important;} }'
    + 'html,body{background:#fff !important;}'
    + 'body{font-family:Arial,sans-serif;padding:16px;}'
    + '.print-area{height:auto !important;overflow:visible !important;}'
    + '.print-hdr{display:flex;align-items:center;justify-content:space-between;'
    +   'padding-bottom:8px;border-bottom:2px solid #0071ce;margin-bottom:14px;}'
    + '.print-hdr .logo{display:flex;align-items:center;gap:8px}'
    + '.print-hdr .wm-text{font-size:1.3rem;font-weight:700;color:#0071ce}'
    + '.print-hdr .tag{font-size:.85rem;color:#666;font-weight:600;margin-left:6px}'
    + '.print-hdr .info{text-align:right;font-size:.74rem;color:#333;line-height:1.7}'
    + '</style>';

  var html = '<!DOCTYPE html><html lang="es"><head>'
    + '<meta charset="UTF-8">'
    + '<title>Walmart CFBC \u00b7 Sem '+semana+' \u00b7 '+tienda+'</title>'
    + stylesHTML + printCSS
    + '</head><body>'
    + '<div class="print-hdr">'
    +   '<div class="logo">'
    +     '<span class="wm-text">Walmart</span>'
    +     '<span class="tag">'+actual.titulo+'</span>'
    +   '</div>'
    +   '<div class="info">'
    +     '<div>'+fecha+'</div>'
    +     '<div>Semana &nbsp;<strong>'+semana+'</strong>'+(tienda ? ' &middot; '+tienda : '')+'</div>'
    +   '</div>'
    + '</div>'
    + clone.outerHTML
    + '<script>'
    + 'window.onload=function(){'
    +   'window.onafterprint=function(){window.close();};'
    +   'setTimeout(function(){window.print();},300);'
    + '};'
    + '<\/script>'
    + '</body></html>';

  // Usar Blob + URL para evitar about:blank en la pesta\u00f1a
  var blob = new Blob([html], {type:'text/html;charset=utf-8'});
  var url  = URL.createObjectURL(blob);
  var win  = window.open(url, '_blank');
  // Liberar URL de objeto cuando la ventana cargue
  if(win){ win.addEventListener('load', function(){ URL.revokeObjectURL(url); }); }
}

window.addEventListener('load', init);

(function fixParent(){
  try {
    var p = window.parent.document;
    var style = p.createElement('style');
    style.textContent = [
      '.main .block-container{padding:0!important;margin:0!important}',
      '.main{padding:0!important}',
      '[data-testid="stAppViewContainer"]{padding:0!important}',
      '[data-testid="stVerticalBlock"]{gap:0!important}',
      'header,[data-testid="stToolbar"],[data-testid="stDecoration"]{display:none!important}',
      'iframe{margin:0!important}',
      'section[data-testid="stMain"]{padding:0!important}',
      '.stMainBlockContainer{padding:0!important}',
      '[data-testid="manage-app-button"]{display:none!important}',
      '.stDeployButton{display:none!important}',
      '#MainMenu{display:none!important}',
      'button[kind="header"]{display:none!important}',
      '.viewerBadge_container__r5tak{display:none!important}',
      '.styles_viewerBadge__CvC9N{display:none!important}',
      'a[href="https://streamlit.io"]{display:none!important}',
      '#stDecoration{display:none!important}',
      'footer{display:none!important}',
      '[data-testid="stBottom"]{display:none!important}',
    ].join('');
    p.head.appendChild(style);
  } catch(e){}
  try {
    var frames = window.parent.document.querySelectorAll('iframe');
    frames.forEach(function(f){
      f.style.height = window.parent.innerHeight + 'px';
      f.style.width  = '100%';
    });
  } catch(e){}
  var _patchResize = function(fn){
    return function(){ fn.apply(this,arguments); setTimeout(resizeIframe,100); };
  };
  window.addEventListener('load', function(){
    if(typeof render!=='undefined') render = _patchResize(render);
    if(typeof renderTienda!=='undefined') renderTienda = _patchResize(renderTienda);
  });
})();

// ── Auto-resize: ajusta el iframe de Streamlit al contenido real ──
function resizeIframe(){
  // El iframe tiene scroll interno (scrolling=True), no necesitamos expandirlo
  // Solo aseguramos que tenga al menos la altura de la ventana
  var h = Math.max(window.innerHeight, 800);
  try {
    var frames = window.parent.document.querySelectorAll('iframe');
    frames.forEach(function(f){ f.style.height = h + 'px'; });
  } catch(e){}
}
// Disparar resize varias veces tras cambios de DOM para capturar expansiones
function resizeAfterChange(){
  resizeIframe();
  setTimeout(resizeIframe, 100);
  setTimeout(resizeIframe, 300);
  setTimeout(resizeIframe, 600);
}
var _resizeScheduled = false;
var _resizeObs = new ResizeObserver(function(){
  if(_resizeScheduled) return;
  _resizeScheduled = true;
  requestAnimationFrame(function(){
    _resizeScheduled = false;
    resizeAfterChange();
  });
});
_resizeObs.observe(document.documentElement);
_resizeObs.observe(document.body);
window.addEventListener('load', function(){ resizeAfterChange(); });
// Parchear toggleDrillT para que también dispare resize
var _origToggleDrillT = toggleDrillT;
toggleDrillT = function(id){ _origToggleDrillT(id); resizeAfterChange(); };
// ── Asistente IA ──
window.aiChatHistory = window.aiChatHistory || [];

function getContextFromDATA() {
  // Lee directamente del objeto DATA (la fuente de verdad), no del DOM
  // Esto evita 100% los problemas de rowspan/colspan del HTML
  try {
    var pivotMode = state.resumenPivot || 'producto';
    var sems      = getSemanasActivas().map(function(s){ return String(s); });
    var tiendas   = getTiendasActivas();
    var prods     = getProductosActivos();
    var resumenSrc = DATA.resumen_diario || {};
    var DIAS_ABREV = ['Dom','Lun','Mar','Mié','Jue','Vie','Sáb'];

    // Acumular por (r1, r2, fecha/semana) → {embarque, ventas, merma}
    // r1 = producto o tienda (según pivotMode)
    var agg = {};     // agg[r1][r2][fecha] = {emb, ven, mer}
    var allFechas = {}; // fecha -> dia_semana_label

    tiendas.forEach(function(t) {
      var tData = resumenSrc[t] || {};
      sems.forEach(function(s) {
        var sData = tData[s] || {};
        Object.keys(sData).forEach(function(p) {
          if (prods.indexOf(p) === -1) return;
          var r1 = pivotMode === 'producto' ? p : t;
          var r2 = pivotMode === 'producto' ? t : p;
          if (!agg[r1]) agg[r1] = {};
          if (!agg[r1][r2]) agg[r1][r2] = {};
          Object.keys(sData[p] || {}).forEach(function(fecha) {
            if (!agg[r1][r2][fecha]) agg[r1][r2][fecha] = {emb:0, ven:0, mer:0};
            var d = sData[p][fecha] || {};
            agg[r1][r2][fecha].emb += d.embarque || 0;
            agg[r1][r2][fecha].ven += d.ventas   || 0;
            agg[r1][r2][fecha].mer += d.merma    || 0;
            // Guardar etiqueta del día
            if (!allFechas[fecha]) {
              var dt = parseResumenDate(fecha);
              allFechas[fecha] = dt ? DIAS_ABREV[dt.getDay()] + ' ' + dt.getDate() + '/' + (dt.getMonth()+1) : fecha;
            }
          });
        });
      });
    });

    var fechasOrdenadas = Object.keys(allFechas).sort();
    if (!fechasOrdenadas.length) {
      // Si no hay días expandidos, mostrar por semana
      return getContextBySemana(pivotMode, sems, tiendas, prods, resumenSrc);
    }

    // Construir tabla plana: r1 | r2 | fecha | Embarque | Ventas | Merma
    var lines = [];
    var view = state.view;
    var filtros = 'Vista:' + view +
      ' | Semanas:' + sems.join('+') +
      ' | Tiendas:' + tiendas.join(', ') +
      ' | Productos:' + prods.join(', ');
    lines.push('FILTROS: ' + filtros);
    lines.push('');

    var col1 = pivotMode === 'producto' ? 'Producto' : 'Tienda';
    var col2 = pivotMode === 'producto' ? 'Tienda'   : 'Producto';
    lines.push(col1 + ' | ' + col2 + ' | Fecha | Dia | Embarque | Ventas_POS | Merma');
    lines.push('---');

    // Totales por fecha para el resumen rápido
    var totFecha = {};
    Object.keys(agg).forEach(function(r1) {
      Object.keys(agg[r1]).forEach(function(r2) {
        fechasOrdenadas.forEach(function(fecha) {
          var d = agg[r1][r2][fecha];
          if (!d) return;
          if (!totFecha[fecha]) totFecha[fecha] = {emb:0, ven:0, mer:0};
          totFecha[fecha].emb += d.emb;
          totFecha[fecha].ven += d.ven;
          totFecha[fecha].mer += d.mer;
          // Solo agregar fila si hay algún valor
          if (d.emb || d.ven || d.mer) {
            lines.push(r1 + ' | ' + r2 + ' | ' + fecha + ' | ' + allFechas[fecha] +
              ' | ' + (d.emb||0) + ' | ' + (d.ven||0) + ' | ' + (d.mer||0));
          }
        });
      });
    });

    lines.push('');
    lines.push('TOTALES POR DIA:');
    fechasOrdenadas.forEach(function(f) {
      var t2 = totFecha[f] || {emb:0,ven:0,mer:0};
      lines.push(allFechas[f] + ' (' + f + '): Embarque=' + t2.emb + ' Ventas=' + t2.ven + ' Merma=' + t2.mer);
    });

    return lines.join('\n');

  } catch(e) {
    return 'Error extrayendo datos: ' + e.message;
  }
}

function getContextBySemana(pivotMode, sems, tiendas, prods, resumenSrc) {
  // Vista por semana (cuando los días no están expandidos en el dashboard)
  var DIAS_ABREV = ['Dom','Lun','Mar','Mié','Jue','Vie','Sáb'];
  var agg = {}; // agg[r1][r2][sem] = {emb, ven, mer}
  var totSem = {};

  tiendas.forEach(function(t) {
    var tData = resumenSrc[t] || {};
    sems.forEach(function(s) {
      var sData = tData[s] || {};
      Object.keys(sData).forEach(function(p) {
        if (prods.indexOf(p) === -1) return;
        var r1 = pivotMode === 'producto' ? p : t;
        var r2 = pivotMode === 'producto' ? t : p;
        if (!agg[r1]) agg[r1] = {};
        if (!agg[r1][r2]) agg[r1][r2] = {};
        if (!agg[r1][r2][s]) agg[r1][r2][s] = {emb:0, ven:0, mer:0};
        if (!totSem[s]) totSem[s] = {emb:0, ven:0, mer:0};
        Object.keys(sData[p] || {}).forEach(function(fecha) {
          var d = sData[p][fecha] || {};
          agg[r1][r2][s].emb += d.embarque || 0;
          agg[r1][r2][s].ven += d.ventas   || 0;
          agg[r1][r2][s].mer += d.merma    || 0;
          totSem[s].emb += d.embarque || 0;
          totSem[s].ven += d.ventas   || 0;
          totSem[s].mer += d.merma    || 0;
        });
      });
    });
  });

  var col1 = pivotMode === 'producto' ? 'Producto' : 'Tienda';
  var col2 = pivotMode === 'producto' ? 'Tienda'   : 'Producto';
  var lines = [];
  lines.push('FILTROS: Vista:' + state.view + ' | Semanas:' + sems.join('+') +
    ' | Tiendas:' + tiendas.join(', ') + ' | Productos:' + prods.join(', '));
  lines.push('');
  lines.push(col1 + ' | ' + col2 + ' | Semana | Embarque | Ventas_POS | Merma');
  lines.push('---');
  Object.keys(agg).forEach(function(r1) {
    Object.keys(agg[r1]).forEach(function(r2) {
      sems.forEach(function(s) {
        var d = agg[r1][r2][s] || {};
        if (d.emb || d.ven || d.mer)
          lines.push(r1 + ' | ' + r2 + ' | Sem ' + s + ' | ' + (d.emb||0) + ' | ' + (d.ven||0) + ' | ' + (d.mer||0));
      });
    });
  });
  lines.push('');
  lines.push('TOTALES POR SEMANA:');
  sems.forEach(function(s) {
    var t2 = totSem[s] || {};
    lines.push('Sem ' + s + ': Embarque=' + (t2.emb||0) + ' Ventas=' + (t2.ven||0) + ' Merma=' + (t2.mer||0));
  });
  return lines.join('\n');
}

function getDynamicContextJSON() {
  return getContextFromDATA();
}

function addMessageToChat(text, sender, isTyping, isHtml) {
  var container = document.getElementById('aiChatMessages');
  if(!container) return;
  var div = document.createElement('div');
  div.style.display = 'flex';
  div.style.flexDirection = 'column';
  div.style.gap = '4px';
  div.style.maxWidth = '85%';
  if(isTyping) div.id = 'aiTypingIndicator';
  
  var bubble = document.createElement('div');
  bubble.style.padding = '8px 12px';
  bubble.style.fontSize = '0.8rem';
  
  if(sender === 'user') {
    div.style.alignSelf = 'flex-end';
    bubble.style.background = '#1565c0';
    bubble.style.color = '#fff';
    bubble.style.borderRadius = '12px 12px 2px 12px';
    bubble.textContent = text;
  } else if (sender === 'ai') {
    div.style.alignSelf = 'flex-start';
    bubble.style.background = '#e2e8f0';
    bubble.style.color = '#334155';
    bubble.style.borderRadius = '12px 12px 12px 2px';
    bubble.style.overflowX = 'auto';
    if (isHtml) {
        bubble.innerHTML = text;
    } else {
        bubble.innerHTML = text.replace(/\*\*(.*?)\*\*/g, '<strong>$1</strong>').replace(/\n/g, '<br>');
    }
  } else {
    div.style.alignSelf = 'center';
    bubble.style.background = 'transparent';
    bubble.style.color = '#94a3b8';
    bubble.style.fontSize = '0.7rem';
    bubble.style.fontStyle = 'italic';
    bubble.textContent = text;
  }
  
  div.appendChild(bubble);
  container.appendChild(div);
  container.scrollTop = container.scrollHeight;
}

function removeTypingIndicator() {
  var el = document.getElementById('aiTypingIndicator');
  if(el) el.remove();
}

function sendChatMessage() {
  var input = document.getElementById('aiChatInput');
  if(!input) return;
  var msg = input.value.trim();
  if(!msg) return;
  input.value = '';
  addMessageToChat(msg, 'user');
  
  var DEFAULT_GROQ_KEY = "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";
  var apiKey = localStorage.getItem('GROQ_API_KEY') || DEFAULT_GROQ_KEY;
  if(!apiKey) {
    apiKey = prompt("Por favor ingresa tu API Key de Groq para usar el asistente:");
    if(apiKey) localStorage.setItem('GROQ_API_KEY', apiKey);
    else {
      addMessageToChat('Para usar el asistente necesitas configurar una API Key.', 'system');
      return;
    }
  }
  
  var contexto = getDynamicContextJSON();
  var systemInst =
    "Eres un analista experto de datos de Walmart México (Walmex).\n" +
    "Se te da la tabla visible en pantalla en formato texto (líneas HDR=encabezado, DAT=fila de dato).\n" +
    "REGLA ABSOLUTA: SOLO menciona valores que aparezcan explícitamente en las líneas DAT. " +
    "Si una celda está vacía ('') o es cero, ese día/tienda NO tiene datos — NO lo menciones como ganador.\n\n" +
    "MÉTRICAS: Cnt POS=ventas, Cntd Embarque=embarques, Cant VC=merma, Venta CFBC=costo facturado, SEM=semana.\n\n" +
    "RESPUESTA: español, **negritas** en cifras clave, cálculos breves (ej: 10+8=**18**).\n" +
    "Si el dato no está visible, dílo explícitamente y sugiere ajustar filtros.\n\n" +
    "DATOS ACTUALES:\n" + contexto;
  
  addMessageToChat('Analizando...', 'system', true);
  
  var systemMsg = { role: "system", content: systemInst };
  var newUserMsg = { role: "user", content: msg };
  var chatHistory = window.aiChatHistory || [];
  
  // Limitar el historial de chat a los últimos 10 mensajes (5 preguntas/respuestas) — Llama 3.3 70B tiene 131k context window
  if (chatHistory.length > 10) {
      chatHistory = chatHistory.slice(chatHistory.length - 10);
  }
  
  var payloadMsgs = [systemMsg].concat(chatHistory).concat([newUserMsg]);
  
  fetch('https://api.groq.com/openai/v1/chat/completions', {
    method: 'POST',
    headers: { 
      'Content-Type': 'application/json',
      'Authorization': 'Bearer ' + apiKey
    },
    body: JSON.stringify({
      model: "llama-3.3-70b-versatile",
      messages: payloadMsgs,
      temperature: 0.1,
      top_p: 0.9,
      max_tokens: 1500
    })
  })
  .then(function(r) { return r.json(); })
  .then(function(data) {
    removeTypingIndicator();
    if(data.error) {
       addMessageToChat('Error: ' + data.error.message, 'system');
       if(data.error.code === 'invalid_api_key' || data.error.type === 'invalid_request_error') {
           localStorage.removeItem('GROQ_API_KEY');
       }
    } else {
       var respuesta = data.choices[0].message.content;
       
       // Procesar tags <think> generados por modelos de razonamiento (DeepSeek)
       respuesta = respuesta.replace(/<think>/gi, '<details style="margin-bottom:10px; padding:10px; background:#f8fafc; border:1px solid #e2e8f0; border-radius:8px; border-left: 4px solid #3b82f6;"><summary style="cursor:pointer; font-weight:600; color:#475569; font-size:0.85rem; outline:none;">🧠 Proceso de Razonamiento (Click para ver)</summary><div style="margin-top:10px; color:#64748b; font-size:0.8rem; font-style:italic;">');
       respuesta = respuesta.replace(/<\/think>/gi, '</div></details>\n');

       var isHtml = typeof marked !== 'undefined';
       var htmlRespuesta = isHtml ? marked.parse(respuesta) : respuesta.replace(/\*\*(.*?)\*\*/g, '<strong>$1</strong>').replace(/\n/g, '<br>');
       addMessageToChat(htmlRespuesta, 'ai', false, isHtml);
       
       window.aiChatHistory = chatHistory;
       window.aiChatHistory.push(newUserMsg);
       window.aiChatHistory.push({ role: "assistant", content: respuesta });
    }
  })
  .catch(function(err) {
    removeTypingIndicator();
    addMessageToChat('Ocurrió un error de red.', 'system');
  });
}
</script>

<!-- ImageViewer Modal -->
<div id="imgViewerOverlay" onclick="closeImageViewer()" style="display:none; position:fixed; inset:0; background:rgba(0,0,0,0.8); z-index:999998; backdrop-filter:blur(3px);"></div>
<div id="imgViewerModal" style="display:none; position:fixed; top:50%; left:50%; transform:translate(-50%, -50%); z-index:999999; max-width:90vw; max-height:90vh; text-align:center;">
    <button onclick="closeImageViewer()" style="position:absolute; top:-40px; right:0; background:transparent; border:none; color:white; font-size:30px; cursor:pointer;">&times;</button>
    <img id="imgViewerImg" src="" style="max-width:100%; max-height:90vh; border-radius:8px; box-shadow:0 10px 40px rgba(0,0,0,0.5);">
</div>
<script>
function openImageViewer(src) {
    document.getElementById('imgViewerImg').src = src;
    document.getElementById('imgViewerOverlay').style.display = 'block';
    document.getElementById('imgViewerModal').style.display = 'block';
}
function closeImageViewer() {
    document.getElementById('imgViewerOverlay').style.display = 'none';
    document.getElementById('imgViewerModal').style.display = 'none';
}
</script>
</body>
</html>"""

def build_html():
    data_json = base64.b64encode(
        json.dumps(DATA, ensure_ascii=True, default=str).encode('utf-8')
    ).decode('ascii')
    
    supabase_data = get_supabase_data()
    supabase_json = base64.b64encode(
        json.dumps(supabase_data, ensure_ascii=True, default=str).encode('utf-8')
    ).decode('ascii')

    devoluciones_data = get_devoluciones_data()
    devoluciones_json = base64.b64encode(
        json.dumps(devoluciones_data, ensure_ascii=True, default=str).encode('utf-8')
    ).decode('ascii')
    
    html_out = HTML.replace('__DATA_JSON__', data_json)
    html_out = html_out.replace('__SUPABASE_JSON__', supabase_json)
    html_out = html_out.replace('__DEVOLUCIONES_JSON__', devoluciones_json)
    
    # Inyectar API Key de Groq por defecto si existe en secrets
    groq_key = ""
    try:
        if "groq" in st.secrets and "api_key" in st.secrets["groq"]:
            groq_key = st.secrets["groq"]["api_key"]
    except Exception:
        pass

    html_out = html_out.replace('__GROQ_API_KEY__', groq_key)

    # ── Inyectar config de Supabase para el navegador ──
    # IMPORTANTE: solo la URL y la "publishable_key" (clave pública, equivalente
    # al antiguo "anon key") van al HTML que ve el navegador. El service_role_key
    # (sb_secret_...) NUNCA debe exponerse del lado del cliente porque otorga
    # acceso total de administrador a toda la base de datos — por eso no se usa
    # aquí; solo está disponible server-side por si se necesita en el futuro.
    sb_url = ""
    sb_publishable_key = ""
    try:
        sb_cfg = dict(st.secrets["supabase"])
        sb_url = sb_cfg.get("url", "")
        sb_publishable_key = sb_cfg.get("publishable_key", "")
    except Exception:
        pass

    html_out = html_out.replace('__SUPABASE_URL__', sb_url)
    html_out = html_out.replace('__SUPABASE_PUBLISHABLE_KEY__', sb_publishable_key)

    return html_out

# Verificar si se pidió recarga vía query params
params = st.query_params
if params.get("reload") == ["1"]:
    if _CACHE_LATEST.exists():
        _CACHE_LATEST.unlink()
    _sharepoint_file_meta.clear()
    cargar_datos.clear()
    st.query_params.clear()
    st.rerun()

# ── Panel de subida: solo ícono, sin ningún cuadro alrededor ────────────────
st.markdown("""
<style>
/* ── Ocultar UI de Streamlit y desactivar el scroll principal (Google scroll) ── */
header[data-testid="stHeader"] { display: none !important; }
[data-testid="stToolbar"] { display: none !important; }
[data-testid="stDecoration"] { display: none !important; }
footer { display: none !important; }

.stApp, section[data-testid="stMain"], .stMainBlockContainer {
    overflow: hidden !important;
    padding: 0 !important;
    margin: 0 !important;
    max-width: 100% !important;
}

iframe {
    height: 100vh !important;
    max-height: 100vh !important;
    border: none !important;
    margin: 0 !important;
}

/* ── Evitar que wrappers de Streamlit rompan el fixed ── */
div[data-testid="stVerticalBlock"] > div:has(div[data-testid="stExpander"]),
section[data-testid="stMain"] > div:has(div[data-testid="stExpander"]) {
    position: static !important;
    overflow: visible !important;
}
/* ── Posición fija ── */
div[data-testid="stExpander"] {
    position: fixed !important;
    top: 6px !important;
    right: 0px !important;
    left: auto !important;
    z-index: 99999 !important;
    width: auto !important;
    min-width: 0 !important;
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    border-radius: 0 !important;
    padding: 0 !important;
    transform: none !important;
    margin: 0 !important;
}

/* ── Matar borde en TODOS los wrappers internos ── */
div[data-testid="stExpander"] *,
div[data-testid="stExpander"] details,
div[data-testid="stExpander"] details > summary,
div[data-testid="stExpander"] > div,
div[data-testid="stExpander"] > div > div,
div[data-testid="stExpander"] > div > details {
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    border-radius: 0 !important;
    outline: none !important;
}

/* ── El summary (el botón visible) ── */
div[data-testid="stExpander"] summary {
    display: inline-flex !important;
    align-items: center !important;
    padding: 0 !important;
    margin: 0 !important;
    font-size: 1.2rem !important;
    background: transparent !important;
    border: none !important;
    box-shadow: none !important;
    outline: none !important;
    border-radius: 0 !important;
    cursor: pointer !important;
    min-height: 0 !important;
    line-height: 1 !important;
    list-style: none !important;
    opacity: 0.7;
    transition: opacity .15s !important;
}
div[data-testid="stExpander"] summary:hover  { opacity: 1 !important; background: transparent !important; border: none !important; }
div[data-testid="stExpander"] summary:focus  { outline: none !important; box-shadow: none !important; background: transparent !important; border: none !important; }
div[data-testid="stExpander"] summary:active { background: transparent !important; border: none !important; }
div[data-testid="stExpander"] summary::-webkit-details-marker { display: none !important; }
div[data-testid="stExpander"] summary::marker { display: none !important; }

/* ── Ocultar el texto/emoji del label, dejar solo la flecha ── */
div[data-testid="stExpander"] summary p,
div[data-testid="stExpander"] summary span.st-emotion-cache-ztfqz8,
div[data-testid="stExpander"] summary [data-testid="stExpanderLabel"] {
    display: none !important;
    width: 0 !important;
    overflow: hidden !important;
}
/* Mostrar y estilizar la flecha nativa */
div[data-testid="stExpander"] summary svg,
div[data-testid="stExpander"] summary [data-testid="stExpanderToggleIcon"] {
    display: inline-block !important;
    width: 20px !important;
    height: 20px !important;
    opacity: 0.7;
}

/* ── Panel desplegable ejecutivo ── */
div[data-testid="stExpander"] [data-testid="stExpanderDetails"] {
    position: absolute !important;
    right: 0 !important;
    top: calc(100% + 6px) !important;
    min-width: 300px !important;
    padding: 0 !important;
    background: #ffffff !important;
    border: 1px solid #d0d7e2 !important;
    border-radius: 8px !important;
    box-shadow: 0 12px 40px rgba(0,20,60,.14), 0 2px 8px rgba(0,0,0,.06) !important;
    overflow: hidden !important;
    font-family: 'Segoe UI', sans-serif !important;
}

/* ── Cabecera oscura ejecutiva ── */
div[data-testid="stExpander"] [data-testid="stExpanderDetails"]::before {
    content: "☁  Sincronizar con SharePoint" !important;
    display: block !important;
    background: #0d1f3c !important;
    color: #e8edf5 !important;
    font-size: .73rem !important;
    font-weight: 600 !important;
    letter-spacing: .06em !important;
    text-transform: uppercase !important;
    padding: 11px 16px !important;
    font-family: 'Segoe UI', sans-serif !important;
    border-bottom: 2px solid #0071ce !important;
}

/* ── Contenido interno ── */
div[data-testid="stExpander"] [data-testid="stExpanderDetails"] > div {
    padding: 14px 16px 14px !important;
    background: #f8fafc !important;
}

/* ── Texto caption ── */
div[data-testid="stExpander"] [data-testid="stExpanderDetails"] p {
    font-size: .68rem !important;
    color: #6b7a90 !important;
    margin: 0 0 10px 0 !important;
    font-family: 'Segoe UI', sans-serif !important;
    letter-spacing: .02em !important;
}
div[data-testid="stExpander"] [data-testid="stExpanderDetails"] small {
    font-size: .65rem !important;
    color: #9aa5b4 !important;
}
div[data-testid="stExpander"] [data-testid="stExpanderDetails"] strong {
    font-size: .72rem !important;
    color: #0d1f3c !important;
    font-weight: 600 !important;
}

/* ── File uploader ejecutivo profesional ── */
.upload-panel {
    background: linear-gradient(145deg, #fafbfc 0%, #f0f3f7 100%) !important;
    border: 1px solid #d0d9e4 !important;
    border-radius: 10px !important;
    padding: 24px !important;
    margin: 8px 0 !important;
    box-shadow: 0 2px 8px rgba(13,31,60,0.08) !important;
}
.upload-header {
    display: flex !important;
    align-items: center !important;
    gap: 12px !important;
    margin-bottom: 16px !important;
    padding-bottom: 16px !important;
    border-bottom: 2px solid #0071ce !important;
}
.upload-icon {
    width: 44px !important;
    height: 44px !important;
    background: linear-gradient(135deg, #0071ce 0%, #004c8c 100%) !important;
    border-radius: 10px !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    color: white !important;
    font-size: 20px !important;
}
.upload-title {
    flex: 1 !important;
}
.upload-title h3 {
    font-size: 1rem !important;
    font-weight: 700 !important;
    color: #0d1f3c !important;
    margin: 0 0 4px 0 !important;
    letter-spacing: -0.2px !important;
}
.upload-title p {
    font-size: .75rem !important;
    color: #6b7a90 !important;
    margin: 0 !important;
}
.upload-requirements {
    background: #ffffff !important;
    border: 1px solid #e1e8ed !important;
    border-radius: 8px !important;
    padding: 16px !important;
    margin-bottom: 16px !important;
}
.upload-requirements h4 {
    font-size: .78rem !important;
    font-weight: 600 !important;
    color: #0d1f3c !important;
    margin: 0 0 12px 0 !important;
    display: flex !important;
    align-items: center !important;
    gap: 8px !important;
}
.upload-requirements h4 i {
    color: #0071ce !important;
}
.req-grid {
    display: grid !important;
    grid-template-columns: repeat(3, 1fr) !important;
    gap: 10px !important;
}
.req-item {
    background: #f7fafc !important;
    border: 1px solid #e8eef3 !important;
    border-radius: 6px !important;
    padding: 10px 12px !important;
    text-align: center !important;
}
.req-item .label {
    font-size: .65rem !important;
    color: #8896a4 !important;
    text-transform: uppercase !important;
    letter-spacing: .5px !important;
    margin-bottom: 4px !important;
}
.req-item .value {
    font-size: .85rem !important;
    font-weight: 700 !important;
    color: #0d1f3c !important;
}
.dropzone-area {
    background: #ffffff !important;
    border: 2px dashed #c2ccd8 !important;
    border-radius: 10px !important;
    padding: 32px 24px !important;
    text-align: center !important;
    transition: all 0.3s ease !important;
    margin-bottom: 16px !important;
}
.dropzone-area:hover {
    border-color: #0071ce !important;
    background: #f0f7ff !important;
}
.dropzone-icon {
    font-size: 40px !important;
    color: #0071ce !important;
    margin-bottom: 12px !important;
}
.dropzone-text {
    font-size: .85rem !important;
    color: #4a5568 !important;
    margin-bottom: 8px !important;
}
.dropzone-hint {
    font-size: .7rem !important;
    color: #8896a4 !important;
}
.file-selected {
    background: #f0fff4 !important;
    border: 1px solid #38a169 !important;
    border-radius: 8px !important;
    padding: 16px !important;
    display: flex !important;
    align-items: center !important;
    gap: 12px !important;
    margin-bottom: 16px !important;
}
.file-selected-icon {
    width: 40px !important;
    height: 40px !important;
    background: #38a169 !important;
    border-radius: 8px !important;
    display: flex !important;
    align-items: center !important;
    justify-content: center !important;
    color: white !important;
    font-size: 18px !important;
}
.file-info {
    flex: 1 !important;
}
.file-info .name {
    font-size: .85rem !important;
    font-weight: 600 !important;
    color: #0d1f3c !important;
    margin-bottom: 2px !important;
}
.file-info .size {
    font-size: .7rem !important;
    color: #6b7a90 !important;
}
/* ── Botón ejecutivo ── */
div[data-testid="stExpander"] .stButton button {
    background: #0d1f3c !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 5px !important;
    font-size: .70rem !important;
    font-weight: 600 !important;
    letter-spacing: .08em !important;
    text-transform: uppercase !important;
    padding: 9px 0 !important;
    width: 100% !important;
    cursor: pointer !important;
    transition: background .18s !important;
    margin-top: 10px !important;
    font-family: 'Segoe UI', sans-serif !important;
}
div[data-testid="stExpander"] .stButton button:hover {
    background: #0071ce !important;
}
.status-indicator {
    display: flex !important;
    align-items: center !important;
    gap: 8px !important;
    padding: 10px 14px !important;
    border-radius: 6px !important;
    font-size: .75rem !important;
    margin-top: 12px !important;
}
.status-indicator.success {
    background: #f0fff4 !important;
    color: #38a169 !important;
    border: 1px solid #9ae6b4 !important;
}
.status-indicator.error {
    background: #fff5f5 !important;
    color: #e53e3e !important;
    border: 1px solid #fc8181 !important;
}
.status-indicator.warning {
    background: #fffff0 !important;
    color: #d69e2e !important;
    border: 1px solid #f6e05e !important;
}
</style>
""", unsafe_allow_html=True)

# Panel de subida profesional
with st.expander("📤 **Actualizar Datos**", expanded=False):
    st.markdown("""
    <div class="upload-panel">
        <div class="upload-header">
            <div class="upload-icon">📊</div>
            <div class="upload-title">
                <h3>Carga de Datos Semanal</h3>
                <p>Sube el archivo Excel actualizado para refrescar el dashboard</p>
            </div>
        </div>

        
    </div>
    """, unsafe_allow_html=True)

    # File uploader con estilo mejorado
    archivo = st.file_uploader(
        "Arrastra tu archivo Excel aquí o haz clic para seleccionar",
        type=["xlsx","xlsm","xls"],
        key="up_sp_profesional",
        help="Formatos permitidos: .xlsx, .xlsm, .xls"
    )

    # Mostrar info del archivo seleccionado
    if archivo is not None:
        file_size = len(archivo.read()) / 1024  # KB
        archivo.seek(0)  # Reset pointer
        st.markdown(f"""
        <div class="file-selected">
            <div class="file-selected-icon">✓</div>
            <div class="file-info">
                <div class="name">{archivo.name}</div>
                <div class="size">{file_size:.1f} KB • Listo para subir</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        st.markdown("""
        <div class="dropzone-area">
            <div class="dropzone-icon">📁</div>
            <div class="dropzone-text">Arrastra tu archivo aquí</div>
            <div class="dropzone-hint">o haz clic para explorar archivos</div>
        </div>
        """, unsafe_allow_html=True)

    # Botón de subida (mantenido igual como pidió)
    if st.button("⬆ Subir a SharePoint", type="primary", use_container_width=True):
        if archivo is None:
            st.markdown("""
            <div class="status-indicator warning">
                <span>⚠️</span> Selecciona un archivo Excel primero
            </div>
            """, unsafe_allow_html=True)
        else:
            with st.spinner("🔄 Subiendo datos a SharePoint..."):
                try:
                    msg = _subir_excel_sharepoint(archivo.read())
                    # ── Limpiar caché para forzar recarga con datos nuevos ──
                    if _CACHE_LATEST.exists():
                        _CACHE_LATEST.unlink()
                    _sharepoint_file_meta.clear()
                    cargar_datos.clear()
                    st.markdown(f"""
                    <div class="status-indicator success">
                        <span>✓</span> {msg}
                    </div>
                    """, unsafe_allow_html=True)
                    time.sleep(1.5)
                    st.rerun()
                except Exception as e:
                    st.markdown(f"""
                    <div class="status-indicator error">
                        <span>✕</span> Error: {e}
                    </div>
                    """, unsafe_allow_html=True)

# ── Render del dashboard ──────────────────────────────────────────────────────
html_content = build_html()
components.html(html_content, height=1080, scrolling=True)
